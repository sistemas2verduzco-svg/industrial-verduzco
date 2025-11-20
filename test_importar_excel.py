#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Script de prueba para validar la importación desde CLAVES.xlsx
Simula lo que hará el endpoint /api/productos/importar-excel
"""

import openpyxl
import os

def test_importar_excel():
    """Prueba la lectura del archivo CLAVES.xlsx"""
    
    filepath = 'CLAVES.xlsx'
    
    if not os.path.exists(filepath):
        print(f"❌ Archivo {filepath} no encontrado")
        return False
    
    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        
        print(f"✓ Archivo cargado correctamente")
        print(f"✓ Sheet: {ws.title}")
        print(f"✓ Total de filas: {ws.max_row}")
        
        # Leer encabezados
        headers = [cell.value for cell in ws[1]]
        print(f"\nEncabezados:")
        for i, h in enumerate(headers, 1):
            print(f"  Columna {i} ({chr(64+i)}): {h}")
        
        # Verificar que existan las columnas necesarias
        print(f"\n📋 Verificando columnas necesarias:")
        
        # Columna C (índice 2) = Clave
        # Columna F (índice 5) = Producto
        
        if headers[2] == 'Clave':
            print(f"  ✓ Columna C (Clave): {headers[2]}")
        else:
            print(f"  ⚠️ Columna C esperaba 'Clave', encontró: {headers[2]}")
        
        if headers[5] == 'Producto':
            print(f"  ✓ Columna F (Producto): {headers[5]}")
        else:
            print(f"  ⚠️ Columna F esperaba 'Producto', encontró: {headers[5]}")
        
        # Leer primeros 5 productos
        print(f"\n📦 Primeros 5 productos a importar:")
        contador = 0
        for row_num, row in enumerate(ws.iter_rows(min_row=2, max_row=6, values_only=True), start=2):
            clave = row[2]  # Columna C
            descripcion = row[5]  # Columna F
            
            if clave and descripcion:
                print(f"  {contador+1}. Clave: '{clave}' -> Descripción: '{descripcion}'")
                contador += 1
        
        # Contar totales
        total_productos = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[2] and row[5]:  # Si tiene clave y descripción
                total_productos += 1
        
        print(f"\n📊 Resumen:")
        print(f"  Total de filas: {ws.max_row}")
        print(f"  Productos a importar (con clave y descripción): {total_productos}")
        
        return True
        
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        return False

if __name__ == '__main__':
    print("=" * 50)
    print("TEST: Importación desde CLAVES.xlsx")
    print("=" * 50)
    success = test_importar_excel()
    print("=" * 50)
    if success:
        print("✓ TEST EXITOSO")
    else:
        print("✗ TEST FALLIDO")
