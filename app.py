from flask import Flask, render_template, request, jsonify, session, redirect, url_for, send_file, send_from_directory, make_response, flash, abort
from models import db, Producto, Proveedor, ProductoProveedor, HistorialPreciosProveedor, Usuario, Ticket, ComentarioTicket, Role, Permission, QCReport, QCItem, QCProduccionRegistro, Máquina, ComponenteMáquina, HojaRutaEntrega, HojaRutaNueva, HojaRutaCargaPiezasHistorial, HojaRutaFlujoLogistica, EntregaRegistro, AlmacenRegistro, FacturacionRegistro, EstacionTrabajo, EstacionPlantilla, ProcesoCatalogo, ClaveProducto, ClaveProceso, EntregaParcial, HojaRutaImpresionParcial, ContpaqSyncRun, ContpaqPedido, ContpaqPedidoDetalle, ContpaqRemision, ContpaqRemisionDetalle, ContpaqNotaVenta, ContpaqSucursalIndice, ContpaqPrecioPublico, ContpaqExistenciaStock, ContpaqSupplierOT, ContpaqSupplierOTDetalle, HojaRutaEntregaOTAsignacion, MaquinariaPedido, MaquinariaContpaqPedido, MaquinariaContpaqPedidoDetalle, MaquinariaBOM, MaquinariaBOMComponente, MaquinariaBOMProceso, MaquinariaOrdenTrabajo, MaquinariaOrdenBOMItem, MaquinariaOrdenProceso, MaquinariaCalidadRegistro, MaquinariaSerie, MaquinariaAlmacenResguardo, OdooSyncRun, OdooPedidoVenta, OdooPedidoVentaLinea, OdooOrdenCompra, OdooOrdenCompraLinea, MaquinariaSolicitud, MaquinariaSolicitudItem, AlertaBuzonGeneral, Tecnico, LogVerificacion, EmpaqueCliente, EmpaquePedido, EmpaquePedidoItem, EmpaqueCaja, EmpaqueMovimiento, EmpaqueLineaProgreso, EmpaqueSeguimientoLog, AlmacenCajaSurtidoSesion, AlmacenCajaSurtidoCaja, AlmacenCajaSurtidoLecturaBascula, AlmacenCajaSurtidoItem
from auth import AuthManager
from email_manager import EmailManager
from odoo_client import OdooClient, OdooError
from odoo_sync import run_odoo_sync
import os
import json
from dotenv import load_dotenv
from sqlalchemy import text, func, inspect, extract, or_
from functools import wraps
import secrets
from werkzeug.utils import secure_filename
from datetime import datetime, timedelta
from datetime import timezone as dt_timezone
from time import time, sleep
import logging
import pandas as pd
from decimal import Decimal, ROUND_HALF_UP
from openpyxl import Workbook
from io import BytesIO
import uuid
import math
import re
import subprocess
import sys
import threading
import requests
import qrcode
from itsdangerous import URLSafeTimedSerializer, BadSignature, SignatureExpired
from types import SimpleNamespace
from zoneinfo import ZoneInfo
from urllib.parse import urlencode

# Configurar logging
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('catalogo_app.log'),
        logging.StreamHandler()  # También mostrar en consola
    ]
)
logger = logging.getLogger(__name__)

# Jobs para importación asíncrona de procesos/claves.
# Persistimos estado en disco para evitar problemas con múltiples workers de gunicorn.
PROCESOS_IMPORT_LOCK = threading.Lock()
PROCESOS_IMPORT_DIR = os.path.join('uploads', 'imports_jobs')
os.makedirs(PROCESOS_IMPORT_DIR, exist_ok=True)
TECNICOS_QR_DIR = os.path.join('uploads', 'tecnicos', 'qr')
os.makedirs(TECNICOS_QR_DIR, exist_ok=True)

MACHINE_SCHEDULE_ENABLED = os.getenv('MACHINE_SCHEDULE_ENABLED', '1').strip().lower() not in ('0', 'false', 'no')
MACHINE_SCHEDULE_POLL_SECONDS = max(5, int(os.getenv('MACHINE_SCHEDULE_POLL_SECONDS', '15') or '15'))
MACHINE_SCHEDULE_TIMEZONE = (os.getenv('MACHINE_SCHEDULE_TIMEZONE') or 'America/Mexico_City').strip()
MACHINE_SCHEDULE_STATE_FILE = os.path.join('uploads', 'machine_schedule_state.json')
MACHINE_SCHEDULE_ADVISORY_LOCK_ID = 94630211
_MACHINE_SCHEDULE_THREAD = None
_MACHINE_SCHEDULE_INIT = False
_MACHINE_SCHEDULE_STOP = threading.Event()
MYE_CATALOG_FILE = os.path.join('uploads', 'maquinaria_estaciones_catalogo.json')
_MYE_CATALOG_LOCK = threading.Lock()
MACHINE_QUEUE_LIMIT = max(1, int(os.getenv('MACHINE_QUEUE_LIMIT', '4') or '4'))
_MACHINE_QUEUE_LIMIT_BY_TYPE_RAW = (os.getenv('MACHINE_QUEUE_LIMIT_BY_TYPE') or '').strip()
MACHINE_QUEUE_LIMIT_BY_TYPE = {}
if _MACHINE_QUEUE_LIMIT_BY_TYPE_RAW:
    try:
        parsed_limits = json.loads(_MACHINE_QUEUE_LIMIT_BY_TYPE_RAW)
        if isinstance(parsed_limits, dict):
            for k, v in parsed_limits.items():
                try:
                    MACHINE_QUEUE_LIMIT_BY_TYPE[str(k).strip().lower()] = max(1, int(v))
                except Exception:
                    continue
    except Exception:
        MACHINE_QUEUE_LIMIT_BY_TYPE = {}

WHATSAPP_ALERTS_ENABLED = os.getenv('WHATSAPP_ALERTS_ENABLED', '0').strip().lower() in ('1', 'true', 'yes', 'on')
WHATSAPP_ALERT_TYPES = {
    x.strip().lower() for x in (os.getenv('WHATSAPP_ALERT_TYPES') or '').split(',') if x.strip()
}
TWILIO_ACCOUNT_SID = (os.getenv('TWILIO_ACCOUNT_SID') or '').strip()
TWILIO_AUTH_TOKEN = (os.getenv('TWILIO_AUTH_TOKEN') or '').strip()
TWILIO_WHATSAPP_FROM = (os.getenv('TWILIO_WHATSAPP_FROM') or '').strip()
WHATSAPP_SUPERVISOR_TO = (os.getenv('WHATSAPP_SUPERVISOR_TO') or '').strip()


def _normalize_whatsapp_address(value):
    raw = (value or '').strip()
    if not raw:
        return ''
    if raw.startswith('whatsapp:'):
        return raw
    if raw.startswith('+'):
        return f'whatsapp:{raw}'
    return ''


def _send_whatsapp_message(body_text):
    if not WHATSAPP_ALERTS_ENABLED:
        return False, 'disabled'

    from_addr = _normalize_whatsapp_address(TWILIO_WHATSAPP_FROM)
    to_addr = _normalize_whatsapp_address(WHATSAPP_SUPERVISOR_TO)
    if not (TWILIO_ACCOUNT_SID and TWILIO_AUTH_TOKEN and from_addr and to_addr):
        return False, 'missing_config'

    url = f'https://api.twilio.com/2010-04-01/Accounts/{TWILIO_ACCOUNT_SID}/Messages.json'
    payload = {
        'From': from_addr,
        'To': to_addr,
        'Body': (body_text or '').strip()[:1500],
    }

    try:
        response = requests.post(
            url,
            data=payload,
            auth=(TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN),
            timeout=15,
        )
        if 200 <= response.status_code < 300:
            return True, 'sent'
        logger.warning(f'WhatsApp Twilio error {response.status_code}: {response.text[:500]}')
        return False, f'http_{response.status_code}'
    except Exception as exc:
        logger.warning(f'WhatsApp send exception: {exc}')
        return False, 'exception'


def _send_alerta_whatsapp_if_enabled(alerta_item):
    if not alerta_item:
        return
    if not WHATSAPP_ALERTS_ENABLED:
        return
    tipo = (alerta_item.tipo or '').strip().lower()
    if WHATSAPP_ALERT_TYPES and tipo and tipo not in WHATSAPP_ALERT_TYPES:
        return

    body = (
        f"ALERTA SISTEMA\n"
        f"Tipo: {alerta_item.tipo or '-'}\n"
        f"Titulo: {alerta_item.titulo or '-'}\n"
        f"Mensaje: {alerta_item.mensaje or '-'}\n"
        f"Fecha: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC"
    )
    ok, reason = _send_whatsapp_message(body)
    if not ok:
        logger.info(f'WhatsApp alert not sent ({reason}) for alerta id={getattr(alerta_item, "id", None)}')


TELEGRAM_ALERTS_ENABLED = os.getenv('TELEGRAM_ALERTS_ENABLED', '0').strip().lower() in ('1', 'true', 'yes', 'on')
TELEGRAM_ALERT_TYPES = {
    x.strip().lower() for x in (os.getenv('TELEGRAM_ALERT_TYPES') or '').split(',') if x.strip()
}
TELEGRAM_BOT_TOKEN = (os.getenv('TELEGRAM_BOT_TOKEN') or '').strip()
TELEGRAM_CHAT_ID = (os.getenv('TELEGRAM_CHAT_ID') or '').strip()
TELEGRAM_CHAT_IDS = [
    x.strip() for x in (os.getenv('TELEGRAM_CHAT_IDS') or '').split(',') if x.strip()
]

GOOGLE_SHEETS_SYNC_ENABLED = os.getenv('GOOGLE_SHEETS_SYNC_ENABLED', '0').strip().lower() in ('1', 'true', 'yes', 'on')
GOOGLE_SHEETS_WEBHOOK_URL = (os.getenv('GOOGLE_SHEETS_WEBHOOK_URL') or '').strip()
GOOGLE_SHEETS_WEBHOOK_TOKEN = (os.getenv('GOOGLE_SHEETS_WEBHOOK_TOKEN') or '').strip()
try:
    GOOGLE_SHEETS_TIMEOUT_SECONDS = max(3, min(60, int(os.getenv('GOOGLE_SHEETS_TIMEOUT_SECONDS', '12') or '12')))
except Exception:
    GOOGLE_SHEETS_TIMEOUT_SECONDS = 12


def _telegram_destinations():
    if TELEGRAM_CHAT_IDS:
        return list(dict.fromkeys(TELEGRAM_CHAT_IDS))
    if TELEGRAM_CHAT_ID:
        return [TELEGRAM_CHAT_ID]
    return []


def _order_machine_queue_items(hojas):
    """Order assigned hojas so active appears first and the rest by recent update."""
    ordered = list(hojas or [])
    ordered.sort(
        key=lambda h: (h.fecha_actualizacion or h.fecha_creacion or datetime.min),
        reverse=True,
    )
    ordered.sort(key=lambda h: 0 if (h.estado or '').lower() == 'activa' else 1)
    return ordered


def _normalize_machine_type_key(value):
    return str(value or '').strip().lower()


def _machine_queue_limit_for_type(maquina_tipo):
    tipo_key = _normalize_machine_type_key(maquina_tipo)
    if not tipo_key:
        return MACHINE_QUEUE_LIMIT
    if tipo_key in MACHINE_QUEUE_LIMIT_BY_TYPE:
        return MACHINE_QUEUE_LIMIT_BY_TYPE[tipo_key]
    wildcard_limit = MACHINE_QUEUE_LIMIT_BY_TYPE.get('*')
    if wildcard_limit:
        return wildcard_limit
    return MACHINE_QUEUE_LIMIT


def _pick_machine_active_hoja(hojas):
    for hoja in hojas or []:
        if (hoja.estado or '').lower() == 'activa':
            return hoja
    return hojas[0] if hojas else None


def _machine_queue_count(model_cls, maquina_id, exclude_hoja_id=None):
    q = model_cls.query.filter(
        model_cls.maquina_id == maquina_id,
        model_cls.estado.in_(['activa', 'pausada'])
    )
    if exclude_hoja_id:
        q = q.filter(model_cls.id != int(exclude_hoja_id))
    return q.count()


def _pause_other_machine_hojas(model_cls, maquina_id, except_hoja_id, now_dt):
    others = model_cls.query.filter(
        model_cls.maquina_id == maquina_id,
        model_cls.estado == 'activa',
        model_cls.id != except_hoja_id,
    ).all()
    for item in others:
        item.estado = 'pausada'
        item.fecha_actualizacion = now_dt


def _resume_or_activate_hoja(hoja, now_dt):
    if (hoja.estado or '').lower() == 'pausada':
        paused_at = hoja.fecha_actualizacion or now_dt
        if hoja.fecha_salida:
            hoja.fecha_salida = hoja.fecha_salida + (now_dt - paused_at)
        else:
            hoja.fecha_salida = now_dt
    elif not hoja.fecha_salida:
        hoja.fecha_salida = now_dt
    hoja.estado = 'activa'
    hoja.fecha_actualizacion = now_dt


def _send_telegram_message(body_text):
    if not TELEGRAM_ALERTS_ENABLED:
        return False, 'disabled'
    destinations = _telegram_destinations()
    if not (TELEGRAM_BOT_TOKEN and destinations):
        return False, 'missing_config'
    url = f'https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage'

    try:
        total = len(destinations)
        sent = 0
        first_error = None
        text_value = (body_text or '').strip()[:4000]

        for chat_id in destinations:
            payload = {
                'chat_id': chat_id,
                'text': text_value,
            }
            response = requests.post(url, json=payload, timeout=15)
            if 200 <= response.status_code < 300:
                sent += 1
                continue
            logger.warning(f'Telegram error {response.status_code} to chat_id={chat_id}: {response.text[:500]}')
            if first_error is None:
                error_suffix = ''
                try:
                    err_payload = response.json() or {}
                    err_desc = str(err_payload.get('description') or '').strip()
                    if err_desc:
                        error_suffix = ':' + err_desc[:160]
                except Exception:
                    pass
                first_error = f'http_{response.status_code}{error_suffix}'

        if sent == total:
            return True, f'sent_{sent}'
        if sent > 0:
            return False, f'partial_{sent}_of_{total}'
        return False, first_error or 'send_failed'
    except Exception as exc:
        logger.warning(f'Telegram send exception: {exc}')
        return False, 'exception'


def _send_alerta_telegram_if_enabled(alerta_item):
    if not alerta_item:
        return
    if not TELEGRAM_ALERTS_ENABLED:
        return
    tipo = (alerta_item.tipo or '').strip().lower()
    if TELEGRAM_ALERT_TYPES and tipo and tipo not in TELEGRAM_ALERT_TYPES:
        return
    body = (
        f"*ALERTA SISTEMA*\n"
        f"Tipo: {alerta_item.tipo or '-'}\n"
        f"Titulo: {alerta_item.titulo or '-'}\n"
        f"Mensaje: {alerta_item.mensaje or '-'}\n"
        f"Fecha: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC"
    )
    ok, reason = _send_telegram_message(body)
    if not ok:
        logger.info(f'Telegram alert not sent ({reason}) for alerta id={getattr(alerta_item, "id", None)}')


def _sync_almacen_liberacion_to_sheets(hoja, flujo, recepcion_id):
    """Sincroniza a Google Sheets (via webhook) cuando Almacen libera una hoja."""
    if not GOOGLE_SHEETS_SYNC_ENABLED:
        return False, 'disabled'
    if not GOOGLE_SHEETS_WEBHOOK_URL:
        return False, 'missing_webhook_url'

    cantidad_total = int((flujo.cantidad_total_piezas if flujo and flujo.cantidad_total_piezas is not None else 0) or 0)
    if cantidad_total <= 0 and hoja and hoja.cantidad_piezas:
        cantidad_total = int(hoja.cantidad_piezas or 0)

    cantidad_entregada = int((flujo.cantidad_entregada if flujo and flujo.cantidad_entregada is not None else 0) or 0)
    if cantidad_entregada < 0:
        cantidad_entregada = 0
    if cantidad_total > 0 and cantidad_entregada > cantidad_total:
        cantidad_entregada = cantidad_total

    payload = {
        'source': 'controlcalidad360',
        'event': 'almacen_liberada',
        'timestamp_utc': datetime.utcnow().isoformat(),
        'hoja_ruta_id': int(hoja.id if hoja else 0),
        'flujo_id': int(flujo.id if flujo else 0),
        'folio_hoja': (hoja.nombre if hoja else '') or '',
        'clave': (hoja.pn if hoja else '') or '',
        'oc_ot': ((hoja.orden_trabajo_hr if hoja else None) or (hoja.orden_trabajo_pt if hoja else '') or ''),
        'recepcion_id': (recepcion_id or '').strip(),
        'estado_actual': (flujo.estado if flujo else '') or '',
        'cantidad_total_piezas': cantidad_total,
        'cantidad_entregada': cantidad_entregada,
        'cantidad_pendiente': max(cantidad_total - cantidad_entregada, 0),
        'usuario': _logistica_username(),
    }

    headers = {'Content-Type': 'application/json'}
    if GOOGLE_SHEETS_WEBHOOK_TOKEN:
        headers['X-Webhook-Token'] = GOOGLE_SHEETS_WEBHOOK_TOKEN

    try:
        response = requests.post(
            GOOGLE_SHEETS_WEBHOOK_URL,
            json=payload,
            headers=headers,
            timeout=GOOGLE_SHEETS_TIMEOUT_SECONDS,
        )
        if 200 <= response.status_code < 300:
            return True, 'sent'
        logger.warning(f'[SHEETS_SYNC] HTTP {response.status_code} body={response.text[:500]} payload={payload}')
        return False, f'http_{response.status_code}'
    except Exception as exc:
        logger.warning(f'[SHEETS_SYNC] Exception syncing almacen liberation: {exc}; payload={payload}')
        return False, 'exception'


def _procesos_import_job_path(job_id):
    safe = re.sub(r'[^a-zA-Z0-9_-]', '', str(job_id or ''))
    return os.path.join(PROCESOS_IMPORT_DIR, f'{safe}.json')


def _get_procesos_import_job(job_id):
    path = _procesos_import_job_path(job_id)
    if not os.path.exists(path):
        return None
    try:
        with open(path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return None


def _set_procesos_import_job(job_id, **fields):
    with PROCESOS_IMPORT_LOCK:
        job = _get_procesos_import_job(job_id) or {}
        job.update(fields)
        path = _procesos_import_job_path(job_id)
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(job, f, ensure_ascii=False)


def _get_machine_schedule_timezone():
    try:
        return ZoneInfo(MACHINE_SCHEDULE_TIMEZONE)
    except Exception:
        logger.warning(
            f"[MAQUINA_SCHEDULE] Zona horaria inválida '{MACHINE_SCHEDULE_TIMEZONE}', usando UTC-06:00 fija"
        )
        return dt_timezone(timedelta(hours=-6))


def _machine_schedule_now():
    return datetime.now(_get_machine_schedule_timezone())


def _get_machine_schedule_window(now_local=None):
    now_local = now_local or _machine_schedule_now()
    weekday = now_local.weekday()
    current_hm = (now_local.hour, now_local.minute)

    if weekday >= 5:
        return {
            'active': False,
            'slot': 'weekend_off',
            'label': 'Fuera de turno fin de semana',
            'now_local': now_local,
        }

    if current_hm < (6, 30):
        return {
            'active': False,
            'slot': 'before_start',
            'label': 'Fuera de turno antes de 06:30',
            'now_local': now_local,
        }
    if current_hm < (12, 0):
        return {
            'active': True,
            'slot': 'turno_manana',
            'label': 'Turno activo 06:30-12:00',
            'now_local': now_local,
        }
    if current_hm < (12, 30):
        return {
            'active': False,
            'slot': 'comida',
            'label': 'Paro programado 12:00-12:30',
            'now_local': now_local,
        }
    if current_hm < (16, 0):
        return {
            'active': True,
            'slot': 'turno_tarde',
            'label': 'Turno activo 12:30-16:00',
            'now_local': now_local,
        }
    return {
        'active': False,
        'slot': 'after_end',
        'label': 'Paro programado después de 16:00',
        'now_local': now_local,
    }


def _machine_schedule_token(schedule_window):
    now_local = schedule_window['now_local']
    return f"{now_local.strftime('%Y-%m-%d')}-{schedule_window['slot']}"


def _load_machine_schedule_state():
    if not os.path.exists(MACHINE_SCHEDULE_STATE_FILE):
        return {}
    try:
        with open(MACHINE_SCHEDULE_STATE_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {}


def _save_machine_schedule_state(payload):
    try:
        os.makedirs(os.path.dirname(MACHINE_SCHEDULE_STATE_FILE), exist_ok=True)
        with open(MACHINE_SCHEDULE_STATE_FILE, 'w', encoding='utf-8') as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
    except Exception as exc:
        logger.warning(f"[MAQUINA_SCHEDULE] No se pudo guardar estado del scheduler: {exc}")


def _pause_machine_hojas(now_dt=None):
    now_dt = now_dt or datetime.utcnow()

    hojas_entrega = HojaRutaEntrega.query.filter(
        HojaRutaEntrega.maquina_id.isnot(None),
        HojaRutaEntrega.estado == 'activa'
    ).all()
    for hoja in hojas_entrega:
        hoja.estado = 'pausada'
        hoja.fecha_actualizacion = now_dt

    hojas_mp = HojaRutaNueva.query.filter(
        HojaRutaNueva.maquina_id.isnot(None),
        HojaRutaNueva.estado == 'activa'
    ).all()
    for hoja in hojas_mp:
        hoja.estado = 'pausada'
        hoja.fecha_actualizacion = now_dt


def _resume_machine_hojas(now_dt=None):
    now_dt = now_dt or datetime.utcnow()

    hojas_entrega = HojaRutaEntrega.query.filter(
        HojaRutaEntrega.maquina_id.isnot(None),
        HojaRutaEntrega.estado == 'pausada'
    ).all()
    for hoja in hojas_entrega:
        paused_at = hoja.fecha_actualizacion or now_dt
        if hoja.fecha_salida:
            hoja.fecha_salida = hoja.fecha_salida + (now_dt - paused_at)
        else:
            hoja.fecha_salida = now_dt
        hoja.estado = 'activa'
        hoja.fecha_actualizacion = now_dt

    hojas_mp = HojaRutaNueva.query.filter(
        HojaRutaNueva.maquina_id.isnot(None),
        HojaRutaNueva.estado == 'pausada'
    ).all()
    for hoja in hojas_mp:
        paused_at = hoja.fecha_actualizacion or now_dt
        if hoja.fecha_salida:
            hoja.fecha_salida = hoja.fecha_salida + (now_dt - paused_at)
        else:
            hoja.fecha_salida = now_dt
        hoja.estado = 'activa'
        hoja.fecha_actualizacion = now_dt


def _apply_machine_schedule(force=False):
    if not MACHINE_SCHEDULE_ENABLED:
        return False

    lock_acquired = False
    try:
        lock_acquired = bool(db.session.execute(
            text('SELECT pg_try_advisory_lock(:lock_id)'),
            {'lock_id': MACHINE_SCHEDULE_ADVISORY_LOCK_ID}
        ).scalar())
    except Exception as exc:
        db.session.rollback()
        logger.warning(f"[MAQUINA_SCHEDULE] No se pudo adquirir lock PostgreSQL: {exc}")
        return False

    if not lock_acquired:
        return False

    try:
        schedule_window = _get_machine_schedule_window()
        token = _machine_schedule_token(schedule_window)
        desired_active = bool(schedule_window['active'])
        saved_state = _load_machine_schedule_state()
        if not force and saved_state.get('token') == token and saved_state.get('desired_active') == desired_active:
            return False

        maquinas = Máquina.query.all()
        now_utc = datetime.utcnow()
        changed = False
        for maq in maquinas:
            if bool(getattr(maq, 'activo', False)) != desired_active:
                maq.activo = desired_active
                changed = True

        if desired_active:
            _resume_machine_hojas(now_dt=now_utc)
        else:
            _pause_machine_hojas(now_dt=now_utc)

        db.session.commit()
        _save_machine_schedule_state({
            'token': token,
            'desired_active': desired_active,
            'label': schedule_window['label'],
            'applied_at_local': schedule_window['now_local'].isoformat(),
            'applied_at_utc': now_utc.isoformat(),
        })
        logger.info(
            f"[MAQUINA_SCHEDULE] Aplicado horario '{schedule_window['label']}' => activo={desired_active} maquinas={len(maquinas)}"
        )
        return changed
    except Exception as exc:
        db.session.rollback()
        logger.error(f"[MAQUINA_SCHEDULE] Error aplicando horario de máquinas: {exc}", exc_info=True)
        return False
    finally:
        try:
            db.session.execute(
                text('SELECT pg_advisory_unlock(:lock_id)'),
                {'lock_id': MACHINE_SCHEDULE_ADVISORY_LOCK_ID}
            )
            db.session.commit()
        except Exception:
            db.session.rollback()


def _machine_schedule_loop():
    logger.info(
        f"[MAQUINA_SCHEDULE] Scheduler activo cada {MACHINE_SCHEDULE_POLL_SECONDS}s timezone={MACHINE_SCHEDULE_TIMEZONE}"
    )
    while not _MACHINE_SCHEDULE_STOP.is_set():
        with app.app_context():
            _apply_machine_schedule()

        waited = 0
        while waited < MACHINE_SCHEDULE_POLL_SECONDS and not _MACHINE_SCHEDULE_STOP.is_set():
            sleep(1)
            waited += 1


def _start_machine_schedule_scheduler_once():
    global _MACHINE_SCHEDULE_THREAD
    global _MACHINE_SCHEDULE_INIT

    if _MACHINE_SCHEDULE_INIT:
        return
    _MACHINE_SCHEDULE_INIT = True

    if not MACHINE_SCHEDULE_ENABLED:
        logger.info('[MAQUINA_SCHEDULE] Scheduler deshabilitado por MACHINE_SCHEDULE_ENABLED=0')
        return

    _MACHINE_SCHEDULE_THREAD = threading.Thread(
        target=_machine_schedule_loop,
        name='machine-shift-scheduler',
        daemon=True,
    )
    _MACHINE_SCHEDULE_THREAD.start()


def _run_procesos_import_job(job_id, filename, abs_saved_path, sheet):
    _set_procesos_import_job(job_id, status='running', started_at=datetime.utcnow().isoformat())
    try:
        cmd = [sys.executable, 'tools/import_procesos.py', '--file', abs_saved_path, '--overwrite']
        if sheet:
            cmd.extend(['--sheet', sheet])

        result = subprocess.run(
            cmd,
            cwd=os.getcwd(),
            capture_output=True,
            text=True,
            timeout=900,
        )
        combined = '\n'.join([x for x in [result.stdout, result.stderr] if x]).strip()

        if result.returncode != 0:
            logger.error(f"[IMPORT_PROCESOS] Error importando {filename}: {combined}")
            _set_procesos_import_job(
                job_id,
                status='error',
                error='Falló la importación de procesos',
                output=combined[-6000:] if combined else '',
                finished_at=datetime.utcnow().isoformat(),
            )
            return

        logger.info(f"[IMPORT_PROCESOS] Importación OK para archivo {filename}")
        _set_procesos_import_job(
            job_id,
            status='success',
            output=combined[-6000:] if combined else 'Importación ejecutada sin salida.',
            finished_at=datetime.utcnow().isoformat(),
        )
    except subprocess.TimeoutExpired:
        logger.error(f"[IMPORT_PROCESOS] Timeout importando {filename}")
        _set_procesos_import_job(
            job_id,
            status='error',
            error='La importación excedió el tiempo límite (900s)',
            finished_at=datetime.utcnow().isoformat(),
        )
    except Exception as e:
        logger.error(f"[IMPORT_PROCESOS] Excepción importando {filename}: {e}")
        _set_procesos_import_job(
            job_id,
            status='error',
            error=f'Error procesando importación: {str(e)}',
            finished_at=datetime.utcnow().isoformat(),
        )
    finally:
        try:
            if os.path.exists(abs_saved_path):
                os.remove(abs_saved_path)
        except Exception:
            pass

# Jornada laboral para planeacion de hojas de ruta
WORKDAY_BLOCKS = ((6, 30, 12, 0), (12, 30, 16, 0))
WORKDAY_SECONDS = (5 * 3600 + 30 * 60) + (3 * 3600 + 30 * 60)  # 9h


def _is_workday(dt_or_date):
    if hasattr(dt_or_date, 'weekday'):
        return dt_or_date.weekday() < 5  # Monday=0 ... Sunday=6
    return False


def _norm_text(value):
    text = str(value or '').upper().strip()
    return (text.replace('Á', 'A')
                .replace('É', 'E')
                .replace('Í', 'I')
                .replace('Ó', 'O')
                .replace('Ú', 'U'))


def _clean_nullable_text(value):
    text = str(value or '').strip()
    if text.lower() in ('none', 'null', 'nan', 'nat', '-'):
        return ''
    return text


def _qc_strip_scrap_summary(text):
    src = str(text or '')
    cleaned = re.sub(
        r'\n?\[QC_SCRAP_SUMMARY_START\].*?\[QC_SCRAP_SUMMARY_END\]\n?',
        '\n',
        src,
        flags=re.S,
    )
    return cleaned.strip()


def _qc_parse_scrap_summary(text):
    src = str(text or '')
    m = re.search(r'\[QC_SCRAP_SUMMARY_START\](.*?)\[QC_SCRAP_SUMMARY_END\]', src, flags=re.S)
    if not m:
        return None

    block = m.group(1)
    values = {}
    for line in block.splitlines():
        line = line.strip()
        if not line or '=' not in line:
            continue
        k, v = line.split('=', 1)
        values[k.strip().upper()] = v.strip()

    def _to_int(key, default=0):
        try:
            return max(0, int(values.get(key, default)))
        except Exception:
            return default

    return {
        'lote_inicial': _to_int('LOTE_INICIAL', 0),
        'total_scrap': _to_int('TOTAL_SCRAP', 0),
        'lote_final': _to_int('LOTE_FINAL', 0),
        'detalle': values.get('DETALLE', ''),
        'actualizado_por': values.get('ACTUALIZADO_POR', ''),
        'fecha': values.get('FECHA', ''),
    }


def _qc_extract_scrap_block(text):
    src = str(text or '')
    m = re.search(r'\[QC_SCRAP_SUMMARY_START\].*?\[QC_SCRAP_SUMMARY_END\]', src, flags=re.S)
    return m.group(0).strip() if m else ''


def _mp_extract_process_state_block(text):
    src = str(text or '')
    m = re.search(r'\[MP_PROCESS_STATE_START\](.*?)\[MP_PROCESS_STATE_END\]', src, flags=re.S)
    return m.group(1).strip() if m else ''


def _mp_strip_process_state_block(text):
    src = str(text or '')
    cleaned = re.sub(
        r'\n?\[MP_PROCESS_STATE_START\].*?\[MP_PROCESS_STATE_END\]\n?',
        '\n',
        src,
        flags=re.S,
    )
    return cleaned.strip()


def _mp_parse_completed_process_ids(text):
    raw = _mp_extract_process_state_block(text)
    if not raw:
        return set()

    try:
        payload = json.loads(raw)
    except Exception:
        return set()

    ids = payload.get('completed_ids') if isinstance(payload, dict) else []
    completed = set()
    for value in ids or []:
        try:
            completed.add(int(value))
        except Exception:
            continue
    return completed


def _mp_upsert_process_state_block(text, completed_ids):
    base = _mp_strip_process_state_block(text)
    normalized = sorted({int(x) for x in (completed_ids or set())})
    if not normalized:
        return base or None

    block_payload = json.dumps({'completed_ids': normalized}, ensure_ascii=False)
    block = f"[MP_PROCESS_STATE_START]\n{block_payload}\n[MP_PROCESS_STATE_END]"
    if base:
        return f"{base}\n\n{block}".strip()
    return block


def _mye_extract_plan_state_block(text):
    src = str(text or '')
    m = re.search(r'\[MYE_PLAN_STATE_START\](.*?)\[MYE_PLAN_STATE_END\]', src, flags=re.S)
    return m.group(1).strip() if m else ''


def _mye_strip_plan_state_block(text):
    src = str(text or '')
    cleaned = re.sub(
        r'\n?\[MYE_PLAN_STATE_START\].*?\[MYE_PLAN_STATE_END\]\n?',
        '\n',
        src,
        flags=re.S,
    )
    return cleaned.strip()


def _mye_parse_plan_state(text):
    raw = _mye_extract_plan_state_block(text)
    if not raw:
        return {}
    try:
        payload = json.loads(raw)
    except Exception:
        return {}
    return payload if isinstance(payload, dict) else {}


def _mye_upsert_plan_state_block(text, state_payload):
    base = _mye_strip_plan_state_block(text)
    payload = state_payload if isinstance(state_payload, dict) else {}
    steps_raw = payload.get('steps')
    steps = steps_raw if isinstance(steps_raw, list) else None
    plans_raw = payload.get('plans')
    plans = plans_raw if isinstance(plans_raw, list) else None

    compact = {
        'operator_id': int(payload.get('operator_id')) if str(payload.get('operator_id') or '').isdigit() else None,
        'operator_username': (str(payload.get('operator_username') or '').strip() or None),
        'station': (str(payload.get('station') or '').strip() or None),
        'process_name': (str(payload.get('process_name') or '').strip() or None),
        'process_status': (str(payload.get('process_status') or '').strip() or None),
        'start_at': (str(payload.get('start_at') or '').strip() or None),
        'duration_hours': float(payload.get('duration_hours') or 0),
        'machine_icon': (str(payload.get('machine_icon') or '').strip() or None),
        'steps': steps,
        'plans': plans,
    }
    compact = {
        k: v for k, v in compact.items()
        if v not in (None, '', 0, 0.0)
    }

    if not compact:
        return base or None

    block_payload = json.dumps(compact, ensure_ascii=False)
    block = f"[MYE_PLAN_STATE_START]\n{block_payload}\n[MYE_PLAN_STATE_END]"
    if base:
        return f"{base}\n\n{block}".strip()
    return block


def _mye_machine_icon(clave_txt):
    key = (clave_txt or '').lower()
    if any(x in key for x in ['sold', 'weld']):
        return '🔥'
    if any(x in key for x in ['laser', 'corte', 'cut']):
        return '✂️'
    if any(x in key for x in ['arnes', 'cable']):
        return '🔌'
    if any(x in key for x in ['ensamble', 'assembly']):
        return '🧩'
    if any(x in key for x in ['lamin', 'roll']):
        return '🛞'
    if any(x in key for x in ['prensa', 'press']):
        return '🗜️'
    if any(x in key for x in ['cnc', 'router']):
        return '⚙️'
    return '🏭'


def _mye_default_catalog():
    return {
        'operators': [],
        'machines': [],
        'processes': [],
    }


def _mye_read_catalog():
    if not os.path.exists(MYE_CATALOG_FILE):
        return _mye_default_catalog()

    try:
        with open(MYE_CATALOG_FILE, 'r', encoding='utf-8') as f:
            payload = json.load(f)
    except Exception:
        return _mye_default_catalog()

    if not isinstance(payload, dict):
        return _mye_default_catalog()

    operators = payload.get('operators')
    machines = payload.get('machines')
    processes = payload.get('processes')
    return {
        'operators': operators if isinstance(operators, list) else [],
        'machines': machines if isinstance(machines, list) else [],
        'processes': processes if isinstance(processes, list) else [],
    }


def _mye_write_catalog(catalog):
    os.makedirs(os.path.dirname(MYE_CATALOG_FILE), exist_ok=True)
    with open(MYE_CATALOG_FILE, 'w', encoding='utf-8') as f:
        json.dump(catalog, f, ensure_ascii=False, indent=2)


def _mye_next_id(items):
    max_id = 0
    for item in items or []:
        try:
            max_id = max(max_id, int(item.get('id') or 0))
        except Exception:
            continue
    return max_id + 1


def _mye_order_plan_entries(order_obj):
    state = _mye_parse_plan_state(order_obj.notas)
    plans_raw = state.get('plans') if isinstance(state, dict) else None
    if isinstance(plans_raw, list):
        return [p for p in plans_raw if isinstance(p, dict)]

    if isinstance(state, dict) and state:
        return [state]
    return []


def _mye_catalog_in_use(kind, value):
    hits = []
    needle = str(value or '').strip().lower()
    if not needle and kind in ('machine', 'process'):
        return hits

    try:
        value_num = int(value)
    except Exception:
        value_num = 0

    ordenes = MaquinariaOrdenTrabajo.query.order_by(MaquinariaOrdenTrabajo.id.desc()).all()
    for orden in ordenes:
        plans = _mye_order_plan_entries(orden)
        for plan in plans:
            matched = False
            if kind == 'operator':
                matched = int(plan.get('operator_id') or 0) == value_num and value_num > 0
            elif kind == 'machine':
                station = str(plan.get('station') or '').strip().lower()
                matched = station == needle
            elif kind == 'process':
                process_name = str(plan.get('process_name') or '').strip().lower()
                names = [x.strip().lower() for x in process_name.split(',') if x.strip()]
                if needle and (needle in names or process_name == needle):
                    matched = True
                if not matched:
                    steps = plan.get('steps')
                    if isinstance(steps, list):
                        for step in steps:
                            if not isinstance(step, dict):
                                continue
                            if str(step.get('name') or '').strip().lower() == needle:
                                matched = True
                                break

            if matched:
                hits.append({'id': orden.id, 'folio_ot': orden.folio_ot})
                break

    return hits


def _qc_parse_review_block(notas_text):
    notas = str(notas_text or '')
    status_match = re.search(r'STATUS=(QC_OK|QC_NOK)', notas)
    status = status_match.group(1) if status_match else None

    block_match = re.search(r'\[QC_REVIEW_START\](.*?)\[QC_REVIEW_END\]', notas, flags=re.S)
    block = block_match.group(1) if block_match else ''

    def _extract(field, default=''):
        m = re.search(rf'^{field}=(.*)$', block, flags=re.M)
        return m.group(1).strip() if m else default

    def _extract_int(field, default=0):
        try:
            return max(0, int(_extract(field, str(default))))
        except Exception:
            return default

    return {
        'status': status,
        'usuario': _extract('USUARIO', ''),
        'fecha': _extract('FECHA', ''),
        'dimensional': _extract('DIMENSIONAL', 'NO') == 'OK',
        'visual': _extract('VISUAL', 'NO') == 'OK',
        'rebaba': _extract('REBABA', 'NO') == 'OK',
        'material': _extract('MATERIAL', 'NO') == 'OK',
        'ajuste': _extract('AJUSTE', 'NO') == 'OK',
        'limpieza': _extract('LIMPIEZA', 'NO') == 'OK',
        'scrap_piezas': _extract_int('SCRAP_PIEZAS', 0),
        'observaciones': _extract('OBSERVACIONES', ''),
    }


def _parse_time_to_seconds(value):
    if value is None:
        return 0
    raw = str(value).strip()
    if not raw:
        return 0

    try:
        parts = [int(p) for p in raw.split(':')]
    except Exception:
        return 0

    if len(parts) == 3:
        h, m, s = parts
        return max(0, h * 3600 + m * 60 + s)
    if len(parts) == 2:
        m, s = parts
        return max(0, m * 60 + s)
    if len(parts) == 1:
        return max(0, parts[0])
    return 0


def _format_seconds_to_hms(total_seconds):
    sec = max(0, int(total_seconds or 0))
    h = sec // 3600
    m = (sec % 3600) // 60
    s = sec % 60
    return f"{h:02d}:{m:02d}:{s:02d}"


def _station_seconds(est):
    for field in ('t_e', 't_tct', 't_tco', 't_to'):
        sec = _parse_time_to_seconds(getattr(est, field, None))
        if sec > 0:
            return sec
    return 0


def _stations_for_machine_type(estaciones, maquina_tipo):
    if not maquina_tipo:
        return estaciones

    tipo = _norm_text(maquina_tipo)
    filtered = []
    for est in estaciones:
        haystack = ' '.join([
            _norm_text(getattr(est, 'centro_trabajo', '')),
            _norm_text(getattr(est, 'nombre', '')),
            _norm_text(getattr(est, 'operacion', '')),
        ])
        if tipo and tipo in haystack:
            filtered.append(est)
    return filtered or estaciones


def _align_to_work_slot(dt):
    base = dt if isinstance(dt, datetime) else datetime.utcnow()

    # Skip weekends: always move to next Monday start block.
    while not _is_workday(base):
        base = (base + timedelta(days=1)).replace(hour=WORKDAY_BLOCKS[0][0], minute=WORKDAY_BLOCKS[0][1], second=0, microsecond=0)

    b1s = base.replace(hour=WORKDAY_BLOCKS[0][0], minute=WORKDAY_BLOCKS[0][1], second=0, microsecond=0)
    b1e = base.replace(hour=WORKDAY_BLOCKS[0][2], minute=WORKDAY_BLOCKS[0][3], second=0, microsecond=0)
    b2s = base.replace(hour=WORKDAY_BLOCKS[1][0], minute=WORKDAY_BLOCKS[1][1], second=0, microsecond=0)
    b2e = base.replace(hour=WORKDAY_BLOCKS[1][2], minute=WORKDAY_BLOCKS[1][3], second=0, microsecond=0)

    if base < b1s:
        return b1s
    if b1s <= base < b1e:
        return base
    if b1e <= base < b2s:
        return b2s
    if b2s <= base < b2e:
        return base
    next_day = base + timedelta(days=1)
    return next_day.replace(hour=WORKDAY_BLOCKS[0][0], minute=WORKDAY_BLOCKS[0][1], second=0, microsecond=0)


def _add_work_seconds(start_dt, seconds):
    remaining = max(0, int(seconds or 0))
    current = _align_to_work_slot(start_dt)

    if remaining == 0:
        return current

    while remaining > 0:
        b1s = current.replace(hour=WORKDAY_BLOCKS[0][0], minute=WORKDAY_BLOCKS[0][1], second=0, microsecond=0)
        b1e = current.replace(hour=WORKDAY_BLOCKS[0][2], minute=WORKDAY_BLOCKS[0][3], second=0, microsecond=0)
        b2s = current.replace(hour=WORKDAY_BLOCKS[1][0], minute=WORKDAY_BLOCKS[1][1], second=0, microsecond=0)
        b2e = current.replace(hour=WORKDAY_BLOCKS[1][2], minute=WORKDAY_BLOCKS[1][3], second=0, microsecond=0)

        if b1s <= current < b1e:
            block_end = b1e
        elif b2s <= current < b2e:
            block_end = b2e
        else:
            current = _align_to_work_slot(current)
            continue

        available = max(0, int((block_end - current).total_seconds()))
        if remaining <= available:
            return current + timedelta(seconds=remaining)

        remaining -= available
        current = _align_to_work_slot(block_end + timedelta(seconds=1))

    return current


def _working_seconds_between(start_dt, end_dt):
    if not isinstance(start_dt, datetime) or not isinstance(end_dt, datetime):
        return 0
    if end_dt <= start_dt:
        return 0

    total = 0
    day = start_dt.date()
    end_day = end_dt.date()

    while day <= end_day:
        if not _is_workday(day):
            day += timedelta(days=1)
            continue

        for h1, m1, h2, m2 in WORKDAY_BLOCKS:
            block_start = datetime.combine(day, datetime.min.time()).replace(hour=h1, minute=m1)
            block_end = datetime.combine(day, datetime.min.time()).replace(hour=h2, minute=m2)

            overlap_start = max(block_start, start_dt)
            overlap_end = min(block_end, end_dt)
            if overlap_end > overlap_start:
                total += int((overlap_end - overlap_start).total_seconds())

        day += timedelta(days=1)

    return max(0, total)


def _bounded_working_seconds(start_dt, end_dt, window_start, window_end):
    if not all(isinstance(x, datetime) for x in (start_dt, end_dt, window_start, window_end)):
        return 0

    overlap_start = max(start_dt, window_start)
    overlap_end = min(end_dt, window_end)
    if overlap_end <= overlap_start:
        return 0

    return _working_seconds_between(overlap_start, overlap_end)


def _apply_hoja_time_plan(hoja, estaciones, maquina_tipo=None, fallback_total_time=None):
    cantidad = max(1, int(hoja.cantidad_piezas or 0))
    estaciones_base = _stations_for_machine_type(estaciones, maquina_tipo)

    per_piece_seconds = sum(_station_seconds(e) for e in estaciones_base)

    # Si no hubo tiempos por proceso, usar último T/O o fallback provisto
    if per_piece_seconds <= 0:
        for est in sorted(estaciones_base, key=lambda x: getattr(x, 'orden', 0), reverse=True):
            sec = _parse_time_to_seconds(getattr(est, 't_to', None))
            if sec > 0:
                per_piece_seconds = sec
                break

    total_seconds = per_piece_seconds * cantidad
    if total_seconds <= 0 and fallback_total_time:
        total_seconds = _parse_time_to_seconds(fallback_total_time)

    if total_seconds <= 0:
        hoja.total_tiempo = None
        hoja.dias_a_laborar = None
        hoja.fecha_termino = None
        return

    hoja.total_tiempo = _format_seconds_to_hms(total_seconds)
    hoja.dias_a_laborar = round(total_seconds / WORKDAY_SECONDS, 2)

    inicio = hoja.fecha_salida or datetime.utcnow()
    hoja.fecha_termino = _add_work_seconds(inicio, total_seconds)


def _resolve_clave_descripcion_by_pn(pn_value):
    """Resolve a human-readable description for a PN.
    Priority: clave.notas -> clave.nombre -> process description/notes/operation -> key code.
    """
    pn = (pn_value or '').strip()
    if not pn:
        return ''

    normalized = pn.upper()
    clave = ClaveProducto.query.filter(func.upper(func.trim(ClaveProducto.clave)) == normalized).first()
    if not clave:
        # Fallback: at least show the key code for legacy hojas.
        return pn

    notas = _clean_nullable_text(getattr(clave, 'notas', None))
    if notas:
        return notas

    nombre_clave = _clean_nullable_text(getattr(clave, 'nombre', None))
    if nombre_clave:
        return nombre_clave

    procesos = ClaveProceso.query.filter_by(clave_id=clave.id).order_by(ClaveProceso.orden.asc()).all()
    for cp in procesos:
        candidates = [
            (cp.proceso.descripcion if getattr(cp, 'proceso', None) else '') or '',
            getattr(cp, 'notas', None) or '',
            getattr(cp, 'operacion', None) or '',
            (cp.proceso.operacion if getattr(cp, 'proceso', None) else '') or '',
            (cp.proceso.nombre if getattr(cp, 'proceso', None) else '') or '',
        ]
        for c in candidates:
            text = _clean_nullable_text(c)
            if text:
                return text

    # Final fallback for legacy data without notes/description.
    return _clean_nullable_text(getattr(clave, 'clave', None)) or pn


def _clave_token_for_serie(clave_txt):
    """Token alfanumerico de la clave para armar el folio HR/HRN."""
    return ''.join(ch for ch in (clave_txt or '') if ch.isalnum())[:10] or 'CLAVE'


def _rebuild_hoja_serie_con_clave(hoja, clave_txt, prefix=None):
    """Sincroniza el folio con la clave del producto.

    Conserva prefijo (HR/HRN), fecha y consecutivo del folio actual cuando existen,
    y solo reemplaza el segmento de clave. Evita desfaces al editar la clave.
    """
    if hoja is None:
        return None

    clave_segura = _clave_token_for_serie(clave_txt)
    current = (getattr(hoja, 'nombre', None) or '').strip()
    match = re.match(r'^(HRN?)-(\d{8})-(.+)-(\d+)$', current, re.IGNORECASE)
    if match:
        pref = match.group(1).upper()
        if pref == 'HRN':
            pref = 'HRN'
        else:
            pref = 'HR'
        hoja.nombre = f"{pref}-{match.group(2)}-{clave_segura}-{match.group(4)}"
        return hoja.nombre

    pref = (prefix or '').strip().upper()
    if pref not in ('HR', 'HRN'):
        pref = 'HRN' if current.upper().startswith('HRN') else 'HR'

    fecha_ref = (
        getattr(hoja, 'fecha_salida', None)
        or getattr(hoja, 'fecha_creacion', None)
        or datetime.utcnow()
    )
    fecha = fecha_ref.strftime('%Y%m%d') if hasattr(fecha_ref, 'strftime') else datetime.utcnow().strftime('%Y%m%d')
    hoja_id = int(getattr(hoja, 'id', 0) or 0)
    hoja.nombre = f"{pref}-{fecha}-{clave_segura}-{hoja_id:04d}"
    return hoja.nombre


def _mp_process_seconds(cp):
    """Retorna segundos base del proceso MP tomando la primera columna de tiempo disponible."""
    if not cp:
        return 0
    for raw in (
        getattr(cp, 't_e', None),
        getattr(cp, 't_tct', None),
        getattr(cp, 't_tco', None),
        getattr(cp, 't_to', None),
        getattr(cp, 'tiempo_estimado', None),
        getattr(getattr(cp, 'proceso', None), 'tiempo_estimado', None),
    ):
        sec = _parse_time_to_seconds(raw)
        if sec > 0:
            return sec
    return 0


def _get_mp_current_process_projection(pn_value, cantidad_piezas, completed_process_ids=None, elapsed_total_seconds=0):
    """Calcula tiempos del proceso actual MP considerando procesos completados previos.
    Devuelve dict con objetivo/transcurrido/restante en segundos y HH:MM:SS.
    """
    pn = (pn_value or '').strip()
    if not pn:
        return None

    completed = {int(x) for x in (completed_process_ids or set())}
    virtual_ests = _build_mp_virtual_estaciones_by_pn(pn, completed)
    current = next((e for e in virtual_ests if e.get('estado') == 'en_curso'), None)
    if not current:
        return None

    cantidad = max(1, int(cantidad_piezas or 1))
    completed_total_sec = 0
    for est in virtual_ests:
        if est.get('id') in completed:
            cp_done = ClaveProceso.query.get(est.get('id'))
            completed_total_sec += (_mp_process_seconds(cp_done) * cantidad)

    cp_current = ClaveProceso.query.get(current.get('id'))
    current_base_sec = _mp_process_seconds(cp_current)
    if current_base_sec <= 0:
        return None

    objetivo_sec = current_base_sec * cantidad
    transcurrido_actual_sec = max(0, int(elapsed_total_seconds or 0) - completed_total_sec)
    restante_sec = max(0, objetivo_sec - transcurrido_actual_sec)

    return {
        'objetivo_sec': objetivo_sec,
        'transcurrido_sec': transcurrido_actual_sec,
        'restante_sec': restante_sec,
        'objetivo_hms': _format_seconds_to_hms(objetivo_sec),
        'transcurrido_hms': _format_seconds_to_hms(transcurrido_actual_sec),
        'restante_hms': _format_seconds_to_hms(restante_sec),
        'proceso_culminado': restante_sec <= 0,
    }


def _get_mp_current_process_objective_time(pn_value, cantidad_piezas, completed_process_ids=None):
    projection = _get_mp_current_process_projection(
        pn_value,
        cantidad_piezas,
        completed_process_ids,
        elapsed_total_seconds=0,
    )
    return projection['objetivo_hms'] if projection else None


def _parse_elapsed_process_seconds(raw_value):
    try:
        seconds = int(raw_value or 0)
    except Exception:
        return None
    if seconds < 0:
        return None
    return min(seconds, 864000)


def _apply_entrega_process_elapsed(hoja, estaciones, target_estacion, elapsed_seconds, now_ref):
    if not hoja or not target_estacion:
        return

    elapsed_seconds = max(0, int(elapsed_seconds or 0))
    process_start = now_ref - timedelta(seconds=elapsed_seconds) if elapsed_seconds > 0 else now_ref
    target_orden = int(target_estacion.orden or 0)

    for estacion in estaciones:
        estacion_orden = int(estacion.orden or 0)
        if estacion.id == target_estacion.id:
            estacion.estado = 'en_curso'
            estacion.fecha_inicio = process_start
            estacion.fecha_finalizacion = None
        elif estacion_orden < target_orden:
            estacion.estado = 'completada'
            if not estacion.fecha_inicio:
                estacion.fecha_inicio = process_start
            if not estacion.fecha_finalizacion:
                estacion.fecha_finalizacion = process_start
        elif (estacion.estado or '').lower() != 'completada':
            estacion.estado = 'pendiente'
            estacion.fecha_finalizacion = None

    if not hoja.fecha_salida or hoja.fecha_salida > process_start:
        hoja.fecha_salida = process_start


def _apply_mp_process_elapsed(hoja, start_process_id, elapsed_seconds, now_ref):
    if not hoja:
        return None

    virtual_ests = _build_mp_virtual_estaciones_by_pn(hoja.pn, _mp_parse_completed_process_ids(hoja.materia_prima))
    if not virtual_ests:
        return None

    ordered = sorted(virtual_ests, key=lambda item: int(item.get('orden') or 0))
    if start_process_id:
        try:
            start_process_id = int(start_process_id)
        except Exception:
            return None
        target = next((item for item in ordered if int(item.get('id') or 0) == start_process_id), None)
    else:
        target = next((item for item in ordered if (item.get('estado') or '').lower() != 'completada'), None)

    if not target:
        return None

    qty = max(1, int(hoja.cantidad_piezas or 1))
    target_orden = int(target.get('orden') or 0)
    completed_ids = set()
    completed_total_sec = 0
    for item in ordered:
        item_orden = int(item.get('orden') or 0)
        item_id = int(item.get('id') or 0)
        if item_orden < target_orden:
            completed_ids.add(item_id)
            cp_prev = ClaveProceso.query.get(item_id)
            completed_total_sec += (_mp_process_seconds(cp_prev) * qty)

    hoja.materia_prima = _mp_upsert_process_state_block(hoja.materia_prima, completed_ids)
    elapsed_seconds = max(0, int(elapsed_seconds or 0))
    hoja.fecha_salida = now_ref - timedelta(seconds=(completed_total_sec + elapsed_seconds))
    return target


def _build_mp_virtual_estaciones_by_pn(pn_value, completed_process_ids=None):
    """Construye una lista virtual de procesos para hojas MP (sin persistir EstacionTrabajo).
    Se usa para visualizacion en UI de hojas MP y panel de asignacion.
    """
    pn = (pn_value or '').strip()
    if not pn:
        return []

    normalized = pn.upper()
    clave = ClaveProducto.query.filter(func.upper(func.trim(ClaveProducto.clave)) == normalized).first()
    if not clave:
        return []

    procesos = ClaveProceso.query.filter_by(clave_id=clave.id).order_by(ClaveProceso.orden.asc()).all()
    completed = {int(x) for x in (completed_process_ids or set())}
    virtual = []
    for idx, cp in enumerate(procesos, start=1):
        operacion = (
            _clean_nullable_text(getattr(cp, 'operacion', None))
            or _clean_nullable_text(getattr(getattr(cp, 'proceso', None), 'operacion', None))
            or _clean_nullable_text(getattr(getattr(cp, 'proceso', None), 'nombre', None))
            or f'Proceso {idx}'
        )
        t_e = _clean_nullable_text(getattr(cp, 't_e', None))
        t_tct = _clean_nullable_text(getattr(cp, 't_tct', None))
        t_tco = _clean_nullable_text(getattr(cp, 't_tco', None))
        t_to = _clean_nullable_text(getattr(cp, 't_to', None))
        if not t_e:
            t_e = _clean_nullable_text(getattr(cp, 'tiempo_estimado', None)) or _clean_nullable_text(getattr(getattr(cp, 'proceso', None), 'tiempo_estimado', None))

        virtual.append({
            'id': cp.id,
            'orden': cp.orden or idx,
            'nombre': operacion,
            'operacion': operacion,
            'centro_trabajo': _clean_nullable_text(getattr(cp, 'centro_trabajo', None))
                             or _clean_nullable_text(getattr(getattr(cp, 'proceso', None), 'centro_trabajo', None)),
            't_e': t_e or '',
            't_tct': t_tct or '',
            't_tco': t_tco or '',
            't_to': t_to or '',
            'estado': 'completada' if cp.id in completed else 'pendiente',
            'origen': 'clave_proceso_mp',
        })

    for est in virtual:
        if est['estado'] != 'completada':
            est['estado'] = 'en_curso'
            break

    return virtual


def _mp_qc_alert_type_meta(alert_type):
    normalized = (str(alert_type or '').strip().lower() or 'verificado')
    if normalized == 'scrat':
        normalized = 'scrap'
    mapping = {
        'verificado': {'label': 'Verificado', 'color': 'verde'},
        'cuarentena': {'label': 'Cuarentena', 'color': 'amarillo'},
        'scrap': {'label': 'Scrap', 'color': 'rojo'},
    }
    return normalized, mapping.get(normalized, mapping['verificado'])


def _mp_qc_alert_payload_from_registro(registro):   
    mediciones = registro.mediciones if isinstance(registro.mediciones, dict) else {}
    if mediciones.get('module') != 'estaciones_t_mp_qc':
        return None

    alert_type, meta = _mp_qc_alert_type_meta(mediciones.get('alert_type') or registro.resultado)
    process_id = 0
    try:
        process_id = int(mediciones.get('process_id') or 0)
    except Exception:
        process_id = 0

    hoja_mp_id = 0
    try:
        hoja_mp_id = int(mediciones.get('hoja_mp_id') or 0)
    except Exception:
        hoja_mp_id = 0

    reviewed_at = registro.creado_en.isoformat() if registro.creado_en else None
    return {
        'id': registro.id,
        'maquina_id': registro.maquina_id,
        'hoja_mp_id': hoja_mp_id,
        'process_id': process_id,
        'process_name': mediciones.get('process_name') or '',
        'tipo': alert_type,
        'label': meta['label'],
        'color': meta['color'],
        'cantidad_revisada': int(registro.cantidad_inspeccionada or 0),
        'usuario': registro.usuario or '',
        'fecha': reviewed_at,
        'maquina_nombre': mediciones.get('maquina_nombre') or '',
        'hoja_nombre': mediciones.get('hoja_nombre') or '',
    }


def _group_mp_qc_alerts_by_maquina_hoja(registros):
    grouped = {}
    for registro in registros or []:
        payload = _mp_qc_alert_payload_from_registro(registro)
        if not payload:
            continue

        key = (int(payload.get('maquina_id') or 0), int(payload.get('hoja_mp_id') or 0))
        bucket = grouped.setdefault(key, {'recent': [], 'latest_by_process': {}})
        if len(bucket['recent']) < 8:
            bucket['recent'].append(payload)

        process_id = int(payload.get('process_id') or 0)
        if process_id > 0 and process_id not in bucket['latest_by_process']:
            bucket['latest_by_process'][process_id] = payload
    return grouped


def _entregas_qc_alert_payload_from_registro(registro):
    mediciones = registro.mediciones if isinstance(registro.mediciones, dict) else {}
    if mediciones.get('module') != 'estaciones_t_entregas_qc':
        return None

    alert_type, meta = _mp_qc_alert_type_meta(mediciones.get('alert_type') or registro.resultado)
    estacion_id = 0
    try:
        estacion_id = int(mediciones.get('estacion_id') or 0)
    except Exception:
        estacion_id = 0

    hoja_ruta_id = 0
    try:
        hoja_ruta_id = int(mediciones.get('hoja_ruta_id') or registro.hoja_ruta_id or 0)
    except Exception:
        hoja_ruta_id = 0

    reviewed_at = registro.creado_en.isoformat() if registro.creado_en else None
    return {
        'id': registro.id,
        'maquina_id': registro.maquina_id,
        'hoja_ruta_id': hoja_ruta_id,
        'estacion_id': estacion_id,
        'process_id': estacion_id,
        'process_name': mediciones.get('process_name') or '',
        'tipo': alert_type,
        'label': meta['label'],
        'color': meta['color'],
        'cantidad_revisada': int(registro.cantidad_inspeccionada or 0),
        'usuario': registro.usuario or '',
        'fecha': reviewed_at,
        'maquina_nombre': mediciones.get('maquina_nombre') or '',
        'hoja_nombre': mediciones.get('hoja_nombre') or '',
    }


def _group_entregas_qc_alerts_by_maquina_hoja(registros):
    grouped = {}
    for registro in registros or []:
        payload = _entregas_qc_alert_payload_from_registro(registro)
        if not payload:
            continue

        key = (int(payload.get('maquina_id') or 0), int(payload.get('hoja_ruta_id') or 0))
        bucket = grouped.setdefault(key, {'recent': [], 'latest_by_process': {}})
        if len(bucket['recent']) < 8:
            bucket['recent'].append(payload)

        estacion_id = int(payload.get('estacion_id') or 0)
        if estacion_id > 0 and estacion_id not in bucket['latest_by_process']:
            bucket['latest_by_process'][estacion_id] = payload
    return grouped


def _query_hojas_entregas_pendientes_estaciones():
    """Hojas entregas disponibles para asignar en Estaciones T.

    Incluye hojas historicas del modulo Entregas aunque tengan maquina_id
    residual, siempre que no esten activas/pausadas en cola de una maquina.

    Solo muestra hojas recientes (por defecto ultimos 15 dias segun fecha de
    captura); el rezago anterior ya no se asigna. Ajustable con la variable
    ESTACIONES_ENTREGAS_DIAS.
    """
    terminal_states = ('completada', 'cancelada')

    actively_on_machine = db.session.query(HojaRutaEntrega.id).filter(
        HojaRutaEntrega.maquina_id.isnot(None),
        HojaRutaEntrega.maquina_id != 0,
        func.lower(func.coalesce(HojaRutaEntrega.estado, 'activa')).in_(('activa', 'pausada')),
    )

    query = HojaRutaEntrega.query.filter(
        func.lower(func.coalesce(HojaRutaEntrega.estado, 'activa')).notin_(terminal_states),
    )

    try:
        dias_window = max(1, int(os.getenv('ESTACIONES_ENTREGAS_DIAS', '15') or '15'))
    except (TypeError, ValueError):
        dias_window = 15
    cutoff = datetime.utcnow() - timedelta(days=dias_window)
    query = query.filter(
        func.coalesce(HojaRutaEntrega.fecha_creacion, HojaRutaEntrega.fecha_salida) >= cutoff,
    )

    active_ids = [row[0] for row in actively_on_machine.all()]
    if active_ids:
        query = query.filter(~HojaRutaEntrega.id.in_(active_ids))

    return query.order_by(HojaRutaEntrega.fecha_creacion.asc()).all()


def _build_mp_time_estaciones_by_clave_id(clave_id):
    """Construye estaciones temporales para estimar tiempos de hojas MP."""
    procesos = ClaveProceso.query.filter_by(clave_id=clave_id).order_by(ClaveProceso.orden.asc()).all()
    estaciones = []
    for idx, cp in enumerate(procesos, start=1):
        estaciones.append(SimpleNamespace(
            orden=cp.orden or idx,
            t_e=cp.t_e or cp.tiempo_estimado or (cp.proceso.tiempo_estimado if getattr(cp, 'proceso', None) else None) or '',
            t_tct=cp.t_tct or '',
            t_tco=cp.t_tco or '',
            t_to=cp.t_to or '',
            nombre=(cp.operacion or (cp.proceso.operacion if getattr(cp, 'proceso', None) else '') or ''),
            operacion=(cp.operacion or (cp.proceso.operacion if getattr(cp, 'proceso', None) else '') or ''),
            centro_trabajo=cp.centro_trabajo or (cp.proceso.centro_trabajo if getattr(cp, 'proceso', None) else '') or '',
        ))
    return estaciones


def _compute_mp_total_time_preview_by_pn(pn_value, cantidad_piezas, fallback_total_time=None):
    """Calcula HH:MM:SS de hoja MP en memoria para mostrar en listados."""
    pn = (pn_value or '').strip()
    if not pn:
        return fallback_total_time

    clave = ClaveProducto.query.filter(func.upper(func.trim(ClaveProducto.clave)) == pn.upper()).first()
    if not clave:
        return fallback_total_time

    estaciones = _build_mp_time_estaciones_by_clave_id(clave.id)
    if not estaciones:
        return fallback_total_time

    hoja_tmp = SimpleNamespace(
        cantidad_piezas=max(1, int(cantidad_piezas or 1)),
        total_tiempo=fallback_total_time,
        dias_a_laborar=None,
        fecha_termino=None,
        fecha_salida=datetime.utcnow(),
    )
    _apply_hoja_time_plan(hoja_tmp, estaciones, maquina_tipo=None, fallback_total_time=fallback_total_time)
    return hoja_tmp.total_tiempo or fallback_total_time


def _recompute_mp_time_plan(hoja):
    """Recalcula tiempos persistidos en una hoja MP usando su clave/procesos."""
    if not hoja or not (hoja.pn or '').strip():
        return

    clave = ClaveProducto.query.filter(func.upper(func.trim(ClaveProducto.clave)) == hoja.pn.upper().strip()).first()
    if not clave:
        return

    estaciones = _build_mp_time_estaciones_by_clave_id(clave.id)
    if not estaciones:
        return

    _apply_hoja_time_plan(hoja, estaciones, maquina_tipo=None, fallback_total_time=hoja.total_tiempo)


def _sync_hoja_estado_with_checks(hoja, estaciones=None, now_dt=None):
    """Sincroniza el estado de hoja contra sus checks de procesos.
    Regla: completada solo si todas las estaciones estan completadas.
    """
    if (hoja.estado or '').strip().lower() == 'cancelada':
        return False

    if estaciones is None:
        estaciones = EstacionTrabajo.query.filter_by(hoja_ruta_id=hoja.id).all()

    total = len(estaciones)
    completadas = sum(1 for e in estaciones if (e.estado or '').lower() == 'completada')

    if total > 0 and completadas == total:
        new_estado = 'completada'
        new_fecha_termino = hoja.fecha_termino or (now_dt or datetime.utcnow())
    else:
        new_estado = 'activa'
        new_fecha_termino = None

    changed = (hoja.estado != new_estado) or (hoja.fecha_termino != new_fecha_termino)
    hoja.estado = new_estado
    hoja.fecha_termino = new_fecha_termino
    return changed


_HOJA_CARGAS_TABLE_READY = False


def _ensure_hoja_cargas_historial_table():
    global _HOJA_CARGAS_TABLE_READY

    if _HOJA_CARGAS_TABLE_READY:
        return True

    try:
        HojaRutaCargaPiezasHistorial.__table__.create(bind=db.engine, checkfirst=True)
        _HOJA_CARGAS_TABLE_READY = True
        return True
    except Exception as exc:
        logger.warning(f"No se pudo asegurar la tabla de historial de cargas de hojas: {exc}")
        try:
            db.session.rollback()
        except Exception:
            pass
        return False


_HOJA_IMPRESIONES_PARCIALES_TABLE_READY = False


def _ensure_hoja_impresiones_parciales_table():
    global _HOJA_IMPRESIONES_PARCIALES_TABLE_READY

    if _HOJA_IMPRESIONES_PARCIALES_TABLE_READY:
        return True

    try:
        HojaRutaImpresionParcial.__table__.create(bind=db.engine, checkfirst=True)
        _HOJA_IMPRESIONES_PARCIALES_TABLE_READY = True
        return True
    except Exception as exc:
        logger.warning(f"No se pudo asegurar la tabla de impresiones parciales de hojas: {exc}")
        try:
            db.session.rollback()
        except Exception:
            pass
        return False


_ALERTAS_BUZON_TABLE_READY = False


def _ensure_alertas_buzon_table():
    global _ALERTAS_BUZON_TABLE_READY

    if _ALERTAS_BUZON_TABLE_READY:
        return True

    try:
        AlertaBuzonGeneral.__table__.create(bind=db.engine, checkfirst=True)
    except Exception as exc:
        logger.warning(f"No se pudo crear la tabla de alertas del buzon: {exc}")
        try:
            db.session.rollback()
        except Exception:
            pass
        return False

    # Ensure nota_atencion column exists (migration for existing tables).
    try:
        with db.engine.connect() as conn:
            conn.execute(db.text(
                "ALTER TABLE alertas_buzon_general ADD COLUMN IF NOT EXISTS nota_atencion TEXT"
            ))
            conn.commit()
    except Exception as exc:
        logger.debug(f"_ensure nota_atencion column: {exc}")
        try:
            db.session.rollback()
        except Exception:
            pass

    _ALERTAS_BUZON_TABLE_READY = True
    return True


_ALMACEN_CAJAS_SURTIDO_TABLES_READY = False


def _ensure_almacen_cajas_surtido_tables():
    global _ALMACEN_CAJAS_SURTIDO_TABLES_READY

    if _ALMACEN_CAJAS_SURTIDO_TABLES_READY:
        return True

    try:
        AlmacenCajaSurtidoSesion.__table__.create(bind=db.engine, checkfirst=True)
        AlmacenCajaSurtidoCaja.__table__.create(bind=db.engine, checkfirst=True)
        AlmacenCajaSurtidoLecturaBascula.__table__.create(bind=db.engine, checkfirst=True)
        AlmacenCajaSurtidoItem.__table__.create(bind=db.engine, checkfirst=True)
        _ALMACEN_CAJAS_SURTIDO_TABLES_READY = True
        return True
    except Exception as exc:
        logger.warning(f"No se pudo asegurar tablas de cajas surtido: {exc}")
        try:
            db.session.rollback()
        except Exception:
            pass
        return False


_EMPAQUE_TABLES_READY = False
_EMPAQUE_SEGUIMIENTO_FAILS = {}  # ip -> (count, first_ts)
_EMPAQUE_ACCESS_ALPHABET = 'ABCDEFGHJKLMNPQRSTUVWXYZ23456789'
_EMPAQUE_PRECIOS_CACHE = {}  # customer_code -> (ts, payload)
_EMPAQUE_PRECIOS_TTL_SEC = 600
_EMPAQUE_PRECIO_MINIMO_REAL = 1.01  # ignora 0 y $1 de relleno; no es regla de venta
_ODOO_LISTAS_PRECIO = (
    ('Predeterminado', ('Predeterminado',)),
    ('PRECIO MAYOREO', ('PRECIO MAYOREO',)),
    ('PRECIO MECÁNICOS', ('PRECIO MECÁNICOS', 'PRECIO MECANICOS')),
    ('PRECIO DISTRIBUIDOR A', ('PRECIO DISTRIBUIDOR A',)),
    ('PRECIO DISTRIBUIDOR B', ('PRECIO DISTRIBUIDOR B',)),
    ('PRECIO SUCURSAL', ('PRECIO SUCURSAL',)),
)


def _ensure_empaque_tables():
    global _EMPAQUE_TABLES_READY
    if _EMPAQUE_TABLES_READY:
        return True
    try:
        EmpaqueCliente.__table__.create(bind=db.engine, checkfirst=True)
        EmpaquePedido.__table__.create(bind=db.engine, checkfirst=True)
        EmpaquePedidoItem.__table__.create(bind=db.engine, checkfirst=True)
        EmpaqueCaja.__table__.create(bind=db.engine, checkfirst=True)
        EmpaqueMovimiento.__table__.create(bind=db.engine, checkfirst=True)
        EmpaqueLineaProgreso.__table__.create(bind=db.engine, checkfirst=True)
        EmpaqueSeguimientoLog.__table__.create(bind=db.engine, checkfirst=True)
        # Columna access_code en logs si la tabla ya existía sin ella
        try:
            insp = inspect(db.engine)
            cols = {c['name'] for c in insp.get_columns('empaque_seguimiento_logs')}
            if 'access_code' not in cols:
                db.session.execute(text(
                    'ALTER TABLE empaque_seguimiento_logs '
                    'ADD COLUMN IF NOT EXISTS access_code VARCHAR(32)'
                ))
                db.session.commit()
        except Exception:
            try:
                db.session.rollback()
            except Exception:
                pass
        _EMPAQUE_TABLES_READY = True
        return True
    except Exception as exc:
        logger.warning(f'No se pudo asegurar tablas Empaque: {exc}')
        try:
            db.session.rollback()
        except Exception:
            pass
        return False


def _normalize_empaque_clave(value):
    return str(value or '').strip()


def _normalize_empaque_access_code(value):
    raw = str(value or '').strip().upper().replace(' ', '')
    if raw.startswith('EV-'):
        raw = raw[3:]
    raw = re.sub(r'[^A-Z0-9]', '', raw)
    return raw


def _generate_empaque_access_code(length=8):
    """Genera clave única de acceso (sin caracteres confusos I/O/0/1)."""
    for _ in range(40):
        body = ''.join(secrets.choice(_EMPAQUE_ACCESS_ALPHABET) for _ in range(length))
        code = f'EV-{body}'
        exists = EmpaqueCliente.query.filter(
            func.upper(EmpaqueCliente.access_code) == code
        ).first()
        if not exists:
            return code
    # Fallback ultra-único
    return f'EV-{secrets.token_hex(5).upper()}'


def _ensure_empaque_cliente(customer_code, customer_name=None):
    """
    Crea cliente Empaque con clave única si no existe.
    Si ya existe, actualiza nombre y conserva su access_code.
    """
    code = _normalize_empaque_clave(customer_code)
    if not code:
        return None, False

    cliente = EmpaqueCliente.query.filter(
        func.lower(EmpaqueCliente.customer_code) == code.lower()
    ).first()
    created = False
    if not cliente:
        cliente = EmpaqueCliente(
            customer_code=code[:80],
            customer_name=(str(customer_name or '').strip()[:255] or None),
            access_code=_generate_empaque_access_code(),
            activo=True,
        )
        db.session.add(cliente)
        created = True
    else:
        name = str(customer_name or '').strip()[:255]
        if name:
            cliente.customer_name = name
        if not cliente.access_code:
            cliente.access_code = _generate_empaque_access_code()
        cliente.updated_at = datetime.utcnow()
    return cliente, created


def _reconcile_empaque_clientes_from_pedidos():
    """
    Automático: por cada customer_code en pedidos espejo, asegura ficha + clave.
    No depende de un botón humano.
    """
    pairs = (
        db.session.query(EmpaquePedido.customer_code, EmpaquePedido.customer_name)
        .filter(EmpaquePedido.customer_code.isnot(None))
        .distinct()
        .all()
    )
    created = 0
    touched = 0
    for code, name in pairs:
        cliente, was_created = _ensure_empaque_cliente(code, name)
        if not cliente:
            continue
        if was_created:
            created += 1
        else:
            touched += 1
    return {'created': created, 'touched': touched, 'total_sources': len(pairs)}


def _find_empaque_cliente_by_access_code(access_code_raw):
    """Resuelve cliente solo por clave de acceso (nunca por customer_code externo)."""
    body = _normalize_empaque_access_code(access_code_raw)
    if not body:
        return None
    candidates = [body, f'EV-{body}']
    for cand in candidates:
        cliente = EmpaqueCliente.query.filter(
            func.upper(EmpaqueCliente.access_code) == cand.upper()
        ).first()
        if cliente:
            return cliente
    return None


def _upsert_empaque_data(payload):
    """Reemplaza espejo por pedido (snapshot) desde agente Empaque360."""
    pedidos = payload.get('pedidos') or []
    stats = {
        'pedidos_upserted': 0,
        'items_upserted': 0,
        'cajas_upserted': 0,
        'movimientos_upserted': 0,
        'progreso_upserted': 0,
        'clientes_creados': 0,
    }

    for row in pedidos:
        if not isinstance(row, dict):
            continue
        order_number = str(row.get('ExternalOrderNumber') or row.get('external_order_number') or '').strip()
        customer_code = _normalize_empaque_clave(row.get('CustomerCode') or row.get('customer_code'))
        if not order_number or not customer_code:
            continue

        customer_name = str(row.get('CustomerName') or row.get('customer_name') or '').strip()[:255] or None
        _cliente, created = _ensure_empaque_cliente(customer_code, customer_name)
        if created:
            stats['clientes_creados'] += 1

        pedido = EmpaquePedido.query.filter_by(external_order_number=order_number).first()
        if not pedido:
            pedido = EmpaquePedido(external_order_number=order_number)
            db.session.add(pedido)

        pedido.customer_code = customer_code[:80]
        pedido.customer_name = customer_name
        pedido.order_date_utc = _parse_datetime(row.get('OrderDateUtc') or row.get('order_date_utc'))
        source_id = row.get('SalesOrderId') or row.get('source_sales_order_id')
        try:
            pedido.source_sales_order_id = int(source_id) if source_id is not None else None
        except Exception:
            pedido.source_sales_order_id = None
        pedido.last_activity_utc = _parse_datetime(row.get('LastActivityUtc') or row.get('last_activity_utc'))
        pedido.synced_at = datetime.utcnow()
        db.session.flush()

        EmpaquePedidoItem.query.filter_by(pedido_id=pedido.id).delete(synchronize_session=False)
        EmpaqueCaja.query.filter_by(pedido_id=pedido.id).delete(synchronize_session=False)
        EmpaqueMovimiento.query.filter_by(pedido_id=pedido.id).delete(synchronize_session=False)
        EmpaqueLineaProgreso.query.filter_by(pedido_id=pedido.id).delete(synchronize_session=False)

        for item in (row.get('items') or []):
            if not isinstance(item, dict):
                continue
            product_code = str(item.get('ProductCode') or item.get('product_code') or '').strip()
            if not product_code:
                continue
            db.session.add(EmpaquePedidoItem(
                pedido_id=pedido.id,
                external_order_number=order_number,
                source_line_number=int(item.get('SourceLineNumber') or item.get('source_line_number') or 0),
                product_code=product_code[:80],
                product_name=str(item.get('ProductName') or item.get('product_name') or '').strip()[:255] or None,
                quantity_ordered=_to_float(item.get('QuantityOrdered') or item.get('quantity_ordered')) or 0.0,
                unit_weight_kg=_to_float(item.get('UnitWeightFromContpaqiKg') or item.get('unit_weight_kg')) or 0.0,
            ))
            stats['items_upserted'] += 1

        for caja in (row.get('cajas') or row.get('boxes') or []):
            if not isinstance(caja, dict):
                continue
            box_code = str(caja.get('BoxCode') or caja.get('box_code') or '').strip()
            if not box_code:
                continue
            source_box_id = caja.get('PackingBoxId') or caja.get('source_packing_box_id')
            try:
                source_box_id = int(source_box_id) if source_box_id is not None else None
            except Exception:
                source_box_id = None
            db.session.add(EmpaqueCaja(
                pedido_id=pedido.id,
                external_order_number=order_number,
                box_code=box_code[:50],
                source_packing_box_id=source_box_id,
                status=int(caja.get('Status') or caja.get('status') or 1),
                max_weight_kg=_to_float(caja.get('MaxWeightKg') or caja.get('max_weight_kg')) or 0.0,
                current_real_weight_kg=_to_float(
                    caja.get('CurrentRealWeightKg') or caja.get('current_real_weight_kg')
                ) or 0.0,
                opened_at_utc=_parse_datetime(caja.get('OpenedAtUtc') or caja.get('opened_at_utc')),
                closed_at_utc=_parse_datetime(caja.get('ClosedAtUtc') or caja.get('closed_at_utc')),
            ))
            stats['cajas_upserted'] += 1

        # Deduplicate by source id; clear global unique collisions before insert
        seen_mov_ids = set()
        mov_rows = []
        for mov in (row.get('movimientos') or row.get('movements') or []):
            if not isinstance(mov, dict):
                continue
            source_mov_id = mov.get('BoxMovementId') or mov.get('source_box_movement_id')
            try:
                source_mov_id = int(source_mov_id or 0)
            except Exception:
                source_mov_id = 0
            box_code = str(mov.get('BoxCode') or mov.get('box_code') or '').strip()
            product_code = str(mov.get('ProductCode') or mov.get('product_code') or '').strip()
            if source_mov_id <= 0 or not box_code or not product_code:
                continue
            if source_mov_id in seen_mov_ids:
                continue
            seen_mov_ids.add(source_mov_id)
            mov_rows.append((source_mov_id, mov, box_code, product_code))
        if seen_mov_ids:
            EmpaqueMovimiento.query.filter(
                EmpaqueMovimiento.source_box_movement_id.in_(list(seen_mov_ids))
            ).delete(synchronize_session=False)
        for source_mov_id, mov, box_code, product_code in mov_rows:
            db.session.add(EmpaqueMovimiento(
                pedido_id=pedido.id,
                external_order_number=order_number,
                source_box_movement_id=source_mov_id,
                box_code=box_code[:50],
                product_code=product_code[:80],
                product_name=str(mov.get('ProductName') or mov.get('product_name') or '').strip()[:255] or None,
                quantity_captured=_to_float(mov.get('QuantityCaptured') or mov.get('quantity_captured')) or 0.0,
                unit_weight_kg=_to_float(mov.get('UnitWeightFromContpaqiKg') or mov.get('unit_weight_kg')) or 0.0,
                real_weight_kg=_to_float(mov.get('RealMovementWeightKg') or mov.get('real_weight_kg')) or 0.0,
                observed_weight_per_piece_kg=_to_float(
                    mov.get('ObservedWeightPerPieceKg') or mov.get('observed_weight_per_piece_kg')
                ) or 0.0,
                operator_name=str(mov.get('OperatorName') or mov.get('operator_name') or '').strip()[:120] or None,
                captured_at_utc=_parse_datetime(mov.get('CapturedAtUtc') or mov.get('captured_at_utc')),
            ))
            stats['movimientos_upserted'] += 1

        for prog in (row.get('progreso') or row.get('progress') or []):
            if not isinstance(prog, dict):
                continue
            product_code = str(prog.get('ProductCode') or prog.get('product_code') or '').strip()
            if not product_code:
                continue
            db.session.add(EmpaqueLineaProgreso(
                pedido_id=pedido.id,
                external_order_number=order_number,
                product_code=product_code[:80],
                source_line_number=int(prog.get('SourceLineNumber') or prog.get('source_line_number') or 0),
                quantity_ordered=_to_float(prog.get('QuantityOrdered') or prog.get('quantity_ordered')) or 0.0,
                quantity_packed=_to_float(prog.get('QuantityPacked') or prog.get('quantity_packed')) or 0.0,
                updated_at_utc=_parse_datetime(prog.get('UpdatedAtUtc') or prog.get('updated_at_utc')),
            ))
            stats['progreso_upserted'] += 1

        db.session.flush()

        # Derivar cajas SUELTO / faltantes desde movimientos si no vinieron en PackingBoxes
        existing_box_codes = {
            c.box_code for c in EmpaqueCaja.query.filter_by(pedido_id=pedido.id).all()
        }
        movs = EmpaqueMovimiento.query.filter_by(pedido_id=pedido.id).all()
        weights_by_box = {}
        for m in movs:
            weights_by_box.setdefault(m.box_code, 0.0)
            weights_by_box[m.box_code] += float(m.real_weight_kg or 0.0)
            if not pedido.last_activity_utc or (m.captured_at_utc and m.captured_at_utc > pedido.last_activity_utc):
                pedido.last_activity_utc = m.captured_at_utc
        for box_code, total_w in weights_by_box.items():
            if box_code in existing_box_codes:
                continue
            db.session.add(EmpaqueCaja(
                pedido_id=pedido.id,
                external_order_number=order_number,
                box_code=box_code[:50],
                source_packing_box_id=None,
                status=2,
                max_weight_kg=0.0,
                current_real_weight_kg=total_w,
                opened_at_utc=None,
                closed_at_utc=pedido.last_activity_utc,
            ))
            stats['cajas_upserted'] += 1

        stats['pedidos_upserted'] += 1

    return stats


def _build_empaque_cliente_view(customer_code, *, access_code=None, customer_name=None):
    """Arma la vista del portal SOLO para el customer_code resuelto por clave de acceso."""
    code = _normalize_empaque_clave(customer_code)
    if not code:
        return None

    pedidos = (
        EmpaquePedido.query
        .filter(func.lower(EmpaquePedido.customer_code) == code.lower())
        .order_by(
            EmpaquePedido.last_activity_utc.desc().nullslast(),
            EmpaquePedido.order_date_utc.desc().nullslast(),
            EmpaquePedido.id.desc(),
        )
        .all()
    )
    # Cliente válido puede no tener pedidos aún: devolver vista vacía (no None).
    # None se reserva para clave inválida.
    pedido_ids = [p.id for p in pedidos]
    if pedido_ids:
        items = EmpaquePedidoItem.query.filter(EmpaquePedidoItem.pedido_id.in_(pedido_ids)).all()
        cajas = EmpaqueCaja.query.filter(EmpaqueCaja.pedido_id.in_(pedido_ids)).all()
        movs = (
            EmpaqueMovimiento.query
            .filter(EmpaqueMovimiento.pedido_id.in_(pedido_ids))
            .order_by(EmpaqueMovimiento.captured_at_utc.asc().nullslast(), EmpaqueMovimiento.id.asc())
            .all()
        )
        progreso = EmpaqueLineaProgreso.query.filter(EmpaqueLineaProgreso.pedido_id.in_(pedido_ids)).all()
    else:
        items, cajas, movs, progreso = [], [], [], []

    items_by = {}
    for it in items:
        items_by.setdefault(it.pedido_id, []).append(it)
    cajas_by = {}
    for c in cajas:
        cajas_by.setdefault(c.pedido_id, []).append(c)
    movs_by = {}
    for m in movs:
        movs_by.setdefault(m.pedido_id, []).append(m)
    prog_by = {}
    for p in progreso:
        prog_by.setdefault(p.pedido_id, []).append(p)

    view_pedidos = []
    for pedido in pedidos:
        p_movs = movs_by.get(pedido.id, [])
        p_cajas = sorted(
            cajas_by.get(pedido.id, []),
            key=lambda c: (0 if (c.box_code or '').upper() == 'SUELTO' else 1, c.box_code or '', c.id),
        )
        movs_por_caja = {}
        for m in p_movs:
            movs_por_caja.setdefault(m.box_code, []).append(m)

        cajas_view = []
        for caja in p_cajas:
            lineas = movs_por_caja.get(caja.box_code, [])
            peso_real = float(caja.current_real_weight_kg or 0.0)
            if peso_real <= 0:
                peso_real = sum(float(m.real_weight_kg or 0.0) for m in lineas)
            piezas = sum(float(m.quantity_captured or 0.0) for m in lineas)
            cajas_view.append({
                'box_code': caja.box_code,
                'status': int(caja.status or 1),
                'status_label': 'Cerrada' if int(caja.status or 0) == 2 else 'En armado',
                'es_suelto': (caja.box_code or '').upper() == 'SUELTO',
                'peso_real_kg': round(peso_real, 3),
                'piezas': piezas,
                'closed_at_utc': caja.closed_at_utc,
                'lineas': [{
                    'product_code': m.product_code,
                    'product_name': m.product_name or m.product_code,
                    'quantity': float(m.quantity_captured or 0.0),
                    'peso_real_kg': round(float(m.real_weight_kg or 0.0), 3),
                    'captured_at_utc': m.captured_at_utc,
                } for m in lineas],
            })

        # Movimientos en cajas no listadas
        listed = {c['box_code'] for c in cajas_view}
        for box_code, lineas in movs_por_caja.items():
            if box_code in listed:
                continue
            cajas_view.append({
                'box_code': box_code,
                'status': 2,
                'status_label': 'Cerrada',
                'es_suelto': (box_code or '').upper() == 'SUELTO',
                'peso_real_kg': round(sum(float(m.real_weight_kg or 0.0) for m in lineas), 3),
                'piezas': sum(float(m.quantity_captured or 0.0) for m in lineas),
                'closed_at_utc': max((m.captured_at_utc for m in lineas if m.captured_at_utc), default=None),
                'lineas': [{
                    'product_code': m.product_code,
                    'product_name': m.product_name or m.product_code,
                    'quantity': float(m.quantity_captured or 0.0),
                    'peso_real_kg': round(float(m.real_weight_kg or 0.0), 3),
                    'captured_at_utc': m.captured_at_utc,
                } for m in lineas],
            })

        p_items = sorted(items_by.get(pedido.id, []), key=lambda x: (x.source_line_number, x.id))
        p_prog = { (pr.product_code, pr.source_line_number): pr for pr in prog_by.get(pedido.id, []) }
        partidas = []
        for it in p_items:
            pr = p_prog.get((it.product_code, it.source_line_number))
            ordered = float(it.quantity_ordered or 0.0)
            packed = float(pr.quantity_packed or 0.0) if pr else 0.0
            partidas.append({
                'product_code': it.product_code,
                'product_name': it.product_name or it.product_code,
                'quantity_ordered': ordered,
                'quantity_packed': packed,
                'quantity_pending': max(0.0, ordered - packed),
            })

        total_peso = sum(c['peso_real_kg'] for c in cajas_view)
        total_cajas = sum(1 for c in cajas_view if not c['es_suelto'])
        view_pedidos.append({
            'external_order_number': pedido.external_order_number,
            'order_date_utc': pedido.order_date_utc,
            'last_activity_utc': pedido.last_activity_utc,
            'customer_name': pedido.customer_name,
            'partidas': partidas,
            'cajas': cajas_view,
            'total_peso_kg': round(total_peso, 3),
            'total_cajas': total_cajas,
            'tiene_suelto': any(c['es_suelto'] for c in cajas_view),
        })

    return {
        'access_code': access_code,
        'customer_name': customer_name or (pedidos[0].customer_name if pedidos else None) or code,
        'customer_code': code,
        'pedidos': view_pedidos,
        'total_pedidos': len(view_pedidos),
    }


def _odoo_search_read_all(client, model, domain, fields, *, batch=400):
    rows = []
    offset = 0
    while True:
        chunk = client.search_read(model, domain, fields, limit=batch, offset=offset)
        if not chunk:
            break
        rows.extend(chunk)
        if len(chunk) < batch:
            break
        offset += batch
        if offset > 20000:
            break
    return rows


def _m2o_id_empaque(value):
    if isinstance(value, (list, tuple)) and value:
        try:
            return int(value[0])
        except (TypeError, ValueError):
            return None
    if isinstance(value, int):
        return value
    return None


def _m2o_name_empaque(value):
    if isinstance(value, (list, tuple)) and len(value) >= 2:
        return str(value[1] or '').strip()
    return ''


def _fetch_odoo_precios_cliente(customer_code, customer_name=None):
    """Precios del cliente: activos con las 6 reglas internas; al cliente solo se muestra su precio."""
    code = _normalize_empaque_clave(customer_code) or ''
    cache_key = f'{code.lower()}::v2'
    now = time()
    cached = _EMPAQUE_PRECIOS_CACHE.get(cache_key)
    if cached and (now - cached[0]) < _EMPAQUE_PRECIOS_TTL_SEC:
        return cached[1]

    empty = {
        'ok': False,
        'error': '',
        'lista_nombre': '',
        'productos': [],
    }
    if not OdooClient.is_configured():
        empty['error'] = 'Los precios no están disponibles por el momento.'
        return empty

    try:
        client = OdooClient.from_env()
        partner = None
        if code.isdigit():
            rows = client.search_read(
                'res.partner',
                [['id', '=', int(code)]],
                ['id', 'name', 'property_product_pricelist'],
                limit=1,
            )
            partner = rows[0] if rows else None
        if not partner and customer_name:
            rows = client.search_read(
                'res.partner',
                [['name', '=', customer_name.strip()]],
                ['id', 'name', 'property_product_pricelist'],
                limit=1,
            )
            partner = rows[0] if rows else None
        if not partner:
            empty['error'] = 'Los precios no están disponibles por el momento.'
            _EMPAQUE_PRECIOS_CACHE[cache_key] = (now, empty)
            return empty

        lista_id = _m2o_id_empaque(partner.get('property_product_pricelist'))
        lista_nombre = _m2o_name_empaque(partner.get('property_product_pricelist')) or ''
        if not lista_id:
            empty['error'] = 'Los precios no están disponibles por el momento.'
            _EMPAQUE_PRECIOS_CACHE[cache_key] = (now, empty)
            return empty

        pl_ids = []
        pl_id_by_canon = {}
        for canon, names in _ODOO_LISTAS_PRECIO:
            found = None
            for name in names:
                rows = client.search_read(
                    'product.pricelist',
                    [['name', '=', name]],
                    ['id', 'name'],
                    limit=1,
                )
                if rows:
                    found = rows[0]
                    break
            if not found:
                logger.warning('[EMPAQUE] Falta lista interna de precios: %s', canon)
                empty['error'] = 'Los precios no están disponibles por el momento.'
                _EMPAQUE_PRECIOS_CACHE[cache_key] = (now, empty)
                return empty
            pl_ids.append(int(found['id']))
            pl_id_by_canon[canon] = int(found['id'])

        items = _odoo_search_read_all(
            client,
            'product.pricelist.item',
            [
                ['pricelist_id', 'in', pl_ids],
                ['applied_on', '=', '1_product'],
                ['product_tmpl_id', '!=', False],
            ],
            ['product_tmpl_id', 'pricelist_id', 'fixed_price'],
        )
        by_tmpl = {}
        for item in items:
            tmpl_id = _m2o_id_empaque(item.get('product_tmpl_id'))
            pl_id = _m2o_id_empaque(item.get('pricelist_id'))
            if not tmpl_id or not pl_id:
                continue
            try:
                price = float(item.get('fixed_price') or 0)
            except (TypeError, ValueError):
                price = 0.0
            by_tmpl.setdefault(tmpl_id, {})[pl_id] = price

        complete_ids = [
            tmpl_id
            for tmpl_id, prices in by_tmpl.items()
            if all(
                pl_id in prices and float(prices.get(pl_id) or 0) >= _EMPAQUE_PRECIO_MINIMO_REAL
                for pl_id in pl_ids
            )
        ]
        extra_prices = {}
        if lista_id not in pl_ids and complete_ids:
            extra_items = _odoo_search_read_all(
                client,
                'product.pricelist.item',
                [
                    ['pricelist_id', '=', lista_id],
                    ['applied_on', '=', '1_product'],
                    ['product_tmpl_id', 'in', complete_ids],
                ],
                ['product_tmpl_id', 'fixed_price'],
            )
            for item in extra_items:
                tmpl_id = _m2o_id_empaque(item.get('product_tmpl_id'))
                if not tmpl_id:
                    continue
                try:
                    extra_prices[tmpl_id] = float(item.get('fixed_price') or 0)
                except (TypeError, ValueError):
                    continue

        productos = []
        for i in range(0, len(complete_ids), 200):
            chunk_ids = complete_ids[i:i + 200]
            templates = client.search_read(
                'product.template',
                [['id', 'in', chunk_ids], ['sale_ok', '=', True], ['active', '=', True]],
                ['id', 'default_code', 'name', 'sale_ok', 'active'],
            )
            for tmpl in templates:
                tmpl_id = int(tmpl['id'])
                prices = by_tmpl.get(tmpl_id) or {}
                client_price = extra_prices.get(tmpl_id, prices.get(lista_id))
                if client_price is None:
                    continue
                try:
                    client_price = float(client_price)
                except (TypeError, ValueError):
                    continue
                if client_price < _EMPAQUE_PRECIO_MINIMO_REAL:
                    continue
                code_txt = str(tmpl.get('default_code') or '').strip()
                name_txt = str(tmpl.get('name') or '').strip() or code_txt
                productos.append({
                    'code': code_txt,
                    'name': name_txt,
                    'precio': round(client_price, 2),
                })

        productos.sort(key=lambda r: ((r['code'] or r['name']).upper(), r['name'].upper()))
        payload = {
            'ok': True,
            'error': '',
            'lista_nombre': lista_nombre,
            'productos': productos,
        }
        _EMPAQUE_PRECIOS_CACHE[cache_key] = (now, payload)
        return payload
    except OdooError as exc:
        empty['error'] = 'Los precios no están disponibles por el momento.'
        logger.warning('[EMPAQUE] Precios Odoo: %s', exc)
        return empty
    except Exception as exc:
        empty['error'] = 'Los precios no están disponibles por el momento.'
        logger.exception('[EMPAQUE] Precios Odoo inesperado: %s', exc)
        return empty


def _empaque_seguimiento_rate_limited(ip):
    now = time()
    count, first = _EMPAQUE_SEGUIMIENTO_FAILS.get(ip, (0, now))
    if now - first > 900:
        _EMPAQUE_SEGUIMIENTO_FAILS.pop(ip, None)
        return False
    return count >= 12


def _empaque_seguimiento_register_fail(ip):
    now = time()
    count, first = _EMPAQUE_SEGUIMIENTO_FAILS.get(ip, (0, now))
    if now - first > 900:
        count, first = 0, now
    _EMPAQUE_SEGUIMIENTO_FAILS[ip] = (count + 1, first)


def _empaque_seguimiento_clear_fail(ip):
    _EMPAQUE_SEGUIMIENTO_FAILS.pop(ip, None)


def _write_empaque_seguimiento_log(customer_code, resultado, access_code=None):
    try:
        if not _ensure_empaque_tables():
            return
        db.session.add(EmpaqueSeguimientoLog(
            customer_code=(customer_code or '')[:80] or None,
            access_code=(access_code or '')[:32] or None,
            ip_cliente=(request.headers.get('X-Forwarded-For') or request.remote_addr or '')[:64],
            user_agent=(request.headers.get('User-Agent') or '')[:1000],
            resultado=(resultado or 'invalido')[:40],
        ))
        db.session.commit()
    except Exception:
        try:
            db.session.rollback()
        except Exception:
            pass


def _serialize_almacen_caja_surtido_sesion(sesion):
    cajas = (
        AlmacenCajaSurtidoCaja.query
        .filter_by(sesion_id=sesion.id)
        .order_by(AlmacenCajaSurtidoCaja.numero_caja.asc(), AlmacenCajaSurtidoCaja.id.asc())
        .all()
    )
    caja_ids = [c.id for c in cajas]
    items = (
        AlmacenCajaSurtidoItem.query
        .filter(AlmacenCajaSurtidoItem.caja_id.in_(caja_ids))
        .order_by(AlmacenCajaSurtidoItem.fecha_creacion.asc(), AlmacenCajaSurtidoItem.id.asc())
        .all()
    ) if caja_ids else []

    items_por_caja = {}
    for it in items:
        items_por_caja.setdefault(it.caja_id, []).append(it.to_dict())

    cajas_data = []
    for caja in cajas:
        caja_dict = caja.to_dict()
        caja_dict['items'] = items_por_caja.get(caja.id, [])
        cajas_data.append(caja_dict)

    lectura = (
        AlmacenCajaSurtidoLecturaBascula.query
        .filter_by(sesion_id=sesion.id)
        .order_by(AlmacenCajaSurtidoLecturaBascula.id.desc())
        .first()
    )

    abierta = next((c for c in cajas_data if c.get('estado') == 'abierta'), None)

    data = sesion.to_dict()
    data.update({
        'cajas': cajas_data,
        'caja_abierta': abierta,
        'lectura_ultima': lectura.to_dict() if lectura else None,
    })
    return data


def _sync_cajas_surtido_to_sheets(sesion):
    if not GOOGLE_SHEETS_SYNC_ENABLED:
        return False, 'disabled'
    if not GOOGLE_SHEETS_WEBHOOK_URL:
        return False, 'missing_webhook_url'

    payload = {
        'event': 'almacen_caja_surtido_finalizada',
        'timestamp': datetime.utcnow().isoformat(),
        'sesion': _serialize_almacen_caja_surtido_sesion(sesion),
    }
    headers = {'Content-Type': 'application/json'}
    if GOOGLE_SHEETS_WEBHOOK_TOKEN:
        headers['Authorization'] = f'Bearer {GOOGLE_SHEETS_WEBHOOK_TOKEN}'

    try:
        res = requests.post(
            GOOGLE_SHEETS_WEBHOOK_URL,
            json=payload,
            headers=headers,
            timeout=GOOGLE_SHEETS_TIMEOUT_SECONDS,
        )
        if 200 <= res.status_code < 300:
            return True, 'ok'
        logger.warning(f"[SHEETS_SYNC] cajas surtido webhook status={res.status_code} body={res.text[:500]}")
        return False, f'http_{res.status_code}'
    except Exception as exc:
        logger.warning(f"[SHEETS_SYNC] cajas surtido exception: {exc}")
        return False, 'exception'


def _serialize_alerta_buzon(item):
    return {
        'id': item.id,
        'evento_clave': item.evento_clave,
        'origen': item.origen,
        'tipo': item.tipo,
        'titulo': item.titulo,
        'mensaje': item.mensaje,
        'maquina_id': item.maquina_id,
        'hoja_id': item.hoja_id,
        'estacion_id': item.estacion_id,
        'atendida': bool(item.atendida),
        'atendida_por': item.atendida_por,
        'atendida_at': item.atendida_at.isoformat() if item.atendida_at else None,
        'nota_atencion': item.nota_atencion or '',
        'created_at': item.created_at.isoformat() if item.created_at else None,
    }


def _crear_alerta_buzon(
    *,
    evento_clave,
    origen,
    tipo,
    titulo,
    mensaje='',
    maquina_id=None,
    hoja_id=None,
    estacion_id=None,
    commit=True,
):
    if not evento_clave:
        return None
    if not _ensure_alertas_buzon_table():
        return None

    try:
        existente = AlertaBuzonGeneral.query.filter_by(evento_clave=evento_clave).first()
        if existente:
            return existente

        alerta = AlertaBuzonGeneral(
            evento_clave=evento_clave,
            origen=origen,
            tipo=tipo,
            titulo=titulo,
            mensaje=mensaje,
            maquina_id=maquina_id,
            hoja_id=hoja_id,
            estacion_id=estacion_id,
            atendida=False,
        )
        db.session.add(alerta)
        if commit:
            db.session.commit()
            _send_alerta_whatsapp_if_enabled(alerta)
            _send_alerta_telegram_if_enabled(alerta)
        return alerta
    except Exception as exc:
        db.session.rollback()
        logger.error(f"Error creando alerta de buzon: {exc}", exc_info=True)
        return None

def _current_username_for_audit(user=None):
    username = getattr(user, 'username', None) if user else None
    if username:
        return username
    return (session.get('user') or 'sistema').strip() or 'sistema'


def _serialize_hoja_carga_historial(item):
    return {
        'id': item.id,
        'cantidad_anterior': item.cantidad_anterior,
        'cantidad_cambio': item.cantidad_cambio,
        'cantidad_nueva': item.cantidad_nueva,
        'tipo_movimiento': item.tipo_movimiento,
        'usuario': item.usuario,
        'fecha_creacion': item.fecha_creacion.isoformat() if item.fecha_creacion else None,
    }


def _registrar_carga_piezas_hoja(hoja, cantidad_anterior, cantidad_nueva, usuario, origen='ajuste'):
    cantidad_anterior = int(cantidad_anterior or 0)
    cantidad_nueva = int(cantidad_nueva or 0)

    if cantidad_nueva == cantidad_anterior:
        return None
    if not _ensure_hoja_cargas_historial_table():
        return None

    delta = cantidad_nueva - cantidad_anterior
    if origen == 'creacion':
        tipo_movimiento = 'inicial'
    elif delta > 0:
        tipo_movimiento = 'incremento'
    elif delta < 0:
        tipo_movimiento = 'decremento'
    else:
        tipo_movimiento = 'ajuste'

    movimiento = HojaRutaCargaPiezasHistorial(
        hoja_ruta_id=hoja.id,
        cantidad_anterior=cantidad_anterior,
        cantidad_cambio=delta,
        cantidad_nueva=cantidad_nueva,
        tipo_movimiento=tipo_movimiento,
        usuario=usuario,
    )
    db.session.add(movimiento)
    return movimiento

# Simple in-memory login rate limiter
# Keys: by IP address. Tracks [attempt_count, first_attempt_ts, locked_until_ts]
FAILED_LOGINS = {}
# Configurable via env
MAX_LOGIN_ATTEMPTS = int(os.getenv('MAX_LOGIN_ATTEMPTS', '5'))
LOCKOUT_SECONDS = int(os.getenv('LOCKOUT_SECONDS', '300'))  # default 5 minutes

# API key for plant sync
SYNC_API_KEY = os.getenv('SYNC_API_KEY', '').strip()

# CONTPAQi sync settings
CONTPAQ_SYNC_ENABLED = os.getenv('CONTPAQ_SYNC_ENABLED', '0').strip().lower() in ('1', 'true', 'yes', 'on')
CONTPAQ_SYNC_INTERVAL_MINUTES = max(5, int(os.getenv('CONTPAQ_SYNC_INTERVAL_MINUTES', '60') or 60))
CONTPAQ_SYNC_STARTUP_DELAY_SECONDS = max(5, int(os.getenv('CONTPAQ_SYNC_STARTUP_DELAY_SECONDS', '30') or 30))
CONTPAQ_CUSTOMER_NAME = os.getenv('CONTPAQ_CUSTOMER_NAME', 'RUTH VERDUZCO SANTOS').strip()
CONTPAQ_START_DATE = os.getenv('CONTPAQ_START_DATE', '2025-01-01').strip()
CONTPAQ_MAQUINARIA_START_DATE = os.getenv('CONTPAQ_MAQUINARIA_START_DATE', '2025-06-01').strip()
CONTPAQ_MM_LOOKBACK_DAYS = max(14, int(os.getenv('CONTPAQ_MM_LOOKBACK_DAYS', '84') or 84))
CONTPAQ_MM_MIN_WEEKS = max(1.0, float(os.getenv('CONTPAQ_MM_MIN_WEEKS', '2.0') or 2.0))
CONTPAQ_MM_MAX_WEEKS = max(CONTPAQ_MM_MIN_WEEKS, float(os.getenv('CONTPAQ_MM_MAX_WEEKS', '6.0') or 6.0))

# Odoo (Maquinaria y Ensamble: pedidos de venta y ordenes de trabajo)
ODOO_SYNC_ENABLED = os.getenv('ODOO_SYNC_ENABLED', '0').strip().lower() in ('1', 'true', 'yes', 'on')
ODOO_SYNC_INTERVAL_MINUTES = max(2, int(os.getenv('ODOO_SYNC_INTERVAL_MINUTES', '15') or 15))
ODOO_SYNC_STARTUP_DELAY_SECONDS = max(5, int(os.getenv('ODOO_SYNC_STARTUP_DELAY_SECONDS', '40') or 40))
_ODOO_SYNC_THREAD = None
_ODOO_SYNC_STOP = threading.Event()
_ODOO_SCHEDULER_INIT = False
_ODOO_SYNC_LOCK = threading.Lock()

_CONTPAQ_SYNC_LOCK = threading.Lock()
_CONTPAQ_SYNC_THREAD = None
_CONTPAQ_SYNC_STOP = threading.Event()
_CONTPAQ_SCHEDULER_INIT = False

load_dotenv()

app = Flask(__name__)

# Permite reflejar cambios de templates sin reinicio manual (util para despliegues por git pull).
templates_auto_reload = os.getenv('TEMPLATES_AUTO_RELOAD', '0').strip().lower() in ('1', 'true', 'yes', 'on')
app.config['TEMPLATES_AUTO_RELOAD'] = templates_auto_reload
app.jinja_env.auto_reload = templates_auto_reload

# Configuración para carga de archivos
UPLOAD_FOLDER = 'uploads/productos'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
MAX_FILE_SIZE = 100 * 1024 * 1024  # 100MB

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = MAX_FILE_SIZE

# Configuración segura
app.secret_key = os.getenv('SECRET_KEY', secrets.token_hex(16))

# Configuración de BD
DATABASE_URL = os.getenv('DATABASE_URL', 'postgresql://catalogo_user:catalogo_pass@localhost:5432/catalogo_db')
app.config['SQLALCHEMY_DATABASE_URI'] = DATABASE_URL
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)

# Nota: la creación de tablas se realiza con `create_db.py` para evitar
# colisiones al arrancar múltiples workers (ej. Gunicorn). Ejecutar:
#   python create_db.py
# una vez antes de arrancar la app en producción.

# Inicializar AuthManager con BD
auth_manager = AuthManager(db=db)

# Inicializar EmailManager para notificaciones
email_manager = EmailManager()


# Helper: check if session user is admin (LEGACY - kept for backwards compatibility)
def is_admin_user():
    username = session.get('user')
    if not username:
        return False
    try:
        user = Usuario.query.filter_by(username=username, activo=True).first()
        return bool(user and user.es_admin)
    except Exception:
        return False


def is_root_user():
    """Return True only if logged-in user is 'root' (exact match)."""
    username = session.get('user')
    return username == 'root'


def _require_sync_key():
    if not SYNC_API_KEY:
        return False, (jsonify({'error': 'SYNC_API_KEY no configurado'}), 500)
    provided = request.headers.get('X-API-KEY', '')
    if not provided or not secrets.compare_digest(provided, SYNC_API_KEY):
        return False, (jsonify({'error': 'No autorizado'}), 401)
    return True, None


def _parse_date(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str):
        try:
            return datetime.fromisoformat(value.replace('Z', '+00:00')).date()
        except ValueError:
            for fmt in ('%Y-%m-%d', '%Y-%m-%d %H:%M:%S'):
                try:
                    return datetime.strptime(value, fmt).date()
                except ValueError:
                    continue
    return None


def _parse_datetime(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        raw = value.strip()
        if not raw:
            return None
        for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%Y/%m/%d %H:%M:%S', '%Y/%m/%d'):
            try:
                return datetime.strptime(raw, fmt)
            except ValueError:
                continue
        try:
            return datetime.fromisoformat(raw.replace('Z', '+00:00'))
        except ValueError:
            return None
    return None


def _to_float(value):
    if value is None:
        return 0.0
    try:
        n = float(str(value).replace(',', '').strip())
        return n if math.isfinite(n) else 0.0
    except Exception:
        return 0.0


def _serie_from_folio(doc_folio):
    folio = (doc_folio or '').strip().upper()
    if folio.startswith('P-'):
        return 'P'
    if folio.startswith('D'):
        return 'D'
    if '-' in folio:
        return folio.split('-', 1)[0][:10]
    return folio[:10] or None


def _contpaq_connection():
    try:
        import pyodbc
    except Exception as exc:
        raise RuntimeError('pyodbc no esta instalado. Agrega pyodbc al entorno.') from exc

    host = os.getenv('CONTPAQ_SQLSERVER_HOST', '').strip()
    port = os.getenv('CONTPAQ_SQLSERVER_PORT', '1433').strip()
    database = os.getenv('CONTPAQ_SQLSERVER_DATABASE', '').strip()
    user = os.getenv('CONTPAQ_SQLSERVER_USER', '').strip()
    password = os.getenv('CONTPAQ_SQLSERVER_PASSWORD', '').strip()
    driver = os.getenv('CONTPAQ_SQLSERVER_DRIVER', 'ODBC Driver 17 for SQL Server').strip()
    trust_cert = os.getenv('CONTPAQ_SQLSERVER_TRUST_CERT', 'yes').strip().lower()
    trusted_connection = os.getenv('CONTPAQ_SQLSERVER_TRUSTED_CONNECTION', '0').strip().lower() in ('1', 'true', 'yes', 'on')

    if not host or not database:
        raise RuntimeError('Faltan variables CONTPAQ_SQLSERVER_HOST o CONTPAQ_SQLSERVER_DATABASE.')

    server = f"{host},{port}" if port else host
    parts = [
        f"DRIVER={{{driver}}}",
        f"SERVER={server}",
        f"DATABASE={database}",
        f"TrustServerCertificate={'yes' if trust_cert in ('1', 'true', 'yes', 'on') else 'no'}",
    ]
    if trusted_connection:
        parts.append('Trusted_Connection=yes')
    else:
        if not user or not password:
            raise RuntimeError('Faltan credenciales CONTPAQ_SQLSERVER_USER o CONTPAQ_SQLSERVER_PASSWORD.')
        parts.append(f"UID={user}")
        parts.append(f"PWD={password}")

    conn_str = ';'.join(parts)
    return pyodbc.connect(conn_str, timeout=30)


def _contpaq_fetch_rows(query, params):
    conn = _contpaq_connection()
    try:
        cursor = conn.cursor()
        cursor.execute(query, params)
        cols = [d[0] for d in cursor.description]
        rows = []
        for item in cursor.fetchall():
            rows.append(dict(zip(cols, item)))
        return rows
    finally:
        try:
            conn.close()
        except Exception:
            pass


CONTPAQ_PEDIDOS_QUERY = """
SELECT
    h.DocumentID,
    h.DocFolio,
    h.BusinessEntityName,
    h.Title,
    h.PeriodWeek,
    h.DateDocument,
    COALESCE(depotDoc.DepotName, h.DepotName, '') AS Sucursal
FROM dbo.vwLBSDocCustomerOrderAllModules h
LEFT JOIN dbo.docDocumentExtra dxe ON h.DocumentID = dxe.DocumentID
LEFT JOIN dbo.orgDepot depotDoc ON dxe.BusinessEntityDepotID = depotDoc.DepotID
WHERE h.ModuleID = 967
  AND h.BusinessEntityName = ?
  AND h.DateDocument >= ?
  AND (h.DocFolio LIKE 'P-%' OR h.DocFolio LIKE 'D%')
"""


CONTPAQ_PEDIDOS_DETALLE_QUERY = """
SELECT
    h.DocumentID,
    h.DocFolio,
    h.BusinessEntityName,
    h.Title,
    h.PeriodWeek,
    h.DateDocument,
    COALESCE(depotDoc.DepotName, h.DepotName, '') AS Sucursal,
    i.LineNumber,
    i.ProductKey,
    i.Description,
    i.Quantity,
    i.UnitPrice
FROM dbo.vwLBSDocCustomerOrderAllModules h
INNER JOIN dbo.docDocumentItem i ON h.DocumentID = i.DocumentID
LEFT JOIN dbo.docDocumentExtra dxe ON h.DocumentID = dxe.DocumentID
LEFT JOIN dbo.orgDepot depotDoc ON dxe.BusinessEntityDepotID = depotDoc.DepotID
WHERE i.DeletedOn IS NULL
  AND h.ModuleID = 967
  AND h.BusinessEntityName = ?
  AND h.DateDocument >= ?
  AND (h.DocFolio LIKE 'P-%' OR h.DocFolio LIKE 'D%')
"""


CONTPAQ_REMISIONES_QUERY = """
SELECT
    d.DocumentID,
    ISNULL(d.FolioPrefix, '') + ISNULL(d.Folio, '') AS DocFolio,
    be.OfficialName AS BusinessEntityName,
    COALESCE(depotDoc.DepotName, dep.DepotName, '') AS Sucursal,
    d.DateDocument,
    d.SourceDocumentID
FROM dbo.docDocument d
LEFT JOIN dbo.vwLBSBusinessEntityList be ON d.BusinessEntityID = be.BusinessEntityID
LEFT JOIN dbo.docDocumentExtra dxe ON d.DocumentID = dxe.DocumentID
LEFT JOIN dbo.orgDepot depotDoc ON dxe.BusinessEntityDepotID = depotDoc.DepotID
LEFT JOIN dbo.orgDepot dep ON d.DepotID = dep.DepotID
WHERE d.ModuleID = 157
  AND be.OfficialName = ?
  AND d.DateDocument >= ?
"""


CONTPAQ_REMISIONES_DETALLE_QUERY = """
SELECT
    h.DocumentID,
    h.DocFolio,
    h.BusinessEntityName,
    h.DateDocument,
    i.LineNumber,
    i.ProductKey,
    i.Description,
    i.Quantity,
    i.CostPrice
FROM dbo.vwLBSDocCustomerDeliveryAllModules h
INNER JOIN dbo.docDocumentItem i ON h.DocumentID = i.DocumentID
WHERE i.DeletedOn IS NULL
  AND h.ModuleID = 157
  AND h.BusinessEntityName = ?
  AND h.DateDocument >= ?
"""


CONTPAQ_NOTAS_VENTA_QUERY = """
SELECT
    d.DocumentID,
    ISNULL(d.FolioPrefix, '') + ISNULL(d.Folio, '') AS DocFolio,
    be.OfficialName AS BusinessEntityName,
    COALESCE(depotDoc.DepotName, dep.DepotName, '') AS Sucursal,
    d.DateDocument,
    d.SourceDocumentID,
    d.DestinationDocumentID,
    CASE WHEN d.CancelledOn IS NULL THEN d.SubTotal ELSE 0 END AS SubTotal,
    CASE WHEN d.CancelledOn IS NULL THEN d.Total ELSE 0 END AS Total,
    CASE WHEN ISNULL(inv.TotalPaid, 0) > 0 THEN inv.TotalPaid ELSE COALESCE(d.TotalPaid, 0) END AS TotalInvoicePaid,
    COALESCE(d.Total, 0) - CASE WHEN ISNULL(inv.TotalPaid, 0) > 0 THEN inv.TotalPaid ELSE COALESCE(d.TotalPaid, 0) END AS TotalInvoiceBalance
FROM dbo.docDocument d
LEFT JOIN dbo.vwLBSBusinessEntityList be ON d.BusinessEntityID = be.BusinessEntityID
LEFT JOIN dbo.docDocumentExtra dxe ON d.DocumentID = dxe.DocumentID
LEFT JOIN dbo.orgDepot depotDoc ON dxe.BusinessEntityDepotID = depotDoc.DepotID
LEFT JOIN dbo.orgDepot dep ON d.DepotID = dep.DepotID
LEFT JOIN dbo.vwLBSDocCustomerSalesTotalInvoiced inv ON d.DocumentID = inv.DocumentID
WHERE d.ModuleID = 158
  AND d.DateDocument >= ?
  AND d.DeletedOn IS NULL
  AND d.CancelledOn IS NULL
"""


CONTPAQ_SUPPLIER_OT_QUERY = """
SELECT
        dbo.docDocument.DocumentID,
        ISNULL(dbo.docDocument.FolioPrefix, N'') + ISNULL(dbo.docDocument.Folio, N'') AS DocFolio,
        dbo.vwLBSBusinessEntityList.OfficialName AS BusinessEntityName,
        COALESCE(depotDoc.DepotName, dbo.orgDepot.DepotName, '') AS DepotName,
        dbo.docDocument.DateDocument,
        dbo.docDocument.DateDocDelivery,
        dbo.docDocument.Title,
        dbo.docDocument.Comments
FROM dbo.docDocumentExtra
RIGHT OUTER JOIN dbo.docDocument ON dbo.docDocumentExtra.DocumentID = dbo.docDocument.DocumentID
LEFT OUTER JOIN dbo.orgDepot ON dbo.docDocument.DepotID = dbo.orgDepot.DepotID
LEFT OUTER JOIN dbo.vwLBSBusinessEntityList ON dbo.docDocument.BusinessEntityID = dbo.vwLBSBusinessEntityList.BusinessEntityID
LEFT OUTER JOIN dbo.orgDepot AS depotDoc ON dbo.docDocumentExtra.BusinessEntityDepotID = depotDoc.DepotID
WHERE dbo.docDocument.ModuleID = 183
    AND dbo.docDocument.DateDocument >= ?
    AND (ISNULL(dbo.docDocument.FolioPrefix, N'') + ISNULL(dbo.docDocument.Folio, N'')) LIKE 'OT%'
"""


CONTPAQ_SUPPLIER_OT_DETALLE_QUERY = """
SELECT
        dbo.vwLBSDocSupplierPOAllModules.DocumentID,
        dbo.vwLBSDocSupplierPOAllModules.DocFolio,
        dbo.vwLBSDocSupplierPOAllModules.BusinessEntityName,
        dbo.vwLBSDocSupplierPOAllModules.DepotName,
        dbo.vwLBSDocSupplierPOAllModules.DateDocument,
        dbo.vwLBSDocSupplierPOAllModules.Title,
        dbo.vwLBSProductsToDeliver.ProductID,
        dbo.vwLBSProductsToDeliver.ProductKey,
        dbo.vwLBSProductsToDeliver.ProductName,
        dbo.vwLBSProductsToDeliver.QtyOrdered,
        dbo.vwLBSProductsToDeliver.QtyDelivered,
        dbo.vwLBSProductsToDeliver.QtyToBeDelivered
FROM dbo.vwLBSDocSupplierPOAllModules
INNER JOIN dbo.vwLBSProductsToDeliver ON dbo.vwLBSDocSupplierPOAllModules.DocumentID = dbo.vwLBSProductsToDeliver.DocumentID
WHERE dbo.vwLBSProductsToDeliver.QtyToBeDelivered > 0
    AND dbo.vwLBSDocSupplierPOAllModules.ModuleID = 183
    AND dbo.vwLBSDocSupplierPOAllModules.DateDocument >= ?
    AND dbo.vwLBSDocSupplierPOAllModules.DocFolio LIKE 'OT%'
"""


def _ensure_contpaq_notas_venta_schema():
    """Aplica migracion minima de notas de venta si falta en BD de produccion."""
    try:
        db.session.execute(db.text("""
            CREATE TABLE IF NOT EXISTS contpaq_notas_venta (
                id SERIAL PRIMARY KEY,
                document_id BIGINT NOT NULL UNIQUE,
                source_document_id BIGINT,
                destination_document_id BIGINT,
                doc_folio VARCHAR(80) NOT NULL,
                cliente VARCHAR(255),
                sucursal VARCHAR(255),
                fecha_documento TIMESTAMP,
                subtotal DOUBLE PRECISION,
                total DOUBLE PRECISION,
                total_paid DOUBLE PRECISION,
                total_invoice_paid DOUBLE PRECISION,
                total_invoice_balance DOUBLE PRECISION,
                balance DOUBLE PRECISION,
                updated_at TIMESTAMP NOT NULL DEFAULT NOW()
            )
        """))
        db.session.execute(db.text("""
            ALTER TABLE contpaq_sync_runs
            ADD COLUMN IF NOT EXISTS notas_venta_upserted INTEGER NOT NULL DEFAULT 0
        """))
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        logger.warning('[CONTPAQ] No se pudo asegurar schema notas_venta: %s', exc)


def _upsert_contpaq_data(payload):
    pedidos = payload.get('pedidos') or []
    pedidos_detalle = payload.get('pedidos_detalle') or []
    remisiones = payload.get('remisiones') or []
    remisiones_detalle = payload.get('remisiones_detalle') or []

    stats = {
        'pedidos_upserted': 0,
        'pedido_detalles_upserted': 0,
        'remisiones_upserted': 0,
        'remision_detalles_upserted': 0,
        'notas_venta_upserted': 0,
    }

    pedidos_map = {}
    for row in pedidos:
        doc_id = int(row.get('DocumentID') or 0)
        if doc_id <= 0:
            continue
        pedido = ContpaqPedido.query.filter_by(document_id=doc_id).first()
        if not pedido:
            pedido = ContpaqPedido(document_id=doc_id)
            db.session.add(pedido)
        pedido.doc_folio = _contpaq_clip(row.get('DocFolio'), 80)
        pedido.serie = _serie_from_folio(pedido.doc_folio)
        pedido.cliente = _contpaq_clip(row.get('BusinessEntityName'), 255)
        pedido.sucursal = _contpaq_clip(row.get('Sucursal'), 255)
        pedido.titulo = _contpaq_clip(_contpaq_norm_text(row.get('Title')), 255)
        pedido.periodo_semana = _contpaq_clip(_contpaq_norm_text(row.get('PeriodWeek')), 30)
        pedido.fecha_documento = _parse_datetime(row.get('DateDocument'))
        pedido.updated_at = datetime.utcnow()
        pedidos_map[doc_id] = pedido
        stats['pedidos_upserted'] += 1

    db.session.flush()

    for row in pedidos_detalle:
        doc_id = int(row.get('DocumentID') or 0)
        if doc_id <= 0 or doc_id not in pedidos_map:
            continue
        line_number = int(row.get('LineNumber') or 0)
        item = ContpaqPedidoDetalle.query.filter_by(document_id=doc_id, line_number=line_number).first()
        if not item:
            item = ContpaqPedidoDetalle(document_id=doc_id, line_number=line_number)
            db.session.add(item)
        cantidad = _to_float(row.get('Quantity'))
        precio = _to_float(row.get('UnitPrice'))
        item.pedido_id = pedidos_map[doc_id].id
        item.clave_producto = str(row.get('ProductKey') or '').strip().upper()[:120]
        item.descripcion = str(row.get('Description') or '').strip()
        item.cantidad = cantidad
        item.precio_unitario = precio
        item.total_partida = round(cantidad * precio, 2)
        item.updated_at = datetime.utcnow()
        stats['pedido_detalles_upserted'] += 1

    remisiones_map = {}
    for row in remisiones:
        doc_id = int(row.get('DocumentID') or 0)
        if doc_id <= 0:
            continue
        remision = ContpaqRemision.query.filter_by(document_id=doc_id).first()
        if not remision:
            remision = ContpaqRemision(document_id=doc_id)
            db.session.add(remision)
        remision.doc_folio = _contpaq_clip(row.get('DocFolio'), 80)
        remision.cliente = _contpaq_clip(row.get('BusinessEntityName'), 255)
        remision.sucursal = _contpaq_clip(row.get('Sucursal'), 255)
        remision.fecha_documento = _parse_datetime(row.get('DateDocument'))
        source_id = row.get('SourceDocumentID')
        try:
            remision.source_document_id = int(source_id) if source_id is not None else None
        except Exception:
            remision.source_document_id = None
        remision.updated_at = datetime.utcnow()
        remisiones_map[doc_id] = remision
        stats['remisiones_upserted'] += 1

    db.session.flush()

    for row in payload.get('notas_venta', []):
        doc_id = int(row.get('DocumentID') or 0)
        if doc_id <= 0:
            continue
        nota = ContpaqNotaVenta.query.filter_by(document_id=doc_id).first()
        if not nota:
            nota = ContpaqNotaVenta(document_id=doc_id)
            db.session.add(nota)
        nota.doc_folio = str(row.get('DocFolio') or '').strip()
        nota.cliente = str(row.get('BusinessEntityName') or '').strip()
        nota.sucursal = str(row.get('Sucursal') or '').strip()
        nota.fecha_documento = _parse_datetime(row.get('DateDocument'))
        source_id = row.get('SourceDocumentID')
        destination_id = row.get('DestinationDocumentID')
        try:
            nota.source_document_id = int(source_id) if source_id is not None else None
        except Exception:
            nota.source_document_id = None
        try:
            nota.destination_document_id = int(destination_id) if destination_id is not None else None
        except Exception:
            nota.destination_document_id = None
        nota.subtotal = _to_float(row.get('SubTotal'))
        nota.total = _to_float(row.get('Total'))
        nota.total_paid = _to_float(row.get('TotalPaid'))
        nota.total_invoice_paid = _to_float(row.get('TotalInvoicePaid'))
        nota.total_invoice_balance = _to_float(row.get('TotalInvoiceBalance'))
        nota.balance = _to_float(row.get('TotalInvoiceBalance'))
        nota.updated_at = datetime.utcnow()
        stats['notas_venta_upserted'] += 1

    db.session.flush()

    for row in remisiones_detalle:
        doc_id = int(row.get('DocumentID') or 0)
        if doc_id <= 0 or doc_id not in remisiones_map:
            continue
        line_number = int(row.get('LineNumber') or 0)
        item = ContpaqRemisionDetalle.query.filter_by(document_id=doc_id, line_number=line_number).first()
        if not item:
            item = ContpaqRemisionDetalle(document_id=doc_id, line_number=line_number)
            db.session.add(item)
        cantidad = _to_float(row.get('Quantity'))
        precio = _to_float(row.get('CostPrice'))
        item.remision_id = remisiones_map[doc_id].id
        item.clave_producto = str(row.get('ProductKey') or '').strip().upper()[:120]
        item.descripcion = str(row.get('Description') or '').strip()
        item.cantidad = cantidad
        item.precio_unitario = precio
        item.total_partida = round(cantidad * precio, 2)
        item.updated_at = datetime.utcnow()
        stats['remision_detalles_upserted'] += 1

    return stats


def _upsert_contpaq_existencia_data(payload):
    existencias = payload.get('existencias') or []
    stats = {
        'existencias_upserted': 0,
    }

    for row in existencias:
        depot_id = int(row.get('DepotID') or 0)
        product_id = int(row.get('ProductID') or 0)
        matrix_key1 = str(row.get('MatrixKey1') or '').strip()
        matrix_key2 = str(row.get('MatrixKey2') or '').strip()
        if depot_id <= 0 or product_id <= 0:
            continue

        item = ContpaqExistenciaStock.query.filter_by(
            depot_id=depot_id,
            product_id=product_id,
            matrix_key1=matrix_key1,
            matrix_key2=matrix_key2,
        ).first()
        if not item:
            item = ContpaqExistenciaStock(
                depot_id=depot_id,
                product_id=product_id,
                matrix_key1=matrix_key1,
                matrix_key2=matrix_key2,
            )
            db.session.add(item)

        item.owned_business_entity_id = int(row.get('OwnedBusinessEntityID') or 0) or None
        item.depot_name = str(row.get('DepotName') or '').strip()
        item.depot_type_id = int(row.get('DepotTypeID') or 0) or None
        item.product_key = str(row.get('ProductKey') or '').strip().upper()[:120]
        item.product_name = str(row.get('ProductName') or '').strip()
        item.category1 = str(row.get('Category1') or '').strip()[:255] or None
        item.category2 = str(row.get('Category2') or '').strip()[:255] or None
        item.unit = str(row.get('Unit') or '').strip()[:60] or None
        item.qty_present = _to_float(row.get('QtyPresent'))
        item.qty_available = _to_float(row.get('QtyAvailable'))
        item.qty_to_deliver_customer = _to_float(row.get('QtyToDeliverToCustomer'))
        item.qty_to_receive_supplier = _to_float(row.get('QtyToReceiveFromSupplier'))
        item.qty_on_transit = _to_float(row.get('QtyOnTransit'))
        item.qty_to_receive = _to_float(row.get('QtyToReceive'))
        item.qty_max_contpaq = _to_float(row.get('QtyMax'))
        item.qty_min_contpaq = _to_float(row.get('QtyMinimum'))
        item.updated_at = datetime.utcnow()
        stats['existencias_upserted'] += 1

    return stats


def _contpaq_max_min_rows(
    q='',
    sucursal='',
    category1='',
    category2='',
    only_alert=False,
    limit=500,
    page=1,
    period_type='week',
    period_value=None,
    period_from=None,
    period_to=None,
    period_year=None,
    date_from=None,
    date_to=None,
):
    now = datetime.utcnow()
    period_type = str(period_type or 'week').strip().lower()
    if period_type not in ('week', 'month', 'date'):
        period_type = 'week'

    try:
        period_year = int(period_year or now.year)
    except Exception:
        period_year = now.year

    if period_type == 'week':
        default_period_value = int(now.isocalendar()[1])
        period_label = 'Semana'
    elif period_type == 'month':
        default_period_value = int(now.month)
        period_label = 'Mes'
    else:
        default_period_value = None
        period_label = 'Fecha'

    try:
        period_value = int(period_value) if period_value not in (None, '') else default_period_value
    except Exception:
        period_value = default_period_value

    def _to_int_or_none(value):
        try:
            if value in (None, ''):
                return None
            return int(value)
        except Exception:
            return None

    period_from_i = _to_int_or_none(period_from)
    period_to_i = _to_int_or_none(period_to)
    if period_value is not None and period_from_i is None and period_to_i is None:
        period_from_i = int(period_value)
        period_to_i = int(period_value)
    if period_from_i is None and period_to_i is not None:
        period_from_i = period_to_i
    if period_to_i is None and period_from_i is not None:
        period_to_i = period_from_i
    if period_from_i is not None and period_to_i is not None and period_from_i > period_to_i:
        period_from_i, period_to_i = period_to_i, period_from_i

    def _parse_date_iso(raw):
        raw_txt = str(raw or '').strip()
        if not raw_txt:
            return None
        try:
            return datetime.strptime(raw_txt, '%Y-%m-%d').date()
        except Exception:
            return None

    date_from_d = _parse_date_iso(date_from)
    date_to_d = _parse_date_iso(date_to)
    if date_from_d and date_to_d and date_from_d > date_to_d:
        date_from_d, date_to_d = date_to_d, date_from_d

    def _extract_period_number(raw_value):
        text = str(raw_value or '').strip()
        if not text:
            return None
        match = re.search(r'(\d+)', text)
        if not match:
            return None
        try:
            return int(match.group(1))
        except Exception:
            return None

    def _excel_round(value, digits=0):
        try:
            quant = Decimal('1') if digits == 0 else Decimal('1').scaleb(-digits)
            return float(Decimal(str(value or 0)).quantize(quant, rounding=ROUND_HALF_UP))
        except Exception:
            return float(value or 0)

    pedido_query = (
        db.session.query(
            ContpaqPedidoDetalle.clave_producto,
            ContpaqPedidoDetalle.descripcion,
            ContpaqPedidoDetalle.cantidad,
            ContpaqPedido.document_id,
            ContpaqPedido.sucursal,
            ContpaqPedido.fecha_documento,
            ContpaqPedido.periodo_semana,
        )
        .join(ContpaqPedido, ContpaqPedido.id == ContpaqPedidoDetalle.pedido_id)
        .filter(ContpaqPedido.fecha_documento.isnot(None))
    )

    # Si filtran por fecha exacta, se respeta rango de fecha; si no, se usa el anio.
    if period_type == 'date' and (date_from_d or date_to_d):
        if date_from_d:
            pedido_query = pedido_query.filter(ContpaqPedido.fecha_documento >= datetime.combine(date_from_d, datetime.min.time()))
        if date_to_d:
            pedido_query = pedido_query.filter(ContpaqPedido.fecha_documento < datetime.combine(date_to_d + timedelta(days=1), datetime.min.time()))
    else:
        pedido_query = pedido_query.filter(extract('year', ContpaqPedido.fecha_documento) == period_year)

    pedido_rows = pedido_query.all()

    demand_map = {}
    docs_map = {}
    desc_map = {}
    for clave_producto, descripcion, cantidad, document_id, pedido_sucursal, fecha_documento, periodo_semana in pedido_rows:
        product_key_norm = str(clave_producto or '').strip().upper()
        if not product_key_norm:
            continue
        if sucursal and sucursal.lower() not in str(pedido_sucursal or '').lower():
            continue

        bucket_label = None
        if period_type == 'week':
            week_number = _extract_period_number(periodo_semana)
            if week_number is None and fecha_documento is not None:
                week_number = int(fecha_documento.isocalendar()[1])
            if week_number is None:
                continue
            if period_from_i is not None and period_to_i is not None and not (period_from_i <= week_number <= period_to_i):
                continue
            bucket_label = f"S{week_number:02d}"
        elif period_type == 'month':
            if fecha_documento is None:
                continue
            month_number = int(fecha_documento.month)
            if period_from_i is not None and period_to_i is not None and not (period_from_i <= month_number <= period_to_i):
                continue
            bucket_label = f"M{month_number:02d}"
        else:
            if fecha_documento is None:
                continue
            fecha_doc = fecha_documento.date()
            if date_from_d and fecha_doc < date_from_d:
                continue
            if date_to_d and fecha_doc > date_to_d:
                continue
            bucket_label = fecha_doc.isoformat()

        map_key = product_key_norm
        bucket_key = (product_key_norm, bucket_label)
        demand_map[bucket_key] = float(demand_map.get(bucket_key, 0.0)) + float(cantidad or 0.0)
        docs_map.setdefault(map_key, set()).add(int(document_id or 0))
        if map_key not in desc_map and str(descripcion or '').strip():
            desc_map[map_key] = str(descripcion or '').strip()

    bucket_order = {}
    for bucket_key in demand_map.keys():
        _, label = bucket_key
        if label.startswith('S') or label.startswith('M'):
            try:
                bucket_order[label] = int(label[1:])
            except Exception:
                bucket_order[label] = 9999
        else:
            bucket_order[label] = label

    period_columns = sorted(bucket_order.keys(), key=lambda label: bucket_order.get(label))

    stocks_q = ContpaqExistenciaStock.query.filter(ContpaqExistenciaStock.product_key.isnot(None))
    if sucursal:
        stocks_q = stocks_q.filter(ContpaqExistenciaStock.depot_name.ilike(f"%{sucursal}%"))
    if category1:
        stocks_q = stocks_q.filter(ContpaqExistenciaStock.category1.ilike(f"%{category1}%"))
    if category2:
        stocks_q = stocks_q.filter(ContpaqExistenciaStock.category2.ilike(f"%{category2}%"))
    if q:
        like = f"%{q}%"
        stocks_q = stocks_q.filter(
            db.or_(
                ContpaqExistenciaStock.product_key.ilike(like),
                ContpaqExistenciaStock.product_name.ilike(like),
                ContpaqExistenciaStock.depot_name.ilike(like),
                ContpaqExistenciaStock.category1.ilike(like),
                ContpaqExistenciaStock.category2.ilike(like),
            )
        )

    stocks = stocks_q.order_by(ContpaqExistenciaStock.product_key.asc(), ContpaqExistenciaStock.depot_name.asc()).all()

    stock_map = {}
    for stock in stocks:
        product_key = str(stock.product_key or '').strip().upper()
        if not product_key:
            continue
        current = stock_map.setdefault(product_key, {
            'product_key': stock.product_key,
            'product_name': stock.product_name,
            'category1': stock.category1,
            'category2': stock.category2,
            'unit': stock.unit,
            'qty_present': 0.0,
            'qty_available': 0.0,
            'qty_to_deliver_customer': 0.0,
            'qty_to_receive_supplier': 0.0,
            'qty_on_transit': 0.0,
            'qty_to_receive': 0.0,
            'updated_at': stock.updated_at.isoformat() if stock.updated_at else None,
        })
        current['product_name'] = current['product_name'] or stock.product_name
        current['category1'] = current['category1'] or stock.category1
        current['category2'] = current['category2'] or stock.category2
        current['unit'] = current['unit'] or stock.unit
        current['qty_present'] += float(stock.qty_present or 0.0)
        current['qty_available'] += float(stock.qty_available or 0.0)
        current['qty_to_deliver_customer'] += float(stock.qty_to_deliver_customer or 0.0)
        current['qty_to_receive_supplier'] += float(stock.qty_to_receive_supplier or 0.0)
        current['qty_on_transit'] += float(stock.qty_on_transit or 0.0)
        current['qty_to_receive'] += float(stock.qty_to_receive or 0.0)

    rows = []
    report_keys = sorted(set([pk for pk, _ in demand_map.keys()]), key=lambda value: value)
    for product_key in report_keys:
        product_buckets = []
        for (pk, label), qty in demand_map.items():
            if pk == product_key:
                product_buckets.append((label, float(qty or 0.0)))
        product_buckets.sort(key=lambda x: bucket_order.get(x[0]))

        period_values = {label: 0.0 for label in period_columns}
        for label, qty in product_buckets:
            period_values[label] = round(float(qty or 0.0), 2)

        positive_period_values = [qty for _, qty in product_buckets if float(qty or 0.0) > 0]
        promedio = _excel_round(sum(positive_period_values) / len(positive_period_values), 0) if positive_period_values else 0.0
        existencia_info = stock_map.get(product_key, {})
        existencia = float(existencia_info.get('qty_present') or 0.0)
        maximo = max(_excel_round(promedio * 1.25, 0), existencia)
        minimo = _excel_round(promedio * 0.75, 0)
        cantidad_periodo = round(sum(float(qty or 0.0) for _, qty in product_buckets), 2)
        pedidos_periodo = len(docs_map.get(product_key, set()))
        detalle_periodos = ' | '.join([f"{lbl}:{round(qty, 2)}" for lbl, qty in product_buckets])

        if minimo > 0 and existencia < minimo:
            status = 'BAJO_MINIMO'
        elif maximo > 0 and existencia > maximo:
            status = 'SOBRE_MAXIMO'
        else:
            status = 'EN_RANGO'

        sugerido_compra = max(0.0, float(maximo) - existencia) if status == 'BAJO_MINIMO' else 0.0

        row = {
            'id': product_key,
            'product_key': product_key,
            'product_name': (existencia_info.get('product_name') or desc_map.get(product_key) or '').strip(),
            'category1': existencia_info.get('category1'),
            'category2': existencia_info.get('category2'),
            'unit': existencia_info.get('unit'),
            'qty_present': round(float(existencia_info.get('qty_present') or 0.0), 2),
            'qty_available': round(float(existencia_info.get('qty_available') or 0.0), 2),
            'qty_to_deliver_customer': round(float(existencia_info.get('qty_to_deliver_customer') or 0.0), 2),
            'qty_to_receive_supplier': round(float(existencia_info.get('qty_to_receive_supplier') or 0.0), 2),
            'qty_on_transit': round(float(existencia_info.get('qty_on_transit') or 0.0), 2),
            'qty_to_receive': round(float(existencia_info.get('qty_to_receive') or 0.0), 2),
            'cantidad_periodo': cantidad_periodo,
            'pedidos_periodo': pedidos_periodo,
            'detalle_periodos': detalle_periodos,
            'promedio': promedio,
            'existencia': round(existencia, 2),
            'maximo': round(float(maximo or 0.0), 2),
            'minimo': round(float(minimo or 0.0), 2),
            'period_values': period_values,
            'sugerido_compra': round(sugerido_compra, 2),
            'status': status,
            'period_type': period_type,
            'period_value': period_value,
            'period_year': period_year,
            'updated_at': existencia_info.get('updated_at'),
        }

        if q:
            q_lower = q.lower()
            haystack = ' '.join([
                str(row.get('product_key') or ''),
                str(row.get('product_name') or ''),
                str(row.get('category1') or ''),
                str(row.get('category2') or ''),
            ]).lower()
            if q_lower not in haystack:
                continue
        if category1 and category1.lower() not in str(row.get('category1') or '').lower():
            continue
        if category2 and category2.lower() not in str(row.get('category2') or '').lower():
            continue

        if only_alert and status == 'EN_RANGO':
            continue
        rows.append(row)

    total = len(rows)
    limit = max(1, min(int(limit or 500), 2000))
    page = max(1, int(page or 1))
    offset = (page - 1) * limit
    paged_rows = rows[offset:offset + limit]

    summary = {
        'total_rows': total,
        'en_rango': sum(1 for r in rows if r['status'] == 'EN_RANGO'),
        'bajo_minimo': sum(1 for r in rows if r['status'] == 'BAJO_MINIMO'),
        'sobre_maximo': sum(1 for r in rows if r['status'] == 'SOBRE_MAXIMO'),
        'sugerido_compra_total': round(sum(float(r['sugerido_compra'] or 0.0) for r in rows), 2),
        'cantidad_periodo_total': round(sum(float(r['cantidad_periodo'] or 0.0) for r in rows), 2),
        'pedidos_periodo_total': sum(int(r['pedidos_periodo'] or 0) for r in rows),
        'promedio_global': round(sum(float(r['promedio'] or 0.0) for r in rows), 2),
        'period_type': period_type,
        'period_label': period_label,
        'period_value': period_value,
        'period_from': period_from_i,
        'period_to': period_to_i,
        'period_year': period_year,
        'date_from': date_from_d.isoformat() if date_from_d else None,
        'date_to': date_to_d.isoformat() if date_to_d else None,
        'period_columns': period_columns,
    }

    return {
        'items': paged_rows,
        'summary': summary,
        'total_records': total,
        'page': page,
        'limit': limit,
        'total_pages': (total + limit - 1) // limit if total else 0,
        'has_prev': page > 1,
        'has_next': offset + limit < total,
    }


def _upsert_contpaq_supplier_ot_data(payload):
    ot_headers = payload.get('supplier_purchase_orders') or []
    ot_details = payload.get('supplier_purchase_order_details') or []

    stats = {
        'supplier_ots_upserted': 0,
        'supplier_ot_detalles_upserted': 0,
    }

    headers_map = {}
    for row in ot_headers:
        doc_id = int(row.get('DocumentID') or 0)
        if doc_id <= 0:
            continue
        header = ContpaqSupplierOT.query.filter_by(document_id=doc_id).first()
        if not header:
            header = ContpaqSupplierOT(document_id=doc_id)
            db.session.add(header)
        header.doc_folio = str(row.get('DocFolio') or '').strip().upper()
        header.serie = _serie_from_folio(header.doc_folio)
        header.proveedor = str(row.get('BusinessEntityName') or '').strip()
        header.sucursal = str(row.get('DepotName') or '').strip()
        header.titulo = _contpaq_norm_text(row.get('Title'))
        header.fecha_documento = _parse_datetime(row.get('DateDocument'))
        header.fecha_entrega = _parse_datetime(row.get('DateDocDelivery'))
        header.comentarios = str(row.get('Comments') or '').strip() or None
        header.updated_at = datetime.utcnow()
        headers_map[doc_id] = header
        stats['supplier_ots_upserted'] += 1

    db.session.flush()

    for row in ot_details:
        doc_id = int(row.get('DocumentID') or 0)
        if doc_id <= 0:
            continue
        header = headers_map.get(doc_id) or ContpaqSupplierOT.query.filter_by(document_id=doc_id).first()
        if not header:
            continue
        product_id = row.get('ProductID')
        try:
            product_id = int(product_id) if product_id is not None else None
        except Exception:
            product_id = None
        product_key = str(row.get('ProductKey') or '').strip().upper()[:120]
        if not product_key:
            continue

        detail = ContpaqSupplierOTDetalle.query.filter_by(document_id=doc_id, product_key=product_key).first()

        if not detail:
            detail = ContpaqSupplierOTDetalle(document_id=doc_id, product_id=product_id)
            db.session.add(detail)

        detail.ot_id = header.id
        detail.product_id = product_id
        detail.product_key = product_key
        detail.product_name = str(row.get('ProductName') or '').strip() or str(row.get('Description') or '').strip() or None
        detail.qty_ordered = _to_float(row.get('QtyOrdered'))
        detail.qty_delivered = _to_float(row.get('QtyDelivered'))
        detail.qty_to_receive = _to_float(row.get('QtyToBeDelivered'))
        detail.updated_at = datetime.utcnow()
        stats['supplier_ot_detalles_upserted'] += 1

    return stats


def _contpaq_supplier_ot_reserved_map(detail_ids):
    ids = [int(x) for x in (detail_ids or []) if x]
    if not ids:
        return {}
    rows = (
        db.session.query(
            HojaRutaEntregaOTAsignacion.supplier_ot_detalle_id,
            func.coalesce(func.sum(HojaRutaEntregaOTAsignacion.qty_assigned), 0),
        )
        .filter(
            HojaRutaEntregaOTAsignacion.status == 'active',
            HojaRutaEntregaOTAsignacion.supplier_ot_detalle_id.in_(ids),
        )
        .group_by(HojaRutaEntregaOTAsignacion.supplier_ot_detalle_id)
        .all()
    )
    return {int(detail_id): float(total or 0) for detail_id, total in rows}


def _contpaq_supplier_ot_options_for_product_key(product_key, requested_qty=None):
    clave = str(product_key or '').strip().upper()
    if not clave:
        return []

    details = (
        ContpaqSupplierOTDetalle.query
        .join(ContpaqSupplierOT, ContpaqSupplierOT.id == ContpaqSupplierOTDetalle.ot_id)
        .filter(ContpaqSupplierOTDetalle.product_key == clave)
        .order_by(ContpaqSupplierOT.fecha_documento.asc(), ContpaqSupplierOT.doc_folio.asc())
        .all()
    )
    reserved_map = _contpaq_supplier_ot_reserved_map([d.id for d in details])
    requested = None
    try:
        requested = float(requested_qty) if requested_qty is not None else None
    except Exception:
        requested = None

    options = []
    for detail in details:
        pendiente = float(detail.qty_to_receive or 0)
        reserved = float(reserved_map.get(detail.id, 0))
        available = max(pendiente - reserved, 0)
        if available <= 0:
            continue
        header = detail.ot
        options.append({
            'detail_id': detail.id,
            'document_id': detail.document_id,
            'doc_folio': header.doc_folio if header else None,
            'serie': header.serie if header else None,
            'proveedor': header.proveedor if header else None,
            'sucursal': header.sucursal if header else None,
            'titulo': header.titulo if header else None,
            'fecha_documento': header.fecha_documento.isoformat() if header and header.fecha_documento else None,
            'product_key': detail.product_key,
            'product_name': detail.product_name,
            'qty_ordered': float(detail.qty_ordered or 0),
            'qty_delivered': float(detail.qty_delivered or 0),
            'qty_to_receive': pendiente,
            'qty_reserved': reserved,
            'qty_available': available,
            'fits_requested_qty': requested is None or requested <= available,
        })
    return options


def _contpaq_supplier_ot_key_status(product_key):
    clave = str(product_key or '').strip().upper()
    if not clave:
        return {
            'exists': False,
            'rows': 0,
            'total_pending': 0.0,
            'total_reserved': 0.0,
            'total_available': 0.0,
        }

    details = ContpaqSupplierOTDetalle.query.filter_by(product_key=clave).all()
    if not details:
        return {
            'exists': False,
            'rows': 0,
            'total_pending': 0.0,
            'total_reserved': 0.0,
            'total_available': 0.0,
        }

    reserved_map = _contpaq_supplier_ot_reserved_map([d.id for d in details])
    total_pending = 0.0
    total_reserved = 0.0
    total_available = 0.0

    for detail in details:
        pendiente = float(detail.qty_to_receive or 0)
        reservado = float(reserved_map.get(detail.id, 0))
        disponible = max(pendiente - reservado, 0)
        total_pending += pendiente
        total_reserved += reservado
        total_available += disponible

    return {
        'exists': True,
        'rows': len(details),
        'total_pending': total_pending,
        'total_reserved': total_reserved,
        'total_available': total_available,
    }


def _contpaq_supplier_ot_top_keys(limit=10):
    rows = (
        db.session.query(
            ContpaqSupplierOTDetalle.product_key,
            func.count(ContpaqSupplierOTDetalle.id),
            func.coalesce(func.sum(ContpaqSupplierOTDetalle.qty_to_receive), 0),
        )
        .group_by(ContpaqSupplierOTDetalle.product_key)
        .order_by(func.coalesce(func.sum(ContpaqSupplierOTDetalle.qty_to_receive), 0).desc())
        .limit(max(1, int(limit or 10)))
        .all()
    )

    return [
        {
            'product_key': product_key,
            'rows': int(rows_count or 0),
            'qty_to_receive_total': float(total_qty or 0),
        }
        for product_key, rows_count, total_qty in rows
        if product_key
    ]


def _create_hoja_entrega_ot_assignment(hoja, detail, qty_assigned, created_by=None):
    assignment = HojaRutaEntregaOTAsignacion(
        hoja_ruta_id=hoja.id,
        supplier_ot_detalle_id=detail.id,
        document_id=detail.document_id,
        doc_folio=(detail.ot.doc_folio if detail.ot else '') or '',
        product_key=detail.product_key,
        qty_assigned=float(qty_assigned or 0),
        status='active',
        created_by=(created_by or '').strip() or None,
    )
    db.session.add(assignment)
    return assignment


def _upsert_maquinaria_contpaq_data(payload):
    maquinaria_pedidos = payload.get('maquinaria_pedidos') or []
    maquinaria_pedidos_detalle = payload.get('maquinaria_pedidos_detalle') or []

    stats = {
        'maquinaria_pedidos_upserted': 0,
        'maquinaria_pedido_detalles_upserted': 0,
    }

    maquinaria_map = {}
    for row in maquinaria_pedidos:
        doc_id = int(row.get('DocumentID') or 0)
        if doc_id <= 0:
            continue
        pedido = MaquinariaContpaqPedido.query.filter_by(document_id=doc_id).first()
        if not pedido:
            pedido = MaquinariaContpaqPedido(document_id=doc_id)
            db.session.add(pedido)
        pedido.owned_business_entity_id = int(row.get('OwnedBusinessEntityID') or 0) or None
        pedido.folio = str(row.get('DocFolio') or '').strip()
        pedido.serie = _serie_from_folio(pedido.folio)
        pedido.business_entity_name = str(row.get('BusinessEntityName') or '').strip()
        pedido.depot_name = str(row.get('DepotName') or '').strip()
        pedido.date_document = _parse_datetime(row.get('DateDocument'))
        pedido.date_doc_delivery = _parse_datetime(row.get('DateDocDelivery'))
        pedido.title = str(row.get('Title') or '').strip()
        pedido.sales_rep = str(row.get('SalesRep') or '').strip()
        pedido.currency = str(row.get('Currency') or '').strip()
        pedido.rate = _to_float(row.get('Rate'))
        pedido.subtotal = _to_float(row.get('SubTotal'))
        pedido.total = _to_float(row.get('Total'))
        pedido.total_tax = _to_float(row.get('TotalTax'))
        pedido.total_discount = _to_float(row.get('TotalDiscount'))
        pedido.total_retention = _to_float(row.get('TotalRetention'))
        pedido.total_cost = _to_float(row.get('TotalCost'))
        pedido.comments = str(row.get('Comments') or '').strip()
        pedido.payment_term_name = str(row.get('PaymentTermName') or '').strip()
        pedido.language_name = str(row.get('LanguageName') or '').strip()
        pedido.cost_center_name = str(row.get('CostCenterName') or '').strip()
        pedido.cost_center_category = str(row.get('CostCenterCategory') or '').strip()
        pedido.period_month = str(row.get('PeriodMonth') or '').strip()
        pedido.period_week = str(row.get('PeriodWeek') or '').strip()
        pedido.period_year = str(row.get('PeriodYear') or '').strip()
        pedido.period_quarter = str(row.get('PeriodQuarter') or '').strip()
        pedido.campaign_name = str(row.get('CampaignName') or '').strip()
        pedido.campaign_id = int(row.get('CampaignID') or 0) or None
        pedido.intl_symbol = str(row.get('IntlSymbol') or '').strip()
        pedido.total_invoiced = _to_float(row.get('TotalInvoiced'))
        pedido.total_invoice_paid = _to_float(row.get('TotalInvoicePaid'))
        pedido.total_invoice_balance = _to_float(row.get('TotalInvoiceBalance'))
        pedido.invoiced = int(row.get('Invoiced') or 0)
        pedido.status_delivery_id = int(row.get('StatusDeliveryID') or 0) or None
        pedido.status_delivery = str(row.get('StatusDelivery') or '').strip()
        pedido.total_paid = _to_float(row.get('TotalPaid'))
        pedido.balance = _to_float(row.get('Balance'))
        pedido.globalizado = bool(row.get('Globalizado')) if row.get('Globalizado') is not None else False
        pedido.rfc_cliente = str(row.get('RFC_Cliente') or '').strip()
        pedido.metodo_pago = str(row.get('MetodoPago') or '').strip()
        pedido.forma_pago = str(row.get('FormaPago') or '').strip()
        pedido.tipo_facturacion = str(row.get('TipoFacturacion') or '').strip()
        pedido.invoice_document_id = int(row.get('InvoiceDocumentID') or 0) or None
        pedido.sucursal = str(row.get('Sucursal') or '').strip()
        pedido.printed = bool(row.get('Printed'))
        pedido.validated = bool(row.get('Validated'))
        pedido.cancelled = bool(row.get('Cancelled'))
        pedido.deleted = bool(row.get('Deleted'))
        pedido.in_use = bool(row.get('InUse'))
        pedido.auth1 = bool(row.get('Auth1'))
        pedido.auth2 = bool(row.get('Auth2'))
        pedido.updated_at = datetime.utcnow()
        maquinaria_map[doc_id] = pedido
        stats['maquinaria_pedidos_upserted'] += 1

    db.session.flush()

    for row in maquinaria_pedidos_detalle:
        doc_id = int(row.get('DocumentID') or 0)
        if doc_id <= 0 or doc_id not in maquinaria_map:
            continue
        line_number = int(row.get('LineNumber') or 0)
        item = MaquinariaContpaqPedidoDetalle.query.filter_by(document_id=doc_id, line_number=line_number).first()
        if not item:
            item = MaquinariaContpaqPedidoDetalle(document_id=doc_id, line_number=line_number)
            db.session.add(item)
        item.pedido_id = maquinaria_map[doc_id].id
        item.quantity = _to_float(row.get('Quantity'))
        item.product_id = int(row.get('ProductID') or 0) or None
        item.product_key = str(row.get('ProductKey') or '').strip().upper()[:120]
        item.description = str(row.get('Description') or '').strip()
        item.discount_perc = _to_float(row.get('DiscountPerc'))
        item.tax_perc = _to_float(row.get('TaxPerc'))
        item.tax_type_name = str(row.get('TaxTypeName') or '').strip()
        item.unit_price = _to_float(row.get('UnitPrice'))
        item.total_item = _to_float(row.get('TotalItem'))
        item.unit = str(row.get('Unit') or '').strip()
        item.clave_unidad = str(row.get('ClaveUnidad') or '').strip()
        item.coef_unit = _to_float(row.get('CoefUnit'))
        item.period_week = str(row.get('PeriodWeek') or '').strip()
        item.period_month = str(row.get('PeriodMonth') or '').strip()
        item.updated_at = datetime.utcnow()
        stats['maquinaria_pedido_detalles_upserted'] += 1

    return stats


def _contpaq_norm_text(value):
    raw = str(value or '').upper().strip()
    # Tolerate inconsistent spacing from CONTPAQ titles, e.g. "25   AL   31".
    return re.sub(r'\s+', ' ', raw)


_CONTPAQ_MONTH_NAMES = (
    'ENERO', 'FEBRERO', 'MARZO', 'ABRIL', 'MAYO', 'JUNIO',
    'JULIO', 'AGOSTO', 'SEPTIEMBRE', 'OCTUBRE', 'NOVIEMBRE', 'DICIEMBRE',
)
_CONTPAQ_MONTH_RE = '|'.join(_CONTPAQ_MONTH_NAMES)
_CONTPAQ_PERIOD_NOISE_RE = re.compile(
    r'\b(PEDIDO|DEL|DE|LA|EL|LOS|LAS|SEMANA|PERIODO|PERÍODO|AL)\b',
    re.IGNORECASE,
)


def _contpaq_strip_accents(value):
    text = str(value or '')
    try:
        import unicodedata
        return ''.join(
            ch for ch in unicodedata.normalize('NFKD', text)
            if not unicodedata.combining(ch)
        )
    except Exception:
        return text


def _contpaq_period_parts(value):
    """
    Extrae (dia_inicio, dia_fin, mes) de titulos/periodos con variantes de texto.
    Acepta: 'PEDIDO DEL 01 AL 07 JULIO', '01 AL 07 DE JULIO', '1-7 JULIO', etc.
    """
    text = _contpaq_norm_text(_contpaq_strip_accents(value))
    if not text:
        return None

    text = text.replace('/', ' ').replace('-', ' ').replace('_', ' ')
    text = re.sub(r'\s+', ' ', text)

    # Forma principal: N AL M [DE] MES
    match = re.search(
        rf'(\d{{1,2}})\s+AL\s+(\d{{1,2}})\s+(?:DE\s+)?({_CONTPAQ_MONTH_RE})\b',
        text,
    )
    if match:
        return int(match.group(1)), int(match.group(2)), match.group(3)

    # Compacta: quita ruido (PEDIDO/DE/DEL/...) y busca N M MES
    compact = _CONTPAQ_PERIOD_NOISE_RE.sub(' ', text)
    compact = re.sub(r'[^0-9A-Z]+', ' ', compact)
    compact = re.sub(r'\s+', ' ', compact).strip()
    match = re.search(
        rf'(\d{{1,2}})\s+(\d{{1,2}})\s+({_CONTPAQ_MONTH_RE})\b',
        compact,
    )
    if match:
        return int(match.group(1)), int(match.group(2)), match.group(3)

    return None


def _contpaq_period_fingerprint(value):
    """
    Huella estable del periodo. Variantes con/sin DE, espacios o tipografia
    distinta caen en la misma clave, p.ej.:
      PEDIDO DEL 01 AL 07 JULIO
      PEDIDO DEL 01 AL 07 DE JULIO
      1 AL 7 DE JULIO
    -> "1 AL 7 DE JULIO"
    """
    parts = _contpaq_period_parts(value)
    if parts:
        start_day, end_day, month = parts
        return f"{start_day} AL {end_day} DE {month}"

    text = _contpaq_norm_text(_contpaq_strip_accents(value))
    if not text:
        return ''
    text = re.sub(r'^(PEDIDO\s+DEL|PEDIDO\s+DE|PEDIDO)\s+', '', text)
    text = _CONTPAQ_PERIOD_NOISE_RE.sub(' ', text)
    text = re.sub(r'[^0-9A-Z]+', ' ', text)
    return re.sub(r'\s+', ' ', text).strip()


def _contpaq_period_compact(value):
    """Solo letras/numeros, para comparar aunque falten o sobren caracteres."""
    text = _contpaq_norm_text(_contpaq_strip_accents(value))
    text = _CONTPAQ_PERIOD_NOISE_RE.sub('', text)
    return re.sub(r'[^0-9A-Z]+', '', text)


def _contpaq_period_matches(filter_value, *candidates):
    """
    True si el filtro de periodo/titulo coincide con algun candidato,
    tolerando caracteres faltantes (DE, espacios, tipografia, etc.).
    La sucursal NO se toca aqui: se filtra aparte y sigue siendo estricta.
    """
    filt_parts = _contpaq_period_parts(filter_value)
    filt_fp = _contpaq_period_fingerprint(filter_value)
    filt_compact = _contpaq_period_compact(filter_value)
    if not filt_parts and not filt_fp and not filt_compact:
        return True

    for candidate in candidates:
        cand_parts = _contpaq_period_parts(candidate)
        if filt_parts and cand_parts and filt_parts == cand_parts:
            return True

        cand_fp = _contpaq_period_fingerprint(candidate)
        if filt_fp and cand_fp and (filt_fp == cand_fp or filt_fp in cand_fp or cand_fp in filt_fp):
            return True

        cand_compact = _contpaq_period_compact(candidate)
        if filt_compact and cand_compact:
            # Contencion tolerante: ignora DE/espacios/signos.
            if filt_compact in cand_compact or cand_compact in filt_compact:
                return True
            # Si solo difieren por 1-2 caracteres en cadenas ya largas, aceptar.
            if len(filt_compact) >= 8 and len(cand_compact) >= 8:
                a, b = (filt_compact, cand_compact) if len(filt_compact) <= len(cand_compact) else (cand_compact, filt_compact)
                if a in b:
                    return True
    return False


def _contpaq_unique_semana_options(labels):
    """Una sola opcion de semana por huella (evita 'JULIO' vs 'DE JULIO' duplicados)."""
    best = {}
    for label in labels:
        text = str(label or '').strip()
        if not text:
            continue
        fp = _contpaq_period_fingerprint(text) or _contpaq_norm_text(text)
        prev = best.get(fp)
        if not prev:
            best[fp] = text
            continue
        # Preferir la forma con "DE" y/o prefijo PEDIDO.
        score = (1 if ' DE ' in _contpaq_norm_text(text) else 0) + (1 if _contpaq_norm_text(text).startswith('PEDIDO') else 0)
        prev_score = (1 if ' DE ' in _contpaq_norm_text(prev) else 0) + (1 if _contpaq_norm_text(prev).startswith('PEDIDO') else 0)
        if score > prev_score or (score == prev_score and len(text) > len(prev)):
            best[fp] = text
    return sorted(best.values())


def _contpaq_clip(value, max_len):
    text = str(value or '').strip()
    if max_len and len(text) > max_len:
        return text[:max_len]
    return text


def _contpaq_norm_qty(value):
    try:
        numeric = float(str(value).replace(',', '').strip())
        if numeric.is_integer():
            return str(int(numeric))
        return f"{numeric:.4f}".rstrip('0').rstrip('.')
    except Exception:
        return _contpaq_norm_text(value)


def _contpaq_week_label_from_date(value):
    dt_value = None
    if isinstance(value, datetime):
        dt_value = value.date()
    else:
        parsed = _parse_date(value)
        if parsed:
            dt_value = parsed

    if not dt_value:
        return ''

    # Semana de negocio observada en archivos del usuario: martes a lunes.
    days_since_tuesday = (dt_value.weekday() - 1) % 7
    start_date = dt_value - timedelta(days=days_since_tuesday)
    end_date = start_date + timedelta(days=6)
    months = {
        1: 'ENERO', 2: 'FEBRERO', 3: 'MARZO', 4: 'ABRIL', 5: 'MAYO', 6: 'JUNIO',
        7: 'JULIO', 8: 'AGOSTO', 9: 'SEPTIEMBRE', 10: 'OCTUBRE', 11: 'NOVIEMBRE', 12: 'DICIEMBRE'
    }
    return f"{start_date.day} AL {end_date.day} DE {months.get(end_date.month, '')}".strip()


def _contpaq_week_match_key(semana_value=None, fecha_value=None):
    semana_fp = _contpaq_period_fingerprint(semana_value)
    if semana_fp and ' AL ' in semana_fp and ' DE ' in semana_fp:
        return semana_fp

    label_from_date = _contpaq_week_label_from_date(fecha_value)
    if label_from_date:
        return _contpaq_period_fingerprint(label_from_date) or _contpaq_norm_text(label_from_date)

    return semana_fp or _contpaq_norm_text(semana_value)


def _contpaq_title_week_key(title_value=None, semana_value=None, fecha_value=None):
    title_fp = _contpaq_period_fingerprint(title_value)
    if title_fp and ' AL ' in title_fp and ' DE ' in title_fp:
        # Preferir etiqueta completa amigable si el titulo trae "PEDIDO..."
        title_norm = _contpaq_norm_text(title_value)
        if title_norm.startswith('PEDIDO'):
            return f"PEDIDO DEL {title_fp}"
        return title_fp
    return _contpaq_week_match_key(semana_value, fecha_value)


def _contpaq_detect_import_column(columns, aliases):
    normalized = [(col, col.strip().lower().replace('-', '').replace('_', '').replace(' ', '')) for col in columns]
    for alias in aliases:
        alias_norm = alias.strip().lower().replace('-', '').replace('_', '').replace(' ', '')
        for original, current in normalized:
            if alias_norm == current or alias_norm in current:
                return original
    return None


def _contpaq_read_indice_dataframe(file_storage, filename, ext, sheet_name=''):
    if ext == '.csv':
        return pd.read_csv(file_storage, dtype=str, encoding='utf-8').fillna('')
    if ext in ('.xlsx', '.xls'):
        kwargs = {'dtype': str}
        if sheet_name:
            kwargs['sheet_name'] = sheet_name
        data = pd.read_excel(file_storage, **kwargs)
        if isinstance(data, dict):
            first_sheet = next(iter(data.values()))
            return first_sheet.fillna('')
        return data.fillna('')
    raise ValueError(f'Formato no soportado: {filename}')


def _excel_col_to_index(column_ref):
    value = str(column_ref or '').strip().upper()
    if not value:
        raise ValueError('Columna Excel vacia')

    idx = 0
    for ch in value:
        if not ('A' <= ch <= 'Z'):
            raise ValueError(f'Columna Excel invalida: {column_ref}')
        idx = (idx * 26) + (ord(ch) - ord('A') + 1)
    return idx - 1


def _contpaq_read_precio_publico_rows(file_storage, filename, ext, sheet_name='Productos', start_row=3, clave_col='D', precio_col='S'):
    if ext not in ('.xlsx', '.xls'):
        raise ValueError('Formato no soportado. Usa .xlsx o .xls')

    try:
        file_storage.stream.seek(0)
    except Exception:
        pass

    data = pd.read_excel(file_storage, sheet_name=sheet_name, header=None, dtype=object)
    if isinstance(data, dict):
        data = next(iter(data.values()))
    df = data.fillna('')

    clave_idx = _excel_col_to_index(clave_col)
    precio_idx = _excel_col_to_index(precio_col)
    start_idx = max(int(start_row) - 1, 0)

    rows = []
    for excel_row_num in range(start_idx, len(df.index)):
        row = df.iloc[excel_row_num]
        clave = str(row.iloc[clave_idx] if clave_idx < len(row) else '').strip().upper()
        precio_raw = row.iloc[precio_idx] if precio_idx < len(row) else ''

        if not clave and (precio_raw is None or str(precio_raw).strip() == ''):
            continue
        if not clave:
            continue

        try:
            precio_publico = float(str(precio_raw).replace(',', '').strip())
        except Exception:
            continue

        rows.append({
            'clave_producto': clave,
            'precio_publico': precio_publico,
            'row_number': excel_row_num + 1,
            'raw_payload': {
                'clave': clave,
                'precio_publico': precio_raw,
            },
        })

    return rows


def run_contpaq_sync(trigger='manual'):
    sync_run = ContpaqSyncRun(
        status='running',
        started_at=datetime.utcnow(),
        message=f'started_by={trigger}'
    )
    db.session.add(sync_run)
    db.session.commit()

    # En la nube no hay conexion directa a SQL Server; el agente Windows hace el push.
    try:
        import pyodbc  # noqa: F401
    except ImportError:
        sync_run.status = 'skipped'
        sync_run.message = 'pyodbc no disponible en este servidor. Usa el agente Windows para sincronizar.'
        sync_run.finished_at = datetime.utcnow()
        db.session.commit()
        return {'ok': True, 'skipped': True, 'run_id': sync_run.id,
                'message': 'Sincronizacion no disponible en la nube. Ejecuta el agente Windows para enviar datos.'}

    if not _CONTPAQ_SYNC_LOCK.acquire(blocking=False):
        sync_run.status = 'skipped'
        sync_run.message = 'Sincronizacion en curso; se omitio ejecucion paralela.'
        sync_run.finished_at = datetime.utcnow()
        db.session.commit()
        return {'ok': True, 'skipped': True, 'run_id': sync_run.id}

    try:
        logger.info('[CONTPAQ] Iniciando sincronizacion hacia nube...')
        _ensure_contpaq_notas_venta_schema()
        if not CONTPAQ_CUSTOMER_NAME:
            raise RuntimeError('CONTPAQ_CUSTOMER_NAME vacio.')

        start_date = _parse_date(CONTPAQ_START_DATE)
        if not start_date:
            raise RuntimeError('CONTPAQ_START_DATE invalida. Usa formato YYYY-MM-DD.')

        params = (CONTPAQ_CUSTOMER_NAME, start_date)
        payload = {
            'pedidos': _contpaq_fetch_rows(CONTPAQ_PEDIDOS_QUERY, params),
            'pedidos_detalle': _contpaq_fetch_rows(CONTPAQ_PEDIDOS_DETALLE_QUERY, params),
            'remisiones': _contpaq_fetch_rows(CONTPAQ_REMISIONES_QUERY, params),
            'remisiones_detalle': _contpaq_fetch_rows(CONTPAQ_REMISIONES_DETALLE_QUERY, params),
            'notas_venta': _contpaq_fetch_rows(CONTPAQ_NOTAS_VENTA_QUERY, (start_date,)),
        }

        stats = _upsert_contpaq_data(payload)
        sync_run.status = 'success'
        sync_run.finished_at = datetime.utcnow()
        sync_run.message = (
            f"cliente={CONTPAQ_CUSTOMER_NAME} desde={start_date.isoformat()} "
            f"pedidos={len(payload['pedidos'])} detalle_pedidos={len(payload['pedidos_detalle'])} "
            f"remisiones={len(payload['remisiones'])} detalle_remisiones={len(payload['remisiones_detalle'])} "
            f"notas_venta={len(payload['notas_venta'])}"
        )
        sync_run.pedidos_upserted = stats['pedidos_upserted']
        sync_run.pedido_detalles_upserted = stats['pedido_detalles_upserted']
        sync_run.remisiones_upserted = stats['remisiones_upserted']
        sync_run.remision_detalles_upserted = stats['remision_detalles_upserted']
        sync_run.notas_venta_upserted = stats['notas_venta_upserted']
        db.session.commit()

        logger.info(f"[CONTPAQ] Sincronizacion OK run={sync_run.id} stats={stats}")
        return {'ok': True, 'run_id': sync_run.id, 'stats': stats}
    except Exception as exc:
        db.session.rollback()
        logger.error(f'[CONTPAQ] Error de sincronizacion: {exc}', exc_info=True)
        run = ContpaqSyncRun.query.get(sync_run.id)
        if run:
            run.status = 'error'
            run.finished_at = datetime.utcnow()
            run.message = str(exc)
            db.session.add(run)
            db.session.commit()
        return {'ok': False, 'error': str(exc), 'run_id': sync_run.id}
    finally:
        _CONTPAQ_SYNC_LOCK.release()


def _contpaq_scheduler_loop():
    logger.info(
        f"[CONTPAQ] Scheduler activo cada {CONTPAQ_SYNC_INTERVAL_MINUTES} min "
        f"(delay inicial {CONTPAQ_SYNC_STARTUP_DELAY_SECONDS}s)"
    )
    waited = 0
    while waited < CONTPAQ_SYNC_STARTUP_DELAY_SECONDS and not _CONTPAQ_SYNC_STOP.is_set():
        sleep(1)
        waited += 1

    while not _CONTPAQ_SYNC_STOP.is_set():
        with app.app_context():
            run_contpaq_sync(trigger='scheduler')

        total_wait = CONTPAQ_SYNC_INTERVAL_MINUTES * 60
        waited = 0
        while waited < total_wait and not _CONTPAQ_SYNC_STOP.is_set():
            sleep(1)
            waited += 1


def _start_contpaq_scheduler_once():
    global _CONTPAQ_SYNC_THREAD
    global _CONTPAQ_SCHEDULER_INIT

    if _CONTPAQ_SCHEDULER_INIT:
        return
    _CONTPAQ_SCHEDULER_INIT = True

    if not CONTPAQ_SYNC_ENABLED:
        logger.info('[CONTPAQ] Scheduler deshabilitado por CONTPAQ_SYNC_ENABLED=0')
        return

    _CONTPAQ_SYNC_THREAD = threading.Thread(
        target=_contpaq_scheduler_loop,
        name='contpaq-sync-scheduler',
        daemon=True,
    )
    _CONTPAQ_SYNC_THREAD.start()


# Registrar accesos (IP, UA, path) en cada petición - evita estáticos
@app.before_request
def log_access_y_cierre_por_hora():
    try:
        _start_contpaq_scheduler_once()
        _start_machine_schedule_scheduler_once()
        _start_odoo_scheduler_once()

        path = request.path
        # skip static files and health checks
        if path.startswith('/static') or path.startswith('/favicon'):
            return
        # sync push: payload grande; no registrar en access_logs
        if path.startswith('/api/contpaq/') and path.endswith('/sync/push'):
            return

        _apply_machine_schedule()

        # get client ip (respect X-Forwarded-For when behind proxy)
        if request.headers.get('X-Forwarded-For'):
            client_ip = request.headers.get('X-Forwarded-For').split(',')[0].strip()
        else:
            client_ip = request.remote_addr

        ua = request.headers.get('User-Agent') or ''
        referer = request.headers.get('Referer')
        username = session.get('user') if 'user' in session else None

        # Skip bots/crawlers — they flood access_logs and fill the disk
        _BOT_KEYWORDS = ('bot', 'crawl', 'spider', 'slurp', 'mediapartners',
                         'facebookexternalhit', 'scan', 'check', 'monitoring')
        if any(kw in ua.lower() for kw in _BOT_KEYWORDS):
            return

        # Create log entry
        from models import AccessLog
        entry = AccessLog(
            ip=client_ip,
            username=username,
            path=path,
            method=request.method,
            user_agent=ua,
            referer=referer
        )
        db.session.add(entry)
        db.session.commit()

        # Periodic cleanup: keep only last 50,000 rows (runs ~1% of requests)
        import random
        if random.random() < 0.01:
            try:
                db.session.execute(db.text(
                    "DELETE FROM access_logs WHERE id NOT IN ("
                    "  SELECT id FROM access_logs ORDER BY timestamp DESC LIMIT 50000"
                    ")"
                ))
                db.session.commit()
            except Exception:
                db.session.rollback()
    except Exception:
        try:
            db.session.rollback()
        except Exception:
            pass
        # do not interrupt request flow on log error
        return

# ==================== DECORADOR DE AUTENTICACIÓN UNIFICADO ====================

def login_required(f):
    """Decorador para proteger rutas que requieren autenticación (cualquier usuario)"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user' not in session:
            # Si la ruta es una API, devolver JSON 401 en lugar de redirigir
            if request.path.startswith('/api/') or request.headers.get('Accept', '').find('application/json') != -1:
                return jsonify({'error': 'Autenticación requerida'}), 401
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def ingeniero_login_required(f):
    """DEPRECATED - usar @login_required en su lugar. Redirige a /login para consistencia."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user' not in session:
            if request.path.startswith('/api/') or request.headers.get('Accept', '').find('application/json') != -1:
                return jsonify({'error': 'Autenticación requerida'}), 401
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# ==================== RUTAS DE AUTENTICACIÓN UNIFICADAS ====================

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Login unificado para todos los usuarios (admin, ingenieros, etc.)
    Redirige automáticamente según el rol del usuario.
    """
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        logger.info(f"[LOGIN UNIFICADO] Intento de login para usuario: {username}")
        
        # Rate-limit by IP (saltarse en modo testing)
        client_ip = request.headers.get('X-Forwarded-For', request.remote_addr)
        
        if not app.config.get('TESTING', False):
            entry = FAILED_LOGINS.get(client_ip, [0, 0, 0])
            attempt_count, first_ts, locked_until = entry
            now = time()

            if first_ts and now - first_ts > LOCKOUT_SECONDS:
                attempt_count, first_ts, locked_until = 0, 0, 0

            if locked_until and now < locked_until:
                remaining = int(locked_until - now)
                logger.warning(f"[LOGIN] IP {client_ip} bloqueada. Faltan {remaining}s")
                return render_template('login.html', error=f'Demasiados intentos. Intenta en {remaining}s.'), 429
        
        credentials_valid = False
        usuario = None
        try:
            logger.debug(f"[LOGIN] Consultando usuario '{username}' en BD...")
            usuario = Usuario.query.filter_by(username=username, activo=True).first()
            if usuario and usuario.check_password(password):
                credentials_valid = True
                logger.info(f"[LOGIN] ✓ Autenticación exitosa para {username}")
            else:
                logger.warning(f"[LOGIN] ✗ Credenciales inválidas para {username}")
        except Exception as bd_error:
            logger.error(f"[BD ERROR] Fallo en login: {bd_error}", exc_info=True)
            # Fallback para testing
            if username == 'admin' and password == 'admin123':
                credentials_valid = True
                logger.info(f"[LOGIN] ✓ Fallback credentials para {username}")
        except Exception as e:
            logger.error(f"[LOGIN ERROR] {e}", exc_info=True)
            return render_template('login.html', error='Error interno del servidor.'), 500
        
        if credentials_valid and usuario:
            # Limpiar contador de intentos fallidos
            if not app.config.get('TESTING', False) and client_ip in FAILED_LOGINS:
                FAILED_LOGINS.pop(client_ip, None)
            
            # Guardar en sesión unificada
            session['user'] = username
            logger.info(f"[LOGIN] ✓ Sesión iniciada para {username}")
            
            endpoint = _resolve_post_login_endpoint(usuario)
            if endpoint:
                logger.info(f"[LOGIN] Redirigiendo usuario '{username}' a /{endpoint}")
                return redirect(url_for(endpoint))

            logger.warning(f"[LOGIN] Usuario '{username}' autenticado sin módulos asignados; redirigiendo a /dashboard")
            return redirect(url_for('dashboard'))
        else:
            # Gestionar intentos fallidos
            if not app.config.get('TESTING', False):
                attempt_count += 1
                if not first_ts:
                    first_ts = now
                if attempt_count >= MAX_LOGIN_ATTEMPTS:
                    locked_until = now + LOCKOUT_SECONDS
                    FAILED_LOGINS[client_ip] = [attempt_count, first_ts, locked_until]
                    logger.warning(f"[LOGIN] IP {client_ip} bloqueada por {MAX_LOGIN_ATTEMPTS} intentos")
                    return render_template('login.html', error='Demasiados intentos. Intenta más tarde.'), 429
                else:
                    FAILED_LOGINS[client_ip] = [attempt_count, first_ts, 0]
            return render_template('login.html', error='Credenciales inválidas', intento=attempt_count if not app.config.get('TESTING') else None), 401
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    """Cerrar sesión (para todos los usuarios)"""
    session.pop('user', None)
    # Limpiar keys antiguas si existen (para compatibilidad)
    session.pop('admin_user', None)
    session.pop('ingeniero_user', None)
    return redirect(url_for('login'))
# ==================== PERMISSION HELPERS (UNIFICADOS) ==================

def get_current_user():
    """Obtiene el usuario actual desde la sesión unificada."""
    username = session.get('user')
    if not username:
        return None
    try:
        return Usuario.query.filter_by(username=username, activo=True).first()
    except Exception:
        return None


def _resolve_post_login_endpoint(user):
    """Choose landing page based on effective module permissions."""
    if not user:
        return None
    if user.es_admin:
        return 'admin'

    if user.has_permission('maquinaria_pedidos', 'view'):
        return 'maquinaria_pedidos_page'
    if user.has_permission('maquinaria_ordenes', 'view'):
        return 'maquinaria_ordenes_page'
    if user.has_permission('maquinaria_boms', 'view'):
        return 'maquinaria_boms_page'
    if user.has_permission('maquinaria_procesos', 'view'):
        return 'maquinaria_procesos_maquina_page'
    if user.has_permission('maquinaria_claves_procesos', 'view'):
        return 'maquinaria_claves_procesos_page'
    if user.has_permission('maquinaria_estaciones', 'view'):
        return 'maquinaria_estaciones_page'
    if user.has_permission('maquinaria_calidad', 'view'):
        return 'maquinaria_calidad_page'
    if user.has_permission('maquinaria_seriado', 'view'):
        return 'maquinaria_seriado_page'
    if user.has_permission('maquinaria_almacen', 'view'):
        return 'maquinaria_almacen_page'

    if user.has_permission('catalog', 'view'):
        return 'index'
    if user.has_permission('estaciones', 'view'):
        return 'hojas_ruta_list'
    if user.has_permission('mapa', 'view'):
        return 'mapa_maquinas'
    if user.has_permission('alertas_buzon', 'view'):
        return 'alertas_buzon_page'
    if user.has_permission('hojas_entregas', 'view') or user.has_permission('hojas_mp', 'view') or user.has_permission('hojas', 'view'):
        return 'hojas_ruta_form'
    if user.has_permission('calidad', 'view'):
        return 'control_calidad_list'
    if user.has_permission('entregas', 'view'):
        return 'entregas_module'
    if user.has_permission('almacen', 'view'):
        return 'almacen_module'
    if user.has_permission('facturacion', 'view'):
        return 'facturacion_module'
    if user.has_permission('contpaq', 'view'):
        return 'contpaq_conciliacion_page'
    if user.has_permission('tickets', 'view'):
        return 'soporte_tecnico'
    if user.has_permission('users', 'view') or user.has_permission('roles', 'view') or user.has_permission('permissions', 'view'):
        return 'admin_users_page'
    if user.has_permission('proveedores', 'view') or user.has_permission('proveedores', 'edit'):
        return 'proveedores'

    return None


@app.context_processor
def inject_user_helpers():
    """Inyecta `current_user`, `has_permission` y `alertas_pendientes_count` en todas las plantillas."""
    user = get_current_user()

    def has_permission(module, action):
        if not user:
            return False
        return user.has_permission(module, action)

    alertas_pendientes_count = 0
    if user and (user.es_admin or (user.has_permission('alertas_buzon', 'view'))):
        try:
            alertas_pendientes_count = db.session.execute(
                db.text('SELECT COUNT(*) FROM alertas_buzon_general WHERE atendida = false')
            ).scalar() or 0
        except Exception:
            alertas_pendientes_count = 0

    return dict(
        current_user=user,
        has_permission=has_permission,
        alertas_pendientes_count=int(alertas_pendientes_count),
    )


def requires_permission(module, action):
    """Decorador para requerir un permiso específico en una ruta o API.
    
    Uso: @requires_permission('tickets', 'view')
    """
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            user = get_current_user()
            if not user:
                if request.path.startswith('/api/'):
                    return jsonify({'error': 'Autenticación requerida'}), 401
                return redirect(url_for('login'))
            if not user.has_permission(module, action):
                if request.path.startswith('/api/'):
                    return jsonify({'error': 'Permiso denegado'}), 403
                return render_template('403.html'), 403
            return f(*args, **kwargs)
        return decorated
    return decorator


def requires_any_permission(permission_pairs):
    """Allow access if user has at least one permission pair from the list."""
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            user = get_current_user()
            if not user:
                if request.path.startswith('/api/'):
                    return jsonify({'error': 'Autenticación requerida'}), 401
                return redirect(url_for('login'))

            pairs = list(permission_pairs or [])
            # En usuarios no admin, evita bypass por permisos de catálogo cuando
            # la ruta también declara permisos de otro módulo.
            if not user.es_admin:
                non_catalog_pairs = [pair for pair in pairs if pair[0] != 'catalog']
                if non_catalog_pairs:
                    pairs = non_catalog_pairs

            if not any(user.has_permission(module, action) for module, action in pairs):
                if request.path.startswith('/api/'):
                    return jsonify({'error': 'Permiso denegado'}), 403
                return render_template('403.html'), 403
            return f(*args, **kwargs)
        return decorated
    return decorator


# Module bundles for admin role-management UI.
ROLE_MODULE_BUNDLES = {
    'catalog': [('catalog', 'view')],
    'admin_catalog': [('catalog', 'view'), ('catalog', 'edit')],
    'admin_panel_view': [('admin', 'view')],
    'users_view': [('users', 'view')],
    'users_edit': [('users', 'view'), ('users', 'edit')],
    'roles_view': [('roles', 'view')],
    'roles_edit': [('roles', 'view'), ('roles', 'edit')],
    'permissions_view': [('permissions', 'view')],
    'permissions_edit': [('permissions', 'view'), ('permissions', 'edit')],
    'admin_user_manager_view': [('admin', 'view'), ('users', 'view'), ('roles', 'view'), ('permissions', 'view')],
    'admin_user_manager_edit': [('admin', 'view'), ('users', 'edit'), ('roles', 'edit'), ('permissions', 'edit')],
    'hojas_readonly': [('catalog', 'view'), ('hojas', 'view')],
    'hojas_generacion': [('catalog', 'view'), ('hojas', 'view'), ('hojas', 'create')],
    'hojas_edicion_basica': [('catalog', 'view'), ('hojas', 'view'), ('hojas', 'edit_basico')],
    'hojas_edicion_firmas': [('catalog', 'view'), ('hojas', 'view'), ('hojas', 'edit_firmas')],
    'hojas_edicion_planeacion': [('catalog', 'view'), ('hojas', 'view'), ('hojas', 'edit_planeacion')],
    'hojas_edicion_estado': [('catalog', 'view'), ('hojas', 'view'), ('hojas', 'edit_estado')],
    'hojas_edicion': [('catalog', 'view'), ('hojas', 'view'), ('hojas', 'edit')],
    'hojas_eliminar': [('catalog', 'view'), ('hojas', 'view'), ('hojas', 'delete')],
    'estaciones_view': [('catalog', 'view'), ('estaciones', 'view')],
    'estaciones_operate': [('catalog', 'view'), ('estaciones', 'view'), ('estaciones', 'operate')],
    'mapa_view': [('catalog', 'view'), ('mapa', 'view')],
    'alertas_buzon_view': [('alertas_buzon', 'view')],
    'calidad_view': [('catalog', 'view'), ('calidad', 'view')],
    'calidad_edit': [('catalog', 'view'), ('calidad', 'view'), ('calidad', 'edit')],
    'entregas_view': [('catalog', 'view'), ('entregas', 'view')],
    'entregas_edit': [('catalog', 'view'), ('entregas', 'view'), ('entregas', 'edit')],
    'almacen_view': [('catalog', 'view'), ('almacen', 'view')],
    'almacen_edit': [('catalog', 'view'), ('almacen', 'view'), ('almacen', 'edit')],
    'facturacion_view': [('catalog', 'view'), ('facturacion', 'view')],
    'facturacion_edit': [('catalog', 'view'), ('facturacion', 'view'), ('facturacion', 'edit')],
    'contpaq_view': [('contpaq', 'view')],
    'contpaq_edit': [('contpaq', 'view'), ('contpaq', 'edit')],
    'procesos_view': [('catalog', 'view'), ('procesos', 'view')],
    'proveedores_view': [('catalog', 'view'), ('proveedores', 'view')],
    'proveedores_edit': [('catalog', 'view'), ('proveedores', 'view'), ('proveedores', 'edit')],
    'soporte': [('tickets', 'view')],
    'soporte_edit': [('tickets', 'view'), ('tickets', 'edit'), ('tickets', 'export')],
    # Perfil sugerido: puede operar calidad/estaciones y solo leer hojas de ruta.
    'supervisor_produccion': [
        ('catalog', 'view'),
        ('hojas', 'view'),
        ('estaciones', 'view'),
        ('estaciones', 'operate'),
        ('calidad', 'view'),
        ('calidad', 'edit'),
    ],
}


DEFAULT_PERMISSION_CATALOG = [
    ('admin', 'view', 'Ver panel de administracion'),
    ('users', 'view', 'Ver gestion de usuarios'),
    ('users', 'edit', 'Editar usuarios'),
    ('roles', 'view', 'Ver roles'),
    ('roles', 'edit', 'Editar roles'),
    ('permissions', 'view', 'Ver permisos'),
    ('permissions', 'edit', 'Editar permisos'),
    ('catalog', 'view', 'Ver catalogo'),
    ('catalog', 'edit', 'Editar catalogo'),
    ('hojas', 'view', 'Ver hojas de ruta'),
    ('hojas', 'create', 'Crear hojas de ruta'),
    ('hojas', 'edit', 'Editar hojas de ruta (total)'),
    ('hojas', 'delete', 'Eliminar hojas de ruta'),
    ('hojas', 'edit_basico', 'Editar hoja: serie/comentarios'),
    ('hojas', 'edit_firmas', 'Editar hoja: firmas'),
    ('hojas', 'edit_planeacion', 'Editar hoja: planeacion'),
    ('hojas', 'edit_estado', 'Editar hoja: estado'),
    ('hojas', 'edit_field_nombre', 'Editar campo hoja: nombre'),
    ('hojas', 'edit_field_descripcion', 'Editar campo hoja: descripcion'),
    ('hojas', 'edit_field_comentarios', 'Editar campo hoja: comentarios'),
    ('hojas', 'edit_field_firma_ing_jose', 'Editar campo hoja: firma ing jose'),
    ('hojas', 'edit_field_firma_ing_rodrigo', 'Editar campo hoja: firma ing rodrigo'),
    ('hojas', 'edit_field_calidad', 'Editar campo hoja: calidad'),
    ('hojas', 'edit_field_almacen', 'Editar campo hoja: almacen'),
    ('hojas', 'edit_field_orden_trabajo', 'Editar campo hoja: orden de trabajo'),
    ('hojas', 'edit_field_cantidad_piezas', 'Editar campo hoja: cantidad de piezas'),
    ('hojas', 'edit_field_estado', 'Editar campo hoja: estado'),
    ('estaciones', 'view', 'Ver estaciones T'),
    ('estaciones', 'operate', 'Operar estaciones/maquinas'),
    ('mapa', 'view', 'Ver mapa de maquinas'),
    ('alertas_buzon', 'view', 'Ver buzon de alertas'),
    ('calidad', 'view', 'Ver control de calidad'),
    ('calidad', 'edit', 'Registrar/editar revision de calidad'),
    ('entregas', 'view', 'Ver módulo de entregas'),
    ('entregas', 'edit', 'Operar módulo de entregas'),
    ('almacen', 'view', 'Ver módulo de almacén'),
    ('almacen', 'edit', 'Operar módulo de almacén'),
    ('facturacion', 'view', 'Ver módulo de facturación'),
    ('facturacion', 'edit', 'Operar módulo de facturación'),
    ('contpaq', 'view', 'Ver módulo de conciliación CONTPAQ'),
    ('contpaq', 'edit', 'Operar conciliación CONTPAQ'),
    ('empaque', 'view', 'Ver clientes Empaque y claves de acceso'),
    ('empaque', 'edit', 'Generar/regenerar claves Empaque'),
    ('procesos', 'view', 'Ver procesos y claves'),
    ('proveedores', 'view', 'Ver proveedores'),
    ('proveedores', 'edit', 'Editar proveedores'),
    ('tickets', 'view', 'Ver tickets'),
    ('tickets', 'edit', 'Editar tickets'),
    ('tickets', 'export', 'Exportar tickets'),
]

# Matriz simple de permisos por modulo: ver/crear/editar/actualizar/borrar.
SIMPLE_PERMISSION_MODULES = [
    ('admin', 'Panel admin'),
    ('users', 'Usuarios'),
    ('roles', 'Roles'),
    ('permissions', 'Permisos'),
    ('catalog', 'Catalogo'),
    ('hojas_entregas', 'Hojas de ruta entregas'),
    ('hojas_mp', 'Hojas de ruta MP'),
    ('estaciones', 'Estaciones T'),
    ('mapa', 'Mapa de maquinas'),
    ('alertas_buzon', 'Buzon de alertas'),
    ('calidad', 'Control de calidad'),
    ('entregas', 'Entregas'),
    ('almacen', 'Almacen'),
    ('facturacion', 'Facturacion'),
    ('contpaq', 'CONTPAQ Conciliacion'),
    ('empaque', 'Empaque - Clientes y seguimiento'),
    ('procesos', 'Procesos y claves'),
    ('proveedores', 'Proveedores'),
    ('tickets', 'Tickets'),
    ('maquinaria_pedidos', 'Maquinaria y ensamble - Pedidos'),
    ('maquinaria_ordenes', 'Maquinaria y ensamble - Ordenes de trabajo'),
    ('maquinaria_boms', 'Maquinaria y ensamble - BOMs'),
    ('maquinaria_procesos', 'Maquinaria y ensamble - Procesos por maquina'),
    ('maquinaria_claves_procesos', 'Maquinaria y ensamble - Claves y procesos'),
    ('maquinaria_estaciones', 'Maquinaria y ensamble - Estaciones de trabajo'),
    ('maquinaria_calidad', 'Maquinaria y ensamble - Calidad'),
    ('maquinaria_seriado', 'Maquinaria y ensamble - Seriado'),
    ('maquinaria_almacen', 'Maquinaria y ensamble - Almacen'),
]

SIMPLE_PERMISSION_ACTIONS = [
    ('view', 'Ver'),
    ('create', 'Crear'),
    ('edit', 'Editar'),
    ('update', 'Actualizar'),
    ('delete', 'Borrar'),
]

for module, module_desc in SIMPLE_PERMISSION_MODULES:
    for action, action_desc in SIMPLE_PERMISSION_ACTIONS:
        ROLE_MODULE_BUNDLES.setdefault(f'{module}_{action}', [(module, action)])
        exists = any((m == module and a == action) for m, a, _ in DEFAULT_PERMISSION_CATALOG)
        if not exists:
            DEFAULT_PERMISSION_CATALOG.append((module, action, f'{action_desc} {module_desc}'))

# Bundle simple por módulo para UI sin desglose por acción.
for module, _module_desc in SIMPLE_PERMISSION_MODULES:
    module_perms = sorted({(m, a) for (m, a, _d) in DEFAULT_PERMISSION_CATALOG if m == module})
    if module_perms:
        ROLE_MODULE_BUNDLES.setdefault(module, module_perms)


HOJA_FIELD_GROUP_ACTIONS = {
    'estado': 'edit_estado',
    'nombre': 'edit_basico',
    'descripcion': 'edit_basico',
    'comentarios': 'edit_basico',
    'firma_ing_jose': 'edit_firmas',
    'firma_ing_rodrigo': 'edit_firmas',
    'calidad': 'edit_planeacion',
    'almacen': 'edit_planeacion',
    'orden_trabajo': 'edit_planeacion',
    'cantidad_piezas': 'edit_planeacion',
}


HOJA_FIELD_SPECIFIC_ACTIONS = {
    'estado': 'edit_field_estado',
    'nombre': 'edit_field_nombre',
    'descripcion': 'edit_field_descripcion',
    'comentarios': 'edit_field_comentarios',
    'firma_ing_jose': 'edit_field_firma_ing_jose',
    'firma_ing_rodrigo': 'edit_field_firma_ing_rodrigo',
    'calidad': 'edit_field_calidad',
    'almacen': 'edit_field_almacen',
    'orden_trabajo': 'edit_field_orden_trabajo',
    'cantidad_piezas': 'edit_field_cantidad_piezas',
}


def _role_modules_from_permissions(role):
    perm_set = {(p.module, p.action) for p in (role.permissions or [])}
    modules = []
    for module_id, required_perms in ROLE_MODULE_BUNDLES.items():
        if all(rp in perm_set for rp in required_perms):
            modules.append(module_id)
    return modules


def _apply_role_modules(role, modules):
    selected = set(modules or [])
    required = set()
    for module_id in selected:
        for perm_pair in ROLE_MODULE_BUNDLES.get(module_id, []):
            required.add(perm_pair)

    # Auto-grant view for any module that has a non-view action assigned.
    # Without view, the user can't even load the page to use edit/create/etc.
    modules_with_action = {module for module, action in required if action != 'view'}
    modules_with_view = {module for module, action in required if action == 'view'}
    for module in modules_with_action - modules_with_view:
        required.add((module, 'view'))

    # Preserve unrelated permissions and ensure required ones are present.
    existing = {(p.module, p.action): p for p in (role.permissions or [])}
    current_perms = list(role.permissions or [])

    for module, action in required:
        if (module, action) in existing:
            continue
        perm = Permission.query.filter_by(module=module, action=action).first()
        if not perm:
            perm = Permission(module=module, action=action, descripcion=f'{module}:{action}')
            db.session.add(perm)
            db.session.flush()
        current_perms.append(perm)

    # Remove permissions managed by bundles if not currently selected.
    bundle_pairs = {pair for pairs in ROLE_MODULE_BUNDLES.values() for pair in pairs}
    filtered = []
    for p in current_perms:
        pair = (p.module, p.action)
        if pair in bundle_pairs and pair not in required:
            continue
        filtered.append(p)

    role.permissions = filtered

def _get_or_create_permission(module, action):
    perm = Permission.query.filter_by(module=module, action=action).first()
    if perm:
        return perm
    perm = Permission(module=module, action=action, descripcion=f'{module}:{action}')
    db.session.add(perm)
    db.session.flush()
    return perm


def _ensure_default_permissions():
    """Create baseline permissions so admin UI can assign fine-grained access."""
    created_any = False
    for module, action, descripcion in DEFAULT_PERMISSION_CATALOG:
        perm = Permission.query.filter_by(module=module, action=action).first()
        if perm:
            continue
        db.session.add(Permission(module=module, action=action, descripcion=descripcion))
        created_any = True
    if created_any:
        db.session.commit()


def _check_field_level_permissions(user, module, payload, field_actions, broad_action='edit'):
    """Return denied fields list for payload updates in a module."""
    denied = []
    for field, required_actions in (field_actions or {}).items():
        if field not in (payload or {}):
            continue
        if user.has_permission(module, broad_action):
            continue
        actions_list = required_actions if isinstance(required_actions, (list, tuple, set)) else [required_actions]
        if any(user.has_permission(module, action) for action in actions_list):
            continue
        denied.append(field)
    return denied


def _hoja_ready_for_signatures(hoja):
    """Validate that hoja is fully completed and quality-approved for non-admin signatures."""
    estaciones = EstacionTrabajo.query.filter_by(hoja_ruta_id=hoja.id).all()
    if not estaciones:
        return False, 'La hoja no tiene procesos para validar.'

    for est in estaciones:
        if (est.estado or '').lower() != 'completada':
            return False, 'Todos los procesos deben estar completados antes de autorizar firmas.'

    for est in estaciones:
        notas = est.notas or ''
        m = re.search(r'STATUS=(QC_OK|QC_NOK)', notas)
        if not m:
            return False, 'Falta liberar procesos en calidad (QC) antes de autorizar firmas.'
        if m.group(1) != 'QC_OK':
            return False, 'Existen procesos con rechazo en calidad. No se puede autorizar la hoja.'

    return True, None


def _permissions_from_modules_and_ids(modules, permission_ids):
    """Resolve final permission set from module bundles and explicit permission IDs."""
    final_pairs = set()

    for module_id in (modules or []):
        for module, action in ROLE_MODULE_BUNDLES.get(module_id, []):
            final_pairs.add((module, action))

    if permission_ids:
        explicit_perms = Permission.query.filter(Permission.id.in_(permission_ids)).all()
        for p in explicit_perms:
            final_pairs.add((p.module, p.action))

    resolved = []
    seen = set()
    for module, action in sorted(final_pairs):
        perm = _get_or_create_permission(module, action)
        pair = (perm.module, perm.action)
        if pair in seen:
            continue
        seen.add(pair)
        resolved.append(perm)
    return resolved


def _can_view_admin_panel(user):
    if not user:
        return False
    if user.es_admin:
        return True
    return user.has_permission('admin', 'view') or user.has_permission('catalog', 'edit')


def _can_view_users_module(user):
    if not user:
        return False
    if user.es_admin:
        return True
    return user.has_permission('users', 'view') or user.has_permission('users', 'edit')


def _can_edit_users_module(user):
    if not user:
        return False
    if user.es_admin:
        return True
    return user.has_permission('users', 'edit')


def _can_view_roles_module(user):
    if not user:
        return False
    if user.es_admin:
        return True
    return user.has_permission('roles', 'view') or user.has_permission('roles', 'edit')


def _can_edit_roles_module(user):
    if not user:
        return False
    if user.es_admin:
        return True
    return user.has_permission('roles', 'edit')


def _can_view_permissions_module(user):
    if not user:
        return False
    if user.es_admin:
        return True
    return user.has_permission('permissions', 'view') or user.has_permission('permissions', 'edit')


def _can_edit_permissions_module(user):
    if not user:
        return False
    if user.es_admin:
        return True
    return user.has_permission('permissions', 'edit')


def _can_view_user_admin(user):
    if not user:
        return False
    return (
        _can_view_users_module(user)
        or _can_view_roles_module(user)
        or _can_view_permissions_module(user)
    )


def _can_edit_user_admin(user):
    if not user:
        return False
    return (
        _can_edit_users_module(user)
        or _can_edit_roles_module(user)
        or _can_edit_permissions_module(user)
    )

# ==================== DASHBOARD / HOME CENTRAL ====================

@app.route('/dashboard')
@login_required
def dashboard():
    """Dashboard central - cada usuario ve su panel según permisos."""
    user = get_current_user()
    if not user:
        return redirect(url_for('login'))

    endpoint = _resolve_post_login_endpoint(user)
    if endpoint and endpoint != 'dashboard':
        return redirect(url_for(endpoint))

    has_any_module = bool(user and user.role and user.role.permissions)
    return render_template('dashboard.html', user=user, has_any_module=has_any_module)

# ==================== RUTAS FRONTEND ====================

@app.route('/')
@login_required
@requires_permission('catalog', 'view')
def index():
    """Página principal - Catálogo (privado)"""
    return render_template('index.html')


@app.route('/hojas_ruta_form')
@login_required
@requires_any_permission([('hojas_entregas', 'view'), ('hojas', 'view'), ('catalog', 'view')])
def hojas_ruta_form():
    """Alias del formulario legacy para mantener compatibilidad con navbar y enlaces existentes."""
    return hojas_ruta_entregas_form()


@app.route('/producto/<int:producto_id>')
@login_required
def producto_detalle(producto_id):
    producto = Producto.query.get_or_404(producto_id)
    # Traer proveedores y precios relacionados
    proveedores = []
    def _pp_date_key(pp):
        return pp.fecha_precio or datetime.min.date()
    proveedores_ordenados = sorted(producto.proveedores, key=_pp_date_key, reverse=True)
    for pp in proveedores_ordenados:
        prov = pp.proveedor.to_dict() if pp.proveedor else None
        historial = [h.to_dict() for h in pp.historial_precios]
        proveedores.append({
            'asignacion_id': pp.id,
            'proveedor': prov,
            'precio_proveedor': pp.precio_proveedor,
            'fecha_precio': pp.fecha_precio.isoformat() if pp.fecha_precio else None,
            'cantidad_minima': pp.cantidad_minima,
            'divisa': pp.divisa,
            'historial_precios': historial
        })
    ultima_compra = proveedores[0] if proveedores else None
    return render_template('producto_detalle.html', producto=producto, proveedores=proveedores, ultima_compra=ultima_compra)


@app.route('/control_calidad')
@login_required
@requires_any_permission([('calidad', 'view'), ('catalog', 'view')])
def control_calidad_list():
    """Lista de hojas de ruta con procesos completados pendientes de validación en calidad."""

    def qc_status(estacion):
        notas = estacion.notas or ''
        m = re.search(r'STATUS=(QC_OK|QC_NOK)', notas)
        return m.group(1) if m else None

    hojas = HojaRutaEntrega.query.order_by(HojaRutaEntrega.fecha_creacion.desc()).all()
    hojas_qc = []
    for hoja in hojas:
        estaciones = EstacionTrabajo.query.filter_by(hoja_ruta_id=hoja.id).order_by(EstacionTrabajo.orden.asc()).all()
        completadas = [e for e in estaciones if (e.estado or '').lower() == 'completada']
        if not completadas:
            continue

        reviewed = 0
        rechazadas = 0
        for e in completadas:
            status = qc_status(e)
            if status in ('QC_OK', 'QC_NOK'):
                reviewed += 1
            if status == 'QC_NOK':
                rechazadas += 1

        pendientes = max(0, len(completadas) - reviewed)
        finalizada_qc = len(completadas) > 0 and pendientes == 0
        certificado_disponible = finalizada_qc
        hojas_qc.append({
            'id': hoja.id,
            'serie': hoja.nombre,
            'clave': hoja.pn,
            'calidad': hoja.calidad,
            'piezas': hoja.cantidad_piezas,
            'estado': hoja.estado,
            'procesos_completados': len(completadas),
            'procesos_revisados': reviewed,
            'procesos_pendientes_qc': pendientes,
            'procesos_rechazados_qc': rechazadas,
            'finalizada_qc': finalizada_qc,
            'certificado_disponible': certificado_disponible,
            'fecha': hoja.fecha_creacion,
        })

    # Prioridad: hojas con pendientes QC arriba.
    hojas_qc = sorted(
        hojas_qc,
        key=lambda h: (0 if h['procesos_pendientes_qc'] > 0 else 1, -(h['procesos_pendientes_qc']), h['fecha'] or datetime.min),
        reverse=False,
    )

    return render_template('control_calidad_list.html', hojas_qc=hojas_qc)


@app.route('/control_calidad/hoja/<int:hoja_id>', methods=['GET', 'POST'])
@login_required
@requires_any_permission([('calidad', 'view'), ('catalog', 'view')])
def control_calidad_hoja(hoja_id):
    """Revisión de calidad por hoja de ruta y por proceso completado."""
    hoja = HojaRutaEntrega.query.get_or_404(hoja_id)
    user = get_current_user()
    can_edit_qc = bool(user and (user.has_permission('calidad', 'edit') or user.has_permission('calidad', 'update')))

    def qc_status(estacion):
        notas = estacion.notas or ''
        m = re.search(r'STATUS=(QC_OK|QC_NOK)', notas)
        return m.group(1) if m else None

    def qc_clean_block(notas_text):
        src = notas_text or ''
        # Reemplaza bloque QC previo para evitar crecimiento infinito de notas.
        return re.sub(r'\n?\[QC_REVIEW_START\].*?\[QC_REVIEW_END\]\n?', '\n', src, flags=re.S).strip()

    def qc_parse_review(estacion):
        notas_text = estacion.notas or ''
        status = qc_status(estacion)
        block_match = re.search(r'\[QC_REVIEW_START\](.*?)\[QC_REVIEW_END\]', notas_text, flags=re.S)
        block = block_match.group(1) if block_match else ''

        def _extract(field, default=''):
            m = re.search(rf'^{field}=(.*)$', block, flags=re.M)
            return m.group(1).strip() if m else default

        scrap_raw = _extract('SCRAP_PIEZAS', '0')
        try:
            scrap_piezas = int(scrap_raw)
        except Exception:
            scrap_piezas = 0
        if scrap_piezas < 0:
            scrap_piezas = 0

        return {
            'status': status,
            'scrap_piezas': scrap_piezas,
            'observaciones': _extract('OBSERVACIONES', ''),
        }

    def qc_extract_lote_inicial(comentarios_text):
        src = comentarios_text or ''
        m = re.search(r'^LOTE_INICIAL=(\d+)$', src, flags=re.M)
        if not m:
            return None
        try:
            return max(0, int(m.group(1)))
        except Exception:
            return None

    informe_guardado = False
    error_guardado = None

    def _is_checked(field_name):
        value = request.form.get(field_name)
        if value is None:
            return False
        return str(value).strip().lower() in ('1', 'true', 'on', 'yes', 'si', 'ok')

    if request.method == 'POST':
        if not can_edit_qc:
            error_guardado = 'Permiso denegado para editar control de calidad'
        else:
            estacion_id = request.form.get('estacion_id', type=int)
            resultado = (request.form.get('resultado') or '').strip().lower()
            observaciones = (request.form.get('observaciones') or '').strip()
            scrap_piezas = request.form.get('scrap_piezas', type=int)
            if scrap_piezas is None:
                scrap_piezas = 0

            estacion = EstacionTrabajo.query.filter_by(id=estacion_id, hoja_ruta_id=hoja.id).first()
            if not estacion:
                error_guardado = 'Proceso no encontrado para esta hoja'
            elif (estacion.estado or '').lower() != 'completada':
                error_guardado = 'Solo se pueden validar procesos completados'
            elif resultado not in ('aprobado', 'rechazado'):
                error_guardado = 'Selecciona un resultado válido (Aprobado/Rechazado)'
            elif scrap_piezas < 0:
                error_guardado = 'Scrap inválido: no puede ser negativo'
            else:
                # Checklist estándar para refacciones (sinfines, engranes, etc.)
                check_dimensional = _is_checked('chk_dimensional')
                check_visual = _is_checked('chk_visual')
                check_rebaba = _is_checked('chk_rebaba')
                check_material = _is_checked('chk_material')
                check_ajuste = _is_checked('chk_ajuste')
                check_limpieza = _is_checked('chk_limpieza')

                status_tag = 'QC_OK' if resultado == 'aprobado' else 'QC_NOK'
                usuario_qc = session.get('user') or 'sistema'
                ahora = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')

                bloque = (
                    "[QC_REVIEW_START]\n"
                    f"STATUS={status_tag}\n"
                    f"USUARIO={usuario_qc}\n"
                    f"FECHA={ahora}\n"
                    f"DIMENSIONAL={'OK' if check_dimensional else 'NO'}\n"
                    f"VISUAL={'OK' if check_visual else 'NO'}\n"
                    f"REBABA={'OK' if check_rebaba else 'NO'}\n"
                    f"MATERIAL={'OK' if check_material else 'NO'}\n"
                    f"AJUSTE={'OK' if check_ajuste else 'NO'}\n"
                    f"LIMPIEZA={'OK' if check_limpieza else 'NO'}\n"
                    f"SCRAP_PIEZAS={scrap_piezas}\n"
                    f"OBSERVACIONES={observaciones}\n"
                    "[QC_REVIEW_END]"
                )

                base_notas = qc_clean_block(estacion.notas)
                estacion.notas = (base_notas + "\n" + bloque).strip() if base_notas else bloque
                estacion.firma_supervisor = status_tag
                estacion.operador = usuario_qc

                # Estado de hoja por consolidación QC de procesos completados
                completadas = EstacionTrabajo.query.filter_by(hoja_ruta_id=hoja.id, estado='completada').all()
                reviews = [qc_parse_review(e) for e in completadas]
                statuses = [r['status'] for r in reviews]
                pendientes = any(s not in ('QC_OK', 'QC_NOK') for s in statuses)
                hay_rechazo = any(s == 'QC_NOK' for s in statuses)

                # Lote base fijo para evitar descuentos acumulados por re-ediciones.
                lote_inicial = qc_extract_lote_inicial(hoja.materia_prima)
                if lote_inicial is None:
                    try:
                        lote_inicial = max(0, int(hoja.cantidad_piezas or 0))
                    except Exception:
                        lote_inicial = 0

                if pendientes:
                    hoja.aprobada = False
                    hoja.rechazada = False
                else:
                    hoja.aprobada = not hay_rechazo
                    hoja.rechazada = hay_rechazo

                if hoja.aprobada and not hoja.rechazada:
                    total_scrap = sum(max(0, int(r['scrap_piezas'] or 0)) for r in reviews)
                    if lote_inicial > 0:
                        total_scrap = min(total_scrap, lote_inicial)
                    cantidad_final = max(lote_inicial - total_scrap, 0)
                    hoja.cantidad_piezas = cantidad_final
                    hoja.scrap = str(total_scrap)

                    detalles = []
                    for est, rev in zip(completadas, reviews):
                        scrap_est = max(0, int(rev['scrap_piezas'] or 0))
                        if scrap_est <= 0:
                            continue
                        etiqueta = (est.centro_trabajo or est.operacion or f'Estacion {est.orden}').strip()
                        detalles.append(f"P{est.orden} - {etiqueta}: {scrap_est}")

                    detalles_txt = '; '.join(detalles) if detalles else 'Sin scrap por proceso.'
                    fecha_resumen = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')
                    bloque_scrap = (
                        "[QC_SCRAP_SUMMARY_START]\n"
                        f"LOTE_INICIAL={lote_inicial}\n"
                        f"TOTAL_SCRAP={total_scrap}\n"
                        f"LOTE_FINAL={cantidad_final}\n"
                        f"DETALLE={detalles_txt}\n"
                        f"ACTUALIZADO_POR={usuario_qc}\n"
                        f"FECHA={fecha_resumen}\n"
                        "[QC_SCRAP_SUMMARY_END]"
                    )

                    base_comentarios = _qc_strip_scrap_summary(hoja.materia_prima)
                    hoja.materia_prima = (
                        (base_comentarios + "\n\n" + bloque_scrap).strip()
                        if base_comentarios else bloque_scrap
                    )
                else:
                    # Si aun no queda aprobada, quitar resumen previo para evitar inconsistencias visuales.
                    base_comentarios = _qc_strip_scrap_summary(hoja.materia_prima)
                    if base_comentarios != (hoja.materia_prima or '').strip():
                        hoja.materia_prima = base_comentarios or None
                    if lote_inicial > 0:
                        hoja.cantidad_piezas = lote_inicial
                    hoja.scrap = None

                try:
                    db.session.commit()
                    informe_guardado = True
                except Exception:
                    db.session.rollback()
                    logger.exception('[QC] Error guardando revision de calidad')
                    error_guardado = 'Ocurrio un error al guardar la revision. Intenta de nuevo.'

    estaciones = EstacionTrabajo.query.filter_by(hoja_ruta_id=hoja.id).order_by(EstacionTrabajo.orden.asc()).all()
    procesos_completados = []
    lote_referencia = qc_extract_lote_inicial(hoja.materia_prima)
    if lote_referencia is None:
        try:
            lote_referencia = max(0, int(hoja.cantidad_piezas or 0))
        except Exception:
            lote_referencia = 0

    for e in estaciones:
        if (e.estado or '').lower() != 'completada':
            continue
        review = qc_parse_review(e)
        procesos_completados.append({
            'id': e.id,
            'orden': e.orden,
            'centro_trabajo': e.centro_trabajo,
            'operacion': e.operacion,
            'status_qc': review['status'],
            'scrap_piezas': review['scrap_piezas'],
            'notas': e.notas,
            'fecha_finalizacion': e.fecha_finalizacion,
        })

    pendientes_qc = sum(1 for p in procesos_completados if p['status_qc'] not in ('QC_OK', 'QC_NOK'))
    qc_finalizada = len(procesos_completados) > 0 and pendientes_qc == 0
    certificado_disponible = qc_finalizada

    return render_template(
        'control_calidad_detalle.html',
        hoja=hoja,
        procesos=procesos_completados,
        can_edit_qc=can_edit_qc,
        lote_referencia=lote_referencia,
        pendientes_qc=pendientes_qc,
        qc_finalizada=qc_finalizada,
        certificado_disponible=certificado_disponible,
        informe_guardado=informe_guardado,
        error_guardado=error_guardado,
    )


@app.route('/control_calidad/hoja/<int:hoja_id>/certificado')
@login_required
@requires_any_permission([('calidad', 'view'), ('catalog', 'view')])
def control_calidad_certificado(hoja_id):
    """Certificado imprimible de control de calidad por hoja de ruta."""
    hoja = HojaRutaEntrega.query.get_or_404(hoja_id)
    estaciones = EstacionTrabajo.query.filter_by(hoja_ruta_id=hoja.id).order_by(EstacionTrabajo.orden.asc()).all()
    completadas = [e for e in estaciones if (e.estado or '').lower() == 'completada']
    reviews = [_qc_parse_review_block(e.notas) for e in completadas]

    qc_finalizada = bool(completadas) and all(r.get('status') in ('QC_OK', 'QC_NOK') for r in reviews)
    if not qc_finalizada:
        return redirect(url_for('control_calidad_hoja', hoja_id=hoja.id))

    checklist_fields = ('dimensional', 'visual', 'rebaba', 'material', 'ajuste', 'limpieza')
    checklist_global = {
        f: bool(reviews) and all(bool(r.get(f)) for r in reviews)
        for f in checklist_fields
    }

    defectos = []
    for est, rev in zip(completadas, reviews):
        status = rev.get('status')
        scrap_piezas = int(rev.get('scrap_piezas') or 0)
        observaciones = (rev.get('observaciones') or '').strip()

        detalle_partes = []
        if observaciones:
            detalle_partes.append(observaciones)
        if scrap_piezas > 0:
            detalle_partes.append(f"Scrap en proceso: {scrap_piezas}")

        if status == 'QC_NOK' or scrap_piezas > 0 or observaciones:
            defecto_detalle = ' | '.join(detalle_partes) if detalle_partes else 'No conforme detectado en inspeccion.'
            etiqueta_proceso = (est.centro_trabajo or est.operacion or f'Proceso {est.orden}').strip()
            defectos.append({
                'proceso': f"P{est.orden} - {etiqueta_proceso}",
                'resultado': 'RECHAZADO' if status == 'QC_NOK' else 'APROBADO',
                'detalle': defecto_detalle,
            })

    comentarios_usuario = _qc_strip_scrap_summary(hoja.materia_prima)
    scrap_qc = _qc_parse_scrap_summary(hoja.materia_prima)

    fecha_qc = ''
    qc_user = ''
    if reviews:
        ult = reviews[-1]
        fecha_qc = ult.get('fecha') or ''
        qc_user = ult.get('usuario') or ''

    fecha_emision = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')
    qc_aprobada = all(r.get('status') == 'QC_OK' for r in reviews)

    return render_template(
        'control_calidad_certificado.html',
        hoja=hoja,
        fecha_emision=fecha_emision,
        fecha_qc=fecha_qc,
        qc_user=qc_user,
        checklist_global=checklist_global,
        defectos=defectos,
        comentarios_usuario=comentarios_usuario,
        scrap_qc=scrap_qc,
        qc_aprobada=qc_aprobada,
    )


@app.route('/control_calidad/<int:maquina_id>')
@login_required
@requires_any_permission([('calidad', 'view'), ('catalog', 'view')])
def control_calidad_legacy_maquina(maquina_id):
    """Compatibilidad con links antiguos de Control Calidad por maquina."""
    return redirect(url_for('control_calidad_list'))


# ==================== FLUJO TEMPORAL: ENTREGAS -> ALMACEN -> ENTREGAS(LISTA) -> FACTURACION ====================

LOGISTICA_IMG_EXTENSIONS = {'png', 'jpg', 'jpeg', 'webp'}
LOGISTICA_MAX_IMAGE_SIDE = 1600
LOGISTICA_WEBP_QUALITY = 80


def _logistica_username():
    return (session.get('user') or 'sistema').strip()


def _logistica_allowed_image(filename):
    if not filename or '.' not in filename:
        return False
    ext = filename.rsplit('.', 1)[-1].lower().strip()
    return ext in LOGISTICA_IMG_EXTENSIONS


def _save_logistica_recepcion_image(file_storage, hoja_ruta_id):
    """Normaliza y comprime capturas de recepción para ahorrar espacio en servidor."""
    from PIL import Image, ImageOps

    image = Image.open(file_storage.stream)
    image = ImageOps.exif_transpose(image)

    # WebP soporta transparencia; convertimos modos incompatibles para un guardado estable.
    if image.mode not in ('RGB', 'RGBA'):
        image = image.convert('RGB')

    width, height = image.size
    max_side = max(width, height)
    if max_side > LOGISTICA_MAX_IMAGE_SIDE:
        scale = LOGISTICA_MAX_IMAGE_SIDE / float(max_side)
        image = image.resize(
            (max(1, int(width * scale)), max(1, int(height * scale))),
            Image.Resampling.LANCZOS,
        )

    nombre = secure_filename(f"recepcion_{hoja_ruta_id}_{uuid.uuid4().hex}.webp")
    rel_dir = os.path.join('logistica_recepciones')
    abs_dir = os.path.join(os.getcwd(), 'uploads', rel_dir)
    os.makedirs(abs_dir, exist_ok=True)
    abs_path = os.path.join(abs_dir, nombre)

    save_kwargs = {
        'format': 'WEBP',
        'quality': LOGISTICA_WEBP_QUALITY,
        'method': 6,
    }
    if image.mode == 'RGBA':
        save_kwargs['lossless'] = False

    image.save(abs_path, **save_kwargs)
    return f"{rel_dir}/{nombre}".replace('\\', '/')


def _serialize_logistica_capture_paths(paths):
    clean_paths = [str(path).strip() for path in (paths or []) if str(path).strip()]
    if not clean_paths:
        return None
    if len(clean_paths) == 1:
        return clean_paths[0]
    return json.dumps(clean_paths, ensure_ascii=False)


def _sync_flujo_parciales(flujo: HojaRutaFlujoLogistica, hoja: HojaRutaEntrega = None):
    """Sincroniza totales/pendientes/porcentaje para entregas parciales."""
    if not flujo:
        return

    hoja_ref = hoja or flujo.hoja_ruta
    total = int(flujo.cantidad_total_piezas or 0)
    if total <= 0 and hoja_ref and hoja_ref.cantidad_piezas:
        total = int(hoja_ref.cantidad_piezas or 0)
    if total < 0:
        total = 0

    entregado = int(flujo.cantidad_entregada or 0)
    if entregado < 0:
        entregado = 0
    if total > 0 and entregado > total:
        entregado = total

    flujo.cantidad_total_piezas = total
    flujo.cantidad_entregada = entregado
    flujo.cantidad_pendiente = max(total - entregado, 0)
    flujo.porcentaje_entregado = round((entregado / total) * 100, 2) if total > 0 else 0.0
    flujo.estado_parciales = 'todas' if total > 0 and entregado == total else 'pendientes'


def _build_logistica_resumen(limit=80):
    flujos = (
        HojaRutaFlujoLogistica.query
        .order_by(HojaRutaFlujoLogistica.fecha_actualizacion.desc())
        .limit(limit)
        .all()
    )
    hoja_ids = [item.hoja_ruta_id for item in flujos]

    devoluciones_almacen = set()
    devoluciones_facturacion = set()
    llego_facturacion = set()
    if hoja_ids:
        entrega_logs = (
            EntregaRegistro.query
            .filter(EntregaRegistro.hoja_ruta_id.in_(hoja_ids))
            .order_by(EntregaRegistro.fecha_creacion.desc())
            .all()
        )
        for log in entrega_logs:
            accion = (log.accion or '').strip().lower()
            if accion == 'devuelta_desde_almacen':
                devoluciones_almacen.add(log.hoja_ruta_id)
            elif accion == 'devuelta_desde_facturacion':
                devoluciones_facturacion.add(log.hoja_ruta_id)
            elif accion == 'enviada_a_facturacion':
                llego_facturacion.add(log.hoja_ruta_id)

    resumen = []
    for flujo in flujos:
        estado = (flujo.estado or '').strip().lower()
        hoja_id = flujo.hoja_ruta_id
        check_almacen = (
            estado in ('almacen', 'entregas_lista_facturacion', 'facturacion', 'finalizada')
            or bool(flujo.almacen_validado or flujo.almacen_recepcion_id)
        )
        check_lista_facturacion = estado in ('entregas_lista_facturacion', 'entregas_revision', 'facturacion', 'finalizada')
        check_facturacion = estado in ('facturacion', 'finalizada') or hoja_id in llego_facturacion
        check_finalizada = estado == 'finalizada' or bool(flujo.facturacion_aprobado)

        estado_label = 'Sin flujo activo'
        responsable_label = 'Sin responsable'
        estado_variant = 'neutral'

        if estado == 'entregas':
            estado_label = 'En espera de Entregas'
            responsable_label = 'Responsable: Entregas'
            estado_variant = 'warn'
        elif estado == 'almacen':
            estado_label = 'En espera de Almacén'
            responsable_label = 'Responsable: Almacén'
            estado_variant = 'warn'
        elif estado == 'entregas_lista_facturacion':
            estado_label = 'Autorizada por Almacén'
            responsable_label = 'Responsable: Entregas para enviar a Facturación'
            estado_variant = 'ok'
        elif estado == 'entregas_revision':
            estado_label = 'Regresada para revisión'
            responsable_label = 'Responsable: Entregas (revisión y reenvío)'
            estado_variant = 'warn'
        elif estado == 'facturacion':
            estado_label = 'En espera de Facturación'
            responsable_label = 'Responsable: Facturación'
            estado_variant = 'warn'
        elif estado == 'finalizada':
            estado_label = 'Autorizada por Facturación'
            responsable_label = 'Proceso cerrado'
            estado_variant = 'ok'

        resumen.append({
            'flujo_id': flujo.id,
            'hoja_ruta_id': hoja_id,
            'hoja_nombre': (flujo.hoja_ruta.nombre if flujo.hoja_ruta else f'HOJA #{hoja_id}'),
            'usuario': flujo.actualizado_por or flujo.creado_por,
            'fecha_ultima': flujo.fecha_actualizacion,
            'check_entregas': True,
            'check_almacen': check_almacen,
            'check_lista_facturacion': check_lista_facturacion,
            'check_facturacion': check_facturacion,
            'check_finalizada': check_finalizada,
            'check_devuelta_almacen': hoja_id in devoluciones_almacen,
            'check_devuelta_facturacion': hoja_id in devoluciones_facturacion,
            'estado_actual': estado,
            'estado_label': estado_label,
            'responsable_label': responsable_label,
            'estado_variant': estado_variant,
            'autorizado_por': flujo.facturacion_aprobado_por or '',
            'autorizado_en': flujo.facturacion_aprobado_en,
            'recepcion_id': flujo.almacen_recepcion_id or '',
        })

    return resumen


def _build_entregas_almacen_hoy(limit=200):
    """Construye un resumen diario de hojas enviadas a Almacen con su estatus actual."""
    now_dt = datetime.utcnow()
    day_start = now_dt.replace(hour=0, minute=0, second=0, microsecond=0)
    day_end = day_start + timedelta(days=1)

    acciones_envio = ['enviada_a_almacen', 'enviada_a_almacen_revision']
    envios_hoy = (
        EntregaRegistro.query
        .filter(EntregaRegistro.accion.in_(acciones_envio))
        .filter(EntregaRegistro.fecha_creacion >= day_start)
        .filter(EntregaRegistro.fecha_creacion < day_end)
        .order_by(EntregaRegistro.fecha_creacion.desc())
        .limit(limit)
        .all()
    )

    if not envios_hoy:
        return {
            'rows': [],
            'summary': {
                'total_envios': 0,
                'en_almacen': 0,
                'liberadas': 0,
                'regresadas': 0,
            },
            'fecha_corte': now_dt.isoformat(),
        }

    hoja_ids = sorted({int(x.hoja_ruta_id) for x in envios_hoy if x.hoja_ruta_id})
    flujo_por_hoja = {
        item.hoja_ruta_id: item
        for item in HojaRutaFlujoLogistica.query.filter(HojaRutaFlujoLogistica.hoja_ruta_id.in_(hoja_ids)).all()
    }

    def _estado_meta(estado_val):
        estado = (estado_val or '').strip().lower()
        if estado == 'almacen':
            return 'En espera en Almacen', 'warn'
        if estado == 'entregas_lista_facturacion':
            return 'Liberada por Almacen', 'ok'
        if estado == 'entregas_revision':
            return 'Regresada por Almacen', 'danger'
        if estado == 'facturacion':
            return 'En Facturacion', 'info'
        if estado == 'finalizada':
            return 'Finalizada', 'ok'
        if estado == 'entregas':
            return 'En Entregas', 'neutral'
        return 'Sin flujo activo', 'neutral'

    rows = []
    en_almacen = 0
    liberadas = 0
    regresadas = 0

    for mov in envios_hoy:
        flujo = flujo_por_hoja.get(mov.hoja_ruta_id)
        hoja = mov.hoja_ruta or (flujo.hoja_ruta if flujo else None)

        total_piezas = int((flujo.cantidad_total_piezas if flujo and flujo.cantidad_total_piezas is not None else 0) or 0)
        if total_piezas <= 0 and hoja and hoja.cantidad_piezas:
            total_piezas = int(hoja.cantidad_piezas or 0)

        entregadas = int((flujo.cantidad_entregada if flujo and flujo.cantidad_entregada is not None else 0) or 0)
        if entregadas < 0:
            entregadas = 0
        if total_piezas > 0 and entregadas > total_piezas:
            entregadas = total_piezas
        pendientes = max(total_piezas - entregadas, 0)

        estado_actual = (flujo.estado if flujo else '').strip().lower()
        estado_label, estado_variant = _estado_meta(estado_actual)

        if estado_actual == 'almacen':
            en_almacen += 1
        if estado_actual in ('entregas_lista_facturacion', 'facturacion', 'finalizada'):
            liberadas += 1
        if estado_actual == 'entregas_revision':
            regresadas += 1

        accion = (mov.accion or '').strip().lower()
        accion_label = 'Enviada a Almacen'
        if accion == 'enviada_a_almacen_revision':
            accion_label = 'Reenviada a Almacen'

        rows.append({
            'movimiento_id': mov.id,
            'hora': mov.fecha_creacion.strftime('%H:%M') if mov.fecha_creacion else '-',
            'hoja_ruta_id': mov.hoja_ruta_id,
            'hoja_nombre': (hoja.nombre if hoja and hoja.nombre else f'HOJA #{mov.hoja_ruta_id}'),
            'clave': (hoja.pn if hoja else '') or '',
            'orden_trabajo': ((hoja.orden_trabajo_hr if hoja else None) or (hoja.orden_trabajo_pt if hoja else '') or ''),
            'cantidad_total_piezas': total_piezas,
            'cantidad_entregada': entregadas,
            'cantidad_pendiente': pendientes,
            'accion': accion,
            'accion_label': accion_label,
            'estado_actual': estado_actual,
            'estado_label': estado_label,
            'estado_variant': estado_variant,
            'recepcion_id': (flujo.almacen_recepcion_id if flujo else '') or '',
            'usuario': mov.usuario or (flujo.actualizado_por if flujo else ''),
            'fecha_movimiento': mov.fecha_creacion.isoformat() if mov.fecha_creacion else None,
        })

    return {
        'rows': rows,
        'summary': {
            'total_envios': len(rows),
            'en_almacen': en_almacen,
            'liberadas': liberadas,
            'regresadas': regresadas,
        },
        'fecha_corte': now_dt.isoformat(),
    }


@app.route('/entregas')
@login_required
@requires_any_permission([('entregas', 'view'), ('catalog', 'edit')])
def entregas_module():
    hojas = (
        HojaRutaEntrega.query
        .join(HojaRutaFlujoLogistica, HojaRutaFlujoLogistica.hoja_ruta_id == HojaRutaEntrega.id)
        .order_by(HojaRutaFlujoLogistica.fecha_actualizacion.desc())
        .all()
    )
    pendientes_entregas = (
        HojaRutaFlujoLogistica.query
        .filter_by(estado='entregas')
        .order_by(HojaRutaFlujoLogistica.fecha_actualizacion.desc())
        .all()
    )
    listas_revision = (
        HojaRutaFlujoLogistica.query
        .filter(HojaRutaFlujoLogistica.estado.in_(['entregas_lista_facturacion', 'entregas_revision']))
        .order_by(HojaRutaFlujoLogistica.fecha_actualizacion.desc())
        .all()
    )
    historial_entregas = _build_logistica_resumen(limit=80)
    return render_template(
        'entregas_module.html',
        hojas=hojas,
        pendientes_entregas=pendientes_entregas,
        listas_revision=listas_revision,
        historial_entregas=historial_entregas,
    )


@app.route('/api/logistica/entregas/agregar', methods=['POST'])
@login_required
@requires_any_permission([('entregas', 'edit'), ('catalog', 'edit')])
def api_logistica_entregas_agregar():
    data = request.get_json() or {}
    hoja_id = data.get('hoja_id')
    try:
        hoja_id = int(hoja_id)
    except Exception:
        return jsonify({'error': 'hoja_id inválido'}), 400

    hoja = HojaRutaEntrega.query.get(hoja_id)
    if not hoja:
        return jsonify({'error': 'Hoja de ruta no encontrada'}), 404

    item = HojaRutaFlujoLogistica.query.filter_by(hoja_ruta_id=hoja_id).first()
    if item:
        if item.estado in ('entregas', 'entregas_lista_facturacion'):
            return jsonify({'ok': True, 'message': 'La hoja ya está en la bandeja de Entregas.'})
        if item.estado in ('almacen', 'facturacion'):
            return jsonify({'error': f'La hoja ya fue transferida a {item.estado}.'}), 409
        if item.estado == 'finalizada':
            return jsonify({'error': 'La hoja ya fue finalizada en Facturación.'}), 409

    item = HojaRutaFlujoLogistica(
        hoja_ruta_id=hoja.id,
        estado='entregas',
        creado_por=_logistica_username(),
        actualizado_por=_logistica_username(),
    )
    db.session.add(item)
    db.session.flush()
    db.session.add(EntregaRegistro(
        hoja_ruta_id=hoja.id,
        flujo_id=item.id,
        accion='agregada_en_entregas',
        usuario=_logistica_username(),
    ))
    db.session.commit()
    return jsonify({'ok': True, 'item': item.to_dict()})


@app.route('/api/entregas/parcial', methods=['POST'])
@login_required
@requires_any_permission([('entregas', 'edit'), ('catalog', 'edit')])
def api_entregas_parcial_registrar():
    data = request.get_json() or {}
    flujo_id = data.get('flujo_id')
    cantidad_entregada = data.get('cantidad_entregada')
    notas = (data.get('notas') or '').strip()

    try:
        flujo_id = int(flujo_id)
        cantidad_entregada = int(cantidad_entregada)
    except Exception:
        return jsonify({'error': 'flujo_id o cantidad_entregada inválidos'}), 400

    if cantidad_entregada <= 0:
        return jsonify({'error': 'La cantidad debe ser mayor a 0'}), 400

    item = HojaRutaFlujoLogistica.query.get(flujo_id)
    if not item:
        return jsonify({'error': 'Flujo de entregas no encontrado'}), 404

    if item.estado != 'entregas':
        return jsonify({'error': 'Solo se permiten parciales en estado entregas'}), 409

    _sync_flujo_parciales(item, hoja=item.hoja_ruta)
    pendiente = int(item.cantidad_pendiente or 0)
    if pendiente <= 0:
        return jsonify({'error': 'La hoja ya no tiene piezas pendientes'}), 409

    if cantidad_entregada > pendiente:
        return jsonify({'error': f'Cantidad excede lo pendiente ({pendiente})'}), 400

    parcial = EntregaParcial(
        flujo_id=item.id,
        hoja_ruta_id=item.hoja_ruta_id,
        cantidad_entregada=cantidad_entregada,
        usuario_entrega=_logistica_username(),
        notas=notas,
    )
    db.session.add(parcial)

    item.cantidad_entregada = int(item.cantidad_entregada or 0) + cantidad_entregada
    item.actualizado_por = _logistica_username()
    _sync_flujo_parciales(item, hoja=item.hoja_ruta)

    db.session.commit()
    return jsonify({'ok': True, 'parcial': parcial.to_dict(), 'flujo': item.to_dict()})


@app.route('/api/entregas/parcial/<int:parcial_id>', methods=['DELETE'])
@login_required
@requires_any_permission([('entregas', 'edit'), ('catalog', 'edit')])
def api_entregas_parcial_eliminar(parcial_id):
    parcial = EntregaParcial.query.get(parcial_id)
    if not parcial:
        return jsonify({'error': 'Entrega parcial no encontrada'}), 404

    item = HojaRutaFlujoLogistica.query.get(parcial.flujo_id)
    if not item:
        db.session.delete(parcial)
        db.session.commit()
        return jsonify({'ok': True, 'message': 'Parcial eliminado'})

    if item.estado != 'entregas':
        return jsonify({'error': 'Solo se pueden deshacer parciales en estado entregas'}), 409

    db.session.delete(parcial)
    db.session.flush()

    nuevo_entregado = db.session.query(func.coalesce(func.sum(EntregaParcial.cantidad_entregada), 0)).filter(
        EntregaParcial.flujo_id == item.id
    ).scalar() or 0

    item.cantidad_entregada = int(nuevo_entregado)
    item.actualizado_por = _logistica_username()
    _sync_flujo_parciales(item, hoja=item.hoja_ruta)

    db.session.commit()
    return jsonify({'ok': True, 'flujo': item.to_dict()})


@app.route('/entregas/mover_almacen/<int:item_id>', methods=['POST'])
@login_required
@requires_any_permission([('entregas', 'edit'), ('catalog', 'edit')])
def entregas_mover_almacen(item_id):
    item = HojaRutaFlujoLogistica.query.get_or_404(item_id)
    if item.estado != 'entregas':
        return redirect(url_for('entregas_module'))

    hoja = item.hoja_ruta
    _sync_flujo_parciales(item, hoja=hoja)

    if (item.cantidad_total_piezas or 0) <= 0:
        flash('La hoja no tiene cantidad total de piezas válida para envío.', 'error')
        return redirect(url_for('entregas_module'))

    # Se permite enviar a Almacén con parcialidad registrada (ley: debe quedar
    # documentada). Solo se bloquea si no hay ninguna entrega capturada.
    if (item.cantidad_entregada or 0) <= 0:
        flash('Registra al menos una entrega parcial antes de enviar a Almacén.', 'error')
        return redirect(url_for('entregas_module'))

    entregado = int(item.cantidad_entregada or 0)
    total = int(item.cantidad_total_piezas or 0)
    es_completa = total > 0 and entregado >= total
    pct = round((entregado / total) * 100, 1) if total > 0 else 0.0

    item.estado = 'almacen'
    item.estado_parciales = 'todas' if es_completa else 'pendientes'
    item.actualizado_por = _logistica_username()
    if es_completa:
        notas_envio = f'Entregas parciales completadas: {entregado} de {total} piezas (100%)'
    else:
        notas_envio = (
            f'Enviada a Almacén con parcialidad registrada: {entregado} de {total} piezas '
            f'({pct}%). Quedan {max(total - entregado, 0)} pendiente(s).'
        )
    db.session.add(EntregaRegistro(
        hoja_ruta_id=item.hoja_ruta_id,
        flujo_id=item.id,
        accion='enviada_a_almacen',
        notas=notas_envio,
        usuario=_logistica_username(),
    ))
    db.session.commit()
    return redirect(url_for('entregas_module'))


@app.route('/entregas/mover_facturacion/<int:item_id>', methods=['POST'])
@login_required
@requires_any_permission([('entregas', 'edit'), ('catalog', 'edit')])

def entregas_mover_facturacion(item_id):
    item = HojaRutaFlujoLogistica.query.get_or_404(item_id)
    if item.estado not in ('entregas_lista_facturacion', 'entregas_revision'):
        return redirect(url_for('entregas_module'))

    _sync_flujo_parciales(item, hoja=item.hoja_ruta)
    if item.cantidad_entregada != item.cantidad_total_piezas:
        return redirect(url_for('entregas_module'))

    item.estado = 'facturacion'
    item.actualizado_por = _logistica_username()
    db.session.add(EntregaRegistro(
        hoja_ruta_id=item.hoja_ruta_id,
        flujo_id=item.id,
        accion='enviada_a_facturacion',
        usuario=_logistica_username(),
    ))
    db.session.commit()
    return redirect(url_for('entregas_module'))


@app.route('/entregas/revision/mover_almacen/<int:item_id>', methods=['POST'])
@login_required
@requires_any_permission([('entregas', 'edit'), ('catalog', 'edit')])
def entregas_revision_mover_almacen(item_id):
    item = HojaRutaFlujoLogistica.query.get_or_404(item_id)
    if item.estado not in ('entregas_lista_facturacion', 'entregas_revision'):
        return redirect(url_for('entregas_module'))

    _sync_flujo_parciales(item, hoja=item.hoja_ruta)
    if item.cantidad_entregada != item.cantidad_total_piezas:
        flash('No se puede reenviar a Almacén: faltan entregas parciales completas.', 'error')
        return redirect(url_for('entregas_module'))

    item.estado = 'almacen'
    item.actualizado_por = _logistica_username()
    db.session.add(EntregaRegistro(
        hoja_ruta_id=item.hoja_ruta_id,
        flujo_id=item.id,
        accion='enviada_a_almacen_revision',
        usuario=_logistica_username(),
        notas='Reenviada a Almacén desde panel de revisión de Entregas.',
    ))
    db.session.commit()
    return redirect(url_for('entregas_module'))



@app.route('/entregas/quitar/<int:item_id>', methods=['POST'])
@login_required
@requires_any_permission([('entregas', 'edit'), ('catalog', 'edit')])
def entregas_quitar_item(item_id):
    item = HojaRutaFlujoLogistica.query.get_or_404(item_id)
    if item.estado == 'entregas':
        db.session.delete(item)
        db.session.commit()
        flash('La hoja fue retirada de Entregas.', 'success')
    else:
        flash('Solo se pueden retirar hojas que siguen en la bandeja de Entregas.', 'error')
    return redirect(url_for('entregas_module'))


@app.route('/almacen')
@login_required
@requires_any_permission([('almacen', 'view'), ('catalog', 'edit')])
def almacen_module():
    pendientes_almacen = (
        HojaRutaFlujoLogistica.query
        .filter_by(estado='almacen')
        .order_by(HojaRutaFlujoLogistica.fecha_actualizacion.desc())
        .all()
    )
    historial_almacen = _build_logistica_resumen(limit=80)

    return render_template(
        'almacen_module.html',
        pendientes_almacen=pendientes_almacen,
        historial_almacen=historial_almacen,
    )
        
        
@app.route('/almacen/recibir/<int:item_id>', methods=['POST'])
@login_required
@requires_any_permission([('almacen', 'edit'), ('catalog', 'edit')])
def almacen_recibir_item(item_id):
    item = HojaRutaFlujoLogistica.query.get_or_404(item_id)
    if item.estado != 'almacen':
        return redirect(url_for('almacen_module'))

    recepcion_id = (request.form.get('recepcion_id') or '').strip()
    entregado = request.form.get('entregado') == 'on'
    capturas = [file for file in request.files.getlist('captura_recepcion') if file and file.filename]

    if not entregado or not recepcion_id:
        return redirect(url_for('almacen_module'))

    if not capturas:
        return redirect(url_for('almacen_module'))

    if any(not _logistica_allowed_image(captura.filename) for captura in capturas):
        return redirect(url_for('almacen_module'))

    try:
        captura_paths = [_save_logistica_recepcion_image(captura, item.hoja_ruta_id) for captura in capturas]
    except Exception:
        flash('No se pudieron procesar las capturas de recepción. Usa imágenes JPG, PNG o WEBP válidas.', 'error')
        return redirect(url_for('almacen_module'))

    captura_path_value = _serialize_logistica_capture_paths(captura_paths)

    item.almacen_validado = True
    item.almacen_recepcion_id = recepcion_id
    item.almacen_captura_path = captura_path_value
    # Almacén solo recepciona/libera y devuelve a Entregas como lista para enviar a Facturación.
    item.estado = 'entregas_lista_facturacion'
    item.actualizado_por = _logistica_username()

    db.session.add(AlmacenRegistro(
        hoja_ruta_id=item.hoja_ruta_id,
        flujo_id=item.id,
        recepcion_id=recepcion_id,
        captura_path=captura_path_value,
        validado=True,
        usuario=_logistica_username(),
        notas=f'Recepción validada en almacén y liberada para Entregas. Capturas: {len(captura_paths)}.',
    ))
    db.session.add(EntregaRegistro(
        hoja_ruta_id=item.hoja_ruta_id,
        flujo_id=item.id,
        accion='lista_para_facturacion',
        usuario=_logistica_username(),
        notas=f'Recepción {recepcion_id} validada en almacén.',
    ))

    db.session.commit()

    ok_sync, reason_sync = _sync_almacen_liberacion_to_sheets(item.hoja_ruta, item, recepcion_id)
    if not ok_sync and reason_sync not in ('disabled', 'missing_webhook_url'):
        logger.info(
            f'[SHEETS_SYNC] No se pudo sincronizar hoja_ruta_id={item.hoja_ruta_id} '
            f'recepcion_id={recepcion_id} reason={reason_sync}'
        )

    return redirect(url_for('almacen_module'))


@app.route('/almacen/regresar_entregas/<int:item_id>', methods=['POST'])
@login_required
@requires_any_permission([('almacen', 'edit'), ('catalog', 'edit')])
def almacen_regresar_entregas(item_id):
    item = HojaRutaFlujoLogistica.query.get_or_404(item_id)
    if item.estado != 'almacen':
        return redirect(url_for('almacen_module'))

    motivo = (request.form.get('motivo_regreso') or '').strip()
    if not motivo:
        motivo = 'Datos incompletos o recepción no válida en almacén.'

    item.estado = 'entregas_revision'
    item.actualizado_por = _logistica_username()

    db.session.add(AlmacenRegistro(
        hoja_ruta_id=item.hoja_ruta_id,
        flujo_id=item.id,
        recepcion_id=item.almacen_recepcion_id,
        captura_path=item.almacen_captura_path,
        validado=False,
        usuario=_logistica_username(),
        notas=f'Devuelta a Entregas: {motivo}',
    ))
    db.session.add(EntregaRegistro(
        hoja_ruta_id=item.hoja_ruta_id,
        flujo_id=item.id,
        accion='devuelta_desde_almacen',
        usuario=_logistica_username(),
        notas=motivo,
    ))

    db.session.commit()
    return redirect(url_for('almacen_module'))


@app.route('/almacen/cajas-surtido')
@login_required
@requires_any_permission([('almacen', 'view'), ('catalog', 'edit')])
def almacen_cajas_surtido_page():
    if not _ensure_almacen_cajas_surtido_tables():
        return render_template('403.html'), 500

    productos = (
        Producto.query
        .order_by(Producto.clave.asc())
        .limit(300)
        .all()
    )
    productos_data = [
        {
            'codigo': (p.clave or '').strip(),
            'nombre': (p.nombre or '').strip(),
        }
        for p in productos
        if (p.clave or '').strip()
    ]

    return render_template('almacen_cajas_surtido.html', productos=productos_data)


@app.route('/api/almacen/cajas-surtido/sesiones', methods=['POST'])
@login_required
@requires_any_permission([('almacen', 'edit'), ('catalog', 'edit')])
def api_almacen_cajas_surtido_crear_sesion():
    if not _ensure_almacen_cajas_surtido_tables():
        return jsonify({'error': 'No se pudo preparar módulo de cajas surtido'}), 500

    data = request.get_json() or {}
    pedido_referencia = (data.get('pedido_referencia') or '').strip()
    notas = (data.get('notas') or '').strip() or None
    if not pedido_referencia:
        return jsonify({'error': 'pedido_referencia es requerido'}), 400

    usuario = _current_username_for_audit(get_current_user())
    sesion = AlmacenCajaSurtidoSesion(
        pedido_referencia=pedido_referencia,
        estado='abierta',
        usuario=usuario,
        notas=notas,
    )
    db.session.add(sesion)
    db.session.flush()

    caja = AlmacenCajaSurtidoCaja(
        sesion_id=sesion.id,
        numero_caja=1,
        estado='abierta',
    )
    db.session.add(caja)
    db.session.commit()

    return jsonify({'ok': True, 'sesion': _serialize_almacen_caja_surtido_sesion(sesion)}), 201


@app.route('/api/almacen/cajas-surtido/sesiones/<int:sesion_id>', methods=['GET'])
@login_required
@requires_any_permission([('almacen', 'view'), ('catalog', 'edit')])
def api_almacen_cajas_surtido_obtener_sesion(sesion_id):
    if not _ensure_almacen_cajas_surtido_tables():
        return jsonify({'error': 'No se pudo preparar módulo de cajas surtido'}), 500

    sesion = AlmacenCajaSurtidoSesion.query.get_or_404(sesion_id)
    return jsonify({'ok': True, 'sesion': _serialize_almacen_caja_surtido_sesion(sesion)})


@app.route('/api/almacen/cajas-surtido/sesiones/<int:sesion_id>/lecturas', methods=['POST'])
@login_required
@requires_any_permission([('almacen', 'edit'), ('catalog', 'edit')])
def api_almacen_cajas_surtido_registrar_lectura(sesion_id):
    if not _ensure_almacen_cajas_surtido_tables():
        return jsonify({'error': 'No se pudo preparar módulo de cajas surtido'}), 500

    sesion = AlmacenCajaSurtidoSesion.query.get_or_404(sesion_id)
    if sesion.estado != 'abierta':
        return jsonify({'error': 'La sesión ya está cerrada'}), 409

    data = request.get_json() or {}
    try:
        peso_kg = float(data.get('peso_kg') or 0)
    except Exception:
        return jsonify({'error': 'peso_kg inválido'}), 400

    if peso_kg <= 0:
        return jsonify({'error': 'peso_kg debe ser mayor a 0'}), 400

    origen = (data.get('origen') or 'manual').strip().lower() or 'manual'
    if origen not in ('manual', 'bascula'):
        origen = 'manual'

    lectura = AlmacenCajaSurtidoLecturaBascula(
        sesion_id=sesion.id,
        peso_kg=peso_kg,
        origen=origen,
        raw_payload=(data.get('raw_payload') or '').strip() or None,
        usuario=_current_username_for_audit(get_current_user()),
    )
    db.session.add(lectura)
    db.session.commit()

    return jsonify({'ok': True, 'lectura': lectura.to_dict()})


@app.route('/api/almacen/cajas-surtido/sesiones/<int:sesion_id>/items', methods=['POST'])
@login_required
@requires_any_permission([('almacen', 'edit'), ('catalog', 'edit')])
def api_almacen_cajas_surtido_agregar_item(sesion_id):
    if not _ensure_almacen_cajas_surtido_tables():
        return jsonify({'error': 'No se pudo preparar módulo de cajas surtido'}), 500

    sesion = AlmacenCajaSurtidoSesion.query.get_or_404(sesion_id)
    if sesion.estado != 'abierta':
        return jsonify({'error': 'La sesión ya está cerrada'}), 409

    caja = (
        AlmacenCajaSurtidoCaja.query
        .filter_by(sesion_id=sesion.id, estado='abierta')
        .order_by(AlmacenCajaSurtidoCaja.id.desc())
        .first()
    )
    if not caja:
        caja = AlmacenCajaSurtidoCaja(sesion_id=sesion.id, numero_caja=max(1, int(sesion.caja_actual_numero or 1)), estado='abierta')
        db.session.add(caja)
        db.session.flush()

    data = request.get_json() or {}
    producto_codigo = (data.get('producto_codigo') or '').strip()
    producto_nombre = (data.get('producto_nombre') or '').strip() or None

    try:
        piezas = int(data.get('piezas') or 0)
    except Exception:
        return jsonify({'error': 'piezas inválidas'}), 400

    if not producto_codigo:
        return jsonify({'error': 'producto_codigo es requerido'}), 400
    if piezas <= 0:
        return jsonify({'error': 'piezas debe ser mayor a 0'}), 400

    lectura_id = data.get('lectura_id')
    lectura = None
    if lectura_id:
        lectura = AlmacenCajaSurtidoLecturaBascula.query.filter_by(id=int(lectura_id), sesion_id=sesion.id).first()
    if not lectura:
        lectura = (
            AlmacenCajaSurtidoLecturaBascula.query
            .filter_by(sesion_id=sesion.id)
            .order_by(AlmacenCajaSurtidoLecturaBascula.id.desc())
            .first()
        )

    if lectura:
        peso_kg = float(lectura.peso_kg or 0.0)
        fuente_peso = lectura.origen or 'manual'
    else:
        try:
            peso_kg = float(data.get('peso_kg') or 0)
        except Exception:
            return jsonify({'error': 'No hay lectura disponible y peso_kg es inválido'}), 400
        if peso_kg <= 0:
            return jsonify({'error': 'No hay lectura disponible y peso_kg debe ser mayor a 0'}), 400
        fuente_peso = 'manual'

    peso_unitario = peso_kg / float(piezas)

    item = AlmacenCajaSurtidoItem(
        sesion_id=sesion.id,
        caja_id=caja.id,
        lectura_id=lectura.id if lectura else None,
        producto_codigo=producto_codigo,
        producto_nombre=producto_nombre,
        piezas=piezas,
        peso_kg=peso_kg,
        peso_unitario_kg=peso_unitario,
        fuente_peso=fuente_peso,
        usuario=_current_username_for_audit(get_current_user()),
    )
    db.session.add(item)

    caja.piezas_totales = int(caja.piezas_totales or 0) + piezas
    caja.peso_total_kg = float(caja.peso_total_kg or 0.0) + peso_kg

    db.session.commit()

    return jsonify({'ok': True, 'sesion': _serialize_almacen_caja_surtido_sesion(sesion)})


@app.route('/api/almacen/cajas-surtido/sesiones/<int:sesion_id>/cerrar-caja', methods=['POST'])
@login_required
@requires_any_permission([('almacen', 'edit'), ('catalog', 'edit')])
def api_almacen_cajas_surtido_cerrar_caja(sesion_id):
    if not _ensure_almacen_cajas_surtido_tables():
        return jsonify({'error': 'No se pudo preparar módulo de cajas surtido'}), 500

    sesion = AlmacenCajaSurtidoSesion.query.get_or_404(sesion_id)
    if sesion.estado != 'abierta':
        return jsonify({'error': 'La sesión ya está cerrada'}), 409

    caja = (
        AlmacenCajaSurtidoCaja.query
        .filter_by(sesion_id=sesion.id, estado='abierta')
        .order_by(AlmacenCajaSurtidoCaja.id.desc())
        .first()
    )
    if not caja:
        return jsonify({'error': 'No hay una caja abierta'}), 409
    if int(caja.piezas_totales or 0) <= 0:
        return jsonify({'error': 'No puedes cerrar una caja vacía'}), 409

    caja.estado = 'cerrada'
    caja.fecha_cierre = datetime.utcnow()
    sesion.total_cajas_cerradas = int(sesion.total_cajas_cerradas or 0) + 1
    sesion.caja_actual_numero = int(caja.numero_caja or 1) + 1

    nueva_caja = AlmacenCajaSurtidoCaja(
        sesion_id=sesion.id,
        numero_caja=sesion.caja_actual_numero,
        estado='abierta',
    )
    db.session.add(nueva_caja)
    db.session.commit()

    return jsonify({'ok': True, 'sesion': _serialize_almacen_caja_surtido_sesion(sesion)})


@app.route('/api/almacen/cajas-surtido/sesiones/<int:sesion_id>/finalizar', methods=['POST'])
@login_required
@requires_any_permission([('almacen', 'edit'), ('catalog', 'edit')])
def api_almacen_cajas_surtido_finalizar(sesion_id):
    if not _ensure_almacen_cajas_surtido_tables():
        return jsonify({'error': 'No se pudo preparar módulo de cajas surtido'}), 500

    sesion = AlmacenCajaSurtidoSesion.query.get_or_404(sesion_id)
    if sesion.estado != 'abierta':
        return jsonify({'ok': True, 'sesion': _serialize_almacen_caja_surtido_sesion(sesion)})

    cajas = (
        AlmacenCajaSurtidoCaja.query
        .filter_by(sesion_id=sesion.id)
        .order_by(AlmacenCajaSurtidoCaja.id.asc())
        .all()
    )
    cajas_cerradas_con_items = [c for c in cajas if c.estado == 'cerrada' and int(c.piezas_totales or 0) > 0]
    caja_abierta = next((c for c in cajas if c.estado == 'abierta'), None)

    if caja_abierta and int(caja_abierta.piezas_totales or 0) > 0:
        caja_abierta.estado = 'cerrada'
        caja_abierta.fecha_cierre = datetime.utcnow()
        cajas_cerradas_con_items.append(caja_abierta)
        sesion.total_cajas_cerradas = max(int(sesion.total_cajas_cerradas or 0), len(cajas_cerradas_con_items))
    elif caja_abierta and int(caja_abierta.piezas_totales or 0) == 0 and len(cajas) > 1:
        db.session.delete(caja_abierta)

    if not cajas_cerradas_con_items:
        return jsonify({'error': 'No hay cajas con contenido para finalizar'}), 409

    sesion.estado = 'cerrada'
    sesion.fecha_cierre = datetime.utcnow()

    ok_sync, reason_sync = _sync_cajas_surtido_to_sheets(sesion)
    sesion.google_sync_estado = 'synced' if ok_sync else 'error'
    sesion.google_sync_error = None if ok_sync else reason_sync

    db.session.commit()

    return jsonify({
        'ok': True,
        'sync': {'ok': ok_sync, 'reason': reason_sync},
        'sesion': _serialize_almacen_caja_surtido_sesion(sesion),
    })


@app.route('/facturacion')
@login_required
@requires_any_permission([('facturacion', 'view'), ('catalog', 'edit')])
def facturacion_module():
    pendientes_facturacion = (
        HojaRutaFlujoLogistica.query
        .filter_by(estado='facturacion')
        .order_by(HojaRutaFlujoLogistica.fecha_actualizacion.desc())
        .all()
    )
    finalizadas = (
        HojaRutaFlujoLogistica.query
        .filter_by(estado='finalizada')
        .order_by(HojaRutaFlujoLogistica.fecha_actualizacion.desc())
        .limit(50)
        .all()
    )
    historial_facturacion = _build_logistica_resumen(limit=80)
    return render_template(
        'facturacion_module.html',
        pendientes_facturacion=pendientes_facturacion,
        finalizadas=finalizadas,
        historial_facturacion=historial_facturacion,
    )


@app.route('/facturacion/aprobar/<int:item_id>', methods=['POST'])
@login_required
@requires_any_permission([('facturacion', 'edit'), ('catalog', 'edit')])
def facturacion_aprobar_item(item_id):
    item = HojaRutaFlujoLogistica.query.get_or_404(item_id)
    if item.estado != 'facturacion':
        return redirect(url_for('facturacion_module'))

    aprobado = request.form.get('aprobado_facturacion') == 'on'
    if not aprobado:
        return redirect(url_for('facturacion_module'))

    item.facturacion_aprobado = True
    item.facturacion_aprobado_por = _logistica_username()
    item.facturacion_aprobado_en = datetime.utcnow()
    item.estado = 'finalizada'
    item.actualizado_por = _logistica_username()

    db.session.add(FacturacionRegistro(
        hoja_ruta_id=item.hoja_ruta_id,
        flujo_id=item.id,
        aprobado=True,
        usuario=_logistica_username(),
        fecha_aprobacion=item.facturacion_aprobado_en,
        notas='Hoja liberada por facturación.',
    ))
    db.session.commit()
    return redirect(url_for('facturacion_module'))


@app.route('/facturacion/regresar_entregas/<int:item_id>', methods=['POST'])
@login_required
@requires_any_permission([('facturacion', 'edit'), ('catalog', 'edit')])
def facturacion_regresar_entregas(item_id):
    item = HojaRutaFlujoLogistica.query.get_or_404(item_id)
    if item.estado != 'facturacion':
        return redirect(url_for('facturacion_module'))

    motivo = (request.form.get('motivo_regreso') or '').strip()
    if not motivo:
        motivo = 'Documentación o recepción no corresponde para facturación.'

    item.estado = 'entregas_revision'
    item.actualizado_por = _logistica_username()

    db.session.add(FacturacionRegistro(
        hoja_ruta_id=item.hoja_ruta_id,
        flujo_id=item.id,
        aprobado=False,
        usuario=_logistica_username(),
        notas=f'Devuelta a Entregas: {motivo}',
        fecha_aprobacion=None,
    ))
    db.session.add(EntregaRegistro(
        hoja_ruta_id=item.hoja_ruta_id,
        flujo_id=item.id,
        accion='devuelta_desde_facturacion',
        usuario=_logistica_username(),
        notas=motivo,
    ))

    db.session.commit()
    return redirect(url_for('facturacion_module'))


@app.route('/uploads/<path:filename>')
def uploaded_file(filename):
    # Sirve archivos desde la carpeta uploads de forma segura
    uploads_root = os.path.join(os.getcwd(), 'uploads')
    try:
        return send_from_directory(uploads_root, filename)
    except Exception:
        return ('', 404)


def _verify_client_ip():
    forwarded = (request.headers.get('X-Forwarded-For') or '').strip()
    if forwarded:
        return forwarded.split(',')[0].strip()
    return (request.remote_addr or '').strip()


def _save_tecnico_qr_image(tecnico):
    base_url = request.url_root.rstrip('/')
    verify_url = f"{base_url}/verificar/{tecnico.token_qr}"
    qr = qrcode.QRCode(version=1, box_size=8, border=4)
    qr.add_data(verify_url)
    qr.make(fit=True)
    img = qr.make_image(fill_color='black', back_color='white')

    filename = f"tecnico_{tecnico.token_qr}.png"
    abs_path = os.path.join(TECNICOS_QR_DIR, filename)
    img.save(abs_path)
    tecnico.qr_imagen = f"/uploads/tecnicos/qr/{filename}"


def _write_verification_log(tecnico, token_raw, resultado):
    entry = LogVerificacion(
        tecnico_id=tecnico.id if tecnico else None,
        ip_cliente=_verify_client_ip(),
        user_agent=(request.headers.get('User-Agent') or '')[:1000],
        token_consultado=(token_raw or '')[:80],
        resultado=resultado,
    )
    db.session.add(entry)
    db.session.commit()


@app.route('/verificar/<token>', methods=['GET'])
def verificar_tecnico_publico(token):
    now = datetime.utcnow()
    token_txt = (token or '').strip()
    tecnico = Tecnico.query.filter_by(token_qr=token_txt).first()

    template_data = {
        'logo_url': '/static/logo.png',
        'tecnico': tecnico,
        'documentos_pdf': _list_tecnico_pdf_docs(tecnico.id) if tecnico else [],
        'estado_visual': 'invalido',
        'mensaje_estado': 'Credencial invalida o expirada',
        'motivo': 'Token no encontrado',
        'verified_at': now,
        'token': token_txt,
    }

    if tecnico:
        if tecnico.estado != Tecnico.ESTADO_ACTIVO:
            template_data.update({
                'estado_visual': 'invalido',
                'mensaje_estado': 'Credencial invalida o expirada',
                'motivo': 'Tecnico suspendido',
            })
            _write_verification_log(tecnico, token_txt, 'suspendido')
            return render_template('verificar_tecnico_publico.html', **template_data), 200

        if now >= tecnico.fecha_expiracion:
            template_data.update({
                'estado_visual': 'invalido',
                'mensaje_estado': 'Credencial invalida o expirada',
                'motivo': 'Credencial expirada',
            })
            _write_verification_log(tecnico, token_txt, 'expirado')
            return render_template('verificar_tecnico_publico.html', **template_data), 200

        if not (tecnico.qr_imagen or '').strip():
            try:
                _save_tecnico_qr_image(tecnico)
                db.session.add(tecnico)
                db.session.commit()
            except Exception as exc:
                db.session.rollback()
                logger.warning(f'No se pudo generar QR para tecnico {tecnico.id}: {exc}')

        template_data.update({
            'estado_visual': 'valido',
            'mensaje_estado': 'Tecnico verificado',
            'motivo': '',
        })
        _write_verification_log(tecnico, token_txt, 'valido')
        return render_template('verificar_tecnico_publico.html', **template_data), 200

    _write_verification_log(None, token_txt, 'token_invalido')
    return render_template('verificar_tecnico_publico.html', **template_data), 404


@app.route('/seguimiento', methods=['GET', 'POST'])
def seguimiento_empaque_cliente():
    """Portal público general: el cliente entra solo con su clave única."""
    _ensure_empaque_tables()
    error = None
    view = None
    clave = ''
    ip = (request.headers.get('X-Forwarded-For') or request.remote_addr or '').split(',')[0].strip()

    def _open_for_cliente(cliente, *, touch_access=False):
        nonlocal view, clave
        if not cliente or not cliente.activo:
            return False
        view = _build_empaque_cliente_view(
            cliente.customer_code,
            access_code=cliente.access_code,
            customer_name=cliente.customer_name,
        )
        if view is None:
            return False
        clave = cliente.access_code
        if touch_access:
            cliente.last_access_at = datetime.utcnow()
            try:
                db.session.commit()
            except Exception:
                db.session.rollback()
        session['empaque_seguimiento_access_code'] = cliente.access_code
        return True

    if request.method == 'POST':
        if _empaque_seguimiento_rate_limited(ip):
            error = 'Demasiados intentos. Espera unos minutos e inténtalo de nuevo.'
            _write_empaque_seguimiento_log('', 'rate_limited')
        else:
            clave_in = _normalize_empaque_clave(
                request.form.get('clave') or request.form.get('access_code')
            )
            if not clave_in:
                error = 'Ingresa tu clave de acceso.'
            else:
                cliente = _find_empaque_cliente_by_access_code(clave_in)
                if not cliente or not cliente.activo or not _open_for_cliente(cliente, touch_access=True):
                    _empaque_seguimiento_register_fail(ip)
                    _write_empaque_seguimiento_log(
                        '',
                        'invalido' if not cliente else ('inactivo' if cliente and not cliente.activo else 'invalido'),
                        access_code=_normalize_empaque_access_code(clave_in)[:32],
                    )
                    error = 'Clave incorrecta o acceso desactivado.'
                    clave = ''
                    view = None
                else:
                    _empaque_seguimiento_clear_fail(ip)
                    _write_empaque_seguimiento_log(
                        cliente.customer_code, 'ok', access_code=cliente.access_code
                    )

    if view is None and request.method == 'GET':
        # Link de acceso: /seguimiento?clave=EV-XXXXXXXX
        clave_qs = _normalize_empaque_clave(
            request.args.get('clave') or request.args.get('c') or request.args.get('access_code')
        )
        if clave_qs:
            if _empaque_seguimiento_rate_limited(ip):
                error = 'Demasiados intentos. Espera unos minutos e inténtalo de nuevo.'
                _write_empaque_seguimiento_log('', 'rate_limited')
            else:
                cliente = _find_empaque_cliente_by_access_code(clave_qs)
                if not cliente or not cliente.activo or not _open_for_cliente(cliente, touch_access=True):
                    _empaque_seguimiento_register_fail(ip)
                    _write_empaque_seguimiento_log(
                        '',
                        'invalido' if not cliente else ('inactivo' if not cliente.activo else 'invalido'),
                        access_code=_normalize_empaque_access_code(clave_qs)[:32],
                    )
                    error = 'Clave incorrecta o acceso desactivado.'
                    clave = ''
                    view = None
                else:
                    _empaque_seguimiento_clear_fail(ip)
                    _write_empaque_seguimiento_log(
                        cliente.customer_code, 'ok', access_code=cliente.access_code
                    )
                    # Evitar dejar la clave en el historial del navegador
                    return redirect(url_for('seguimiento_empaque_cliente'))

        if view is None:
            saved = _normalize_empaque_clave(session.get('empaque_seguimiento_access_code'))
            if saved:
                cliente = _find_empaque_cliente_by_access_code(saved)
                if not cliente or not cliente.activo or not _open_for_cliente(cliente, touch_access=False):
                    session.pop('empaque_seguimiento_access_code', None)
                    session.pop('empaque_seguimiento_clave', None)

    section = (request.args.get('sec') or 'pedidos').strip().lower()
    if section not in ('pedidos', 'precios', 'cuenta'):
        section = 'pedidos'
    selected_pedido = (request.args.get('pedido') or '').strip()
    selected_caja = (request.args.get('caja') or '').strip()
    selected_pedido_view = None
    selected_caja_view = None
    if view and selected_pedido:
        for pedido in view.get('pedidos') or []:
            if (pedido.get('external_order_number') or '') == selected_pedido:
                selected_pedido_view = pedido
                if selected_caja:
                    for caja in pedido.get('cajas') or []:
                        if (caja.get('box_code') or '') == selected_caja:
                            selected_caja_view = caja
                            break
                break
    precios = None
    if view and section == 'precios':
        precios = _fetch_odoo_precios_cliente(
            view.get('customer_code'),
            view.get('customer_name'),
        )

    return render_template(
        'seguimiento_empaque.html',
        error=error,
        view=view,
        clave=clave,
        logo_url='/static/logo.png',
        portal_url='/seguimiento',
        section=section,
        selected_pedido=selected_pedido,
        selected_caja=selected_caja,
        selected_pedido_view=selected_pedido_view,
        selected_caja_view=selected_caja_view,
        precios=precios,
    )


@app.route('/seguimiento/salir', methods=['POST', 'GET'])
def seguimiento_empaque_salir():
    session.pop('empaque_seguimiento_access_code', None)
    session.pop('empaque_seguimiento_clave', None)
    return redirect(url_for('seguimiento_empaque_cliente'))


@app.route('/verificar/<token>/reportar-problema', methods=['POST'])
def reportar_problema_verificacion(token):
    token_txt = (token or '').strip()
    tecnico = Tecnico.query.filter_by(token_qr=token_txt).first()
    comentario = (request.form.get('comentario') or '').strip()

    log = LogVerificacion(
        tecnico_id=tecnico.id if tecnico else None,
        ip_cliente=_verify_client_ip(),
        user_agent=((request.headers.get('User-Agent') or '') + f' | reportar:{comentario}')[:1000],
        token_consultado=token_txt[:80],
        resultado='reporte_problema',
    )
    db.session.add(log)
    db.session.commit()
    flash('Problema reportado. Gracias por avisarnos.', 'success')
    return redirect(url_for('verificar_tecnico_publico', token=token_txt))


# ==================== ADMIN TÉCNICOS QR ====================

TECNICOS_FOTO_DIR = os.path.join('uploads', 'tecnicos', 'fotos')
os.makedirs(TECNICOS_FOTO_DIR, exist_ok=True)
TECNICOS_DOCS_DIR = os.path.join('uploads', 'tecnicos', 'docs')
os.makedirs(TECNICOS_DOCS_DIR, exist_ok=True)
TECNICOS_FIRMAS_DIR = os.path.join('uploads', 'tecnicos', 'firmas')
os.makedirs(TECNICOS_FIRMAS_DIR, exist_ok=True)
TECNICOS_SIGNATURE_LINKS_DIR = os.path.join('uploads', 'tecnicos', 'signature_links')
os.makedirs(TECNICOS_SIGNATURE_LINKS_DIR, exist_ok=True)

SIGNATURE_ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'webp'}
TECNICO_SIGNATURE_TOKEN_SALT = 'tecnicos-public-signature'
TECNICO_SIGNATURE_LINK_MAX_AGE_SECONDS = max(
    3600,
    int(os.getenv('TECNICO_SIGNATURE_LINK_MAX_AGE_SECONDS', '604800') or '604800')
)
TECNICO_SIGNATURE_LABELS = {
    'cri': 'Firma de Validación CRI',
    'supervision': 'Firma de Validación Supervisión',
}


def _allowed_pdf_file(filename):
    if not filename or '.' not in filename:
        return False
    return filename.rsplit('.', 1)[1].lower() == 'pdf'


def _allowed_signature_file(filename):
    if not filename or '.' not in filename:
        return False
    return filename.rsplit('.', 1)[1].lower() in SIGNATURE_ALLOWED_EXTENSIONS


def _valid_tecnico_signature_type(sig_type):
    return (sig_type or '').strip().lower() in TECNICO_SIGNATURE_LABELS


def _tecnico_signature_label(sig_type):
    return TECNICO_SIGNATURE_LABELS.get((sig_type or '').strip().lower(), 'Firma')


def _tecnico_signature_serializer():
    return URLSafeTimedSerializer(app.secret_key)


def _tecnico_signature_link_state_path(tecnico_id, sig_type):
    sig_type_txt = (sig_type or '').strip().lower()
    if not _valid_tecnico_signature_type(sig_type_txt):
        return None
    return os.path.join(TECNICOS_SIGNATURE_LINKS_DIR, f'tec_{int(tecnico_id)}_{sig_type_txt}.json')


def _load_tecnico_signature_link_state(tecnico_id, sig_type):
    path = _tecnico_signature_link_state_path(tecnico_id, sig_type)
    if not path or not os.path.exists(path):
        return None
    try:
        with open(path, 'r', encoding='utf-8') as f:
            payload = json.load(f)
        return payload if isinstance(payload, dict) else None
    except Exception:
        return None


def _save_tecnico_signature_link_state(tecnico_id, sig_type, payload):
    path = _tecnico_signature_link_state_path(tecnico_id, sig_type)
    if not path:
        return False
    try:
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(payload or {}, f, ensure_ascii=False, indent=2)
        return True
    except Exception:
        return False


def _build_tecnico_signature_token(tecnico_id, sig_type):
    sig_type_txt = (sig_type or '').strip().lower()
    token_id = uuid.uuid4().hex
    _save_tecnico_signature_link_state(
        tecnico_id,
        sig_type_txt,
        {
            'token_id': token_id,
            'tecnico_id': int(tecnico_id),
            'sig_type': sig_type_txt,
            'created_at': datetime.utcnow().isoformat(),
            'used_at': None,
        },
    )
    serializer = _tecnico_signature_serializer()
    return serializer.dumps(
        {
            'tid': int(tecnico_id),
            'sig_type': sig_type_txt,
            'token_id': token_id,
        },
        salt=TECNICO_SIGNATURE_TOKEN_SALT,
    )


def _resolve_tecnico_signature_token(token, max_age=None):
    serializer = _tecnico_signature_serializer()
    try:
        payload = serializer.loads(
            token,
            salt=TECNICO_SIGNATURE_TOKEN_SALT,
            max_age=max_age or TECNICO_SIGNATURE_LINK_MAX_AGE_SECONDS,
        )
    except SignatureExpired:
        return None, 'expirado'
    except BadSignature:
        return None, 'invalido'

    try:
        tecnico_id = int(payload.get('tid') or 0)
    except Exception:
        tecnico_id = 0

    sig_type = (payload.get('sig_type') or '').strip().lower()
    token_id = (payload.get('token_id') or '').strip()
    if tecnico_id <= 0 or not _valid_tecnico_signature_type(sig_type):
        return None, 'invalido'
    if not token_id:
        return None, 'invalido'

    state = _load_tecnico_signature_link_state(tecnico_id, sig_type)
    if not state:
        return None, 'invalido'
    if (state.get('token_id') or '').strip() != token_id:
        return None, 'invalido'
    if (state.get('used_at') or '').strip():
        return None, 'consumido'

    tecnico = Tecnico.query.get(tecnico_id)
    if not tecnico:
        return None, 'invalido'

    return {
        'tecnico': tecnico,
        'sig_type': sig_type,
        'sig_label': _tecnico_signature_label(sig_type),
        'token_id': token_id,
    }, None


def _build_tecnico_signature_public_url(tecnico_id, sig_type):
    token = _build_tecnico_signature_token(tecnico_id, sig_type)
    return url_for('tecnico_firma_publica', token=token, _external=True)


def _mark_tecnico_signature_token_consumed(tecnico_id, sig_type, token_id):
    state = _load_tecnico_signature_link_state(tecnico_id, sig_type)
    if not state:
        return False
    if (state.get('token_id') or '').strip() != (token_id or '').strip():
        return False
    if (state.get('used_at') or '').strip():
        return True

    state['used_at'] = datetime.utcnow().isoformat()
    return _save_tecnico_signature_link_state(tecnico_id, sig_type, state)


def _save_tecnico_signature_file(tecnico_id, file_obj, sig_type):
    if not file_obj or not (file_obj.filename or '').strip():
        return None

    original = secure_filename(file_obj.filename)
    ext = original.rsplit('.', 1)[1].lower() if '.' in original else 'png'
    if ext not in SIGNATURE_ALLOWED_EXTENSIONS:
        return None

    prefix = f"tec_{int(tecnico_id)}_{sig_type}."
    try:
        for name in os.listdir(TECNICOS_FIRMAS_DIR):
            if name.startswith(prefix):
                try:
                    os.remove(os.path.join(TECNICOS_FIRMAS_DIR, name))
                except Exception:
                    pass
    except Exception:
        pass

    # Normaliza y recorta espacio vacío para que la firma se vea grande en la credencial.
    try:
        from PIL import Image, ImageOps

        file_obj.stream.seek(0)
        img = Image.open(file_obj.stream)
        img = ImageOps.exif_transpose(img)
        rgba = img.convert('RGBA')

        alpha = rgba.split()[-1]
        alpha_bbox = alpha.getbbox()

        gray = rgba.convert('L')
        dark_mask = gray.point(lambda p: 255 if p < 245 else 0)
        dark_bbox = dark_mask.getbbox()

        bbox = alpha_bbox or dark_bbox
        if bbox:
            pad = 16
            left = max(0, bbox[0] - pad)
            top = max(0, bbox[1] - pad)
            right = min(rgba.width, bbox[2] + pad)
            bottom = min(rgba.height, bbox[3] + pad)
            rgba = rgba.crop((left, top, right, bottom))

        # Escala consistente para impresión nítida.
        target_h = 280
        if rgba.height > 0 and rgba.height != target_h:
            target_w = max(1, int((rgba.width * target_h) / rgba.height))
            rgba = rgba.resize((target_w, target_h), Image.Resampling.LANCZOS)

        final_name = f"tec_{int(tecnico_id)}_{sig_type}.png"
        abs_path = os.path.join(TECNICOS_FIRMAS_DIR, final_name)
        rgba.save(abs_path, format='PNG', optimize=True)
        return f'/uploads/tecnicos/firmas/{final_name}'
    except Exception:
        # Fallback simple si PIL falla por cualquier razón.
        file_obj.stream.seek(0)
        final_name = f"tec_{int(tecnico_id)}_{sig_type}.{ext}"
        abs_path = os.path.join(TECNICOS_FIRMAS_DIR, final_name)
        file_obj.save(abs_path)
        return f'/uploads/tecnicos/firmas/{final_name}'


def _get_tecnico_signature_url(tecnico_id, sig_type):
    for ext in ('png', 'jpg', 'jpeg', 'webp'):
        name = f"tec_{int(tecnico_id)}_{sig_type}.{ext}"
        abs_path = os.path.join(TECNICOS_FIRMAS_DIR, name)
        if os.path.exists(abs_path):
            try:
                v = int(os.path.getmtime(abs_path))
            except Exception:
                v = int(time())
            return f'/uploads/tecnicos/firmas/{name}?v={v}'
    return None


def _list_tecnico_pdf_docs(tecnico_id):
    prefix = f"tec_{int(tecnico_id)}_"
    docs = []
    try:
        for name in os.listdir(TECNICOS_DOCS_DIR):
            if not name.lower().endswith('.pdf'):
                continue
            if not name.startswith(prefix):
                continue
            abs_path = os.path.join(TECNICOS_DOCS_DIR, name)
            mtime = os.path.getmtime(abs_path)
            docs.append({
                'filename': name,
                'url': f'/uploads/tecnicos/docs/{name}',
                'updated_at': datetime.utcfromtimestamp(mtime).isoformat(),
            })
    except Exception:
        return []

    docs.sort(key=lambda d: d['updated_at'], reverse=True)
    return docs


def _save_tecnico_pdf_files(tecnico_id, files):
    saved = []
    if not files:
        return saved

    for f in files:
        if not f or not (f.filename or '').strip():
            continue
        original = secure_filename(f.filename)
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        suffix = uuid.uuid4().hex[:8]
        final_name = f"tec_{int(tecnico_id)}_{ts}_{suffix}_{original}"
        abs_path = os.path.join(TECNICOS_DOCS_DIR, final_name)
        f.save(abs_path)
        saved.append(final_name)

    return saved


def _tecnico_to_api(tecnico):
    data = tecnico.to_dict()
    data['documentos_pdf'] = _list_tecnico_pdf_docs(tecnico.id)
    data['firma_cri_url'] = _get_tecnico_signature_url(tecnico.id, 'cri')
    data['firma_supervision_url'] = _get_tecnico_signature_url(tecnico.id, 'supervision')
    return data

@app.route('/tecnicos')
@login_required
def tecnicos_admin():
    user = get_current_user()
    if not (user and (user.es_admin or user.has_permission('catalog', 'edit'))):
        return render_template('403.html'), 403
    return render_template('tecnicos.html')


@app.route('/api/tecnicos', methods=['GET'])
@login_required
def api_list_tecnicos():
    user = get_current_user()
    if not (user and (user.es_admin or user.has_permission('catalog', 'edit'))):
        return jsonify({'error': 'Permiso denegado'}), 403
    tecnicos = Tecnico.query.order_by(Tecnico.creado_en.desc()).all()
    return jsonify({'tecnicos': [_tecnico_to_api(t) for t in tecnicos]})


@app.route('/api/tecnicos', methods=['POST'])
@login_required
def api_crear_tecnico():
    user = get_current_user()
    if not (user and (user.es_admin or user.has_permission('catalog', 'edit'))):
        return jsonify({'error': 'Permiso denegado'}), 403

    nombre = (request.form.get('nombre') or '').strip()
    empresa = (request.form.get('empresa') or '').strip()
    numero_empleado = (request.form.get('numero_empleado') or '').strip()
    fecha_exp_str = (request.form.get('fecha_expiracion') or '').strip()

    if not nombre or not empresa or not numero_empleado or not fecha_exp_str:
        return jsonify({'error': 'nombre, empresa, numero_empleado y fecha_expiracion son requeridos'}), 400

    pdf_files = [
        f for f in request.files.getlist('documentos_pdf')
        if f and (f.filename or '').strip()
    ]
    for pf in pdf_files:
        if not _allowed_pdf_file(pf.filename):
            return jsonify({'error': 'Solo se permiten archivos PDF en documentos del técnico'}), 400

    firma_cri_file = request.files.get('firma_cri')
    firma_supervision_file = request.files.get('firma_supervision')
    if firma_cri_file and (firma_cri_file.filename or '').strip() and not _allowed_signature_file(firma_cri_file.filename):
        return jsonify({'error': 'Firma CRI inválida. Solo PNG/JPG/JPEG/WEBP'}), 400
    if firma_supervision_file and (firma_supervision_file.filename or '').strip() and not _allowed_signature_file(firma_supervision_file.filename):
        return jsonify({'error': 'Firma Supervisión inválida. Solo PNG/JPG/JPEG/WEBP'}), 400

    if Tecnico.query.filter_by(numero_empleado=numero_empleado).first():
        return jsonify({'error': f'Ya existe un técnico con número de empleado {numero_empleado}'}), 409

    try:
        fecha_exp = datetime.strptime(fecha_exp_str, '%Y-%m-%d')
    except ValueError:
        return jsonify({'error': 'fecha_expiracion debe ser YYYY-MM-DD'}), 400

    foto_url = None
    if 'foto' in request.files:
        f = request.files['foto']
        if f and f.filename and allowed_file(f.filename):
            fname = secure_filename(f.filename)
            ts = datetime.now().strftime('%Y%m%d_%H%M%S_')
            fname = ts + fname
            f.save(os.path.join(TECNICOS_FOTO_DIR, fname))
            foto_url = f'/uploads/tecnicos/fotos/{fname}'

    def _opt(key):
        return (request.form.get(key) or '').strip() or None

    def _bool_opt(key):
        raw = (request.form.get(key) or '').strip().lower()
        return raw in ('1', 'true', 'si', 'sí', 'on', 'x')

    tecnico = Tecnico(
        nombre=nombre,
        empresa=empresa,
        numero_empleado=numero_empleado,
        puesto=_opt('puesto'),
        nss=_opt('nss'),
        curp=_opt('curp'),
        tipo_sangre=_opt('tipo_sangre'),
        alergias=_opt('alergias'),
        contacto_emergencia=_opt('contacto_emergencia'),
        antiguedad=_opt('antiguedad'),
        esp_alturas=_bool_opt('esp_alturas'),
        esp_maniobras_baja=_bool_opt('esp_maniobras_baja'),
        esp_electricos=_bool_opt('esp_electricos'),
        esp_trabajos_caliente=_bool_opt('esp_trabajos_caliente'),
        esp_espacios_confinados=_bool_opt('esp_espacios_confinados'),
        esp_excavaciones=_bool_opt('esp_excavaciones'),
        esp_maquinaria=_bool_opt('esp_maquinaria'),
        foto=foto_url,
        fecha_expiracion=fecha_exp,
    )
    db.session.add(tecnico)
    db.session.flush()  # genera el id y token_qr

    if pdf_files:
        _save_tecnico_pdf_files(tecnico.id, pdf_files)

    if firma_cri_file and (firma_cri_file.filename or '').strip():
        _save_tecnico_signature_file(tecnico.id, firma_cri_file, 'cri')
    if firma_supervision_file and (firma_supervision_file.filename or '').strip():
        _save_tecnico_signature_file(tecnico.id, firma_supervision_file, 'supervision')

    try:
        _save_tecnico_qr_image(tecnico)
    except Exception as exc:
        logger.warning(f'No se pudo pre-generar QR para técnico nuevo: {exc}')

    db.session.commit()
    return jsonify({'mensaje': 'Técnico creado', 'tecnico': _tecnico_to_api(tecnico)}), 201


@app.route('/api/tecnicos/<int:tid>', methods=['PUT', 'POST'])
@login_required
def api_editar_tecnico(tid):
    user = get_current_user()
    if not (user and (user.es_admin or user.has_permission('catalog', 'edit'))):
        return jsonify({'error': 'Permiso denegado'}), 403

    tecnico = Tecnico.query.get_or_404(tid)
    nombre = (request.form.get('nombre') or '').strip()
    empresa = (request.form.get('empresa') or '').strip()
    numero_empleado = (request.form.get('numero_empleado') or '').strip()
    fecha_exp_str = (request.form.get('fecha_expiracion') or '').strip()

    if nombre:
        tecnico.nombre = nombre
    if empresa:
        tecnico.empresa = empresa
    if numero_empleado:
        dup = Tecnico.query.filter(Tecnico.numero_empleado == numero_empleado, Tecnico.id != tid).first()
        if dup:
            return jsonify({'error': 'Número de empleado ya existe en otro técnico'}), 409
        tecnico.numero_empleado = numero_empleado
    if fecha_exp_str:
        try:
            tecnico.fecha_expiracion = datetime.strptime(fecha_exp_str, '%Y-%m-%d')
        except ValueError:
            return jsonify({'error': 'fecha_expiracion debe ser YYYY-MM-DD'}), 400

    pdf_files = [
        f for f in request.files.getlist('documentos_pdf')
        if f and (f.filename or '').strip()
    ]
    for pf in pdf_files:
        if not _allowed_pdf_file(pf.filename):
            return jsonify({'error': 'Solo se permiten archivos PDF en documentos del técnico'}), 400

    firma_cri_file = request.files.get('firma_cri')
    firma_supervision_file = request.files.get('firma_supervision')
    if firma_cri_file and (firma_cri_file.filename or '').strip() and not _allowed_signature_file(firma_cri_file.filename):
        return jsonify({'error': 'Firma CRI inválida. Solo PNG/JPG/JPEG/WEBP'}), 400
    if firma_supervision_file and (firma_supervision_file.filename or '').strip() and not _allowed_signature_file(firma_supervision_file.filename):
        return jsonify({'error': 'Firma Supervisión inválida. Solo PNG/JPG/JPEG/WEBP'}), 400

    for field in ('puesto', 'nss', 'curp', 'tipo_sangre', 'alergias', 'contacto_emergencia', 'antiguedad'):
        val = (request.form.get(field) or '').strip()
        if val:
            setattr(tecnico, field, val)

    for field in (
        'esp_alturas',
        'esp_maniobras_baja',
        'esp_electricos',
        'esp_trabajos_caliente',
        'esp_espacios_confinados',
        'esp_excavaciones',
        'esp_maquinaria',
    ):
        if field in request.form:
            raw = (request.form.get(field) or '').strip().lower()
            setattr(tecnico, field, raw in ('1', 'true', 'si', 'sí', 'on', 'x'))

    if 'foto' in request.files:
        f = request.files['foto']
        if f and f.filename and allowed_file(f.filename):
            fname = secure_filename(f.filename)
            ts = datetime.now().strftime('%Y%m%d_%H%M%S_')
            fname = ts + fname
            f.save(os.path.join(TECNICOS_FOTO_DIR, fname))
            tecnico.foto = f'/uploads/tecnicos/fotos/{fname}'

    if pdf_files:
        _save_tecnico_pdf_files(tecnico.id, pdf_files)

    if firma_cri_file and (firma_cri_file.filename or '').strip():
        _save_tecnico_signature_file(tecnico.id, firma_cri_file, 'cri')
    if firma_supervision_file and (firma_supervision_file.filename or '').strip():
        _save_tecnico_signature_file(tecnico.id, firma_supervision_file, 'supervision')

    db.session.commit()
    return jsonify({'mensaje': 'Técnico actualizado', 'tecnico': _tecnico_to_api(tecnico)})


@app.route('/api/tecnicos/<int:tid>/estado', methods=['POST'])
@login_required
def api_toggle_tecnico_estado(tid):
    user = get_current_user()
    if not (user and (user.es_admin or user.has_permission('catalog', 'edit'))):
        return jsonify({'error': 'Permiso denegado'}), 403
    tecnico = Tecnico.query.get_or_404(tid)
    nuevo = Tecnico.ESTADO_SUSPENDIDO if tecnico.estado == Tecnico.ESTADO_ACTIVO else Tecnico.ESTADO_ACTIVO
    tecnico.estado = nuevo
    db.session.commit()
    return jsonify({'mensaje': f'Estado cambiado a {nuevo}', 'estado': nuevo})


@app.route('/api/tecnicos/<int:tid>/qr')
@login_required
def api_descargar_qr_tecnico(tid):
    user = get_current_user()
    if not (user and (user.es_admin or user.has_permission('catalog', 'edit'))):
        return jsonify({'error': 'Permiso denegado'}), 403
    tecnico = Tecnico.query.get_or_404(tid)
    if not (tecnico.qr_imagen or '').strip():
        try:
            _save_tecnico_qr_image(tecnico)
            db.session.commit()
        except Exception as exc:
            return jsonify({'error': f'No se pudo generar QR: {exc}'}), 500

    filename = os.path.basename(tecnico.qr_imagen)
    abs_path = os.path.join(TECNICOS_QR_DIR, filename)
    if not os.path.exists(abs_path):
        return jsonify({'error': 'Archivo QR no encontrado'}), 404
    return send_from_directory(os.path.abspath(TECNICOS_QR_DIR), filename, as_attachment=True,
                               download_name=f'QR_{tecnico.numero_empleado}.png')


@app.route('/api/tecnicos/<int:tid>/signature-link', methods=['POST'])
@login_required
def api_generar_link_firma_tecnico(tid):
    user = get_current_user()
    if not (user and (user.es_admin or user.has_permission('catalog', 'edit'))):
        return jsonify({'error': 'Permiso denegado'}), 403

    tecnico = Tecnico.query.get_or_404(tid)
    payload = request.get_json(silent=True) or request.form or {}
    sig_type = (payload.get('sig_type') or '').strip().lower()
    if not _valid_tecnico_signature_type(sig_type):
        return jsonify({'error': 'Tipo de firma inválido'}), 400

    return jsonify({
        'ok': True,
        'sig_type': sig_type,
        'sig_label': _tecnico_signature_label(sig_type),
        'url': _build_tecnico_signature_public_url(tecnico.id, sig_type),
        'expires_in_seconds': TECNICO_SIGNATURE_LINK_MAX_AGE_SECONDS,
    }), 200


@app.route('/tecnicos/firma/<token>', methods=['GET', 'POST'])
def tecnico_firma_publica(token):
    resolved, error_code = _resolve_tecnico_signature_token(token)

    if request.method == 'POST':
        if error_code == 'expirado':
            return jsonify({'error': 'El link para firmar ya expiró. Solicita uno nuevo.'}), 410
        if error_code == 'consumido':
            return jsonify({'error': 'Este link de firma ya fue utilizado. Solicita uno nuevo si necesitas recapturarla.'}), 410
        if error_code:
            return jsonify({'error': 'El link para firmar es inválido.'}), 404

        signature_file = request.files.get('signature_file')
        if not signature_file or not (signature_file.filename or '').strip():
            return jsonify({'error': 'Adjunta una firma válida antes de guardar.'}), 400
        if not _allowed_signature_file(signature_file.filename):
            return jsonify({'error': 'Formato inválido. Usa PNG/JPG/JPEG/WEBP.'}), 400

        tecnico = resolved['tecnico']
        sig_type = resolved['sig_type']
        token_id = resolved['token_id']
        try:
            saved_url = _save_tecnico_signature_file(tecnico.id, signature_file, sig_type)
            if not saved_url:
                return jsonify({'error': 'No se pudo guardar la firma.'}), 500
            tecnico.actualizado_en = datetime.utcnow()
            db.session.commit()
            _mark_tecnico_signature_token_consumed(tecnico.id, sig_type, token_id)
            return jsonify({
                'ok': True,
                'message': f"{resolved['sig_label']} guardada correctamente.",
                'signature_url': _get_tecnico_signature_url(tecnico.id, sig_type),
            }), 200
        except Exception as exc:
            db.session.rollback()
            logger.error(f'Error guardando firma pública técnico={tecnico.id} tipo={sig_type}: {exc}', exc_info=True)
            return jsonify({'error': 'No se pudo guardar la firma en este momento.'}), 500

    template_data = {
        'page_status': 'ready',
        'tecnico': None,
        'sig_type': None,
        'sig_label': 'Firma',
        'signature_url': None,
        'max_age_hours': max(1, int(round(TECNICO_SIGNATURE_LINK_MAX_AGE_SECONDS / 3600))),
    }

    if error_code == 'expirado':
        template_data.update({
            'page_status': 'expired',
            'error_message': 'El link para firmar ya expiró. Solicita uno nuevo al área responsable.',
        })
        return render_template('tecnico_firma_publica.html', **template_data), 410

    if error_code == 'consumido':
        template_data.update({
            'page_status': 'used',
            'error_message': 'Este link de firma ya fue utilizado y no puede volver a usarse.',
        })
        return render_template('tecnico_firma_publica.html', **template_data), 410

    if error_code:
        template_data.update({
            'page_status': 'invalid',
            'error_message': 'El link para firmar es inválido o ya no está disponible.',
        })
        return render_template('tecnico_firma_publica.html', **template_data), 404

    tecnico = resolved['tecnico']
    sig_type = resolved['sig_type']
    template_data.update({
        'tecnico': tecnico,
        'sig_type': sig_type,
        'sig_label': resolved['sig_label'],
        'signature_url': _get_tecnico_signature_url(tecnico.id, sig_type),
    })
    return render_template('tecnico_firma_publica.html', **template_data), 200


@app.route('/uploads/tecnicos/fotos/<filename>')
def servir_foto_tecnico(filename):
    return send_from_directory(os.path.abspath(TECNICOS_FOTO_DIR), filename)


@app.route('/uploads/tecnicos/qr/<filename>')
def servir_qr_tecnico(filename):
    return send_from_directory(os.path.abspath(TECNICOS_QR_DIR), filename)


@app.route('/uploads/tecnicos/docs/<filename>')
def servir_doc_tecnico(filename):
    return send_from_directory(os.path.abspath(TECNICOS_DOCS_DIR), filename)


@app.route('/uploads/tecnicos/firmas/<filename>')
def servir_firma_tecnico(filename):
    return send_from_directory(os.path.abspath(TECNICOS_FIRMAS_DIR), filename)


def _get_tecnico_signature_abs_path(tecnico_id, sig_type):
    for ext in ('png', 'jpg', 'jpeg', 'webp'):
        name = f'tec_{int(tecnico_id)}_{sig_type}.{ext}'
        abs_path = os.path.join(TECNICOS_FIRMAS_DIR, name)
        if os.path.exists(abs_path):
            return abs_path
    return None


def _send_credencial_jpeg(tecnico, variant, lado, vigente=True):
    from credencial_export import render_credencial_jpeg

    base_dir = os.path.dirname(os.path.abspath(__file__))
    try:
        data = render_credencial_jpeg(
            tecnico,
            variant,
            lado,
            base_dir,
            firma_cri_path=_get_tecnico_signature_abs_path(tecnico.id, 'cri'),
            firma_supervision_path=_get_tecnico_signature_abs_path(tecnico.id, 'supervision'),
            vigente=vigente,
        )
    except ValueError:
        abort(404)
    except Exception as exc:
        logger.exception('Error generando JPG de credencial: %s', exc)
        abort(500)
    num = getattr(tecnico, 'numero_empleado', None) or tecnico.id
    prefix = 'credencial_seguridad' if variant == 'seguridad' else 'credencial'
    fname = f'{prefix}_{num}_{lado}.jpg'
    return send_file(
        BytesIO(data),
        mimetype='image/jpeg',
        as_attachment=True,
        download_name=fname,
    )


@app.route('/tecnicos/<int:tid>/credencial')
@login_required
def credencial_tecnico(tid):
    user = get_current_user()
    if not (user and (user.es_admin or user.has_permission('catalog', 'edit'))):
        return render_template('403.html'), 403
    tecnico = Tecnico.query.get_or_404(tid)
    # Generar QR si no existe
    if not (tecnico.qr_imagen or '').strip():
        try:
            _save_tecnico_qr_image(tecnico)
            db.session.commit()
        except Exception as exc:
            logger.warning(f'No se pudo generar QR para credencial: {exc}')
    now = datetime.utcnow()
    vigente = tecnico.esta_vigente(now)
    return render_template('credencial_tecnico.html', tecnico=tecnico, vigente=vigente, now=now)


@app.route('/tecnicos/<int:tid>/credencial/export/<lado>.jpg')
@login_required
def credencial_tecnico_export_jpg(tid, lado):
    user = get_current_user()
    if not (user and (user.es_admin or user.has_permission('catalog', 'edit'))):
        return render_template('403.html'), 403
    tecnico = Tecnico.query.get_or_404(tid)
    now = datetime.utcnow()
    vigente = tecnico.esta_vigente(now)
    return _send_credencial_jpeg(tecnico, 'corporativa', lado, vigente=vigente)


@app.route('/tecnicos/<int:tid>/credencial-seguridad')
@login_required
def credencial_walmart(tid):
    user = get_current_user()
    if not (user and (user.es_admin or user.has_permission('catalog', 'edit'))):
        return render_template('403.html'), 403
    tecnico = Tecnico.query.get_or_404(tid)
    if not (tecnico.qr_imagen or '').strip():
        try:
            _save_tecnico_qr_image(tecnico)
            db.session.commit()
        except Exception as exc:
            logger.warning(f'No se pudo generar QR para credencial seguridad: {exc}')
    now = datetime.utcnow()
    vigente = tecnico.esta_vigente(now)
    firma_cri_url = _get_tecnico_signature_url(tecnico.id, 'cri')
    firma_supervision_url = _get_tecnico_signature_url(tecnico.id, 'supervision')
    return render_template(
        'credencial_walmart.html',
        tecnico=tecnico,
        vigente=vigente,
        now=now,
        firma_cri_url=firma_cri_url,
        firma_supervision_url=firma_supervision_url,
    )


@app.route('/tecnicos/<int:tid>/credencial-seguridad/export/<lado>.jpg')
@login_required
def credencial_seguridad_export_jpg(tid, lado):
    user = get_current_user()
    if not (user and (user.es_admin or user.has_permission('catalog', 'edit'))):
        return render_template('403.html'), 403
    tecnico = Tecnico.query.get_or_404(tid)
    now = datetime.utcnow()
    vigente = tecnico.esta_vigente(now)
    return _send_credencial_jpeg(tecnico, 'seguridad', lado, vigente=vigente)


# ==================== MÓDULO HOJAS DE RUTA NUEVO ====================

@app.route('/hojas_ruta_nuevo')
@login_required
@requires_any_permission([('hojas_mp', 'view'), ('hojas', 'view'), ('catalog', 'view')])
def hojas_ruta_nuevo_list():
    """Acceso principal del módulo MP: mismo flujo visual que Hojas de ruta, independiente."""
    return hojas_ruta_nuevo_form()

@app.route('/hojas_ruta_nuevo_form')
@login_required
@requires_any_permission([('hojas_mp', 'view'), ('hojas', 'view'), ('catalog', 'view')])
def hojas_ruta_nuevo_form():
    """Formulario para crear hojas de ruta nuevas."""
    almacenes = ['AlmacenPT', 'AlmacenMP', 'Maquinaria', 'Walmart', 'ALMACEN 3']
    hojas = HojaRutaNueva.query.order_by(HojaRutaNueva.fecha_creacion.desc()).all()
    hojas_data = []
    for h in hojas:
        item = h.to_dict()
        item['serie'] = item.get('nombre')
        item['clave'] = item.get('pn')
        item['descripcion_clave'] = _resolve_clave_descripcion_by_pn(h.pn)
        item['qr_payload'] = f"HRNID:{h.id};SERIE:{h.nombre or ''}"
        item['qr_deeplink'] = request.url_root.rstrip('/') + f"/hoja_nuevo/{h.id}"
        item['orden_trabajo'] = item.get('orden_trabajo_hr')
        item['comentarios'] = _mp_strip_process_state_block(_clean_nullable_text(item.get('materia_prima')))
        item['firma_ing_jose'] = item.get('supervisor')
        item['firma_ing_rodrigo'] = item.get('operador')
        item['historial_cargas'] = []
        item['estaciones'] = _build_mp_virtual_estaciones_by_pn(h.pn, _mp_parse_completed_process_ids(h.materia_prima))
        hojas_data.append(item)

    companion_hojas = []
    hojas_entregas = HojaRutaEntrega.query.order_by(HojaRutaEntrega.fecha_creacion.desc()).limit(120).all()
    for h in hojas_entregas:
        comentarios_bruto = _clean_nullable_text(h.materia_prima)
        companion_hojas.append({
            'id': h.id,
            'serie': h.nombre,
            'qr_payload': f"HRID:{h.id};SERIE:{h.nombre or ''}",
            'qr_deeplink': request.url_root.rstrip('/') + f"/hoja/{h.id}",
            'clave': h.pn,
            'descripcion_clave': _resolve_clave_descripcion_by_pn(h.pn),
            'calidad': h.calidad,
            'fecha_salida': h.fecha_salida.isoformat() if h.fecha_salida else None,
            'cantidad_piezas': h.cantidad_piezas,
            'almacen': h.almacen,
            'orden_trabajo': h.orden_trabajo_hr,
            'comentarios': _qc_strip_scrap_summary(comentarios_bruto),
            'estado': h.estado,
        })

    return render_template(
        'hojas_ruta_form.html',
        hojas=hojas_data,
        almacenes=almacenes,
        nuevo_modulo=True,
        api_hojas_base='/api/hojas_ruta_nuevo',
        hoja_view_base='/hoja_nuevo',
        modulo_titulo='HOJAS DE RUTA MP',
        companion_modulo_titulo='HOJAS DE RUTA ENTREGAS',
        companion_hojas=companion_hojas,
    )

@app.route('/hojas_ruta_nuevo/<int:maquina_id>')
@login_required
def hojas_ruta_nuevo_detalle(maquina_id):
    """Detalle de hojas de ruta nuevas para una máquina específica."""
    maquina = Máquina.query.get_or_404(maquina_id)
    hojas = HojaRutaNueva.query.filter_by(maquina_id=maquina_id).order_by(HojaRutaNueva.fecha_creacion.desc()).all()
    hojas_data = []
    for hoja in hojas:
        h = hoja.to_dict()
        h['estaciones'] = _build_mp_virtual_estaciones_by_pn(
            hoja.pn,
            _mp_parse_completed_process_ids(hoja.materia_prima),
        )
        hojas_data.append(h)
    facturadas_info = {}
    return render_template(
        'hojas_ruta_detalle.html',
        maquina=maquina,
        hojas=hojas_data,
        facturadas_info=facturadas_info,
        nuevo_modulo=True,
        volver_url='/hojas_ruta_nuevo',
        nueva_hoja_url='/hojas_ruta_nuevo_form',
        hoja_view_base='/hoja_nuevo'
    )

@app.route('/hoja_nuevo/<int:hoja_id>')
@login_required
@requires_any_permission([('hojas_mp', 'view'), ('hojas', 'view'), ('catalog', 'view'), ('entregas', 'view'), ('almacen', 'view'), ('facturacion', 'view')])
def hoja_ruta_nuevo_ver(hoja_id):
    """Vista independiente para ver una hoja nueva por ID, sin requerir máquina."""
    hoja = HojaRutaNueva.query.get_or_404(hoja_id)
    h = hoja.to_dict()
    h['comentarios_usuario'] = _mp_strip_process_state_block(_clean_nullable_text(h.get('materia_prima')))
    h['scrap_qc'] = None
    h['descripcion_clave'] = _resolve_clave_descripcion_by_pn(hoja.pn)
    h['qr_payload'] = f"HRNID:{hoja.id};SERIE:{hoja.nombre or ''}"
    h['qr_deeplink'] = request.url_root.rstrip('/') + f"/hoja_nuevo/{hoja.id}"
    h['estaciones'] = _build_mp_virtual_estaciones_by_pn(hoja.pn, _mp_parse_completed_process_ids(hoja.materia_prima))
    registros_qc = QCProduccionRegistro.query.filter_by(maquina_id=hoja.maquina_id).order_by(QCProduccionRegistro.creado_en.desc()).limit(120).all() if hoja.maquina_id else []
    grouped_qc = _group_mp_qc_alerts_by_maquina_hoja(registros_qc)
    qc_bundle = grouped_qc.get((int(hoja.maquina_id or 0), int(hoja.id)), {'recent': [], 'latest_by_process': {}})
    for est in h['estaciones']:
        est['qc_alert'] = qc_bundle['latest_by_process'].get(int(est.get('id') or 0))
    h['qc_alerts_recent'] = qc_bundle['recent']
    return render_template('hoja_ruta_ver.html', hoja=h, volver_url='/hojas_ruta_nuevo_form')

@app.route('/hojas_ruta')
@login_required
@requires_any_permission([('estaciones', 'view'), ('hojas_entregas', 'view'), ('catalog', 'view')])
def hojas_ruta_entregas_list():
    """Estaciones T - producción por máquina usando hojas de ruta entregas."""
    maquinas = Máquina.query.order_by(Máquina.nombre.asc()).all()

    hojas_asignadas = HojaRutaEntrega.query.filter(
        HojaRutaEntrega.maquina_id.isnot(None),
        HojaRutaEntrega.estado.in_(['activa', 'pausada'])
    ).order_by(HojaRutaEntrega.fecha_actualizacion.desc(), HojaRutaEntrega.fecha_creacion.desc()).all()

    hojas_asignadas_por_maquina = {}
    for hoja in hojas_asignadas:
        hojas_asignadas_por_maquina.setdefault(hoja.maquina_id, []).append(hoja)

    hoja_activa_por_maquina = {}
    for maq_id, items in hojas_asignadas_por_maquina.items():
        ordered = _order_machine_queue_items(items)
        hojas_asignadas_por_maquina[maq_id] = ordered
        hoja_activa_por_maquina[maq_id] = _pick_machine_active_hoja(ordered)

    qc_registros = QCProduccionRegistro.query.filter(
        QCProduccionRegistro.maquina_id.isnot(None)
    ).order_by(QCProduccionRegistro.creado_en.desc()).limit(500).all()
    qc_alerts_grouped = _group_entregas_qc_alerts_by_maquina_hoja(qc_registros)

    maquinas_data = []
    for maq in maquinas:
        hojas_asignadas_maq = hojas_asignadas_por_maquina.get(maq.id, [])
        hoja_activa = hoja_activa_por_maquina.get(maq.id)
        hoja_activa_dict = hoja_activa.to_dict() if hoja_activa else None
        tiempo_objetivo_proceso = None
        tiempo_transcurrido_proceso = None
        estaciones_db = []

        if hoja_activa:
            estaciones_db = EstacionTrabajo.query.filter_by(
                hoja_ruta_id=hoja_activa.id
            ).order_by(EstacionTrabajo.orden).all()
            if hoja_activa_dict is not None:
                hoja_activa_dict['estaciones'] = [e.to_dict() for e in estaciones_db]
                qc_bundle = qc_alerts_grouped.get(
                    (int(maq.id), int(hoja_activa.id)),
                    {'recent': [], 'latest_by_process': {}},
                )
                for est in hoja_activa_dict['estaciones']:
                    est['nombre'] = est.get('operacion') or ''
                    est['qc_alert'] = qc_bundle['latest_by_process'].get(int(est.get('id') or 0))
                hoja_activa_dict['qc_alerts_recent'] = qc_bundle['recent']

        estacion_actual = 'Sin produccion'
        tiempo_real = '00:00:00'

        if hoja_activa and estaciones_db:
            estacion_en_curso = next(
                (e for e in estaciones_db if (e.estado or '').lower() == 'en_curso'),
                None,
            )
            if estacion_en_curso:
                estacion_actual = estacion_en_curso.operacion or estacion_en_curso.nombre or 'En curso'
                cantidad = max(1, int(hoja_activa.cantidad_piezas or 0))
                sec_por_pieza = _station_seconds(estacion_en_curso)
                objetivo_sec = max(0, sec_por_pieza * cantidad)
                inicio = estacion_en_curso.fecha_inicio or hoja_activa.fecha_salida
                if hoja_activa.estado == 'pausada' and hoja_activa.fecha_actualizacion and inicio:
                    transcurrido_sec = _working_seconds_between(inicio, hoja_activa.fecha_actualizacion)
                else:
                    transcurrido_sec = _working_seconds_between(inicio, datetime.utcnow()) if inicio else 0
                if objetivo_sec > 0:
                    tiempo_objetivo_proceso = _format_seconds_to_hms(objetivo_sec)
                    tiempo_transcurrido_proceso = _format_seconds_to_hms(transcurrido_sec)
                    tiempo_real = tiempo_transcurrido_proceso

        maquinas_data.append({
            'id': maq.id,
            'nombre': maq.nombre,
            'descripcion': maq.descripcion,
            'imagen_url': maq.imagen_url,
            'hoja_activa': hoja_activa_dict,
            'hojas_asignadas': [
                {
                    'id': qh.id,
                    'nombre': qh.nombre,
                    'serie': qh.nombre,
                    'pn': qh.pn,
                    'estado': qh.estado,
                    'cantidad_piezas': qh.cantidad_piezas,
                    'es_activa': bool(hoja_activa and qh.id == hoja_activa.id),
                }
                for qh in hojas_asignadas_maq
            ],
            'hojas_asignadas_count': len(hojas_asignadas_maq),
            'cola_maxima': _machine_queue_limit_for_type(getattr(maq, 'tipo', None)),
            'activo': getattr(maq, 'activo', False),
            'estacion_actual': estacion_actual,
            'tiempo_real': tiempo_real,
            'tiempo_objetivo_proceso': tiempo_objetivo_proceso,
            'tiempo_transcurrido_proceso': tiempo_transcurrido_proceso,
            'tipo': getattr(maq, 'tipo', None),
            'plantilla_default': getattr(maq, 'plantilla_default', None),
        })

    hojas_pendientes = _query_hojas_entregas_pendientes_estaciones()

    pendientes_data = []
    for hoja in hojas_pendientes:
        estaciones = EstacionTrabajo.query.filter_by(
            hoja_ruta_id=hoja.id
        ).order_by(EstacionTrabajo.orden).all()
        pendientes_data.append({
            'id': hoja.id,
            'serie': hoja.nombre,
            'clave': hoja.pn,
            'estado': hoja.estado,
            'cantidad_piezas': hoja.cantidad_piezas,
            'tiempo_total': hoja.total_tiempo,
            'fecha_creacion': hoja.fecha_creacion.isoformat() if hoja.fecha_creacion else None,
            'estaciones': [
                {**e.to_dict(), 'nombre': e.operacion or e.nombre or ''}
                for e in estaciones
            ],
        })

    facturadas_info = {}

    resp = make_response(render_template(
        'hojas_ruta_list.html',
        maquinas=maquinas_data,
        hojas_pendientes=pendientes_data,
        facturadas_info=facturadas_info,
        nuevo_modulo=False,
        hoja_detalle_base='/hojas_ruta',
        api_resolver_codigo='/api/hojas_ruta/resolver_codigo',
        api_maquina_base='/api/maquinas',
        api_hojas_base='/api/hojas_ruta',
        modulo_titulo='Estaciones T - Produccion por Maquina',
        modulo_subtitulo='Gestion centralizada de activacion, OT y seguimiento de hojas de ruta entregas por equipo.',
    ))
    resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    resp.headers['Pragma'] = 'no-cache'
    resp.headers['Expires'] = '0'
    return resp


@app.route('/mapa_maquinas')
@login_required
@requires_any_permission([('mapa', 'view'), ('catalog', 'view')])
def mapa_maquinas():
    """Vista de mapa de maquinas con estado en tiempo real."""
    return render_template('mapa_maquinas.html')


@app.route('/produccion/reportes')
@login_required
@requires_any_permission([('mapa', 'view'), ('estaciones', 'view'), ('catalog', 'view')])
def produccion_reportes_page():
    """Modulo de reportes de produccion con consultas interactivas en tiempo real."""
    return render_template('produccion_reportes.html')


@app.route('/api/mapa_maquinas')
@login_required
@requires_any_permission([('mapa', 'view'), ('catalog', 'view')])
def api_mapa_maquinas():
    """Datos para el mapa de maquinas (estado, hoja activa, pieza, tiempo)."""
    todas_maquinas = Máquina.query.order_by(Máquina.nombre.asc()).all()
    hojas_activas = HojaRutaEntrega.query.filter(
        HojaRutaEntrega.maquina_id.isnot(None),
        HojaRutaEntrega.estado.in_(['activa', 'pausada'])
    ).order_by(HojaRutaEntrega.fecha_creacion.desc()).all()
    hoja_activa_por_maquina: dict = {}
    hoja_entrega_por_maquina: dict = {}
    for h in hojas_activas:
        existing_entrega = hoja_entrega_por_maquina.get(h.maquina_id)
        if existing_entrega is None:
            hoja_entrega_por_maquina[h.maquina_id] = h
        elif existing_entrega.estado != 'activa' and h.estado == 'activa':
            hoja_entrega_por_maquina[h.maquina_id] = h

        existing = hoja_activa_por_maquina.get(h.maquina_id)
        if existing is None:
            hoja_activa_por_maquina[h.maquina_id] = h
        elif existing.estado != 'activa' and h.estado == 'activa':
            hoja_activa_por_maquina[h.maquina_id] = h

    schedule_window = _get_machine_schedule_window()

    # Regla operativa: fuera de turno, sin hoja activa asignada => maquina desactivada por default.
    # Durante turno activo, el scheduler de horario manda para respetar 06:30 / 12:00 / 12:30 / 16:00.
    estado_maquina_changed = False
    if not schedule_window['active']:
        for maq in todas_maquinas:
            if (maq.id not in hoja_activa_por_maquina
                    and bool(getattr(maq, 'activo', False))):
                maq.activo = False
                estado_maquina_changed = True

    if estado_maquina_changed:
        try:
            db.session.commit()
        except Exception:
            db.session.rollback()

    maquinas = list(todas_maquinas)
    data = []
    now_dt = datetime.utcnow()

    window_hour_start = now_dt - timedelta(hours=1)
    window_day_start = now_dt.replace(hour=0, minute=0, second=0, microsecond=0)
    window_week_start = (window_day_start - timedelta(days=window_day_start.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
    machine_available_hour = int((now_dt - window_hour_start).total_seconds())
    machine_available_day = int((now_dt - window_day_start).total_seconds())
    machine_available_week = int((now_dt - window_week_start).total_seconds())

    productive_hour = 0
    productive_day = 0
    productive_week = 0
    producing_count = 0
    assigned_count = 0
    progreso_sum = 0.0
    progreso_count = 0
    ritmo_sum = 0.0
    ritmo_count = 0
    alertas_buzon_nuevas = False
    alertas_buzon_pendientes_envio = []

    def _bounded_seconds(start_dt, end_dt, window_start, window_end):
        if not start_dt or not end_dt:
            return 0
        bounded_start = max(start_dt, window_start)
        bounded_end = min(end_dt, window_end)
        if bounded_end <= bounded_start:
            return 0
        return int((bounded_end - bounded_start).total_seconds())

    def _safe_pct(prod_sec, avail_sec):
        if avail_sec <= 0:
            return 0.0
        return round((float(prod_sec) * 100.0) / float(avail_sec), 1)

    for idx, maq in enumerate(maquinas):
        hoja_activa = hoja_activa_por_maquina.get(maq.id)
        hoja_entrega_activa = hoja_entrega_por_maquina.get(maq.id)
        maquina_productive_hour = 0
        maquina_productive_day = 0
        maquina_productive_week = 0
        hoja_total_tiempo = None
        hoja_modulo = 'entregas'
        estacion_actual = None
        tiempo_objetivo = None
        tiempo_transcurrido = None
        tiempo_restante = None
        proceso_culminado = False
        progreso_pct = 0
        tiempo_proceso_pieza = None
        objetivo_sec_metric = 0
        transcurrido_sec_metric = 0
        if hoja_activa:
            hoja_total_tiempo = hoja_activa.total_tiempo
            estacion_actual = EstacionTrabajo.query.filter_by(
                hoja_ruta_id=hoja_activa.id,
                estado='en_curso'
            ).order_by(EstacionTrabajo.orden).first()

            if estacion_actual:
                cantidad = max(1, int(hoja_activa.cantidad_piezas or 0))
                sec_por_pieza = _station_seconds(estacion_actual)
                objetivo_sec = max(0, sec_por_pieza * cantidad)

                inicio = estacion_actual.fecha_inicio or hoja_activa.fecha_salida
                if hoja_activa.estado == 'pausada' and hoja_activa.fecha_actualizacion and inicio:
                    transcurrido_sec = _working_seconds_between(inicio, hoja_activa.fecha_actualizacion)
                else:
                    transcurrido_sec = _working_seconds_between(inicio, datetime.utcnow()) if inicio else 0
                restante_sec = max(0, objetivo_sec - transcurrido_sec)

                if sec_por_pieza > 0:
                    tiempo_proceso_pieza = _format_seconds_to_hms(sec_por_pieza)
                if objetivo_sec > 0:
                    objetivo_sec_metric = objetivo_sec
                    transcurrido_sec_metric = transcurrido_sec
                    tiempo_objetivo = _format_seconds_to_hms(objetivo_sec)
                    tiempo_transcurrido = _format_seconds_to_hms(transcurrido_sec)
                    tiempo_restante = _format_seconds_to_hms(restante_sec)
                    proceso_culminado = restante_sec <= 0
                    progreso_pct = min(100, int((transcurrido_sec * 100) / objetivo_sec))

            # Eficiencia planta: computar por hoja activa en maquina.
            start_ref = hoja_activa.fecha_salida
            if start_ref:
                if hoja_activa.estado == 'pausada' and hoja_activa.fecha_actualizacion:
                    interval_end = hoja_activa.fecha_actualizacion
                else:
                    interval_end = now_dt
                if interval_end > start_ref:
                    maquina_productive_hour = _bounded_seconds(start_ref, interval_end, window_hour_start, now_dt)
                    maquina_productive_day = _bounded_seconds(start_ref, interval_end, window_day_start, now_dt)
                    maquina_productive_week = _bounded_seconds(start_ref, interval_end, window_week_start, now_dt)
                    productive_hour += maquina_productive_hour
                    productive_day += maquina_productive_day
                    productive_week += maquina_productive_week

            assigned_count += 1
            if objetivo_sec_metric > 0:
                progreso_sum += float(progreso_pct)
                progreso_count += 1
                if transcurrido_sec_metric <= objetivo_sec_metric:
                    ritmo_actual = 100.0
                else:
                    ritmo_actual = max(0.0, (objetivo_sec_metric * 100.0) / max(1, transcurrido_sec_metric))
                ritmo_sum += min(100.0, ritmo_actual)
                ritmo_count += 1

        if hoja_activa:
            if estacion_actual:
                estado_code = 'produciendo'
                estado_label = 'En curso'
            else:
                estado_code = 'activa'
                estado_label = 'Activa'
        else:
            estado_code = 'sin_hoja'
            estado_label = 'Sin hoja'

        if proceso_culminado and hoja_activa:
            estado_code = 'mover'
            estado_label = 'MOVER'
            origen_modulo = 'mapa_maquinas'
            tipo_alerta = 'proceso_culminado'
            hoja_modulo = 'entregas'
            estacion_key = str(estacion_actual.id) if estacion_actual else 'sin_estacion'
            evento_clave = f"mapa:{hoja_modulo}:maq:{maq.id}:hoja:{hoja_activa.id}:est:{estacion_key}:culminado"
            alerta = _crear_alerta_buzon(
                evento_clave=evento_clave,
                origen=origen_modulo,
                tipo=tipo_alerta,
                titulo=f"Proceso culminado en {maq.nombre}",
                mensaje=f"Hoja: {hoja_activa.nombre or 'N/A'} | Pieza: {hoja_activa.pn or 'N/A'}",
                maquina_id=maq.id,
                hoja_id=hoja_activa.id,
                estacion_id=estacion_actual.id if estacion_actual else None,
                commit=False,
            )
            if alerta is not None and alerta in db.session.new:
                alertas_buzon_nuevas = True
                alertas_buzon_pendientes_envio.append(alerta)

        # Máquina inactiva → estado PARO (MOVER tiene prioridad)
        if not bool(getattr(maq, 'activo', False)) and estado_code != 'mover':
            estado_code = 'paro'
            estado_label = 'Paro de turno' if not schedule_window['active'] else 'Paro'

        if estado_code == 'produciendo':
            producing_count += 1

        data.append({
            'id': maq.id,
            'nombre': maq.nombre,
            'descripcion': maq.descripcion,
            'tipo': maq.tipo,
            'plantilla_default': maq.plantilla_default,
            'activo': bool(getattr(maq, 'activo', False)),
            'pos_x': getattr(maq, 'pos_x', None),
            'pos_y': getattr(maq, 'pos_y', None),
            'estado_code': estado_code,
            'estado_label': estado_label,
            'estacion_actual': estacion_actual.nombre if estacion_actual else None,
            'hoja_actual_modulo': hoja_modulo if hoja_activa else None,
            'hoja_serie': hoja_activa.nombre if hoja_activa else None,
            'pieza': hoja_activa.pn if hoja_activa else None,
            'pieza_descripcion': _resolve_clave_descripcion_by_pn(hoja_activa.pn) if hoja_activa else None,
            'cantidad_piezas': int(hoja_activa.cantidad_piezas) if hoja_activa else None,
            'orden_trabajo_hr': hoja_activa.orden_trabajo_hr if hoja_activa else None,
            'orden_trabajo_pt': hoja_activa.orden_trabajo_pt if hoja_activa else None,
            'hoja_entregas_serie': hoja_entrega_activa.nombre if hoja_entrega_activa else None,
            'hoja_entregas_pieza': hoja_entrega_activa.pn if hoja_entrega_activa else None,
            'hoja_entregas_ot': ((hoja_entrega_activa.orden_trabajo_hr or '') if hoja_entrega_activa else '') or ((hoja_entrega_activa.orden_trabajo_pt or '') if hoja_entrega_activa else ''),
            'tiene_hoja_entregas': bool(hoja_entrega_activa),
            'tiempo_total': hoja_total_tiempo if hoja_activa else None,
            'fecha_termino': hoja_activa.fecha_termino.isoformat() if (hoja_activa and hoja_activa.fecha_termino) else None,
            'tiempo_proceso_pieza': tiempo_proceso_pieza,
            'tiempo_objetivo_proceso': tiempo_objetivo,
            'tiempo_transcurrido_proceso': tiempo_transcurrido,
            'tiempo_restante_proceso': tiempo_restante,
            'proceso_culminado': proceso_culminado,
            'progreso_proceso_pct': progreso_pct,
            'estacion_actual_id': estacion_actual.id if estacion_actual else None,
            'hoja_modulo': hoja_modulo,
            'detalle_url': f"/hojas_ruta/{maq.id}",
            'tiempo_productivo_hora_sec': maquina_productive_hour,
            'tiempo_productivo_dia_sec': maquina_productive_day,
            'tiempo_productivo_semana_sec': maquina_productive_week,
            'tiempo_productivo_hora_hms': _format_seconds_to_hms(maquina_productive_hour),
            'tiempo_productivo_dia_hms': _format_seconds_to_hms(maquina_productive_day),
            'tiempo_productivo_semana_hms': _format_seconds_to_hms(maquina_productive_week),
            'tiempo_disponible_hora_sec': machine_available_hour,
            'tiempo_disponible_dia_sec': machine_available_day,
            'tiempo_disponible_semana_sec': machine_available_week,
            'tiempo_disponible_hora_hms': _format_seconds_to_hms(machine_available_hour),
            'tiempo_disponible_dia_hms': _format_seconds_to_hms(machine_available_day),
            'tiempo_disponible_semana_hms': _format_seconds_to_hms(machine_available_week),
            'eficiencia_hora_pct': _safe_pct(maquina_productive_hour, machine_available_hour),
            'eficiencia_dia_pct': _safe_pct(maquina_productive_day, machine_available_day),
            'eficiencia_semana_pct': _safe_pct(maquina_productive_week, machine_available_week),
        })

    estado_priority = {
        'mover': 0,
        'produciendo': 1,
        'activa': 2,
        'paro': 3,
        'sin_hoja': 4,
    }
    data.sort(
        key=lambda item: (
            estado_priority.get(item.get('estado_code'), 9),
            0 if bool(item.get('activo')) else 1,
            (item.get('nombre') or '').lower(),
        )
    )

    # Eficiencia de planta: usar alcance del tablero (todas las maquinas visibles en el mapa).
    machine_count = max(1, len(maquinas))
    available_hour = machine_count * int((now_dt - window_hour_start).total_seconds())
    available_day = machine_count * int((now_dt - window_day_start).total_seconds())
    available_week = machine_count * int((now_dt - window_week_start).total_seconds())

    def _pct(prod, avail):
        if avail <= 0:
            return 0
        return round((prod * 100.0) / avail, 1)

    hora_pct = _pct(productive_hour, available_hour)
    dia_pct = _pct(productive_day, available_day)
    semana_pct = _pct(productive_week, available_week)

    utilizacion_actual_pct = round((producing_count * 100.0) / machine_count, 1)
    asignacion_actual_pct = round((assigned_count * 100.0) / machine_count, 1)
    avance_promedio_pct = round((progreso_sum / progreso_count), 1) if progreso_count else 0.0
    ritmo_objetivo_pct = round((ritmo_sum / ritmo_count), 1) if ritmo_count else 0.0
    oee_lite_pct = round((dia_pct * ritmo_objetivo_pct) / 100.0, 1)

    if alertas_buzon_nuevas:
        try:
            db.session.commit()
            for alerta in alertas_buzon_pendientes_envio:
                _send_alerta_whatsapp_if_enabled(alerta)
                _send_alerta_telegram_if_enabled(alerta)
        except Exception:
            db.session.rollback()

    eficiencia_planta = {
        'hora': {
            'porcentaje': hora_pct,
            'productivo_hms': _format_seconds_to_hms(productive_hour),
            'disponible_hms': _format_seconds_to_hms(available_hour),
            'maquinas_base': machine_count,
        },
        'dia': {
            'porcentaje': dia_pct,
            'productivo_hms': _format_seconds_to_hms(productive_day),
            'disponible_hms': _format_seconds_to_hms(available_day),
            'maquinas_base': machine_count,
        },
        'semana': {
            'porcentaje': semana_pct,
            'productivo_hms': _format_seconds_to_hms(productive_week),
            'disponible_hms': _format_seconds_to_hms(available_week),
            'maquinas_base': machine_count,
        },
        'kpis': {
            'utilizacion_actual_pct': utilizacion_actual_pct,
            'asignacion_actual_pct': asignacion_actual_pct,
            'avance_promedio_pct': avance_promedio_pct,
            'ritmo_objetivo_pct': ritmo_objetivo_pct,
            'oee_lite_pct': oee_lite_pct,
            'maquinas_produciendo': producing_count,
            'maquinas_asignadas': assigned_count,
            'procesos_medidos': progreso_count,
            'alcance_maquinas': machine_count,
        },
    }

    logistica_almacen_hoy = _build_entregas_almacen_hoy(limit=200)

    # Insights operativos ligeros para tablero de reportes (sin IA pesada).
    hoy_inicio = now_dt.replace(hour=0, minute=0, second=0, microsecond=0)
    semana_inicio = (hoy_inicio - timedelta(days=hoy_inicio.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
    mes_inicio = hoy_inicio.replace(day=1)
    if mes_inicio.month == 12:
        siguiente_mes_inicio = mes_inicio.replace(year=mes_inicio.year + 1, month=1)
    else:
        siguiente_mes_inicio = mes_inicio.replace(month=mes_inicio.month + 1)

    semana_fin = semana_inicio + timedelta(days=7)
    minutos_transcurridos_hoy = max(1, int((now_dt - hoy_inicio).total_seconds() // 60))
    minutos_dia = 24 * 60
    minutos_semana = 7 * 24 * 60
    minutos_mes = max(1, int((siguiente_mes_inicio - mes_inicio).total_seconds() // 60))

    avance_dia_ratio = min(1.0, max(0.0, float(minutos_transcurridos_hoy) / float(minutos_dia)))
    minutos_transcurridos_semana = max(1, int((now_dt - semana_inicio).total_seconds() // 60))
    minutos_transcurridos_mes = max(1, int((now_dt - mes_inicio).total_seconds() // 60))
    avance_semana_ratio = min(1.0, max(0.0, float(minutos_transcurridos_semana) / float(minutos_semana)))
    avance_mes_ratio = min(1.0, max(0.0, float(minutos_transcurridos_mes) / float(minutos_mes)))

    mp_hoy = HojaRutaNueva.query.filter(HojaRutaNueva.fecha_creacion >= hoy_inicio).count()
    entregas_hoy = HojaRutaEntrega.query.filter(HojaRutaEntrega.fecha_creacion >= hoy_inicio).count()
    mp_semana = HojaRutaNueva.query.filter(HojaRutaNueva.fecha_creacion >= semana_inicio).count()
    entregas_semana = HojaRutaEntrega.query.filter(HojaRutaEntrega.fecha_creacion >= semana_inicio).count()
    mp_mes = HojaRutaNueva.query.filter(HojaRutaNueva.fecha_creacion >= mes_inicio).count()
    entregas_mes = HojaRutaEntrega.query.filter(HojaRutaEntrega.fecha_creacion >= mes_inicio).count()

    def _hms_to_seconds(value):
        if not value:
            return 0
        parts = str(value).split(':')
        if len(parts) != 3:
            return 0
        try:
            hh = int(parts[0])
            mm = int(parts[1])
            ss = int(parts[2])
            return max(0, hh * 3600 + mm * 60 + ss)
        except Exception:
            return 0

    mover_rows = [r for r in data if (r.get('estado_code') or '') == 'mover']
    mover_count = len(mover_rows)
    mover_transcurrido_sec = sum(_hms_to_seconds(r.get('tiempo_transcurrido_proceso')) for r in mover_rows)

    sin_hoja_count = len([r for r in data if (r.get('estado_code') or '') == 'sin_hoja'])
    inactivas_count = len([r for r in data if not bool(r.get('activo'))])

    no_productivo_dia_sec = max(0, available_day - productive_day)
    productive_day_ratio = 1.0 - (float(no_productivo_dia_sec) / float(max(1, available_day)))

    envios_hoy = int((logistica_almacen_hoy or {}).get('total_envios') or 0)
    acciones_envio = ['enviada_a_almacen', 'enviada_a_almacen_revision']
    envios_semana = int(
        EntregaRegistro.query
        .filter(EntregaRegistro.accion.in_(acciones_envio))
        .filter(EntregaRegistro.fecha_creacion >= semana_inicio)
        .filter(EntregaRegistro.fecha_creacion < semana_fin)
        .count() or 0
    )
    envios_mes = int(
        EntregaRegistro.query
        .filter(EntregaRegistro.accion.in_(acciones_envio))
        .filter(EntregaRegistro.fecha_creacion >= mes_inicio)
        .filter(EntregaRegistro.fecha_creacion < siguiente_mes_inicio)
        .count() or 0
    )

    envios_proyectados_hoy = int(round(float(envios_hoy) / max(0.2, avance_dia_ratio)))
    envios_proyectados_semana = int(round(float(envios_semana) / max(0.2, avance_semana_ratio)))
    envios_proyectados_mes = int(round(float(envios_mes) / max(0.2, avance_mes_ratio)))
    mp_proyectado_hoy = int(round(float(mp_hoy) / max(0.2, avance_dia_ratio)))
    mp_proyectado_semana = int(round(float(mp_semana) / max(0.2, avance_semana_ratio)))
    mp_proyectado_mes = int(round(float(mp_mes) / max(0.2, avance_mes_ratio)))
    entregas_proyectado_hoy = int(round(float(entregas_hoy) / max(0.2, avance_dia_ratio)))
    entregas_proyectado_semana = int(round(float(entregas_semana) / max(0.2, avance_semana_ratio)))
    entregas_proyectado_mes = int(round(float(entregas_mes) / max(0.2, avance_mes_ratio)))

    sugerencias = []
    if mover_count >= max(2, int(machine_count * 0.2)):
        sugerencias.append('Revisar cuellos de botella: hay muchas máquinas en estado mover.')
    if sin_hoja_count > 0:
        sugerencias.append('Hay máquinas sin hoja activa; prioriza asignación de hojas para evitar ociosidad.')
    if productive_day_ratio < 0.6:
        sugerencias.append('La eficiencia diaria está baja; valida tiempos muertos y cambios de estado extensos.')
    if envios_hoy > 0 and envios_proyectados_hoy < envios_hoy:
        sugerencias.append('Ritmo de entregas irregular; valida tiempos de liberación y confirmación a almacén.')

    insights_operativos = {
        'fecha_utc': now_dt.isoformat(),
        'hoy': {
            'mp_generadas': int(mp_hoy),
            'entregas_generadas': int(entregas_hoy),
            'envios_almacen': int(envios_hoy),
            'envios_proyectados_cierre': int(envios_proyectados_hoy),
            'avance_dia_pct': round(avance_dia_ratio * 100.0, 1),
        },
        'proyeccion': {
            'dia': {
                'mp_actual': int(mp_hoy),
                'mp_cierre': int(mp_proyectado_hoy),
                'entregas_actual': int(entregas_hoy),
                'entregas_cierre': int(entregas_proyectado_hoy),
                'envios_actual': int(envios_hoy),
                'envios_cierre': int(envios_proyectados_hoy),
                'avance_pct': round(avance_dia_ratio * 100.0, 1),
            },
            'semana': {
                'mp_actual': int(mp_semana),
                'mp_cierre': int(mp_proyectado_semana),
                'entregas_actual': int(entregas_semana),
                'entregas_cierre': int(entregas_proyectado_semana),
                'envios_actual': int(envios_semana),
                'envios_cierre': int(envios_proyectados_semana),
                'avance_pct': round(avance_semana_ratio * 100.0, 1),
            },
            'mes': {
                'mp_actual': int(mp_mes),
                'mp_cierre': int(mp_proyectado_mes),
                'entregas_actual': int(entregas_mes),
                'entregas_cierre': int(entregas_proyectado_mes),
                'envios_actual': int(envios_mes),
                'envios_cierre': int(envios_proyectados_mes),
                'avance_pct': round(avance_mes_ratio * 100.0, 1),
            },
        },
        'operacion_actual': {
            'maquinas_en_mover': int(mover_count),
            'maquinas_sin_hoja': int(sin_hoja_count),
            'maquinas_inactivas': int(inactivas_count),
            'tiempo_mover_hms': _format_seconds_to_hms(mover_transcurrido_sec),
            'tiempo_no_productivo_dia_hms': _format_seconds_to_hms(no_productivo_dia_sec),
            'eficiencia_productiva_dia_pct': round(productive_day_ratio * 100.0, 1),
        },
        'sugerencias': sugerencias,
    }

    return jsonify({
        'maquinas': data,
        'eficiencia_planta': eficiencia_planta,
        'logistica_almacen_hoy': logistica_almacen_hoy,
        'insights_operativos': insights_operativos,
    })


@app.route('/alertas_buzon')
@login_required
@requires_permission('alertas_buzon', 'view')
def alertas_buzon_page():
    _ensure_alertas_buzon_table()
    return render_template('alertas_buzon.html')


@app.route('/api/alertas_buzon/count')
@login_required
@requires_permission('alertas_buzon', 'view')
def api_alertas_buzon_count():
    """Lightweight endpoint for sidebar badge. Always returns 200 with a safe count."""
    try:
        _ensure_alertas_buzon_table()
        count = db.session.execute(
            db.text('SELECT COUNT(*) FROM alertas_buzon_general WHERE atendida = false')
        ).scalar() or 0
        return jsonify({'pendientes': int(count)})
    except Exception:
        return jsonify({'pendientes': 0})


@app.route('/api/alertas_buzon')
@login_required
@requires_permission('alertas_buzon', 'view')
def api_alertas_buzon_list():
    if not _ensure_alertas_buzon_table():
        return jsonify({'items': [], 'total': 0, 'pendientes': 0})

    status = (request.args.get('status') or 'pendientes').strip().lower()
    limit = max(1, min(500, int(request.args.get('limit') or 200)))
    query = AlertaBuzonGeneral.query
    if status in ('pendientes', 'pending'):
        query = query.filter(AlertaBuzonGeneral.atendida.is_(False))
    elif status in ('atendidas', 'done'):
        query = query.filter(AlertaBuzonGeneral.atendida.is_(True))

    items = query.order_by(AlertaBuzonGeneral.created_at.desc()).limit(limit).all()
    pendientes_count = AlertaBuzonGeneral.query.filter(AlertaBuzonGeneral.atendida.is_(False)).count()

    return jsonify({
        'items': [_serialize_alerta_buzon(x) for x in items],
        'total': len(items),
        'pendientes': pendientes_count,
    })


@app.route('/api/alertas_buzon/<int:alerta_id>/atender', methods=['POST'])
@login_required
@requires_permission('alertas_buzon', 'view')
def api_alertas_buzon_atender(alerta_id):
    if not _ensure_alertas_buzon_table():
        return jsonify({'error': 'No se pudo acceder al buzón de alertas'}), 500

    data = request.get_json(silent=True) or {}
    nota = (data.get('nota') or '').strip()
    if not nota:
        return jsonify({'error': 'La nota de atención es obligatoria. Describe qué acción tomaste.'}), 422

    alerta = AlertaBuzonGeneral.query.get_or_404(alerta_id)
    if not alerta.atendida:
        alerta.atendida = True
        alerta.atendida_por = _current_username_for_audit(get_current_user())
        alerta.atendida_at = datetime.utcnow()
        alerta.nota_atencion = nota
        db.session.commit()

    return jsonify({'ok': True, 'item': _serialize_alerta_buzon(alerta)})


@app.route('/api/alertas_buzon/atender_todas', methods=['POST'])
@login_required
@requires_permission('alertas_buzon', 'view')
def api_alertas_buzon_atender_todas():
    if not _ensure_alertas_buzon_table():
        return jsonify({'error': 'No se pudo acceder al buzón de alertas'}), 500

    data = request.get_json(silent=True) or {}
    nota = (data.get('nota') or '').strip()
    if not nota:
        return jsonify({'error': 'La nota de atención es obligatoria. Describe qué acción tomaste.'}), 422

    user_name = _current_username_for_audit(get_current_user())
    now_dt = datetime.utcnow()
    q = AlertaBuzonGeneral.query.filter(AlertaBuzonGeneral.atendida.is_(False)).all()
    for alerta in q:
        alerta.atendida = True
        alerta.atendida_por = user_name
        alerta.atendida_at = now_dt
        alerta.nota_atencion = nota
    db.session.commit()
    return jsonify({'ok': True, 'updated': len(q)})


@app.route('/api/alertas_buzon/test_whatsapp', methods=['POST'])
@login_required
@requires_permission('alertas_buzon', 'view')
def api_alertas_buzon_test_whatsapp():
    data = request.get_json(silent=True) or {}
    mensaje = (data.get('mensaje') or '').strip()
    if not mensaje:
        mensaje = f"Prueba WhatsApp alertas desde sistema ({datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC)"

    ok, reason = _send_whatsapp_message(mensaje)
    code = 200 if ok else 400
    return jsonify({'ok': ok, 'reason': reason}), code


@app.route('/api/alertas_buzon/test_telegram', methods=['POST'])
@login_required
@requires_permission('alertas_buzon', 'view')
def api_alertas_buzon_test_telegram():
    data = request.get_json(silent=True) or {}
    mensaje = (data.get('mensaje') or '').strip()
    if not mensaje:
        mensaje = f"Prueba Telegram alertas desde sistema ({datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC)"

    ok, reason = _send_telegram_message(mensaje)
    code = 200 if ok else 400
    return jsonify({'ok': ok, 'reason': reason}), code


@app.route('/api/alertas_buzon/<int:alerta_id>/reenviar_telegram', methods=['POST'])
@login_required
@requires_permission('alertas_buzon', 'view')
def api_alertas_buzon_reenviar_telegram(alerta_id):
    if not _ensure_alertas_buzon_table():
        return jsonify({'error': 'No se pudo acceder al buzón de alertas'}), 500

    alerta = AlertaBuzonGeneral.query.get_or_404(alerta_id)
    body = (
        f"*ALERTA SISTEMA*\n"
        f"Tipo: {alerta.tipo or '-'}\n"
        f"Titulo: {alerta.titulo or '-'}\n"
        f"Mensaje: {alerta.mensaje or '-'}\n"
        f"Fecha: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC"
    )

    ok, reason = _send_telegram_message(body)
    code = 200 if ok else 400
    return jsonify({'ok': ok, 'reason': reason, 'alerta_id': alerta.id}), code


@app.route('/api/hojas_ruta/lista_para_entregas')
@login_required
@requires_any_permission([('hojas_entregas', 'view'), ('entregas', 'view'), ('catalog', 'view')])
def api_hojas_ruta_lista_para_entregas():
    """Devuelve todas las hojas de ruta como JSON para el selector de entregas del dia."""
    hojas = HojaRutaEntrega.query.order_by(HojaRutaEntrega.fecha_creacion.desc()).all()
    hoja_ids = [h.id for h in hojas]

    # Hojas que ya estan en el flujo de entregas (no finalizada/facturacion)
    flujos = HojaRutaFlujoLogistica.query.filter(
        HojaRutaFlujoLogistica.hoja_ruta_id.in_(hoja_ids)
    ).all() if hoja_ids else []
    flujo_por_hoja = {f.hoja_ruta_id: f.estado for f in flujos}

    result = []
    for h in hojas:
        descripcion = _resolve_clave_descripcion_by_pn(h.pn)
        estado_flujo = flujo_por_hoja.get(h.id)
        result.append({
            'id': h.id,
            'serie': h.nombre or '',
            'clave': h.pn or '',
            'descripcion': descripcion or '',
            'calidad': h.calidad or '',
            'cantidad_piezas': h.cantidad_piezas or 0,
            'almacen': h.almacen or '',
            'orden_trabajo': (h.orden_trabajo_hr or h.orden_trabajo_pt or ''),
            'estado': h.estado or '',
            'estado_flujo': estado_flujo or '',
            'fecha_creacion': h.fecha_creacion.strftime('%Y-%m-%d %H:%M') if h.fecha_creacion else '',
        })
    return jsonify({'hojas': result, 'total': len(result)})


@app.route('/hojas_ruta_entregas_form')
@login_required
@requires_any_permission([('hojas_entregas', 'view'), ('hojas', 'view'), ('catalog', 'view')])
def hojas_ruta_entregas_form():
    """Formulario simplificado para crear hojas de ruta de produccion."""
    almacenes = ['AlmacenPT', 'AlmacenMP', 'Maquinaria', 'Walmart', 'ALMACEN 3']
    # Listado completo para consulta
    hojas = HojaRutaEntrega.query.order_by(HojaRutaEntrega.fecha_creacion.desc()).all()

    hoja_ids = [h.id for h in hojas]
    flujos_facturacion = HojaRutaFlujoLogistica.query.filter(
        HojaRutaFlujoLogistica.hoja_ruta_id.in_(hoja_ids),
        HojaRutaFlujoLogistica.estado == 'finalizada'
    ).all() if hoja_ids else []
    facturacion_por_hoja = {f.hoja_ruta_id: f for f in flujos_facturacion}

    registros_facturacion = FacturacionRegistro.query.filter(
        FacturacionRegistro.hoja_ruta_id.in_(hoja_ids)
    ).order_by(
        FacturacionRegistro.fecha_aprobacion.desc().nullslast(),
        FacturacionRegistro.fecha_creacion.desc()
    ).all() if hoja_ids else []
    ultimo_registro_facturacion = {}
    for registro in registros_facturacion:
        if registro.hoja_ruta_id not in ultimo_registro_facturacion:
            ultimo_registro_facturacion[registro.hoja_ruta_id] = registro

    historial_cargas_por_hoja = {}
    if hoja_ids and _ensure_hoja_cargas_historial_table():
        historial_cargas = HojaRutaCargaPiezasHistorial.query.filter(
            HojaRutaCargaPiezasHistorial.hoja_ruta_id.in_(hoja_ids)
        ).order_by(
            HojaRutaCargaPiezasHistorial.fecha_creacion.desc(),
            HojaRutaCargaPiezasHistorial.id.desc()
        ).all()
        for item in historial_cargas:
            historial_cargas_por_hoja.setdefault(item.hoja_ruta_id, []).append(_serialize_hoja_carga_historial(item))

    impresion_parcial_totales = {}
    if hoja_ids and _ensure_hoja_impresiones_parciales_table():
        resumen_impresiones = db.session.query(
            HojaRutaImpresionParcial.hoja_ruta_id,
            func.coalesce(func.sum(HojaRutaImpresionParcial.cantidad_impresa), 0).label('total_impreso'),
            func.count(HojaRutaImpresionParcial.id).label('movimientos')
        ).filter(
            HojaRutaImpresionParcial.hoja_ruta_id.in_(hoja_ids)
        ).group_by(HojaRutaImpresionParcial.hoja_ruta_id).all()

        for fila in resumen_impresiones:
            impresion_parcial_totales[int(fila.hoja_ruta_id)] = {
                'total_impreso': int(fila.total_impreso or 0),
                'movimientos': int(fila.movimientos or 0),
            }

    claves_all = ClaveProducto.query.all()
    claves_idx = {
        (str(c.clave or '').strip().upper()): c
        for c in claves_all
        if str(c.clave or '').strip()
    }

    hojas_data = []
    for h in hojas:
        flujo_fact = facturacion_por_hoja.get(h.id)
        registro_fact = ultimo_registro_facturacion.get(h.id)
        qr_payload = f"HRID:{h.id};SERIE:{h.nombre or ''}"
        descripcion_clave = _resolve_clave_descripcion_by_pn(h.pn)
        clave_obj = claves_idx.get(str(h.pn or '').strip().upper())
        comentarios_bruto = _clean_nullable_text(h.materia_prima)
        scrap_summary = _qc_parse_scrap_summary(comentarios_bruto)
        comentarios_val = _qc_strip_scrap_summary(comentarios_bruto)
        hojas_data.append({
            'id': h.id,
            'maquina_id': h.maquina_id,
            'serie': h.nombre,
            'qr_payload': qr_payload,
            'clave': h.pn,
            'clave_id': clave_obj.id if clave_obj else None,
            'descripcion_clave': descripcion_clave,
            'calidad': h.calidad,
            'almacen': h.almacen,
            'orden_trabajo': h.orden_trabajo_hr,
            'comentarios': comentarios_val,
            'scrap_qc': scrap_summary,
            'firma_ing_jose': h.supervisor,
            'firma_ing_rodrigo': h.operador,
            'estado': h.estado,
            'liberada_facturacion': bool(flujo_fact),
            'facturacion_aprobado_por': (flujo_fact.facturacion_aprobado_por if flujo_fact else None),
            'facturacion_aprobado_en': (flujo_fact.facturacion_aprobado_en.isoformat() if flujo_fact and flujo_fact.facturacion_aprobado_en else None),
            'facturacion_registro_estado': (
                'APROBADA' if registro_fact and registro_fact.aprobado else 'REGRESADA'
            ) if registro_fact else None,
            'facturacion_registro_usuario': registro_fact.usuario if registro_fact else None,
            'facturacion_registro_fecha': (
                registro_fact.fecha_aprobacion.isoformat()
                if registro_fact and registro_fact.fecha_aprobacion
                else (registro_fact.fecha_creacion.isoformat() if registro_fact and registro_fact.fecha_creacion else None)
            ),
            'facturacion_registro_notas': registro_fact.notas if registro_fact else None,
            'cantidad_piezas': h.cantidad_piezas,
            'impresion_parcial_total': (impresion_parcial_totales.get(h.id, {}).get('total_impreso', 0)),
            'impresion_parcial_movs': (impresion_parcial_totales.get(h.id, {}).get('movimientos', 0)),
            'historial_cargas': historial_cargas_por_hoja.get(h.id, []),
            'hoja_en_produccion': bool(h.hoja_en_produccion),
            'fecha_salida': h.fecha_salida.isoformat() if h.fecha_salida else None,
            'fecha_creacion': h.fecha_creacion.isoformat() if h.fecha_creacion else None,
        })

    companion_hojas = []
    hojas_mp = HojaRutaNueva.query.order_by(HojaRutaNueva.fecha_creacion.desc()).limit(120).all()
    for h in hojas_mp:
        companion_hojas.append({
            'id': h.id,
            'serie': h.nombre,
            'qr_payload': f"HRNID:{h.id};SERIE:{h.nombre or ''}",
            'qr_deeplink': request.url_root.rstrip('/') + f"/hoja_nuevo/{h.id}",
            'clave': h.pn,
            'descripcion_clave': _resolve_clave_descripcion_by_pn(h.pn),
            'calidad': h.calidad,
            'fecha_salida': h.fecha_salida.isoformat() if h.fecha_salida else None,
            'cantidad_piezas': h.cantidad_piezas,
            'almacen': h.almacen,
            'orden_trabajo': h.orden_trabajo_hr,
            'comentarios': _clean_nullable_text(h.materia_prima),
            'estado': h.estado,
        })

    return render_template(
        'hojas_ruta_form.html',
        hojas=hojas_data,
        almacenes=almacenes,
        allow_tipo_switch=True,
        default_tipo_hoja='entrega',
        modulo_titulo='HOJAS DE RUTA ENTREGAS',
        companion_modulo_titulo='HOJAS DE RUTA MP',
        companion_hojas=companion_hojas,
    )


@app.route('/hoja/<int:hoja_id>')
@login_required
@requires_any_permission([('hojas_entregas', 'view'), ('hojas', 'view'), ('catalog', 'view'), ('entregas', 'view'), ('almacen', 'view'), ('facturacion', 'view')])
def hoja_ruta_entregas_ver(hoja_id):
    """Vista independiente para ver una hoja por ID, sin requerir máquina."""
    hoja = HojaRutaEntrega.query.get_or_404(hoja_id)
    h = hoja.to_dict()
    print_qty = request.args.get('print_qty', type=int)
    h['es_impresion_parcial'] = bool(print_qty and print_qty > 0)
    h['cantidad_piezas_impresion'] = int(print_qty) if (print_qty and print_qty > 0) else (h.get('cantidad_piezas') or 0)
    h['auto_print'] = request.args.get('auto_print') == '1'
    comentarios_bruto = _clean_nullable_text(h.get('materia_prima'))
    h['comentarios_usuario'] = _qc_strip_scrap_summary(comentarios_bruto)
    h['scrap_qc'] = _qc_parse_scrap_summary(comentarios_bruto)
    h['descripcion_clave'] = _resolve_clave_descripcion_by_pn(hoja.pn)
    h['qr_payload'] = f"HRID:{hoja.id};SERIE:{hoja.nombre or ''}"
    h['qr_deeplink'] = request.url_root.rstrip('/') + f"/hoja/{hoja.id}"
    estaciones = EstacionTrabajo.query.filter_by(hoja_ruta_id=hoja.id).order_by(EstacionTrabajo.orden).all()
    h['estaciones'] = [e.to_dict() for e in estaciones]
    return render_template('hoja_ruta_ver.html', hoja=h)



@app.route('/api/hojas_ruta/<int:hoja_id>/impresion_parcial', methods=['POST'])
@login_required
@requires_any_permission([('hojas_entregas', 'edit'), ('hojas', 'edit'), ('catalog', 'edit')])
def api_hoja_ruta_impresion_parcial(hoja_id):
    """Registra una impresion parcial solo para fines de impresion (sin alterar lotes/hoja base)."""
    hoja = HojaRutaEntrega.query.get_or_404(hoja_id)
    data = request.get_json() or {}

    cantidad_impresion = data.get('cantidad_impresion')
    try:
        cantidad_impresion = int(cantidad_impresion)
    except Exception:
        return jsonify({'error': 'cantidad_impresion invalida'}), 400

    if cantidad_impresion <= 0:
        return jsonify({'error': 'La cantidad a imprimir debe ser mayor a cero'}), 400

    if not _ensure_hoja_impresiones_parciales_table():
        return jsonify({'error': 'No se pudo preparar tabla de impresiones parciales'}), 500

    usuario = _current_username_for_audit(get_current_user())
    movimiento = HojaRutaImpresionParcial(
        hoja_ruta_id=hoja.id,
        cantidad_impresa=cantidad_impresion,
        usuario=usuario,
    )
    db.session.add(movimiento)

    try:
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        logger.error(f"Error registrando impresion parcial de hoja {hoja.id}: {exc}", exc_info=True)
        return jsonify({'error': 'No se pudo registrar la impresion parcial'}), 500

    total_impreso = db.session.query(
        func.coalesce(func.sum(HojaRutaImpresionParcial.cantidad_impresa), 0)
    ).filter(HojaRutaImpresionParcial.hoja_ruta_id == hoja.id).scalar() or 0

    print_url = url_for(
        'hoja_ruta_entregas_ver',
        hoja_id=hoja.id,
        print_qty=cantidad_impresion,
        auto_print=1,
        _external=False,
    )

    return jsonify({
        'ok': True,
        'hoja_id': hoja.id,
        'serie': hoja.nombre,
        'cantidad_impresion': cantidad_impresion,
        'impresion_parcial_total': int(total_impreso),
        'impresion_parcial_movs': HojaRutaImpresionParcial.query.filter_by(hoja_ruta_id=hoja.id).count(),
        'print_url': print_url,
    }), 200

@app.route('/api/hojas_ruta/resolver_codigo', methods=['POST'])
@login_required
@requires_any_permission([('hojas_entregas', 'view'), ('hojas', 'view'), ('catalog', 'view')])
def api_resolver_codigo_hoja_ruta():
    """Resuelve texto escaneado (QR/codigo) a una hoja de ruta."""
    data = request.get_json() or {}
    raw_value = (data.get('value') or '').strip()
    if not raw_value:
        return jsonify({'error': 'Codigo vacio'}), 400

    value = raw_value.strip()
    upper_value = value.upper()

    hoja_id = None

    # 1) URL tipo /hoja/<id>
    m_url = re.search(r'/hoja/(\d+)', value)
    if m_url:
        hoja_id = int(m_url.group(1))

    # 2) Payload tipo HRID:<id>
    if hoja_id is None:
        m_hrid = re.search(r'HRID\s*[:=]\s*(\d+)', upper_value)
        if m_hrid:
            hoja_id = int(m_hrid.group(1))

    # 3) Si viene solo el numero
    if hoja_id is None and value.isdigit():
        hoja_id = int(value)

    hoja = None
    if hoja_id is not None:
        hoja = HojaRutaEntrega.query.get(hoja_id)

    # 4) Fallback por serie exacta
    if hoja is None:
        hoja = HojaRutaEntrega.query.filter_by(nombre=value).first()

    if hoja is None:
        return jsonify({'error': 'No se encontro hoja para el codigo escaneado'}), 404

    return jsonify({
        'ok': True,
        'hoja': {
            'id': hoja.id,
            'serie': hoja.nombre,
            'clave': hoja.pn,
            'estado': hoja.estado,
            'maquina_id': hoja.maquina_id,
        }
    }), 200


@app.route('/api/hojas_ruta_nuevo/resolver_codigo', methods=['POST'])
@login_required
@requires_any_permission([('hojas_mp', 'view'), ('hojas', 'view'), ('catalog', 'view')])
def api_resolver_codigo_hoja_ruta_nuevo():
    """Resuelve texto escaneado (QR/codigo) a una hoja de ruta del módulo nuevo."""
    data = request.get_json() or {}
    raw_value = (data.get('value') or '').strip()
    if not raw_value:
        return jsonify({'error': 'Codigo vacio'}), 400

    value = raw_value.strip()
    upper_value = value.upper()
    hoja_id = None

    m_url = re.search(r'/hoja_nuevo/(\d+)', value)
    if m_url:
        hoja_id = int(m_url.group(1))

    if hoja_id is None:
        m_hrid = re.search(r'HRNID\s*[:=]\s*(\d+)', upper_value)
        if m_hrid:
            hoja_id = int(m_hrid.group(1))

    if hoja_id is None and value.isdigit():
        hoja_id = int(value)

    hoja = HojaRutaNueva.query.get(hoja_id) if hoja_id is not None else None
    if hoja is None:
        hoja = HojaRutaNueva.query.filter_by(nombre=value).first()

    if hoja is None:
        return jsonify({'error': 'No se encontro hoja para el codigo escaneado'}), 404

    return jsonify({
        'ok': True,
        'hoja': {
            'id': hoja.id,
            'serie': hoja.nombre,
            'clave': hoja.pn,
            'estado': hoja.estado,
            'maquina_id': hoja.maquina_id,
        }
    }), 200


@app.route('/hojas_ruta/<int:maquina_id>')
@login_required
def hojas_ruta_entregas_detalle(maquina_id):
    """Detalle de hojas de ruta para una máquina específica."""
    maquina = Máquina.query.get_or_404(maquina_id)
    hojas = HojaRutaEntrega.query.filter_by(maquina_id=maquina_id).order_by(HojaRutaEntrega.fecha_creacion.desc()).all()
    
    hojas_data = []
    for hoja in hojas:
        # Usa to_dict() completo para incluir todos los campos (calidad, pn, tiempos, etc.)
        h = hoja.to_dict()
        # Asegura orden de estaciones
        estaciones = EstacionTrabajo.query.filter_by(hoja_ruta_id=hoja.id).order_by(EstacionTrabajo.orden).all()
        h['estaciones'] = [e.to_dict() for e in estaciones]
        hojas_data.append(h)

    # Hojas liberadas por Facturación para esta máquina
    hoja_ids = [h['id'] for h in hojas_data]
    facturadas_flujos = HojaRutaFlujoLogistica.query.filter(
        HojaRutaFlujoLogistica.hoja_ruta_id.in_(hoja_ids),
        HojaRutaFlujoLogistica.estado == 'finalizada'
    ).all() if hoja_ids else []
    facturadas_info = {
        f.hoja_ruta_id: {
            'aprobado_por': f.facturacion_aprobado_por or '',
            'aprobado_en': f.facturacion_aprobado_en.strftime('%d/%m/%Y %H:%M') if f.facturacion_aprobado_en else ''
        }
        for f in facturadas_flujos
    }

    return render_template('hojas_ruta_detalle.html', maquina=maquina, hojas=hojas_data, facturadas_info=facturadas_info)


@app.route('/qc_estaciones/<int:maquina_id>')
@login_required
def qc_estaciones_maquina(maquina_id):
    """Vista independiente de control de calidad para producción por máquina."""
    maquina = Máquina.query.get_or_404(maquina_id)
    hoja_activa = HojaRutaEntrega.query.filter_by(maquina_id=maquina_id, estado='activa').first()
    registros = QCProduccionRegistro.query.filter_by(maquina_id=maquina_id).order_by(QCProduccionRegistro.creado_en.desc()).limit(50).all()
    return render_template('qc_estaciones.html', maquina=maquina, hoja_activa=hoja_activa, registros=registros)


# API para crear / actualizar hojas de ruta

@app.route('/api/hojas_ruta', methods=['POST'])
@login_required
@requires_any_permission([('hojas_entregas', 'create'), ('hojas', 'create'), ('catalog', 'edit')])
def api_crear_hoja_ruta():
    """Crear una hoja de ruta con formato simplificado y procesos desde la clave."""
    data = request.get_json() or {}
    clave_id = data.get('clave_id')
    calidad = (data.get('calidad') or '').strip()
    almacen = (data.get('almacen') or '').strip()
    orden_trabajo = (data.get('orden_trabajo') or '').strip()
    comentarios = (data.get('comentarios') or '').strip()
    contpaq_ot_detail_id = data.get('contpaq_ot_detail_id')
    firma_ing_jose = (data.get('firma_ing_jose') or '').strip()
    firma_ing_rodrigo = (data.get('firma_ing_rodrigo') or '').strip()
    cantidad_piezas = data.get('cantidad_piezas')
    user = get_current_user()

    firma_jose_aut = firma_ing_jose.upper() == 'AUTORIZADO'
    firma_rodrigo_aut = firma_ing_rodrigo.upper() == 'AUTORIZADO'
    if (firma_jose_aut or firma_rodrigo_aut) and not (user and user.es_admin):
        return jsonify({
            'error': 'Solo admin puede autorizar firmas al crear una hoja nueva.',
            'firmas_forzadas': True,
        }), 403

    if not clave_id:
        return jsonify({'error': 'clave_id requerido'}), 400
    if not calidad:
        return jsonify({'error': 'calidad requerida'}), 400
    if not almacen:
        return jsonify({'error': 'almacen requerido'}), 400
    if not cantidad_piezas or cantidad_piezas <= 0:
        return jsonify({'error': 'cantidad_piezas debe ser mayor a 0'}), 400

    # Obtener la clave y sus procesos
    clave = ClaveProducto.query.get(clave_id)
    if not clave:
        return jsonify({'error': 'Clave no encontrada'}), 404

    # Politica de duplicados por clave (configurable)
    # none/off/allow: permite crear hojas aunque ya exista la clave (default).
    # active: bloquea si existe hoja activa/pausada con la misma clave.
    # day: bloquea si ya existe hoja de esa clave creada hoy.
    # week: bloquea si ya existe hoja de esa clave creada en la semana actual.
    duplicate_scope = (os.getenv('HOJA_RUTA_DUPLICATE_SCOPE', 'none') or 'none').strip().lower()
    if duplicate_scope in ('active', 'day', 'week'):
        existing_q = HojaRutaEntrega.query.filter(HojaRutaEntrega.pn == clave.clave)
        if duplicate_scope == 'day':
            now_dt = datetime.utcnow()
            day_start = now_dt.replace(hour=0, minute=0, second=0, microsecond=0)
            existing_q = existing_q.filter(HojaRutaEntrega.fecha_creacion >= day_start)
        elif duplicate_scope == 'week':
            now_dt = datetime.utcnow()
            day_start = now_dt.replace(hour=0, minute=0, second=0, microsecond=0)
            week_start = day_start - timedelta(days=day_start.weekday())
            existing_q = existing_q.filter(HojaRutaEntrega.fecha_creacion >= week_start)
        else:
            existing_q = existing_q.filter(HojaRutaEntrega.estado.in_(['activa', 'pausada']))

        hoja_existente = existing_q.order_by(HojaRutaEntrega.fecha_creacion.desc()).first()
        if hoja_existente:
            return jsonify({
                'error': 'Ya existe una hoja de ruta con esta clave.',
                'code': 'duplicate_clave',
                'scope': duplicate_scope,
                'existing_hoja': {
                    'id': hoja_existente.id,
                    'folio': hoja_existente.nombre,
                    'fecha': hoja_existente.fecha_creacion.isoformat() if hoja_existente.fecha_creacion else None,
                    'estado': hoja_existente.estado,
                }
            }), 409

    maquina_id = int(data.get('maquina_id')) if data.get('maquina_id') else None
    if maquina_id:
        maq_for_limit = Máquina.query.get(maquina_id)
        queue_limit = _machine_queue_limit_for_type(getattr(maq_for_limit, 'tipo', None) if maq_for_limit else None)
        hojas_en_cola = HojaRutaEntrega.query.filter(
            HojaRutaEntrega.maquina_id == maquina_id,
            HojaRutaEntrega.estado.in_(['activa', 'pausada'])
        ).order_by(HojaRutaEntrega.fecha_actualizacion.desc(), HojaRutaEntrega.fecha_creacion.desc()).all()
        if len(hojas_en_cola) >= queue_limit:
            hoja_ocupada = hojas_en_cola[0]
            return jsonify({
                'error': f'La máquina ya alcanzó su cola máxima ({queue_limit} hojas).',
                'code': 'machine_busy',
                'existing_hoja': {
                    'id': hoja_ocupada.id,
                    'folio': hoja_ocupada.nombre,
                    'estado': hoja_ocupada.estado,
                }
            }), 409

    veces_previas_maquina = 0
    if maquina_id:
        veces_previas_maquina = HojaRutaEntrega.query.filter_by(maquina_id=maquina_id, pn=clave.clave).count()

    procesos = ClaveProceso.query.filter_by(clave_id=clave_id).order_by(ClaveProceso.orden).all()
    if not procesos:
        return jsonify({'error': 'La clave seleccionada no tiene procesos definidos'}), 400

    supplier_ot_detail = None
    if contpaq_ot_detail_id not in (None, '', 0, '0'):
        try:
            supplier_ot_detail = ContpaqSupplierOTDetalle.query.get(int(contpaq_ot_detail_id))
        except Exception:
            supplier_ot_detail = None
        if not supplier_ot_detail:
            return jsonify({'error': 'La OT seleccionada ya no existe.', 'code': 'ot_not_found'}), 404
        if (supplier_ot_detail.product_key or '').strip().upper() != (clave.clave or '').strip().upper():
            return jsonify({'error': 'La OT seleccionada no corresponde a la clave capturada.', 'code': 'ot_clave_mismatch'}), 409
        options = _contpaq_supplier_ot_options_for_product_key(clave.clave, requested_qty=cantidad_piezas)
        selected_option = next((item for item in options if int(item.get('detail_id') or 0) == int(supplier_ot_detail.id)), None)
        if not selected_option:
            return jsonify({'error': 'La OT seleccionada ya no tiene disponibilidad.', 'code': 'ot_not_available'}), 409
        if not selected_option.get('fits_requested_qty'):
            return jsonify({
                'error': 'La OT seleccionada no tiene cantidad suficiente disponible.',
                'code': 'ot_insufficient_qty',
                'qty_available': selected_option.get('qty_available') or 0,
            }), 409
        orden_trabajo = (selected_option.get('doc_folio') or '').strip() or orden_trabajo

    try:
        fecha_actual = datetime.utcnow()
        maquina = Máquina.query.get(int(data.get('maquina_id'))) if data.get('maquina_id') else None
        descripcion_hoja = _resolve_clave_descripcion_by_pn(clave.clave)
        audit_username = _current_username_for_audit(user)

        hoja = HojaRutaEntrega(
            maquina_id=maquina_id,
            nombre='PENDIENTE_SERIE',
            descripcion=descripcion_hoja,
            estado='activa',
            producto=clave.nombre,
            calidad=calidad,
            pn=clave.clave,
            revision=None,
            fecha_salida=fecha_actual,
            cantidad_piezas=int(cantidad_piezas),
            orden_trabajo_hr=orden_trabajo or None,
            orden_trabajo_pt=None,
            almacen=almacen,
            no_sin_orden=None,
            materia_prima=(comentarios or '').strip() or None,
            total_tiempo=None,
            dias_a_laborar=None,
            fecha_termino=None,
            aprobada=False,
            rechazada=False,
            hoja_en_produccion=str(data.get('hoja_en_produccion') or '').strip().lower() in ('1', 'true', 'yes', 'on'),
            scrap=None,
            retrabajo=None,
            supervisor=firma_ing_jose or None,
            operador=firma_ing_rodrigo or None,
            eficiencia=None,
        )
        db.session.add(hoja)
        db.session.flush()

        # Serie automatica: HR-YYYYMMDD-CLAVE-####
        _rebuild_hoja_serie_con_clave(hoja, clave.clave, prefix='HR')

        if maquina_id:
            _pause_other_machine_hojas(HojaRutaEntrega, maquina_id, hoja.id, fecha_actual)
            _resume_or_activate_hoja(hoja, fecha_actual)

        estaciones_creadas = []
        for idx, cp in enumerate(procesos, start=1):
            estacion = EstacionTrabajo(
                hoja_ruta_id=hoja.id,
                nombre=f"{cp.operacion or cp.proceso.operacion or cp.proceso.nombre}",
                pro_c=str(idx),
                centro_trabajo=cp.centro_trabajo or cp.proceso.centro_trabajo or '',
                operacion=cp.operacion or cp.proceso.operacion or cp.proceso.nombre or '',
                orden=cp.orden,
                t_e=cp.t_e or cp.proceso.tiempo_estimado or '',
                t_tct=cp.t_tct or '',
                t_tco=cp.t_tco or '',
                t_to=cp.t_to or '',
                total_piezas=None,
                operador=None,
                eficiencia=None,
                firma_supervisor=None,
                estado='pendiente'
            )
            db.session.add(estacion)
            estaciones_creadas.append(estacion)

        _apply_hoja_time_plan(
            hoja,
            estaciones_creadas,
            maquina_tipo=maquina.tipo if maquina else None,
            fallback_total_time=hoja.total_tiempo,
        )
        _registrar_carga_piezas_hoja(hoja, 0, int(cantidad_piezas), audit_username, origen='creacion')
        if supplier_ot_detail:
            _create_hoja_entrega_ot_assignment(
                hoja,
                supplier_ot_detail,
                qty_assigned=int(cantidad_piezas),
                created_by=audit_username,
            )

        db.session.commit()
        logger.info(
            f"[HOJAS_RUTA] Nueva hoja creada {hoja.nombre} ({hoja.id}) con {len(procesos)} estaciones para clave {clave.clave}"
        )
        result = hoja.to_dict()
        result['ya_paso_por_maquina'] = veces_previas_maquina > 0
        result['veces_previas_maquina'] = veces_previas_maquina
        if supplier_ot_detail and supplier_ot_detail.ot:
            result['ot_asignada'] = {
                'detail_id': supplier_ot_detail.id,
                'document_id': supplier_ot_detail.document_id,
                'doc_folio': supplier_ot_detail.ot.doc_folio,
                'qty_assigned': int(cantidad_piezas),
            }
        return jsonify(result), 201
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error creando hoja de ruta: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/hojas_ruta/<int:hoja_id>', methods=['PUT'])
@login_required
@requires_any_permission([('hojas_entregas', 'edit'), ('hojas', 'edit'), ('catalog', 'edit')])
def api_actualizar_hoja_ruta(hoja_id):
    """Actualizar campos editables de una hoja de ruta."""
    hoja = HojaRutaEntrega.query.get_or_404(hoja_id)
    data = request.get_json() or {}
    estado_change_dt = None

    user = get_current_user()
    hoja_field_permissions = {
        field: [HOJA_FIELD_SPECIFIC_ACTIONS[field], HOJA_FIELD_GROUP_ACTIONS[field]]
        for field in HOJA_FIELD_GROUP_ACTIONS
    }
    denied_fields = _check_field_level_permissions(
        user=user,
        module='hojas_entregas',
        payload=data,
        field_actions=hoja_field_permissions,
        broad_action='edit',
    )
    if denied_fields and not user.has_permission('catalog', 'edit'):
        return jsonify({
            'error': 'No tienes permiso para editar algunos campos',
            'denied_fields': denied_fields,
        }), 403

    firma_jose_in = (data.get('firma_ing_jose') or '').strip() if 'firma_ing_jose' in data else None
    firma_rodrigo_in = (data.get('firma_ing_rodrigo') or '').strip() if 'firma_ing_rodrigo' in data else None
    wants_authorize = (
        (firma_jose_in is not None and firma_jose_in.upper() == 'AUTORIZADO') or
        (firma_rodrigo_in is not None and firma_rodrigo_in.upper() == 'AUTORIZADO')
    )

    if wants_authorize and not (user and user.es_admin):
        ready, reason = _hoja_ready_for_signatures(hoja)
        if not ready:
            return jsonify({
                'error': reason or 'No se puede autorizar la hoja en este momento.',
                'requires_admin_override': True,
            }), 409
    
    if 'estado' in data:
        estado_in = (data.get('estado') or '').strip().lower()
        allowed_estados = {'activa', 'pausada', 'completada', 'cancelada'}
        if estado_in not in allowed_estados:
            return jsonify({'error': 'estado invalido'}), 400

        estado_change_dt = datetime.utcnow()
        hoja.estado = estado_in
        if estado_in in {'completada', 'cancelada'}:
            hoja.fecha_termino = hoja.fecha_termino or estado_change_dt
        else:
            hoja.fecha_termino = None
    if 'clave_id' in data and data.get('clave_id') is not None:
        try:
            clave_id = int(data.get('clave_id'))
        except Exception:
            return jsonify({'error': 'clave_id inválido'}), 400

        if clave_id <= 0:
            return jsonify({'error': 'clave_id inválido'}), 400

        clave_obj = ClaveProducto.query.get(clave_id)
        if not clave_obj:
            return jsonify({'error': 'Clave no encontrada'}), 404

        hoja.pn = clave_obj.clave
        hoja.producto = _clean_nullable_text(clave_obj.nombre) or clave_obj.clave
        hoja.descripcion = _resolve_clave_descripcion_by_pn(clave_obj.clave)
        # Mantener folio alineado con la clave (corrige desfaces al editar).
        if 'nombre' not in data:
            _rebuild_hoja_serie_con_clave(hoja, clave_obj.clave, prefix='HR')
    if 'nombre' in data:
        hoja.nombre = data['nombre']
    if 'descripcion' in data:
        hoja.descripcion = data['descripcion']
    if 'comentarios' in data:
        comentarios_usuario = (data.get('comentarios') or '').strip()
        scrap_block = _qc_extract_scrap_block(hoja.materia_prima)
        if scrap_block:
            hoja.materia_prima = (
                (comentarios_usuario + "\n\n" + scrap_block).strip()
                if comentarios_usuario else scrap_block
            )
        else:
            hoja.materia_prima = comentarios_usuario or None
    if 'firma_ing_jose' in data:
        hoja.supervisor = (data.get('firma_ing_jose') or '').strip() or None
    if 'firma_ing_rodrigo' in data:
        hoja.operador = (data.get('firma_ing_rodrigo') or '').strip() or None
    if 'calidad' in data:
        hoja.calidad = (data.get('calidad') or '').strip() or None
    if 'almacen' in data:
        hoja.almacen = (data.get('almacen') or '').strip() or None
    if 'orden_trabajo' in data:
        hoja.orden_trabajo_hr = (data.get('orden_trabajo') or '').strip() or None
    if 'cantidad_piezas' in data:
        try:
            cantidad = int(data.get('cantidad_piezas') or 0)
        except Exception:
            return jsonify({'error': 'cantidad_piezas inválida'}), 400
        if cantidad <= 0:
            return jsonify({'error': 'cantidad_piezas debe ser mayor a 0'}), 400
        cantidad_anterior = int(hoja.cantidad_piezas or 0)
        hoja.cantidad_piezas = cantidad
        _registrar_carga_piezas_hoja(hoja, cantidad_anterior, cantidad, _current_username_for_audit(user), origen='edicion')
    if 'hoja_en_produccion' in data:
        hoja.hoja_en_produccion = str(data.get('hoja_en_produccion') or '').strip().lower() in ('1', 'true', 'yes', 'on')

    # Regla de negocio: descripcion de hoja siempre proviene de la clave actual.
    if hoja.pn:
        hoja.descripcion = _resolve_clave_descripcion_by_pn(hoja.pn)

    estaciones = EstacionTrabajo.query.filter_by(hoja_ruta_id=hoja.id).order_by(EstacionTrabajo.orden).all()
    maquina = Máquina.query.get(hoja.maquina_id) if hoja.maquina_id else None
    _apply_hoja_time_plan(
        hoja,
        estaciones,
        maquina_tipo=maquina.tipo if maquina else None,
        fallback_total_time=hoja.total_tiempo,
    )

    if (hoja.estado or '').lower() == 'cancelada' and hoja.maquina_id:
        maq = Máquina.query.get(hoja.maquina_id)
        if maq and bool(getattr(maq, 'activo', False)):
            restantes = HojaRutaEntrega.query.filter(
                HojaRutaEntrega.maquina_id == maq.id,
                HojaRutaEntrega.id != hoja.id,
                HojaRutaEntrega.estado.in_(['activa', 'pausada'])
            ).all()
            restantes = _order_machine_queue_items(restantes)
            siguiente = _pick_machine_active_hoja(restantes)
            if siguiente:
                _pause_other_machine_hojas(HojaRutaEntrega, maq.id, siguiente.id, estado_change_dt or datetime.utcnow())
                _resume_or_activate_hoja(siguiente, estado_change_dt or datetime.utcnow())
            else:
                maq.activo = False
    
    db.session.commit()
    logger.info(f"[HOJAS_RUTA] Hoja actualizada: {hoja_id}")
    return jsonify(hoja.to_dict()), 200


@app.route('/api/claves_procesos', methods=['GET'])
@login_required
@requires_any_permission([('hojas_entregas', 'view'), ('hojas_mp', 'view'), ('hojas', 'view'), ('catalog', 'view')])
def api_claves_procesos():
    """Obtener todas las claves con sus procesos y tiempo total T/O."""
    try:
        claves = ClaveProducto.query.filter_by(activo=True).order_by(ClaveProducto.clave.asc()).all()
        result = []
        for clave in claves:
            try:
                # Obtener todos los procesos de la clave ordenados
                procesos = ClaveProceso.query.filter_by(clave_id=clave.id).order_by(ClaveProceso.orden).all()

                # El T/O es el del último proceso (suma acumulada)
                tiempo_to = "00:00:00"
                if procesos:
                    # Buscar el último proceso que tenga T/O
                    for cp in reversed(procesos):
                        if cp.t_to:
                            tiempo_to = cp.t_to
                            break

                procesos_payload = []
                for p in procesos:
                    try:
                        base = p.to_dict()
                        base['proceso_descripcion'] = (p.proceso.descripcion if getattr(p, 'proceso', None) else '') or ''
                        procesos_payload.append(base)
                    except Exception as p_err:
                        logger.warning(f"Proceso invalido en clave_id={clave.id}, proceso_id={getattr(p, 'id', '?')}: {p_err}")

                result.append({
                    'id': clave.id,
                    'clave': clave.clave,
                    'nombre': _clean_nullable_text(clave.nombre) or clave.clave,
                    'notas': _clean_nullable_text(getattr(clave, 'notas', None)),
                    'tiempo_to': tiempo_to,
                    'procesos': procesos_payload,
                })
            except Exception as clave_err:
                logger.warning(f"Clave invalida omitida clave_id={getattr(clave, 'id', '?')}: {clave_err}")
                continue
        return jsonify(result), 200
    except Exception as e:
        logger.error(f"Error obteniendo claves/procesos: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/estaciones', methods=['POST'])
@login_required
@requires_any_permission([('hojas_entregas', 'edit'), ('hojas', 'edit'), ('catalog', 'edit')])
def api_crear_estacion():
    """Crear una nueva estación de trabajo en una hoja de ruta."""
    data = request.get_json()
    hoja_ruta_id = data.get('hoja_ruta_id')
    nombre = data.get('nombre')
    
    if not hoja_ruta_id or not nombre:
        return jsonify({'error': 'hoja_ruta_id y nombre requeridos'}), 400
    
    # Obtener orden máxima actual
    max_orden = db.session.query(db.func.max(EstacionTrabajo.orden)).filter_by(
        hoja_ruta_id=hoja_ruta_id
    ).scalar() or 0
    
    estacion = EstacionTrabajo(
        hoja_ruta_id=hoja_ruta_id,
        nombre=nombre,
        descripcion=data.get('descripcion'),
        orden=max_orden + 1,
        estado='pendiente'
    )
    db.session.add(estacion)
    db.session.commit()
    
    logger.info(f"[HOJAS_RUTA] Nueva estación creada: {estacion.id}")
    return jsonify(estacion.to_dict()), 201


# ==== Producción / flujo operativo ==== 
@app.route('/api/produccion/aprobar_ot', methods=['POST'])
@login_required
@requires_any_permission([('estaciones', 'operate'), ('catalog', 'edit')])
def api_aprobar_ot():
    data = request.get_json() or {}
    maquina_id = data.get('maquina_id')
    ot = data.get('orden_trabajo')
    # Sólo registramos en logs por ahora
    logger.info(f"[PRODUCCION] OT aprobada para maquina={maquina_id} OT={ot}")
    return jsonify({'ok': True, 'message': 'OT aprobada'}), 200


@app.route('/api/maquinas/<int:maquina_id>/activar', methods=['POST'])
@login_required
@requires_any_permission([('estaciones', 'operate'), ('catalog', 'edit')])
def api_activar_maquina(maquina_id):
    maq = Máquina.query.get_or_404(maquina_id)
    try:
        now_dt = datetime.utcnow()
        maq.activo = True
        cola = HojaRutaEntrega.query.filter(
            HojaRutaEntrega.maquina_id == maq.id,
            HojaRutaEntrega.estado.in_(['activa', 'pausada'])
        ).all()
        cola = _order_machine_queue_items(cola)
        hoja_actual = _pick_machine_active_hoja(cola)
        if hoja_actual:
            _pause_other_machine_hojas(HojaRutaEntrega, maq.id, hoja_actual.id, now_dt)
            _resume_or_activate_hoja(hoja_actual, now_dt)
        db.session.commit()
        logger.info(f"[MAQUINA] Activada maquina {maquina_id}")
        return jsonify({'ok': True, 'maquina_id': maquina_id, 'activo': True}), 200
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error activando maquina: {e}", exc_info=True)
        return jsonify({'error': 'No se pudo activar la máquina. Ejecuta ALTER TABLE para agregar columna activo si no existe.'}), 500


@app.route('/api/maquinas/<int:maquina_id>/desactivar', methods=['POST'])
@login_required
@requires_any_permission([('estaciones', 'operate'), ('catalog', 'edit')])
def api_desactivar_maquina(maquina_id):
    maq = Máquina.query.get_or_404(maquina_id)
    try:
        now_dt = datetime.utcnow()
        maq.activo = False
        hojas_activas = HojaRutaEntrega.query.filter_by(maquina_id=maq.id, estado='activa').all()
        for hoja_activa in hojas_activas:
            hoja_activa.estado = 'pausada'
            hoja_activa.fecha_actualizacion = now_dt
        db.session.commit()
        logger.info(f"[MAQUINA] Desactivada maquina {maquina_id}")
        return jsonify({'ok': True, 'maquina_id': maquina_id, 'activo': False}), 200
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error desactivando maquina: {e}", exc_info=True)
        return jsonify({'error': 'No se pudo desactivar la máquina.'}), 500


@app.route('/api/maquinas/<int:maquina_id>/paro_mantenimiento', methods=['POST'])
@login_required
@requires_any_permission([('estaciones', 'operate'), ('catalog', 'edit')])
def api_paro_mantenimiento(maquina_id):
    """Poner máquina en paro por mantenimiento (desactivada)."""
    maq = Máquina.query.get_or_404(maquina_id)
    try:
        now_dt = datetime.utcnow()
        maq.activo = False
        hojas_activas = HojaRutaEntrega.query.filter_by(maquina_id=maq.id, estado='activa').all()
        for hoja_activa in hojas_activas:
            hoja_activa.estado = 'pausada'
            hoja_activa.fecha_actualizacion = now_dt
        db.session.commit()
        logger.info(f"[MAQUINA] Paro mantenimiento maquina {maquina_id}")
        return jsonify({'ok': True, 'maquina_id': maquina_id, 'activo': False, 'motivo': 'mantenimiento'}), 200
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error en paro mantenimiento: {e}", exc_info=True)
        return jsonify({'error': 'No se pudo registrar el paro por mantenimiento'}), 500


@app.route('/api/maquinas/<int:maquina_id>/asignar_hoja', methods=['POST'])
@login_required
@requires_any_permission([('estaciones', 'operate'), ('catalog', 'edit')])
def api_asignar_hoja_maquina(maquina_id):
    """Asignar hoja de ruta pendiente (sin máquina) a una máquina."""
    maq = Máquina.query.get_or_404(maquina_id)
    data = request.get_json() or {}
    hoja_id = data.get('hoja_id')
    start_estacion_id = data.get('start_estacion_id')
    elapsed_process_seconds = _parse_elapsed_process_seconds(data.get('elapsed_process_seconds'))
    if not hoja_id:
        return jsonify({'error': 'hoja_id requerido'}), 400
    if elapsed_process_seconds is None:
        return jsonify({'error': 'elapsed_process_seconds invalido'}), 400

    hoja = HojaRutaEntrega.query.get_or_404(int(hoja_id))
    if hoja.maquina_id and hoja.maquina_id != maq.id:
        return jsonify({'error': 'La hoja ya está asignada a otra máquina'}), 409

    queue_limit = _machine_queue_limit_for_type(getattr(maq, 'tipo', None))
    queue_count = _machine_queue_count(HojaRutaEntrega, maq.id, exclude_hoja_id=hoja.id)
    if queue_count >= queue_limit:
        return jsonify({'error': f'La máquina ya alcanzó su cola máxima ({queue_limit} hojas).'}), 409

    try:
        now_ref = datetime.utcnow()
        hoja.maquina_id = maq.id
        _pause_other_machine_hojas(HojaRutaEntrega, maq.id, hoja.id, now_ref)
        _resume_or_activate_hoja(hoja, now_ref)
        maq.activo = True  # Activar la máquina al recibir una hoja

        estaciones = EstacionTrabajo.query.filter_by(hoja_ruta_id=hoja.id).order_by(EstacionTrabajo.orden).all()

        target_estacion = None
        # Modo temporal: permitir iniciar desde un proceso avanzado al asignar.
        if start_estacion_id:
            try:
                start_estacion_id = int(start_estacion_id)
            except Exception:
                return jsonify({'error': 'start_estacion_id invalido'}), 400

            objetivo = next((e for e in estaciones if e.id == start_estacion_id), None)
            if not objetivo:
                return jsonify({'error': 'El proceso inicial seleccionado no pertenece a la hoja'}), 409
            target_estacion = objetivo

            for e in estaciones:
                if (e.orden or 0) < (objetivo.orden or 0):
                    # Adelanto temporal: asumir completados los anteriores.
                    e.estado = 'completada'
                    if not e.fecha_inicio:
                        e.fecha_inicio = now_ref
                    if not e.fecha_finalizacion:
                        e.fecha_finalizacion = now_ref

                    # Enviar automaticamente a calidad como pendiente de revision.
                    notas_src = e.notas or ''
                    has_qc_status = re.search(r'STATUS=(QC_OK|QC_NOK)', notas_src or '') is not None
                    if not has_qc_status:
                        qc_pending_block = (
                            "[AUTO_ADVANCE_START]\n"
                            "STATUS=QC_PENDING\n"
                            "ORIGEN=Adelanto_Proceso_Asignacion\n"
                            f"FECHA={now_ref.strftime('%Y-%m-%d %H:%M:%S')}\n"
                            "NOTA=Proceso marcado como completado por adelanto temporal; requiere revision de Calidad.\n"
                            "[AUTO_ADVANCE_END]"
                        )
                        e.notas = (notas_src.strip() + "\n" + qc_pending_block).strip() if notas_src else qc_pending_block
                elif e.id == objetivo.id:
                    e.estado = 'en_curso'
                    if not e.fecha_inicio:
                        e.fecha_inicio = now_ref
                    e.fecha_finalizacion = None
                else:
                    if (e.estado or '').lower() != 'completada':
                        e.estado = 'pendiente'
                        e.fecha_finalizacion = None

            # Mantener hoja en espera de liberacion QC (no auto-liberar por adelanto).
            hoja.aprobada = False
            hoja.rechazada = False
        else:
            # Si no hay proceso en curso, arrancar el siguiente pendiente al asignar.
            en_curso = next((e for e in estaciones if (e.estado or '').lower() == 'en_curso'), None)
            if not en_curso:
                siguiente = next((e for e in estaciones if (e.estado or 'pendiente').lower() != 'completada'), None)
                if siguiente:
                    siguiente.estado = 'en_curso'
                    if not siguiente.fecha_inicio:
                        siguiente.fecha_inicio = now_ref
                    target_estacion = siguiente
            else:
                target_estacion = en_curso

        if target_estacion and elapsed_process_seconds > 0:
            _apply_entrega_process_elapsed(hoja, estaciones, target_estacion, elapsed_process_seconds, now_ref)

        _apply_hoja_time_plan(
            hoja,
            estaciones,
            maquina_tipo=maq.tipo,
            fallback_total_time=hoja.total_tiempo,
        )

        db.session.commit()
        logger.info(f"[HOJAS_RUTA] Hoja {hoja.id} asignada a maquina {maquina_id}")
        return jsonify({'success': True, 'hoja': hoja.to_dict()}), 200
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error asignando hoja a maquina: {e}", exc_info=True)
        return jsonify({'error': 'No se pudo asignar la hoja a la máquina'}), 500


@app.route('/api/maquinas/<int:maquina_id>/ajustar_tiempo_proceso', methods=['POST'])
@login_required
@requires_any_permission([('estaciones', 'operate'), ('catalog', 'edit')])
def api_ajustar_tiempo_proceso_maquina(maquina_id):
    maq = Máquina.query.get_or_404(maquina_id)
    data = request.get_json() or {}
    elapsed_process_seconds = _parse_elapsed_process_seconds(data.get('elapsed_process_seconds'))
    if elapsed_process_seconds is None:
        return jsonify({'error': 'elapsed_process_seconds invalido'}), 400

    cola = HojaRutaEntrega.query.filter(
        HojaRutaEntrega.maquina_id == maq.id,
        HojaRutaEntrega.estado.in_(['activa', 'pausada'])
    ).all()
    cola = _order_machine_queue_items(cola)
    hoja = _pick_machine_active_hoja(cola)
    if not hoja:
        return jsonify({'error': 'La máquina no tiene hoja activa o pausada'}), 404

    estaciones = EstacionTrabajo.query.filter_by(hoja_ruta_id=hoja.id).order_by(EstacionTrabajo.orden).all()
    target = next((e for e in estaciones if (e.estado or '').lower() == 'en_curso'), None)
    if not target:
        target = next((e for e in estaciones if (e.estado or 'pendiente').lower() != 'completada'), None)
    if not target:
        return jsonify({'error': 'No hay proceso actual ajustable en la hoja'}), 409

    try:
        now_ref = datetime.utcnow()
        _apply_entrega_process_elapsed(hoja, estaciones, target, elapsed_process_seconds, now_ref)
        _apply_hoja_time_plan(hoja, estaciones, maquina_tipo=maq.tipo, fallback_total_time=hoja.total_tiempo)
        db.session.commit()
        return jsonify({'ok': True, 'hoja_id': hoja.id, 'estacion_id': target.id, 'elapsed_process_seconds': elapsed_process_seconds}), 200
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error ajustando tiempo de proceso: {e}", exc_info=True)
        return jsonify({'error': 'No se pudo ajustar el tiempo del proceso'}), 500


@app.route('/api/maquinas/<int:maquina_id>/retirar_hoja', methods=['POST'])
@login_required
@requires_any_permission([('estaciones', 'operate'), ('catalog', 'edit')])
def api_retirar_hoja_maquina(maquina_id):
    """Retirar/desasignar hoja activa de la máquina y devolverla a pendientes."""
    maq = Máquina.query.get_or_404(maquina_id)
    data = request.get_json() or {}
    hoja_id = data.get('hoja_id')

    if hoja_id:
        hoja = HojaRutaEntrega.query.get_or_404(int(hoja_id))
        if hoja.maquina_id != maq.id:
            return jsonify({'error': 'La hoja no pertenece a esta máquina'}), 409
    else:
        hoja = HojaRutaEntrega.query.filter(
            HojaRutaEntrega.maquina_id == maq.id,
            HojaRutaEntrega.estado.in_(['activa', 'pausada'])
        ).order_by(HojaRutaEntrega.fecha_creacion.desc()).first()
        if not hoja:
            return jsonify({'error': 'No hay hoja activa o pausada asignada a esta máquina'}), 404

    try:
        now_ref = datetime.utcnow()
        hoja.estado = 'activa'
        hoja.maquina_id = None
        # Al volver a pendientes, reiniciar ventanas de tiempo para futura reasignacion.
        hoja.fecha_salida = None
        hoja.fecha_termino = None

        if bool(getattr(maq, 'activo', False)):
            restantes = HojaRutaEntrega.query.filter(
                HojaRutaEntrega.maquina_id == maq.id,
                HojaRutaEntrega.estado.in_(['activa', 'pausada'])
            ).all()
            restantes = _order_machine_queue_items(restantes)
            siguiente = _pick_machine_active_hoja(restantes)
            if siguiente:
                _pause_other_machine_hojas(HojaRutaEntrega, maq.id, siguiente.id, now_ref)
                _resume_or_activate_hoja(siguiente, now_ref)

        db.session.commit()
        logger.info(f"[HOJAS_RUTA] Hoja {hoja.id} retirada de maquina {maquina_id}")
        return jsonify({'success': True, 'hoja': hoja.to_dict()}), 200
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error retirando hoja de maquina: {e}", exc_info=True)
        return jsonify({'error': 'No se pudo retirar la hoja de la máquina'}), 500


@app.route('/api/maquinas/<int:maquina_id>/activar_hoja', methods=['POST'])
@login_required
@requires_any_permission([('estaciones', 'operate'), ('catalog', 'edit')])
def api_activar_hoja_maquina(maquina_id):
    maq = Máquina.query.get_or_404(maquina_id)
    data = request.get_json() or {}
    hoja_id = int(data.get('hoja_id') or 0)
    if not hoja_id:
        return jsonify({'error': 'hoja_id requerido'}), 400

    hoja = HojaRutaEntrega.query.get_or_404(hoja_id)
    if hoja.maquina_id != maq.id:
        return jsonify({'error': 'La hoja no está asignada a esta máquina'}), 409
    if (hoja.estado or '').strip().lower() == 'cancelada':
        return jsonify({'error': 'La hoja está cancelada. Cambia su estado antes de volver a activarla.'}), 409

    try:
        now_dt = datetime.utcnow()
        _pause_other_machine_hojas(HojaRutaEntrega, maq.id, hoja.id, now_dt)
        _resume_or_activate_hoja(hoja, now_dt)
        maq.activo = True
        db.session.commit()
        return jsonify({'ok': True, 'maquina_id': maq.id, 'hoja_id': hoja.id}), 200
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error activando hoja en cola de maquina: {e}", exc_info=True)
        return jsonify({'error': 'No se pudo priorizar la hoja en la máquina'}), 500


@app.route('/api/estaciones/<int:estacion_id>/check_proceso', methods=['POST'])
@login_required
@requires_any_permission([('estaciones', 'operate'), ('catalog', 'edit')])
def api_check_proceso_estacion(estacion_id):
    """Marcar/desmarcar proceso de estación y avanzar automáticamente al siguiente pendiente."""
    estacion = EstacionTrabajo.query.get_or_404(estacion_id)
    hoja = HojaRutaEntrega.query.get_or_404(estacion.hoja_ruta_id)
    if (hoja.estado or '').strip().lower() == 'cancelada':
        return jsonify({'error': 'La hoja está cancelada. Reactivala desde Hojas Entregas antes de mover procesos.'}), 409
    data = request.get_json() or {}
    completada = bool(data.get('completada', False))

    alertas_pendientes_envio = []

    try:
        ahora = datetime.utcnow()

        if completada:
            if not estacion.fecha_inicio:
                estacion.fecha_inicio = ahora
            estacion.estado = 'completada'
            estacion.fecha_finalizacion = ahora

            evento_clave_est = f"estaciones_t:estacion:{estacion.id}:hoja:{hoja.id}:fin:{ahora.strftime('%Y%m%d%H%M%S')}"
            alerta_est = _crear_alerta_buzon(
                evento_clave=evento_clave_est,
                origen='estaciones_t',
                tipo='estacion_completada',
                titulo=f"Estacion completada: {estacion.operacion or estacion.nombre or 'Proceso'}",
                mensaje=f"Hoja {hoja.nombre or hoja.id} | Maquina ID: {hoja.maquina_id or 'N/A'}",
                maquina_id=hoja.maquina_id,
                hoja_id=hoja.id,
                estacion_id=estacion.id,
                commit=False,
            )
            if alerta_est is not None and alerta_est in db.session.new:
                alertas_pendientes_envio.append(alerta_est)

            # Limpiar cualquier en_curso previo para mantener un solo proceso activo.
            otras_en_curso = EstacionTrabajo.query.filter(
                EstacionTrabajo.hoja_ruta_id == hoja.id,
                EstacionTrabajo.id != estacion.id,
                EstacionTrabajo.estado == 'en_curso'
            ).all()
            for e in otras_en_curso:
                e.estado = 'pendiente'

            siguiente = EstacionTrabajo.query.filter(
                EstacionTrabajo.hoja_ruta_id == hoja.id,
                EstacionTrabajo.estado == 'pendiente'
            ).order_by(EstacionTrabajo.orden.asc()).first()
            if siguiente:
                siguiente.estado = 'en_curso'
                if not siguiente.fecha_inicio:
                    siguiente.fecha_inicio = ahora
            else:
                # No auto-cerrar hoja: puede requerir pasar por otras maquinas/procesos.
                hoja.estado = 'activa'
                hoja.fecha_termino = None

        else:
            estacion.estado = 'pendiente'
            estacion.fecha_finalizacion = None

            # Si se reabre un proceso, la hoja vuelve a activa.
            hoja.estado = 'activa'
            hoja.fecha_termino = None

            en_curso = EstacionTrabajo.query.filter_by(hoja_ruta_id=hoja.id, estado='en_curso').first()
            if not en_curso:
                proximo = EstacionTrabajo.query.filter(
                    EstacionTrabajo.hoja_ruta_id == hoja.id,
                    EstacionTrabajo.estado == 'pendiente'
                ).order_by(EstacionTrabajo.orden.asc()).first()
                if proximo:
                    proximo.estado = 'en_curso'
                    if not proximo.fecha_inicio:
                        proximo.fecha_inicio = ahora

        if not hoja.fecha_salida:
            hoja.fecha_salida = ahora

        # Sincronizar estado final segun checks reales.
        _sync_hoja_estado_with_checks(hoja, now_dt=ahora)

        if (hoja.estado or '').lower() == 'completada':
            evento_clave_hoja = f"estaciones_t:hoja:{hoja.id}:completada:{(hoja.fecha_termino or ahora).strftime('%Y%m%d%H%M%S')}"
            alerta_hoja = _crear_alerta_buzon(
                evento_clave=evento_clave_hoja,
                origen='estaciones_t',
                tipo='hoja_completada',
                titulo=f"Hoja completada: {hoja.nombre or hoja.id}",
                mensaje=f"Todos los procesos de la hoja fueron completados.",
                maquina_id=hoja.maquina_id,
                hoja_id=hoja.id,
                estacion_id=None,
                commit=False,
            )
            if alerta_hoja is not None and alerta_hoja in db.session.new:
                alertas_pendientes_envio.append(alerta_hoja)

        db.session.commit()
        for alerta in alertas_pendientes_envio:
            _send_alerta_whatsapp_if_enabled(alerta)
            _send_alerta_telegram_if_enabled(alerta)

        pendientes = EstacionTrabajo.query.filter_by(hoja_ruta_id=hoja.id, estado='pendiente').count()
        completadas = EstacionTrabajo.query.filter_by(hoja_ruta_id=hoja.id, estado='completada').count()
        total = EstacionTrabajo.query.filter_by(hoja_ruta_id=hoja.id).count()

        return jsonify({
            'success': True,
            'estacion': estacion.to_dict(),
            'hoja': {'id': hoja.id, 'estado': hoja.estado, 'fecha_termino': hoja.fecha_termino.isoformat() if hoja.fecha_termino else None},
            'resumen': {'total': total, 'completadas': completadas, 'pendientes': pendientes}
        }), 200
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error marcando proceso estacion={estacion_id}: {e}", exc_info=True)
        return jsonify({'error': 'No se pudo actualizar el proceso'}), 500


@app.route('/api/hojas_ruta/<int:hoja_id>', methods=['DELETE'])
@login_required
@requires_any_permission([('hojas_entregas', 'delete'), ('hojas', 'delete'), ('catalog', 'edit')])
def api_eliminar_hoja_ruta(hoja_id):
    """Eliminar una hoja de ruta. Solo permite borrar hojas no asignadas a maquina."""
    hoja = HojaRutaEntrega.query.get_or_404(hoja_id)

    if hoja.maquina_id is not None:
        return jsonify({'error': 'No puedes eliminar una hoja asignada a una maquina. Primero desasignala.'}), 409

    try:
        db.session.delete(hoja)
        db.session.commit()
        logger.info(f"[HOJAS_RUTA] Hoja eliminada: {hoja_id}")
        return jsonify({'success': True, 'id': hoja_id}), 200
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error eliminando hoja {hoja_id}: {e}", exc_info=True)
        return jsonify({'error': 'No se pudo eliminar la hoja de ruta'}), 500


@app.route('/api/hojas_ruta_nuevo', methods=['POST'])
@login_required
@requires_any_permission([('hojas_mp', 'create'), ('hojas', 'create'), ('catalog', 'edit')])
def api_crear_hoja_ruta_nuevo():
    """Crear hoja de ruta del módulo nuevo (tabla hojas_ruta_nueva)."""
    data = request.get_json() or {}
    clave_id = data.get('clave_id')
    calidad = (data.get('calidad') or '').strip()
    almacen = (data.get('almacen') or '').strip()
    orden_trabajo = (data.get('orden_trabajo') or '').strip()
    comentarios = (data.get('comentarios') or '').strip()
    cantidad_piezas = data.get('cantidad_piezas')

    if not clave_id:
        return jsonify({'error': 'clave_id requerido'}), 400
    if not calidad:
        return jsonify({'error': 'calidad requerida'}), 400
    if not almacen:
        return jsonify({'error': 'almacen requerido'}), 400
    if not cantidad_piezas or int(cantidad_piezas) <= 0:
        return jsonify({'error': 'cantidad_piezas debe ser mayor a 0'}), 400

    clave = ClaveProducto.query.get(clave_id)
    if not clave:
        return jsonify({'error': 'Clave no encontrada'}), 404

    procesos = ClaveProceso.query.filter_by(clave_id=clave.id).order_by(ClaveProceso.orden.asc()).all()
    if not procesos:
        return jsonify({'error': 'La clave seleccionada no tiene procesos definidos'}), 400

    try:
        fecha_actual = datetime.utcnow()
        hoja = HojaRutaNueva(
            maquina_id=int(data.get('maquina_id')) if data.get('maquina_id') else None,
            nombre='PENDIENTE_SERIE',
            descripcion=_resolve_clave_descripcion_by_pn(clave.clave),
            estado='activa',
            producto=clave.nombre,
            calidad=calidad,
            pn=clave.clave,
            # El reloj de producción inicia al asignar a máquina, no al crear pendiente.
            fecha_salida=None,
            cantidad_piezas=int(cantidad_piezas),
            orden_trabajo_hr=orden_trabajo or None,
            almacen=almacen,
            materia_prima=comentarios or None,
        )
        db.session.add(hoja)
        db.session.flush()

        clave_segura = _clave_token_for_serie(clave.clave)
        hoja.nombre = f"HRN-{fecha_actual.strftime('%Y%m%d')}-{clave_segura}-{hoja.id:04d}"

        _recompute_mp_time_plan(hoja)

        db.session.commit()
        result = hoja.to_dict()
        result['estaciones'] = _build_mp_virtual_estaciones_by_pn(hoja.pn)
        return jsonify(result), 201
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error creando hoja nueva: {e}", exc_info=True)
        return jsonify({'error': 'No se pudo crear la hoja de ruta nueva'}), 500


@app.route('/api/hojas_ruta_nuevo/<int:hoja_id>', methods=['PUT'])
@login_required
@requires_any_permission([('hojas_mp', 'edit'), ('hojas', 'edit'), ('catalog', 'edit')])
def api_actualizar_hoja_ruta_nuevo(hoja_id):
    """Actualizar hoja de ruta del módulo nuevo."""
    hoja = HojaRutaNueva.query.get_or_404(hoja_id)
    data = request.get_json() or {}

    if 'estado' in data:
        hoja.estado = data['estado']
    if 'clave_id' in data and data.get('clave_id') is not None:
        clave_obj = ClaveProducto.query.get(int(data.get('clave_id')))
        if not clave_obj:
            return jsonify({'error': 'Clave no encontrada'}), 404
        hoja.pn = clave_obj.clave
        hoja.producto = _clean_nullable_text(clave_obj.nombre) or clave_obj.clave
        hoja.descripcion = _resolve_clave_descripcion_by_pn(clave_obj.clave)
        if 'nombre' not in data:
            _rebuild_hoja_serie_con_clave(hoja, clave_obj.clave, prefix='HRN')
    if 'calidad' in data:
        hoja.calidad = (data.get('calidad') or '').strip() or None
    if 'cantidad_piezas' in data and data.get('cantidad_piezas') is not None:
        try:
            hoja.cantidad_piezas = max(1, int(data.get('cantidad_piezas')))
        except Exception:
            return jsonify({'error': 'cantidad_piezas invalida'}), 400
    if 'almacen' in data:
        hoja.almacen = (data.get('almacen') or '').strip() or None
    if 'orden_trabajo' in data:
        hoja.orden_trabajo_hr = (data.get('orden_trabajo') or '').strip() or None
    if 'comentarios' in data:
        comentarios_usuario = (data.get('comentarios') or '').strip()
        process_state = _mp_extract_process_state_block(hoja.materia_prima)
        if process_state:
            hoja.materia_prima = _mp_upsert_process_state_block(comentarios_usuario, _mp_parse_completed_process_ids(hoja.materia_prima))
        else:
            hoja.materia_prima = comentarios_usuario or None

    _recompute_mp_time_plan(hoja)

    try:
        db.session.commit()
        return jsonify(hoja.to_dict()), 200
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error actualizando hoja nueva {hoja_id}: {e}", exc_info=True)
        return jsonify({'error': 'No se pudo actualizar la hoja de ruta nueva'}), 500


@app.route('/api/hojas_ruta_nuevo/<int:hoja_id>/check_proceso', methods=['POST'])
@login_required
@requires_any_permission([('estaciones', 'operate'), ('hojas_mp', 'edit'), ('catalog', 'edit')])
def api_check_proceso_hoja_ruta_nuevo(hoja_id):
    """Marcar/desmarcar proceso virtual de hoja MP.
    Persiste el avance en bloque de estado dentro de materia_prima.
    """
    hoja = HojaRutaNueva.query.get_or_404(hoja_id)
    data = request.get_json() or {}

    proceso_id = data.get('proceso_id')
    completada = bool(data.get('completada', False))
    if not proceso_id:
        return jsonify({'error': 'proceso_id requerido'}), 400

    try:
        proceso_id = int(proceso_id)
    except Exception:
        return jsonify({'error': 'proceso_id invalido'}), 400

    clave = None
    if hoja.pn:
        clave = ClaveProducto.query.filter(func.upper(func.trim(ClaveProducto.clave)) == hoja.pn.upper().strip()).first()
    if not clave:
        return jsonify({'error': 'No se pudo resolver la clave de la hoja'}), 409

    cp = ClaveProceso.query.filter_by(id=proceso_id, clave_id=clave.id).first()
    if not cp:
        return jsonify({'error': 'El proceso no pertenece a la clave de la hoja'}), 409

    try:
        completed_ids = _mp_parse_completed_process_ids(hoja.materia_prima)
        if completada:
            completed_ids.add(proceso_id)
        else:
            completed_ids.discard(proceso_id)

        hoja.materia_prima = _mp_upsert_process_state_block(hoja.materia_prima, completed_ids)

        virtual = _build_mp_virtual_estaciones_by_pn(hoja.pn, completed_ids)
        total = len(virtual)
        completadas = sum(1 for e in virtual if (e.get('estado') or '').lower() == 'completada')

        ahora = datetime.utcnow()
        if total > 0 and completadas == total:
            hoja.estado = 'completada'
            if not hoja.fecha_termino:
                hoja.fecha_termino = ahora
        else:
            hoja.estado = 'activa'
            hoja.fecha_termino = None

        if not hoja.fecha_salida:
            hoja.fecha_salida = ahora

        db.session.commit()

        return jsonify({
            'success': True,
            'hoja': hoja.to_dict(),
            'resumen': {
                'total': total,
                'completadas': completadas,
                'pendientes': max(0, total - completadas),
            }
        }), 200
    except Exception as exc:
        db.session.rollback()
        logger.error(f"Error actualizando proceso MP hoja={hoja_id}: {exc}", exc_info=True)
        return jsonify({'error': 'No se pudo actualizar el proceso MP'}), 500


@app.route('/api/hojas_ruta_nuevo/<int:hoja_id>', methods=['DELETE'])
@login_required
@requires_any_permission([('hojas_mp', 'delete'), ('hojas', 'delete'), ('catalog', 'edit')])
def api_eliminar_hoja_ruta_nuevo(hoja_id):
    """Eliminar hoja de ruta nueva no asignada."""
    hoja = HojaRutaNueva.query.get_or_404(hoja_id)
    if hoja.maquina_id is not None:
        return jsonify({'error': 'No puedes eliminar una hoja asignada a una maquina. Primero desasignala.'}), 409

    try:
        db.session.delete(hoja)
        db.session.commit()
        return jsonify({'success': True, 'id': hoja_id}), 200
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error eliminando hoja nueva {hoja_id}: {e}", exc_info=True)
        return jsonify({'error': 'No se pudo eliminar la hoja de ruta nueva'}), 500


@app.route('/api/maquinas_nuevo/<int:maquina_id>/activar', methods=['POST'])
@login_required
@requires_any_permission([('estaciones', 'operate'), ('catalog', 'edit')])
def api_activar_maquina_nuevo(maquina_id):
    maq = Máquina.query.get_or_404(maquina_id)
    now_dt = datetime.utcnow()
    maq.activo = True
    cola = HojaRutaNueva.query.filter(
        HojaRutaNueva.maquina_id == maq.id,
        HojaRutaNueva.estado.in_(['activa', 'pausada'])
    ).all()
    cola = _order_machine_queue_items(cola)
    hoja_actual = _pick_machine_active_hoja(cola)
    if hoja_actual:
        _pause_other_machine_hojas(HojaRutaNueva, maq.id, hoja_actual.id, now_dt)
        _resume_or_activate_hoja(hoja_actual, now_dt)
    db.session.commit()
    return jsonify({'ok': True, 'maquina_id': maquina_id, 'activo': True}), 200


@app.route('/api/maquinas_nuevo/<int:maquina_id>/desactivar', methods=['POST'])
@login_required
@requires_any_permission([('estaciones', 'operate'), ('catalog', 'edit')])
def api_desactivar_maquina_nuevo(maquina_id):
    maq = Máquina.query.get_or_404(maquina_id)
    now_dt = datetime.utcnow()
    maq.activo = False
    hojas_activas = HojaRutaNueva.query.filter_by(maquina_id=maq.id, estado='activa').all()
    for hoja_activa in hojas_activas:
        hoja_activa.estado = 'pausada'
        hoja_activa.fecha_actualizacion = now_dt
    db.session.commit()
    return jsonify({'ok': True, 'maquina_id': maquina_id, 'activo': False}), 200


@app.route('/api/maquinas_nuevo/<int:maquina_id>/paro_mantenimiento', methods=['POST'])
@login_required
@requires_any_permission([('estaciones', 'operate'), ('catalog', 'edit')])
def api_paro_mantenimiento_nuevo(maquina_id):
    maq = Máquina.query.get_or_404(maquina_id)
    now_dt = datetime.utcnow()
    maq.activo = False
    hojas_activas = HojaRutaNueva.query.filter_by(maquina_id=maq.id, estado='activa').all()
    for hoja_activa in hojas_activas:
        hoja_activa.estado = 'pausada'
        hoja_activa.fecha_actualizacion = now_dt
    db.session.commit()
    return jsonify({'ok': True, 'maquina_id': maquina_id, 'activo': False, 'motivo': 'mantenimiento'}), 200


@app.route('/api/maquinas_nuevo/<int:maquina_id>/asignar_hoja', methods=['POST'])
@login_required
@requires_any_permission([('estaciones', 'operate'), ('catalog', 'edit')])
def api_asignar_hoja_maquina_nuevo(maquina_id):
    maq = Máquina.query.get_or_404(maquina_id)
    data = request.get_json() or {}
    hoja_id = data.get('hoja_id')
    start_estacion_id = data.get('start_estacion_id')
    elapsed_process_seconds = _parse_elapsed_process_seconds(data.get('elapsed_process_seconds'))
    if not hoja_id:
        return jsonify({'error': 'hoja_id requerido'}), 400
    if elapsed_process_seconds is None:
        return jsonify({'error': 'elapsed_process_seconds invalido'}), 400

    hoja = HojaRutaNueva.query.get_or_404(int(hoja_id))
    if hoja.maquina_id and hoja.maquina_id != maq.id:
        return jsonify({'error': 'La hoja ya está asignada a otra máquina'}), 409

    queue_limit = _machine_queue_limit_for_type(getattr(maq, 'tipo', None))
    queue_count = _machine_queue_count(HojaRutaNueva, maq.id, exclude_hoja_id=hoja.id)
    if queue_count >= queue_limit:
        return jsonify({'error': f'La máquina ya alcanzó su cola máxima ({queue_limit} hojas).'}), 409

    try:
        now_ref = datetime.utcnow()
        hoja.maquina_id = maq.id
        _pause_other_machine_hojas(HojaRutaNueva, maq.id, hoja.id, now_ref)
        _resume_or_activate_hoja(hoja, now_ref)
        target = _apply_mp_process_elapsed(hoja, start_estacion_id, elapsed_process_seconds, now_ref)
        if target is None and not hoja.fecha_salida:
            hoja.fecha_salida = now_ref - timedelta(seconds=max(0, int(elapsed_process_seconds or 0)))
        _recompute_mp_time_plan(hoja)
        maq.activo = True
        db.session.commit()
        return jsonify({'success': True, 'hoja': hoja.to_dict()}), 200
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error asignando hoja MP a maquina: {e}", exc_info=True)
        return jsonify({'error': 'No se pudo asignar la hoja MP a la máquina'}), 500


@app.route('/api/maquinas_nuevo/<int:maquina_id>/ajustar_tiempo_proceso', methods=['POST'])
@login_required
@requires_any_permission([('estaciones', 'operate'), ('catalog', 'edit')])
def api_ajustar_tiempo_proceso_maquina_nuevo(maquina_id):
    maq = Máquina.query.get_or_404(maquina_id)
    data = request.get_json() or {}
    elapsed_process_seconds = _parse_elapsed_process_seconds(data.get('elapsed_process_seconds'))
    if elapsed_process_seconds is None:
        return jsonify({'error': 'elapsed_process_seconds invalido'}), 400

    cola = HojaRutaNueva.query.filter(
        HojaRutaNueva.maquina_id == maq.id,
        HojaRutaNueva.estado.in_(['activa', 'pausada'])
    ).all()
    cola = _order_machine_queue_items(cola)
    hoja = _pick_machine_active_hoja(cola)
    if not hoja:
        return jsonify({'error': 'La máquina no tiene hoja activa o pausada'}), 404

    try:
        target = _apply_mp_process_elapsed(hoja, None, elapsed_process_seconds, datetime.utcnow())
        if target is None:
            return jsonify({'error': 'No hay proceso MP ajustable en la hoja'}), 409
        _recompute_mp_time_plan(hoja)
        db.session.commit()
        return jsonify({'ok': True, 'hoja_id': hoja.id, 'proceso_id': target.get('id'), 'elapsed_process_seconds': elapsed_process_seconds}), 200
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error ajustando tiempo de proceso MP: {e}", exc_info=True)
        return jsonify({'error': 'No se pudo ajustar el tiempo del proceso MP'}), 500


@app.route('/api/maquinas_nuevo/<int:maquina_id>/retirar_hoja', methods=['POST'])
@login_required
@requires_any_permission([('estaciones', 'operate'), ('catalog', 'edit')])
def api_retirar_hoja_maquina_nuevo(maquina_id):
    maq = Máquina.query.get_or_404(maquina_id)
    data = request.get_json() or {}
    hoja_id = data.get('hoja_id')

    if hoja_id:
        hoja = HojaRutaNueva.query.get_or_404(int(hoja_id))
        if hoja.maquina_id != maq.id:
            return jsonify({'error': 'La hoja no pertenece a esta máquina'}), 409
    else:
        hoja = HojaRutaNueva.query.filter(
            HojaRutaNueva.maquina_id == maq.id,
            HojaRutaNueva.estado.in_(['activa', 'pausada'])
        ).order_by(HojaRutaNueva.fecha_creacion.desc()).first()
        if not hoja:
            return jsonify({'error': 'No hay hoja activa o pausada asignada a esta máquina'}), 404

    now_ref = datetime.utcnow()
    hoja.maquina_id = None
    hoja.estado = 'activa'
    hoja.fecha_salida = None
    hoja.fecha_termino = None

    if bool(getattr(maq, 'activo', False)):
        restantes = HojaRutaNueva.query.filter(
            HojaRutaNueva.maquina_id == maq.id,
            HojaRutaNueva.estado.in_(['activa', 'pausada'])
        ).all()
        restantes = _order_machine_queue_items(restantes)
        siguiente = _pick_machine_active_hoja(restantes)
        if siguiente:
            _pause_other_machine_hojas(HojaRutaNueva, maq.id, siguiente.id, now_ref)
            _resume_or_activate_hoja(siguiente, now_ref)

    db.session.commit()
    return jsonify({'success': True, 'hoja': hoja.to_dict()}), 200


@app.route('/api/maquinas_nuevo/<int:maquina_id>/activar_hoja', methods=['POST'])
@login_required
@requires_any_permission([('estaciones', 'operate'), ('catalog', 'edit')])
def api_activar_hoja_maquina_nuevo(maquina_id):
    maq = Máquina.query.get_or_404(maquina_id)
    data = request.get_json() or {}
    hoja_id = int(data.get('hoja_id') or 0)
    if not hoja_id:
        return jsonify({'error': 'hoja_id requerido'}), 400

    hoja = HojaRutaNueva.query.get_or_404(hoja_id)
    if hoja.maquina_id != maq.id:
        return jsonify({'error': 'La hoja no está asignada a esta máquina'}), 409

    try:
        now_dt = datetime.utcnow()
        _pause_other_machine_hojas(HojaRutaNueva, maq.id, hoja.id, now_dt)
        _resume_or_activate_hoja(hoja, now_dt)
        maq.activo = True
        db.session.commit()
        return jsonify({'ok': True, 'maquina_id': maq.id, 'hoja_id': hoja.id}), 200
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error activando hoja MP en cola de maquina: {e}", exc_info=True)
        return jsonify({'error': 'No se pudo priorizar la hoja en la máquina'}), 500


@app.route('/api/produccion/ingresar_piezas', methods=['POST'])
@login_required
def api_ingresar_piezas():
    """Recibe: maquina_id, cantidad, clave, tiempo_total (HH:MM:SS opcional), producto (opcional).
    Crea una HojaRuta con esos datos y la marca como activa.
    """
    data = request.get_json() or {}
    maquina_id = data.get('maquina_id')
    cantidad = data.get('cantidad')
    clave = data.get('clave')
    tiempo_total = data.get('tiempo_total')
    producto = data.get('producto')

    if not maquina_id or not cantidad or not clave:
        return jsonify({'error': 'maquina_id, cantidad y clave son requeridos'}), 400

    try:
        maquina_id_int = int(maquina_id)
    except Exception:
        return jsonify({'error': 'maquina_id inválido'}), 400

    maq_for_limit = Máquina.query.get(maquina_id_int)
    queue_limit = _machine_queue_limit_for_type(getattr(maq_for_limit, 'tipo', None) if maq_for_limit else None)

    hojas_en_cola = HojaRutaEntrega.query.filter(
        HojaRutaEntrega.maquina_id == maquina_id_int,
        HojaRutaEntrega.estado.in_(['activa', 'pausada'])
    ).order_by(HojaRutaEntrega.fecha_actualizacion.desc(), HojaRutaEntrega.fecha_creacion.desc()).all()
    if len(hojas_en_cola) >= queue_limit:
        return jsonify({'error': f'La máquina ya alcanzó su cola máxima ({queue_limit} hojas).'}), 409

    veces_previas_maquina = HojaRutaEntrega.query.filter_by(maquina_id=maquina_id_int, pn=clave).count()

    try:
        nombre = f"Producción {clave}"
        maquina = Máquina.query.get(maquina_id_int)
        hoja = HojaRutaEntrega(
            maquina_id=maquina_id_int,
            nombre=nombre,
            producto=producto or clave,
            pn=clave,
            cantidad_piezas=int(cantidad),
            total_tiempo=tiempo_total,
            fecha_salida=datetime.utcnow(),
            estado='activa'
        )
        db.session.add(hoja)
        db.session.flush()  # obtener id sin commit
        now_ref = datetime.utcnow()
        _pause_other_machine_hojas(HojaRutaEntrega, maquina_id_int, hoja.id, now_ref)
        _resume_or_activate_hoja(hoja, now_ref)

        estaciones_creadas = []

        # Preferir procesos por clave (modulo procesos y claves)
        clave_obj = ClaveProducto.query.filter_by(clave=clave).first()
        if clave_obj:
            procesos = ClaveProceso.query.filter_by(clave_id=clave_obj.id).order_by(ClaveProceso.orden).all()
            for idx, cp in enumerate(procesos, start=1):
                est = EstacionTrabajo(
                    hoja_ruta_id=hoja.id,
                    nombre=f"{cp.operacion or cp.proceso.operacion or cp.proceso.nombre}",
                    pro_c=str(idx),
                    centro_trabajo=cp.centro_trabajo or cp.proceso.centro_trabajo or '',
                    operacion=cp.operacion or cp.proceso.operacion or cp.proceso.nombre or '',
                    orden=cp.orden,
                    t_e=cp.t_e or cp.proceso.tiempo_estimado or '',
                    t_tct=cp.t_tct or '',
                    t_tco=cp.t_tco or '',
                    t_to=cp.t_to or '',
                    estado='pendiente'
                )
                db.session.add(est)
                estaciones_creadas.append(est)

        # Fallback: plantillas por tipo si la clave no existe o no tiene procesos
        if not estaciones_creadas:
            plantilla_nombre = data.get('plantilla_nombre')
            if plantilla_nombre:
                plantillas = EstacionPlantilla.query.filter_by(maquina_tipo=maquina.tipo if maquina else None, plantilla_nombre=plantilla_nombre).order_by(EstacionPlantilla.orden).all()
            else:
                plantilla_tipo = maquina.tipo if maquina else None
                plantillas = EstacionPlantilla.query.filter_by(maquina_tipo=plantilla_tipo).order_by(EstacionPlantilla.orden).all() if plantilla_tipo else []

            for p in plantillas:
                est = EstacionTrabajo(
                    hoja_ruta_id=hoja.id,
                    nombre=(p.operacion or p.pro_c or 'Estación'),
                    pro_c=p.pro_c,
                    centro_trabajo=p.centro_trabajo,
                    operacion=p.operacion,
                    orden=p.orden,
                    t_e=p.t_e,
                    t_tct=p.t_tct,
                    t_tco=p.t_tco,
                    t_to=p.t_to
                )
                db.session.add(est)
                estaciones_creadas.append(est)

        _apply_hoja_time_plan(
            hoja,
            estaciones_creadas,
            maquina_tipo=maquina.tipo if maquina else None,
            fallback_total_time=tiempo_total,
        )

        db.session.commit()
        logger.info(f"[PRODUCCION] Hoja creada {hoja.id} para maquina {maquina_id} con {len(estaciones_creadas)} procesos")
        return jsonify({
            'success': True,
            'hoja': hoja.to_dict(),
            'ya_paso_por_maquina': veces_previas_maquina > 0,
            'veces_previas_maquina': veces_previas_maquina
        }), 201
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error creando hoja de ruta desde ingreso de piezas: {e}", exc_info=True)
        return jsonify({'error': 'Error creando hoja de ruta'}), 500


@app.route('/api/qc_estaciones', methods=['POST'])
@login_required
def api_qc_estaciones_registro():
    """Registrar control de calidad de producción (independiente del QC de maquinaria)."""
    payload = request.get_json(silent=True) or request.form
    maquina_id = payload.get('maquina_id')
    clave_pieza = payload.get('clave_pieza')
    resultado = payload.get('resultado')

    if not maquina_id or not clave_pieza or not resultado:
        return jsonify({'error': 'maquina_id, clave_pieza y resultado son requeridos'}), 400

    def to_int(value):
        try:
            return int(value) if value is not None and value != '' else None
        except Exception:
            return None

    mediciones = None
    mediciones_raw = payload.get('mediciones')
    if mediciones_raw:
        if isinstance(mediciones_raw, (dict, list)):
            mediciones = mediciones_raw
        else:
            try:
                mediciones = json.loads(mediciones_raw)
            except Exception:
                mediciones = {'valor': str(mediciones_raw)}

    try:
        registro = QCProduccionRegistro(
            maquina_id=int(maquina_id),
            hoja_ruta_id=to_int(payload.get('hoja_ruta_id')),
            clave_pieza=clave_pieza,
            lote=payload.get('lote'),
            cantidad_inspeccionada=to_int(payload.get('cantidad_inspeccionada')),
            cantidad_aprobada=to_int(payload.get('cantidad_aprobada')),
            cantidad_rechazada=to_int(payload.get('cantidad_rechazada')),
            resultado=resultado,
            notas=payload.get('notas'),
            mediciones=mediciones,
            usuario=session.get('user')
        )
        db.session.add(registro)
        db.session.commit()
        logger.info(f"[QC ESTACIONES] Registro creado {registro.id} para maquina {maquina_id}")
        return jsonify({'success': True, 'registro': registro.to_dict()}), 201
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error guardando QC estaciones: {e}", exc_info=True)
        return jsonify({'error': 'No se pudo guardar el control de calidad'}), 500


@app.route('/api/qc_estaciones/mp_alert', methods=['POST'])
@login_required
@requires_any_permission([('calidad', 'edit'), ('estaciones', 'operate'), ('catalog', 'edit')])
def api_qc_estaciones_mp_alert():
    payload = request.get_json(silent=True) or request.form

    try:
        maquina_id = int(payload.get('maquina_id') or 0)
        hoja_mp_id = int(payload.get('hoja_mp_id') or 0)
        process_id = int(payload.get('process_id') or 0)
        cantidad_revisada = int(payload.get('cantidad_revisada') or 0)
    except Exception:
        return jsonify({'error': 'Datos invalidos para registrar alerta QC'}), 422

    alert_type, meta = _mp_qc_alert_type_meta(payload.get('alert_type'))
    if maquina_id <= 0 or hoja_mp_id <= 0 or process_id <= 0 or cantidad_revisada <= 0:
        return jsonify({'error': 'maquina_id, hoja_mp_id, process_id y cantidad_revisada son requeridos'}), 400

    hoja = HojaRutaNueva.query.get_or_404(hoja_mp_id)
    if int(hoja.maquina_id or 0) != maquina_id:
        return jsonify({'error': 'La hoja no esta asignada a esa maquina'}), 409

    procesos = _build_mp_virtual_estaciones_by_pn(hoja.pn, _mp_parse_completed_process_ids(hoja.materia_prima))
    proceso = next((p for p in procesos if int(p.get('id') or 0) == process_id), None)
    if not proceso:
        return jsonify({'error': 'Proceso no encontrado para esta hoja'}), 404

    maquina = Máquina.query.get(maquina_id)
    reviewed_at = datetime.utcnow().isoformat()

    try:
        registro = QCProduccionRegistro(
            maquina_id=maquina_id,
            hoja_ruta_id=None,
            clave_pieza=hoja.pn or hoja.nombre or 'SIN_CLAVE',
            lote=hoja.nombre or None,
            cantidad_inspeccionada=cantidad_revisada,
            cantidad_aprobada=cantidad_revisada if alert_type == 'verificado' else 0,
            cantidad_rechazada=cantidad_revisada if alert_type == 'scrap' else 0,
            resultado=alert_type,
            notas=payload.get('notas') or '',
            mediciones={
                'module': 'estaciones_t_mp_qc',
                'alert_type': alert_type,
                'alert_label': meta['label'],
                'alert_color': meta['color'],
                'hoja_mp_id': hoja.id,
                'hoja_nombre': hoja.nombre or '',
                'process_id': process_id,
                'process_name': proceso.get('operacion') or proceso.get('nombre') or '',
                'station_order': proceso.get('orden'),
                'maquina_nombre': maquina.nombre if maquina else '',
                'reviewed_at': reviewed_at,
            },
            usuario=session.get('user') or 'sistema',
        )
        db.session.add(registro)
        db.session.commit()
        return jsonify({'success': True, 'registro': _mp_qc_alert_payload_from_registro(registro)}), 201
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error guardando QC MP semaforo: {e}", exc_info=True)
        return jsonify({'error': 'No se pudo guardar la alerta de calidad'}), 500


@app.route('/api/qc_estaciones/entregas_alert', methods=['POST'])
@login_required
@requires_any_permission([('calidad', 'edit'), ('estaciones', 'operate'), ('catalog', 'edit')])
def api_qc_estaciones_entregas_alert():
    payload = request.get_json(silent=True) or request.form

    try:
        maquina_id = int(payload.get('maquina_id') or 0)
        hoja_ruta_id = int(payload.get('hoja_ruta_id') or 0)
        estacion_id = int(payload.get('estacion_id') or payload.get('process_id') or 0)
        cantidad_revisada = int(payload.get('cantidad_revisada') or 0)
    except Exception:
        return jsonify({'error': 'Datos invalidos para registrar alerta QC'}), 422

    alert_type, meta = _mp_qc_alert_type_meta(payload.get('alert_type'))
    if maquina_id <= 0 or hoja_ruta_id <= 0 or estacion_id <= 0 or cantidad_revisada <= 0:
        return jsonify({'error': 'maquina_id, hoja_ruta_id, estacion_id y cantidad_revisada son requeridos'}), 400

    hoja = HojaRutaEntrega.query.get_or_404(hoja_ruta_id)
    if int(hoja.maquina_id or 0) != maquina_id:
        return jsonify({'error': 'La hoja no esta asignada a esa maquina'}), 409

    estacion = EstacionTrabajo.query.get_or_404(estacion_id)
    if int(estacion.hoja_ruta_id or 0) != hoja.id:
        return jsonify({'error': 'La estacion no pertenece a esa hoja'}), 409

    maquina = Máquina.query.get(maquina_id)
    reviewed_at = datetime.utcnow().isoformat()

    try:
        registro = QCProduccionRegistro(
            maquina_id=maquina_id,
            hoja_ruta_id=hoja.id,
            clave_pieza=hoja.pn or hoja.nombre or 'SIN_CLAVE',
            lote=hoja.nombre or None,
            cantidad_inspeccionada=cantidad_revisada,
            cantidad_aprobada=cantidad_revisada if alert_type == 'verificado' else 0,
            cantidad_rechazada=cantidad_revisada if alert_type == 'scrap' else 0,
            resultado=alert_type,
            notas=payload.get('notas') or '',
            mediciones={
                'module': 'estaciones_t_entregas_qc',
                'alert_type': alert_type,
                'alert_label': meta['label'],
                'alert_color': meta['color'],
                'hoja_ruta_id': hoja.id,
                'hoja_nombre': hoja.nombre or '',
                'estacion_id': estacion.id,
                'process_id': estacion.id,
                'process_name': estacion.operacion or estacion.nombre or '',
                'station_order': estacion.orden,
                'maquina_nombre': maquina.nombre if maquina else '',
                'reviewed_at': reviewed_at,
            },
            usuario=session.get('user') or 'sistema',
        )
        db.session.add(registro)
        db.session.commit()
        return jsonify({'success': True, 'registro': _entregas_qc_alert_payload_from_registro(registro)}), 201
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error guardando QC entregas semaforo: {e}", exc_info=True)
        return jsonify({'error': 'No se pudo guardar la alerta de calidad'}), 500


@app.route('/api/maquinas/<int:maquina_id>/plantilla_default', methods=['POST'])
@login_required
def api_set_plantilla_default(maquina_id):
    """Asigna una plantilla por defecto a una máquina (campo plantilla_default)."""
    data = request.get_json() or {}
    plantilla = data.get('plantilla_nombre')
    if plantilla is None:
        return jsonify({'error': 'plantilla_nombre es requerido'}), 400
    try:
        maquina = Máquina.query.get_or_404(maquina_id)
        maquina.plantilla_default = plantilla
        db.session.commit()
        logger.info(f"[MAQUINA] plantilla_default set {plantilla} for maquina {maquina_id}")
        return jsonify({'success': True, 'maquina': maquina.to_dict()}), 200
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error asignando plantilla_default: {e}", exc_info=True)
        return jsonify({'error': 'No se pudo asignar plantilla_default'}), 500


@app.route('/api/estaciones/<int:estacion_id>', methods=['PUT'])
@login_required
def api_actualizar_estacion(estacion_id):
    """Actualizar estado y detalles de una estación."""
    estacion = EstacionTrabajo.query.get_or_404(estacion_id)
    data = request.get_json()
    
    if 'estado' in data:
        estacion.estado = data['estado']
        if data['estado'] == 'en_curso' and not estacion.fecha_inicio:
            estacion.fecha_inicio = datetime.utcnow()
        elif data['estado'] == 'completada' and not estacion.fecha_finalizacion:
            estacion.fecha_finalizacion = datetime.utcnow()
    
    if 'nombre' in data:
        estacion.nombre = data['nombre']
    if 'descripcion' in data:
        estacion.descripcion = data['descripcion']
    if 'notas' in data:
        estacion.notas = data['notas']
    
    db.session.commit()
    logger.info(f"[HOJAS_RUTA] Estación actualizada: {estacion_id}")
    return jsonify(estacion.to_dict()), 200


@app.route('/api/plantillas_estaciones', methods=['GET'])
@login_required
def api_list_plantillas():
    """Listar plantillas; opcionalmente filtrar por `maquina_tipo` query param."""
    tipo = request.args.get('maquina_tipo')
    if tipo:
        plantillas = EstacionPlantilla.query.filter_by(maquina_tipo=tipo).order_by(EstacionPlantilla.plantilla_nombre, EstacionPlantilla.orden).all()
    else:
        plantillas = EstacionPlantilla.query.order_by(EstacionPlantilla.maquina_tipo, EstacionPlantilla.plantilla_nombre, EstacionPlantilla.orden).all()
    return jsonify({'plantillas': [p.to_dict() for p in plantillas]})


@app.route('/api/plantillas_estaciones/nombres')
@login_required
def api_plantilla_nombres():
    """Devuelve nombres de plantillas (distinct) para un tipo de máquina dado."""
    tipo = request.args.get('maquina_tipo')
    if not tipo:
        return jsonify({'error': 'maquina_tipo requerido'}), 400
    try:
        rows = db.session.query(EstacionPlantilla.plantilla_nombre).filter_by(maquina_tipo=tipo).distinct().all()
        nombres = [r[0] for r in rows if r[0]]
        return jsonify({'nombres': nombres})
    except Exception as e:
        logger.error(f"Error fetch plantilla nombres: {e}", exc_info=True)
        return jsonify({'error': 'Error interno'}), 500


@app.route('/api/plantillas_estaciones', methods=['POST'])
@login_required
def api_create_plantilla():
    data = request.get_json() or {}
    required = ['maquina_tipo', 'operacion']
    for r in required:
        if r not in data:
            return jsonify({'error': f'{r} es requerido'}), 400
    try:
        p = EstacionPlantilla(
            plantilla_nombre=data.get('plantilla_nombre'),
            maquina_tipo=data.get('maquina_tipo'),
            pro_c=data.get('pro_c'),
            centro_trabajo=data.get('centro_trabajo'),
            operacion=data.get('operacion'),
            orden=int(data.get('orden') or 0),
            t_e=data.get('t_e'),
            t_tct=data.get('t_tct'),
            t_tco=data.get('t_tco'),
            t_to=data.get('t_to')
        )
        db.session.add(p)
        db.session.commit()
        return jsonify({'success': True, 'plantilla': p.to_dict()}), 201
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error creando plantilla: {e}", exc_info=True)
        return jsonify({'error': 'Error creando plantilla'}), 500


@app.route('/api/plantillas_estaciones/<int:pid>', methods=['PUT'])
@login_required
def api_update_plantilla(pid):
    p = EstacionPlantilla.query.get_or_404(pid)
    data = request.get_json() or {}
    for k in ['plantilla_nombre', 'maquina_tipo', 'pro_c', 'centro_trabajo', 'operacion', 'orden', 't_e', 't_tct', 't_tco', 't_to']:
        if k in data:
            setattr(p, k, data[k])
    try:
        db.session.commit()
        return jsonify({'success': True, 'plantilla': p.to_dict()})
    except Exception:
        db.session.rollback()
        return jsonify({'error': 'No se pudo actualizar plantilla'}), 500


@app.route('/api/plantillas_estaciones/<int:pid>', methods=['DELETE'])
@login_required
def api_delete_plantilla(pid):
    p = EstacionPlantilla.query.get_or_404(pid)
    try:
        db.session.delete(p)
        db.session.commit()
        return jsonify({'success': True})
    except Exception:
        db.session.rollback()
        return jsonify({'error': 'No se pudo eliminar plantilla'}), 500


@app.route('/plantillas_estaciones')
@login_required
def plantillas_page():
    if not is_admin_user():
        return render_template('403.html'), 403
    return render_template('plantillas_estaciones.html')


@app.route('/admin')
@login_required
def admin():
    """Panel de administración"""
    user = get_current_user()
    if not _can_view_admin_panel(user):
        return render_template('403.html'), 403
    return render_template('admin.html')


@app.route('/admin/users')
@login_required
def admin_users_page():
    """Página de administración de usuarios (solo admin)."""
    user = get_current_user()
    if not _can_view_user_admin(user):
        return render_template('403.html'), 403
    can_view_users = _can_view_users_module(user)
    can_edit_users = _can_edit_users_module(user)
    can_view_roles = _can_view_roles_module(user)
    can_edit_roles = _can_edit_roles_module(user)
    can_view_permissions = _can_view_permissions_module(user)
    can_edit_permissions = _can_edit_permissions_module(user)
    can_manage_roles_permissions = (
        can_view_roles or can_edit_roles or can_view_permissions or can_edit_permissions
    )

    return render_template(
        'admin_users.html',
        can_view_users=can_view_users,
        can_edit_users=can_edit_users,
        can_view_roles=can_view_roles,
        can_edit_roles=can_edit_roles,
        can_view_permissions=can_view_permissions,
        can_edit_permissions=can_edit_permissions,
        can_manage_roles_permissions=can_manage_roles_permissions,
    )


# ======= API: Usuarios (admin only) ======
@app.route('/api/users')
@login_required
def api_list_users():
    user = get_current_user()
    if not _can_view_users_module(user):
        return jsonify({'error': 'Permiso denegado'}), 403
    users = Usuario.query.order_by(Usuario.id.asc()).all()
    data = []
    for u in users:
        data.append({
            'id': u.id,
            'username': u.username,
            'correo': u.correo,
            'activo': u.activo,
            'es_admin': u.es_admin,
            'role': u.role.name if u.role else None,
            'fecha_creacion': u.fecha_creacion.isoformat() if u.fecha_creacion else None
        })
    return jsonify({'users': data})


@app.route('/api/permissions')
@login_required
def api_list_permissions():
    user = get_current_user()
    if not _can_view_permissions_module(user):
        return jsonify({'error': 'Permiso denegado'}), 403
    _ensure_default_permissions()
    perms = Permission.query.order_by(Permission.module, Permission.action).all()
    return jsonify({'permissions': [p.to_dict() for p in perms]})


@app.route('/api/roles')
@login_required
def api_list_roles():
    user = get_current_user()
    if not _can_view_roles_module(user):
        return jsonify({'error': 'Permiso denegado'}), 403
    _ensure_default_permissions()
    roles = Role.query.order_by(Role.name).all()
    data = []
    for r in roles:
        item = r.to_dict()
        item['description'] = item.get('descripcion')
        item['modules'] = _role_modules_from_permissions(r)
        data.append(item)
    return jsonify({'roles': data})


@app.route('/api/roles', methods=['POST'])
@login_required
def api_create_role():
    user = get_current_user()
    if not _can_edit_roles_module(user):
        return jsonify({'error': 'Permiso denegado'}), 403

    _ensure_default_permissions()

    payload = request.get_json() or {}
    name = (payload.get('name') or '').strip()
    description = (payload.get('description') or payload.get('descripcion') or '').strip() or None
    modules = payload.get('modules', []) or []
    permission_ids = payload.get('permission_ids', []) or []

    if not name:
        return jsonify({'error': 'name requerido'}), 400

    existing = Role.query.filter_by(name=name).first()
    if existing:
        return jsonify({'error': 'Ya existe un role con ese nombre'}), 409

    role = Role(name=name, descripcion=description)
    role.permissions = _permissions_from_modules_and_ids(modules, permission_ids)
    db.session.add(role)
    db.session.commit()

    role_data = role.to_dict()
    role_data['description'] = role_data.get('descripcion')
    role_data['modules'] = _role_modules_from_permissions(role)
    return jsonify({'ok': True, 'role': role_data}), 201


@app.route('/api/roles/<int:role_id>/permissions', methods=['PUT'])
@login_required
def api_set_role_permissions(role_id):
    user = get_current_user()
    if not _can_edit_permissions_module(user):
        return jsonify({'error': 'Permiso denegado'}), 403
    payload = request.get_json() or {}
    perm_ids = payload.get('permission_ids', [])
    role = Role.query.get_or_404(role_id)
    # fetch permissions
    perms = Permission.query.filter(Permission.id.in_(perm_ids)).all() if perm_ids else []
    role.permissions = perms
    db.session.add(role)
    db.session.commit()
    return jsonify({'ok': True, 'role': role.to_dict()})


@app.route('/api/roles/<int:role_id>/modules', methods=['PUT'])
@login_required
def api_set_role_modules(role_id):
    user = get_current_user()
    if not (_can_edit_roles_module(user) or _can_edit_permissions_module(user)):
        return jsonify({'error': 'Permiso denegado'}), 403
    _ensure_default_permissions()
    payload = request.get_json() or {}
    modules = payload.get('modules', [])
    permission_ids = payload.get('permission_ids')
    role = Role.query.get_or_404(role_id)

    # If explicit permission IDs are provided, fully synchronize role permissions.
    if permission_ids is not None:
        role.permissions = _permissions_from_modules_and_ids(modules, permission_ids)
    else:
        _apply_role_modules(role, modules)

    db.session.add(role)
    db.session.commit()
    role_data = role.to_dict()
    role_data['description'] = role_data.get('descripcion')
    role_data['modules'] = _role_modules_from_permissions(role)
    return jsonify({'ok': True, 'role': role_data})


@app.route('/api/users', methods=['POST'])
@login_required
def api_create_user():
    user = get_current_user()
    if not _can_edit_users_module(user):
        return jsonify({'error': 'Permiso denegado'}), 403
    data = request.get_json() or {}
    username = data.get('username')
    correo = data.get('correo') or None  # empty string -> None to avoid unique constraint
    password = data.get('password')
    role_name = data.get('role')
    custom_role_name = (data.get('custom_role_name') or '').strip()
    custom_role_description = (data.get('custom_role_description') or '').strip() or None
    modules = data.get('modules', []) or []
    permission_ids = data.get('permission_ids', []) or []
    es_admin = bool(data.get('es_admin', False))
    if not username or not password:
        return jsonify({'error': 'username y password requeridos'}), 400
    if Usuario.query.filter_by(username=username).first():
        return jsonify({'error': 'username ya existe'}), 409
    if correo and Usuario.query.filter_by(correo=correo).first():
        return jsonify({'error': 'correo ya está registrado'}), 409
    u = Usuario(username=username, correo=correo, es_admin=es_admin, activo=True)
    u.set_password(password)

    wants_custom_profile = bool(modules or permission_ids or custom_role_name)
    if role_name and wants_custom_profile:
        return jsonify({'error': 'Usa role existente o perfil personalizado, no ambos al mismo tiempo'}), 400

    if wants_custom_profile:
        role_target_name = custom_role_name or f'perfil_{username}'
        if Role.query.filter_by(name=role_target_name).first():
            return jsonify({'error': f'Ya existe un role con nombre {role_target_name}'}), 409
        role = Role(name=role_target_name, descripcion=custom_role_description or f'Perfil personalizado de {username}')
        role.permissions = _permissions_from_modules_and_ids(modules, permission_ids)
        db.session.add(role)
        db.session.flush()
        u.role = role
    elif role_name:
        role = Role.query.filter_by(name=role_name).first()
        if role:
            u.role = role

    db.session.add(u)
    db.session.commit()
    user_data = u.to_dict()
    user_data['role'] = u.role.name if u.role else None
    return jsonify({'ok': True, 'user': user_data}), 201


@app.route('/api/users/<int:user_id>/toggle_active', methods=['PUT'])
@login_required
def api_toggle_active(user_id):
    user = get_current_user()
    if not _can_edit_users_module(user):
        return jsonify({'error': 'Permiso denegado'}), 403
    u = Usuario.query.get_or_404(user_id)
    if u.username == 'admin':
        return jsonify({'error': 'No se puede desactivar el usuario admin'}), 400
    u.activo = not bool(u.activo)
    db.session.add(u)
    db.session.commit()
    return jsonify({'ok': True, 'activo': u.activo})


@app.route('/api/users/<int:user_id>/role', methods=['PUT'])
@login_required
def api_set_role(user_id):
    user = get_current_user()
    if not _can_edit_users_module(user):
        return jsonify({'error': 'Permiso denegado'}), 403
    payload = request.get_json() or {}
    role_name = payload.get('role')
    u = Usuario.query.get_or_404(user_id)
    if u.username == 'admin' and role_name != 'admin':
        return jsonify({'error': 'El usuario admin debe conservar role admin'}), 400
    if role_name:
        role = Role.query.filter_by(name=role_name).first()
        if not role:
            return jsonify({'error': 'Role no encontrado'}), 404
        u.role = role
    else:
        u.role = None
    db.session.add(u)
    db.session.commit()
    return jsonify({'ok': True, 'role': u.role.name if u.role else None})


@app.route('/api/users/<int:user_id>', methods=['DELETE'])
@login_required
def api_delete_user(user_id):
    user = get_current_user()
    if not _can_edit_users_module(user):
        return jsonify({'error': 'Permiso denegado'}), 403
    u = Usuario.query.get_or_404(user_id)
    if u.username == 'admin':
        return jsonify({'error': 'No se puede borrar el usuario admin'}), 400
    db.session.delete(u)
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/api/users/<int:user_id>', methods=['GET'])
@login_required
def api_get_user(user_id):
    user = get_current_user()
    if not _can_view_users_module(user):
        return jsonify({'error': 'Permiso denegado'}), 403
    u = Usuario.query.get_or_404(user_id)
    data = {
        'id': u.id,
        'username': u.username,
        'correo': u.correo,
        'activo': u.activo,
        'es_admin': u.es_admin,
        'role': u.role.name if u.role else None,
        'fecha_creacion': u.fecha_creacion.isoformat() if u.fecha_creacion else None
    }
    data['password'] = None
    if user and user.es_admin:
        try:
            # devolver contraseña desencriptada solo a administradores.
            data['password'] = u.decrypt_password()
        except Exception:
            data['password'] = None
    return jsonify({'user': data})


@app.route('/api/users/<int:user_id>', methods=['PUT'])
@login_required
def api_update_user(user_id):
    user = get_current_user()
    if not _can_edit_users_module(user):
        return jsonify({'error': 'Permiso denegado'}), 403
    u = Usuario.query.get_or_404(user_id)
    data = request.get_json() or {}
    correo = data.get('correo')
    password = data.get('password')
    es_admin = data.get('es_admin')
    role_name = data.get('role')

    if correo is not None:
        u.correo = correo or None  # empty string -> None
    if password:
        u.set_password(password)
    if es_admin is not None:
        u.es_admin = bool(es_admin)
    if role_name is not None:
        if role_name == '':
            u.role = None
        else:
            role = Role.query.filter_by(name=role_name).first()
            if role:
                u.role = role
    db.session.add(u)
    db.session.commit()
    return jsonify({'ok': True, 'user': u.to_dict()})


@app.route('/delete_nonadmin_users', methods=['POST'])
@login_required
def delete_nonadmin_users():
    """Endpoint conveniente que borra todos los usuarios salvo 'admin'. Protegido a admin."""
    if not is_admin_user():
        return jsonify({'error': 'Permiso denegado'}), 403
    users = Usuario.query.filter(Usuario.username != 'admin').all()
    deleted = 0
    for u in users:
        db.session.delete(u)
        deleted += 1
    db.session.commit()
    return f"Usuarios eliminados: {deleted}", 200


@app.route('/admin/test-telegram', methods=['POST'])
@login_required
def test_telegram():
    """Envía un mensaje de prueba a Telegram. Solo admin."""
    if not is_admin_user():
        return jsonify({'error': 'Permiso denegado'}), 403
    
    mensaje_prueba = (
        "🎉 *Bienvenido al Portal de Notificaciones* 🎉\n\n"
        "Laboratorio Grupo Industrial Verduzco - Control de Calidad\n\n"
        "✅ Integración de alertas de Telegram activada correctamente.\n"
        "Este es un mensaje de prueba enviado desde la plataforma.\n\n"
        f"Fecha: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC"
    )
    success, reason = _send_telegram_message(mensaje_prueba)
    return jsonify({
        'ok': success,
        'reason': reason,
        'message': 'Mensaje de prueba enviado a Telegram' if success else f'Error: {reason}'
    })


@app.route('/proveedores')
@login_required
@requires_permission('catalog', 'view')
def proveedores():
    """Página de gestión de proveedores"""
    return render_template('proveedores.html')
@app.route('/reportar')
def reportar_incidencia():
    """Página pública para reportar incidencias (sin login)"""
    return render_template('reportar_incidencia.html')

@app.route('/soporte')
@login_required
@requires_permission('tickets', 'view')
def soporte_tecnico():
    """Panel de ingenieros para gestionar tickets de soporte"""
    return render_template('soporte_tecnico.html')


# ========== Paneles de Tickets (rutas asignadas) ===========
@app.route('/tickets')
@login_required
@requires_permission('tickets', 'view')
def tickets_panel():
    """Panel de tickets para usuarios (mis tickets)."""
    return render_template('tickets.html')


@app.route('/tickets/admin')
@login_required
def tickets_admin_panel():
    """Panel de administración de tickets (solo admin)."""
    if not is_admin_user():
        return render_template('403.html'), 403
    return render_template('tickets_admin.html')


@app.route('/tickets/ingeniero')
@login_required
def tickets_ingeniero_panel():
    """Panel específico para ingenieros de soporte."""
    user = get_current_user()
    if not user:
        return redirect(url_for('login'))
    if not (user.es_admin or (user.role and user.role.name == 'support')):
        return render_template('403.html'), 403
    return render_template('tickets_ingeniero.html')



# ========== Public read-only endpoints and consulta page ===========
@app.route('/catalogo_consulta')
def catalogo_consulta():
    """Página pública de consulta para accesos directos (sin login)."""
    return render_template('catalogo_consulta.html')


def _producto_sanitizado(p):
    ultimo_pp = None
    if p.proveedores:
        def _pp_key(pp):
            return pp.fecha_precio or datetime.min.date()
        try:
            ultimo_pp = max(p.proveedores, key=_pp_key)
        except ValueError:
            ultimo_pp = None
    return {
        'id': p.id,
        'clave': p.clave,
        'nombre': p.nombre,
        'descripcion': p.descripcion,
        'precio': p.precio,
        'divisa_venta': p.divisa_venta,
        'cantidad': p.cantidad,
        'imagen_url': p.imagen_url,
        'categoria': p.categoria,
        'unidad': p.unidad,
        'linea': p.linea,
        'clasificacion': p.clasificacion,
        'clasificacion_departamento': p.clasificacion_departamento,
        'divisa_ultima': ultimo_pp.divisa if ultimo_pp else None,
        'precio_compra_ultimo': ultimo_pp.precio_proveedor if ultimo_pp else None,
        'proveedor_ultimo': ultimo_pp.proveedor.nombre if ultimo_pp and ultimo_pp.proveedor else None,
        'fecha_precio_ultimo': ultimo_pp.fecha_precio.isoformat() if ultimo_pp and ultimo_pp.fecha_precio else None
    }


def _catalogo_productos_filtrados(query='', categoria='', clasificacion=''):
    productos_q = Producto.query
    q = (query or '').strip().lower()
    cat = (categoria or '').strip().lower()
    cla = (clasificacion or '').strip().lower()

    if q:
        productos_q = productos_q.filter(
            db.or_(
                Producto.nombre.ilike(f'%{q}%'),
                Producto.descripcion.ilike(f'%{q}%'),
                Producto.clave.ilike(f'%{q}%')
            )
        )
    if cat:
        productos_q = productos_q.filter(Producto.categoria.ilike(f'%{cat}%'))
    if cla:
        productos_q = productos_q.filter(Producto.clasificacion.ilike(f'%{cla}%'))

    return productos_q


def _valor_bool(value):
    return str(value or '').strip().lower() in ('1', 'true', 'yes', 'si', 'on')


def _catalogo_reporte_compra_data(query='', categoria='', clasificacion='', solo_china=False):
    productos_q = _catalogo_productos_filtrados(
        query=query,
        categoria=categoria,
        clasificacion=clasificacion,
    )
    if solo_china:
        productos_q = productos_q.filter(Producto.clasificacion.ilike('%CH%'))

    productos = productos_q.all()
    productos = sorted(
        productos,
        key=lambda p: ((p.clave or '').strip().lower(), (p.nombre or '').strip().lower())
    )

    filas = []
    for p in productos:
        data = _producto_sanitizado(p)
        clasificacion_producto = (data.get('clasificacion') or '').strip()
        es_china = 'SI' if 'CH' in clasificacion_producto.upper() else 'NO'
        filas.append({
            'clave': data.get('clave') or '',
            'descripcion': data.get('descripcion') or data.get('nombre') or '',
            'proveedor': data.get('proveedor_ultimo') or '',
            'precio_compra': data.get('precio_compra_ultimo'),
            'divisa': data.get('divisa_ultima') or '',
            'fecha_precio': data.get('fecha_precio_ultimo'),
            'clasificacion': clasificacion_producto,
            'china': es_china,
        })

    return filas


@app.route('/catalogo_consulta/reporte_compra')
def catalogo_consulta_reporte_compra():
    """Reporte imprimible de compra por producto (clave, descripcion, proveedor, precio y divisa)."""
    try:
        query = request.args.get('q', '')
        categoria = request.args.get('categoria', '')
        clasificacion = request.args.get('clasificacion', '')
        solo_china = _valor_bool(request.args.get('solo_china'))
        if solo_china and not (clasificacion or '').strip():
            clasificacion = 'CH'

        filas = _catalogo_reporte_compra_data(
            query=query,
            categoria=categoria,
            clasificacion=clasificacion,
            solo_china=solo_china,
        )

        params_reporte = {}
        if query:
            params_reporte['q'] = query
        if categoria:
            params_reporte['categoria'] = categoria
        if clasificacion:
            params_reporte['clasificacion'] = clasificacion
        if solo_china:
            params_reporte['solo_china'] = '1'

        return render_template(
            'catalogo_consulta_reporte_compra.html',
            filas=filas,
            total_filas=len(filas),
            filtros={
                'q': query,
                'categoria': categoria,
                'clasificacion': clasificacion,
                'solo_china': solo_china,
            },
            fecha_generacion=datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            reporte_query_string=urlencode(params_reporte)
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/catalogo_consulta/reporte_compra_excel')
def catalogo_consulta_reporte_compra_excel():
    """Descarga en Excel del reporte de compra por producto."""
    try:
        query = request.args.get('q', '')
        categoria = request.args.get('categoria', '')
        clasificacion = request.args.get('clasificacion', '')
        solo_china = _valor_bool(request.args.get('solo_china'))
        if solo_china and not (clasificacion or '').strip():
            clasificacion = 'CH'

        filas = _catalogo_reporte_compra_data(
            query=query,
            categoria=categoria,
            clasificacion=clasificacion,
            solo_china=solo_china,
        )

        df = pd.DataFrame(filas)
        if df.empty:
            df = pd.DataFrame(columns=['clave', 'descripcion', 'proveedor', 'precio_compra', 'divisa', 'fecha_precio', 'clasificacion', 'china'])

        df = df.rename(columns={
            'clave': 'CLAVE',
            'descripcion': 'DESCRIPCION',
            'proveedor': 'PROVEEDOR',
            'precio_compra': 'PRECIO_COMPRA',
            'divisa': 'DIVISA',
            'fecha_precio': 'FECHA_PRECIO',
            'clasificacion': 'CLASIFICACION',
            'china': 'CHINA',
        })

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='reporte_compra')
        output.seek(0)

        sufijo = 'CH_' if solo_china else ''
        stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'reporte_compra_{sufijo}{stamp}.xlsx'
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/public/categorias', methods=['GET'])
def public_categorias():
    """Public endpoint para obtener categorías distintas."""
    try:
        cats = db.session.query(Producto.categoria).filter(Producto.categoria != None).distinct().all()
        # cats is list of tuples
        lista = sorted([c[0] for c in cats if c[0]])
        return jsonify(lista)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/public/productos/buscar', methods=['GET'])
def public_buscar_productos():
    """Public endpoint para buscar productos (sin login)."""
    try:
        query = request.args.get('q', '').lower()
        categoria = request.args.get('categoria', '').lower()
        clasificacion = request.args.get('clasificacion', '').lower()

        productos_q = _catalogo_productos_filtrados(
            query=query,
            categoria=categoria,
            clasificacion=clasificacion,
        )

        productos = productos_q.all()
        return jsonify([_producto_sanitizado(p) for p in productos])
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/public/productos', methods=['GET'])
def public_get_productos():
    """Public endpoint para listar todos los productos (sin login)."""
    try:
        productos = Producto.query.all()
        return jsonify([_producto_sanitizado(p) for p in productos])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ==================== API CRUD ====================

# GET - Obtener todos los productos
@app.route('/api/productos', methods=['GET'])
@login_required
def get_productos():
    productos = Producto.query.all()
    return jsonify([p.to_dict() for p in productos])

# GET - Obtener producto por ID
@app.route('/api/productos/<int:id>', methods=['GET'])
@login_required
def get_producto(id):
    producto = Producto.query.get(id)
    if not producto:
        return jsonify({'error': 'Producto no encontrado'}), 404
    return jsonify(producto.to_dict())

# POST - Crear producto
@app.route('/api/productos', methods=['POST'])
@login_required
def crear_producto():
    data = request.get_json()
    
    if not data or not data.get('nombre') or not data.get('precio'):
        return jsonify({'error': 'Nombre y precio son obligatorios'}), 400
    
    nuevo_producto = Producto(
        nombre=data.get('nombre'),
        descripcion=data.get('descripcion', ''),
        precio=float(data.get('precio')),
        cantidad=int(data.get('cantidad', 0)),
        imagen_url=data.get('imagen_url', ''),
        categoria=data.get('categoria', '')
    )
    
    db.session.add(nuevo_producto)
    db.session.commit()
    
    return jsonify(nuevo_producto.to_dict()), 201

# PUT - Actualizar producto
@app.route('/api/productos/<int:id>', methods=['PUT'])
@login_required
def actualizar_producto(id):
    producto = Producto.query.get(id)
    
    if not producto:
        return jsonify({'error': 'Producto no encontrado'}), 404
    
    data = request.get_json()
    
    producto.nombre = data.get('nombre', producto.nombre)
    producto.descripcion = data.get('descripcion', producto.descripcion)
    producto.precio = float(data.get('precio', producto.precio))
    producto.cantidad = int(data.get('cantidad', producto.cantidad))
    producto.imagen_url = data.get('imagen_url', producto.imagen_url)
    producto.categoria = data.get('categoria', producto.categoria)
    
    db.session.commit()
    
    return jsonify(producto.to_dict())

# DELETE - Eliminar producto
@app.route('/api/productos/<int:id>', methods=['DELETE'])
@login_required
def eliminar_producto(id):
    producto = Producto.query.get(id)
    
    if not producto:
        return jsonify({'error': 'Producto no encontrado'}), 404
    
    db.session.delete(producto)
    db.session.commit()
    
    return jsonify({'mensaje': 'Producto eliminado correctamente'})

# ==================== RUTAS ADICIONALES ====================

@app.route('/api/estadisticas', methods=['GET'])
def estadisticas():
    """Estadísticas del catálogo"""
    total_productos = Producto.query.count()
    valor_total_inventario = db.session.query(db.func.sum(Producto.precio * Producto.cantidad)).scalar() or 0
    
    # Estadísticas por categoría
    categorias = db.session.query(
        Producto.categoria,
        db.func.count(Producto.id).label('cantidad'),
        db.func.avg(Producto.precio).label('precio_promedio')
    ).group_by(Producto.categoria).all()
    
    stats_por_categoria = [
        {
            'categoria': cat[0] or 'Sin categoría',
            'cantidad': cat[1],
            'precio_promedio': float(cat[2]) if cat[2] else 0
        }
        for cat in categorias
    ]
    
    return jsonify({
        'total_productos': total_productos,
        'valor_total_inventario': float(valor_total_inventario),
        'categorias': stats_por_categoria,
        'producto_mas_caro': db.session.query(Producto).order_by(Producto.precio.desc()).first().to_dict() if total_productos > 0 else None,
        'producto_mas_barato': db.session.query(Producto).order_by(Producto.precio.asc()).first().to_dict() if total_productos > 0 else None,
        'stock_total': db.session.query(db.func.sum(Producto.cantidad)).scalar() or 0
    })

@app.route('/api/productos/buscar', methods=['GET'])
@login_required
def buscar_productos():
    """Buscar productos por nombre, categoría o descripción"""
    query = request.args.get('q', '').lower()
    categoria = request.args.get('categoria', '').lower()
    precio_min = request.args.get('precio_min', type=float)
    precio_max = request.args.get('precio_max', type=float)
    
    productos = Producto.query
    
    if query:
        productos = productos.filter(
            db.or_(
                Producto.nombre.ilike(f'%{query}%'),
                Producto.descripcion.ilike(f'%{query}%'),
                Producto.clave.ilike(f'%{query}%')
            )
        )
    
    if categoria:
        productos = productos.filter(Producto.categoria.ilike(f'%{categoria}%'))
    
    if precio_min is not None:
        productos = productos.filter(Producto.precio >= precio_min)
    
    if precio_max is not None:
        productos = productos.filter(Producto.precio <= precio_max)
    
    return jsonify([p.to_dict() for p in productos.all()])

@app.route('/api/productos/exportar', methods=['GET'])
@login_required
def exportar_productos():
    """Exportar productos a CSV"""
    import csv
    from io import StringIO
    from flask import make_response
    
    productos = Producto.query.all()
    
    output = StringIO()
    writer = csv.DictWriter(output, fieldnames=['id', 'nombre', 'descripcion', 'precio', 'cantidad', 'categoria', 'fecha_creacion'])
    writer.writeheader()
    
    for p in productos:
        writer.writerow({
            'id': p.id,
            'nombre': p.nombre,
            'descripcion': p.descripcion,
            'precio': p.precio,
            'cantidad': p.cantidad,
            'categoria': p.categoria,
            'fecha_creacion': p.fecha_creacion.isoformat()
        })
    
    response = make_response(output.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=productos.csv'
    response.headers['Content-Type'] = 'text/csv'
    return response


@app.route('/api/productos/exportar-excel', methods=['GET'])
@login_required
def exportar_excel():
    """Exportar productos a XLSX con imágenes incrustadas"""
    try:
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        import requests
        from PIL import Image
    except ImportError:
        return jsonify({'error': 'Dependencias faltantes (openpyxl, Pillow)'}), 500

    productos = Producto.query.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Catálogo"

    # Header con estilos
    headers = ['ID', 'Nombre', 'Descripción', 'Categoría', 'Precio', 'Stock', 'Imagen', 'Fecha Creación']
    ws.append(headers)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Anchos de columnas
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 18

    # Datos
    for idx, p in enumerate(productos, start=2):
        ws[f'A{idx}'] = p.id
        ws[f'B{idx}'] = p.nombre
        ws[f'C{idx}'] = p.descripcion or ''
        ws[f'D{idx}'] = p.categoria or ''
        ws[f'E{idx}'] = p.precio
        ws[f'F{idx}'] = p.cantidad
        ws[f'H{idx}'] = p.fecha_creacion.isoformat() if p.fecha_creacion else ''

        # Descargar e insertar imagen si existe
        if p.imagen_url:
            try:
                # Si es URL completa, descargar; si es ruta local, usarla directo
                if p.imagen_url.startswith('http'):
                    img_response = requests.get(p.imagen_url, timeout=5)
                    img_data = BytesIO(img_response.content)
                else:
                    # Ruta local (relative a UPLOAD_FOLDER)
                    img_path = os.path.join(UPLOAD_FOLDER, os.path.basename(p.imagen_url))
                    if os.path.exists(img_path):
                        img_data = img_path
                    else:
                        img_data = None

                if img_data:
                    # Insertar imagen redimensionada
                    if isinstance(img_data, BytesIO):
                        pil_img = Image.open(img_data)
                    else:
                        pil_img = Image.open(img_data)

                    # Redimensionar a máx 200x200 para que quepan en Excel
                    pil_img.thumbnail((200, 200), Image.Resampling.LANCZOS)

                    # Guardar en BytesIO
                    img_bytes = BytesIO()
                    pil_img.save(img_bytes, format='PNG')
                    img_bytes.seek(0)

                    # Insertar en Excel
                    xl_img = XLImage(img_bytes)
                    xl_img.width = 150
                    xl_img.height = 150
                    ws.add_image(xl_img, f'G{idx}')
                    ws.row_dimensions[idx].height = 120
            except Exception as e:
                # Si falla descargar, dejar URL como texto
                ws[f'G{idx}'] = p.imagen_url
                print(f"Warning: No se pudo insertar imagen de {p.nombre}: {e}")
        else:
            ws[f'G{idx}'] = 'Sin imagen'

    # Guardar a BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = make_response(output.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=catalogo_productos.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response


@app.route('/api/productos/importar-excel', methods=['POST'])
@login_required
def importar_excel():
    """Importar productos desde Excel (CLAVES.xlsx)
    Mapea: Columna C (Clave) -> nombre, Columna F (Producto) -> descripción
    """
    if not is_admin_user():
        return jsonify({'error': 'Solo admins pueden importar'}), 403
    
    try:
        import openpyxl
    except ImportError:
        return jsonify({'error': 'openpyxl no instalado'}), 500
    
    # Verificar si hay archivo en request
    if 'file' not in request.files:
        return jsonify({'error': 'No se encontró archivo'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Archivo vacío'}), 400
    
    if not file.filename.lower().endswith('.xlsx'):
        return jsonify({'error': 'Solo se aceptan archivos .xlsx'}), 400
    
    try:
        # Cargar workbook
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        
        # Mapear columnas: Clave (C=3), Producto (F=6)
        # Row 1 es encabezado, comenzar desde row 2
        
        stats = {
            'creados': 0,
            'actualizados': 0,
            'errores': 0,
            'detalles': []
        }
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # row es tupla con valores: índice 2 = columna C (Clave), índice 5 = columna F (Producto)
                clave = row[2]  # Columna C
                descripcion = row[5]  # Columna F
                
                # Validar que ambos campos existan
                if not clave or not descripcion:
                    stats['detalles'].append(f"Fila {row_num}: Falta Clave o Producto")
                    continue
                
                # Convertir a string y limpiar
                clave = str(clave).strip()
                descripcion = str(descripcion).strip()
                
                # Buscar si el producto ya existe por nombre (clave)
                producto = Producto.query.filter_by(nombre=clave).first()
                
                if producto:
                    # ACTUALIZAR
                    producto.descripcion = descripcion
                    stats['actualizados'] += 1
                    stats['detalles'].append(f"Actualizado: {clave}")
                else:
                    # CREAR NUEVO
                    nuevo_producto = Producto(
                        nombre=clave,
                        descripcion=descripcion,
                        precio=0.0,
                        cantidad=0,
                        categoria='Importado',
                        imagen_url=None
                    )
                    db.session.add(nuevo_producto)
                    stats['creados'] += 1
                    stats['detalles'].append(f"Creado: {clave}")
                
            except Exception as e:
                stats['errores'] += 1
                stats['detalles'].append(f"Fila {row_num}: Error - {str(e)}")
                continue
        
        # Guardar cambios
        try:
            db.session.commit()
            stats['mensaje'] = 'Importación completada'
        except Exception as e:
            db.session.rollback()
            stats['mensaje'] = f'Error al guardar: {str(e)}'
            stats['errores'] += 1
        
        return jsonify(stats), 200
    
    except Exception as e:
        return jsonify({'error': f'Error procesando Excel: {str(e)}'}), 500


@app.route('/api/procesos/importar-excel', methods=['POST'])
@login_required
def importar_procesos_excel():
    """Inicia importación asíncrona de claves/procesos desde Excel/CSV."""
    if not is_admin_user():
        return jsonify({'error': 'Solo admins pueden importar procesos'}), 403

    if 'file' not in request.files:
        return jsonify({'error': 'No se encontró archivo'}), 400

    file = request.files['file']
    if not file or file.filename == '':
        return jsonify({'error': 'Archivo vacío'}), 400

    filename = secure_filename(file.filename)
    ext = os.path.splitext(filename)[1].lower()
    if ext not in ('.xlsx', '.xls', '.csv'):
        return jsonify({'error': 'Formato no soportado. Usa .xlsx, .xls o .csv'}), 400

    sheet = (request.form.get('sheet') or '').strip()
    imports_dir = os.path.join('uploads', 'imports')
    os.makedirs(imports_dir, exist_ok=True)

    tmp_name = f"procesos_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}{ext}"
    saved_path = os.path.join(imports_dir, tmp_name)
    abs_saved_path = os.path.abspath(saved_path)
    job_id = uuid.uuid4().hex

    try:
        file.save(abs_saved_path)

        _set_procesos_import_job(
            job_id,
            status='queued',
            filename=filename,
            created_at=datetime.utcnow().isoformat(),
            output='',
        )

        worker = threading.Thread(
            target=_run_procesos_import_job,
            args=(job_id, filename, abs_saved_path, sheet),
            daemon=True,
        )
        worker.start()

        return jsonify({
            'ok': True,
            'job_id': job_id,
            'mensaje': 'Importación iniciada'
        }), 202
    except Exception as e:
        try:
            if os.path.exists(abs_saved_path):
                os.remove(abs_saved_path)
        except Exception:
            pass
        logger.error(f"[IMPORT_PROCESOS] Excepción importando {filename}: {e}")
        return jsonify({'error': f'Error procesando importación: {str(e)}'}), 500


@app.route('/api/procesos/importar-excel/status/<job_id>', methods=['GET'])
@login_required
def importar_procesos_excel_status(job_id):
    if not is_admin_user():
        return jsonify({'error': 'Solo admins pueden consultar importaciones'}), 403

    job = _get_procesos_import_job(job_id)

    if not job:
        return jsonify({'error': 'Job no encontrado'}), 404

    return jsonify(job), 200


@app.route('/api/contpaq/indice/upload', methods=['POST'])
@login_required
@requires_permission('contpaq', 'edit')
def contpaq_indice_upload():
    if 'file' not in request.files:
        return jsonify({'error': 'No se encontró archivo'}), 400

    file = request.files['file']
    if not file or file.filename == '':
        return jsonify({'error': 'Archivo vacío'}), 400

    filename = secure_filename(file.filename)
    ext = os.path.splitext(filename)[1].lower()
    if ext not in ('.xlsx', '.xls', '.csv'):
        return jsonify({'error': 'Formato no soportado. Usa .xlsx, .xls o .csv'}), 400

    try:
        df = _contpaq_read_indice_dataframe(file, filename, ext, request.form.get('sheet', '').strip())
    except Exception as exc:
        return jsonify({'error': f'No se pudo leer archivo: {exc}'}), 400

    if df.empty:
        return jsonify({'error': 'El archivo no contiene filas'}), 400

    semana_col = _contpaq_detect_import_column(df.columns, ['SEMANA'])
    clave_col = _contpaq_detect_import_column(df.columns, ['CLAVE', 'DETALLECLAVE', 'CLAVEPRODUCTO'])
    cantidad_col = _contpaq_detect_import_column(df.columns, ['CANTIDAD', 'DETALLECANTIDAD'])
    sucursal_col = _contpaq_detect_import_column(df.columns, ['SUCURSAL'])
    descripcion_col = _contpaq_detect_import_column(df.columns, ['DESCRIPCION', 'DETALLEDESCRIPCION'])
    folio_col = _contpaq_detect_import_column(df.columns, ['FOLIO'])
    fecha_col = _contpaq_detect_import_column(df.columns, ['FECHA', 'FECHADOCUMENTO'])

    missing = []
    if not semana_col:
        missing.append('SEMANA')
    if not clave_col:
        missing.append('CLAVE')
    if not cantidad_col:
        missing.append('CANTIDAD')
    if not sucursal_col:
        missing.append('SUCURSAL')

    if missing:
        return jsonify({'error': f'Faltan columnas obligatorias: {", ".join(missing)}'}), 400

    imported = 0
    db.session.query(ContpaqSucursalIndice).delete()

    for _, row in df.iterrows():
        semana = str(row.get(semana_col) or '').strip()
        clave = str(row.get(clave_col) or '').strip().upper()
        cantidad = str(row.get(cantidad_col) or '').strip()
        sucursal = str(row.get(sucursal_col) or '').strip()
        if not any([semana, clave, cantidad, sucursal]):
            continue

        item = ContpaqSucursalIndice(
            semana=semana,
            sucursal=sucursal,
            clave_producto=clave,
            descripcion=str(row.get(descripcion_col) or '').strip() if descripcion_col else '',
            cantidad=cantidad,
            folio=str(row.get(folio_col) or '').strip() if folio_col else '',
            fecha_documento=str(row.get(fecha_col) or '').strip() if fecha_col else '',
            source_filename=filename,
            raw_payload=json.dumps({str(col): str(row.get(col) or '') for col in df.columns}, ensure_ascii=False),
        )
        db.session.add(item)
        imported += 1

    db.session.commit()
    return jsonify({
        'ok': True,
        'imported': imported,
        'filename': filename,
        'detected_columns': {
            'semana': semana_col,
            'clave': clave_col,
            'cantidad': cantidad_col,
            'sucursal': sucursal_col,
            'descripcion': descripcion_col,
            'folio': folio_col,
            'fecha': fecha_col,
        }
    }), 200


@app.route('/api/contpaq/indice/summary', methods=['GET'])
@login_required
@requires_permission('contpaq', 'view')
def contpaq_indice_summary():
    total = ContpaqSucursalIndice.query.count()
    latest = ContpaqSucursalIndice.query.order_by(ContpaqSucursalIndice.imported_at.desc()).first()
    return jsonify({
        'total': total,
        'latest': latest.to_dict() if latest else None,
    }), 200


@app.route('/api/contpaq/indice/clear', methods=['POST'])
@login_required
@requires_permission('contpaq', 'edit')
def contpaq_indice_clear():
    try:
        deleted = ContpaqSucursalIndice.query.delete(synchronize_session=False)
        db.session.commit()
        return jsonify({'ok': True, 'deleted': int(deleted)}), 200
    except Exception as exc:
        db.session.rollback()
        logger.error(f"Error limpiando indice CONTPAQ: {exc}", exc_info=True)
        return jsonify({'error': str(exc)}), 500


@app.route('/api/contpaq/precio-publico/upload', methods=['POST'])
@login_required
@requires_permission('contpaq', 'edit')
def contpaq_precio_publico_upload():
    if 'file' not in request.files:
        return jsonify({'error': 'No se encontró archivo'}), 400

    file = request.files['file']
    if not file or file.filename == '':
        return jsonify({'error': 'Archivo vacío'}), 400

    filename = secure_filename(file.filename)
    ext = os.path.splitext(filename)[1].lower()
    if ext not in ('.xlsx', '.xls'):
        return jsonify({'error': 'Formato no soportado. Usa .xlsx o .xls'}), 400

    sheet_name = (request.form.get('sheet') or 'Productos').strip() or 'Productos'
    clave_col = (request.form.get('clave_col') or 'D').strip() or 'D'
    precio_col = (request.form.get('precio_col') or 'S').strip() or 'S'
    start_row = request.form.get('start_row', default=3, type=int) or 3

    try:
        rows = _contpaq_read_precio_publico_rows(
            file, filename, ext,
            sheet_name=sheet_name,
            start_row=start_row,
            clave_col=clave_col,
            precio_col=precio_col,
        )
    except Exception as exc:
        return jsonify({'error': f'No se pudo leer archivo: {exc}'}), 400

    if not rows:
        return jsonify({'error': 'No se encontraron claves/precios válidos en el archivo'}), 400

    imported = 0
    db.session.query(ContpaqPrecioPublico).delete()

    for row in rows:
        item = ContpaqPrecioPublico(
            clave_producto=row['clave_producto'],
            precio_publico=row['precio_publico'],
            source_filename=filename,
            source_sheet=sheet_name,
            raw_payload=json.dumps(row['raw_payload'], ensure_ascii=False),
        )
        db.session.add(item)
        imported += 1

    db.session.commit()
    return jsonify({
        'ok': True,
        'imported': imported,
        'filename': filename,
        'config': {
            'sheet': sheet_name,
            'start_row': start_row,
            'clave_col': clave_col,
            'precio_col': precio_col,
        }
    }), 200


@app.route('/api/contpaq/precio-publico/summary', methods=['GET'])
@login_required
@requires_permission('contpaq', 'view')
def contpaq_precio_publico_summary():
    total = ContpaqPrecioPublico.query.count()
    latest = ContpaqPrecioPublico.query.order_by(ContpaqPrecioPublico.imported_at.desc()).first()
    return jsonify({
        'total': total,
        'latest': latest.to_dict() if latest else None,
    }), 200


@app.route('/api/contpaq/precio-publico/clear', methods=['POST'])
@login_required
@requires_permission('contpaq', 'edit')
def contpaq_precio_publico_clear():
    try:
        deleted = ContpaqPrecioPublico.query.delete(synchronize_session=False)
        db.session.commit()
        return jsonify({'ok': True, 'deleted': int(deleted)}), 200
    except Exception as exc:
        db.session.rollback()
        logger.error(f"Error limpiando precios publicos CONTPAQ: {exc}", exc_info=True)
        return jsonify({'error': str(exc)}), 500


@app.route('/api/productos/bajo-stock', methods=['GET'])
@login_required
def bajo_stock():
    """Obtener productos con bajo stock (< 5 unidades)"""
    productos = Producto.query.filter(Producto.cantidad < 5).all()
    return jsonify([p.to_dict() for p in productos])

@app.route('/api/categorias', methods=['GET'])
def obtener_categorias():
    """Obtener todas las categorías únicas"""
    categorias = db.session.query(Producto.categoria).distinct().filter(
        Producto.categoria != None
    ).all()
    return jsonify([cat[0] for cat in categorias])


@app.route('/api/logs', methods=['GET'])
@login_required
def get_logs():
    """Obtener logs de acceso (solo admin). Parámetro opcional: limit (default 50)"""
    if not is_admin_user():
        return jsonify({'error': 'Prohibido'}), 403

    limit = min(int(request.args.get('limit', 50)), 500)
    try:
        logs = db.session.query('access_logs').from_statement(db.text('SELECT * FROM access_logs ORDER BY timestamp DESC LIMIT :lim')).params(lim=limit).all()
        # Fallback to ORM query if direct text not supported
    except Exception:
        from models import AccessLog
        logs = AccessLog.query.order_by(AccessLog.timestamp.desc()).limit(limit).all()

    # Convert to dicts
    result = []
    for l in logs:
        try:
            result.append(l.to_dict())
        except Exception:
            # if row is a tuple from raw query, try mapping
            try:
                d = {
                    'id': l.id,
                    'ip': l.ip,
                    'username': l.username,
                    'path': l.path,
                    'method': l.method,
                    'user_agent': l.user_agent,
                    'referer': l.referer,
                    'timestamp': l.timestamp.isoformat() if hasattr(l, 'timestamp') else None
                }
                result.append(d)
            except Exception:
                continue

    return jsonify(result)


@app.route('/admin/puerta')
@login_required
def puerta():
    """Vista web exclusiva para el usuario 'root' con registros de acceso."""
    # Permitimos ver la 'Puerta' a cualquier administrador (es_admin)
    if not is_admin_user():
        return render_template('login.html', error='Acceso restringido'), 403

    from models import AccessLog
    limit = min(int(request.args.get('limit', 200)), 2000)
    logs = AccessLog.query.order_by(AccessLog.timestamp.desc()).limit(limit).all()
    # convert timestamps to ISO for template
    logs_serialized = []
    for l in logs:
        logs_serialized.append({
            'timestamp': l.timestamp.isoformat(),
            'username': l.username,
            'ip': l.ip,
            'path': l.path,
            'method': l.method,
            'user_agent': l.user_agent,
            'referer': l.referer
        })
    return render_template('puerta.html', logs=logs_serialized)


# ==================== ENDPOINTS DE PROVEEDORES ====================

# GET - Obtener todos los proveedores
@app.route('/api/proveedores', methods=['GET'])
@login_required
def get_proveedores():
    proveedores = Proveedor.query.all()
    return jsonify([p.to_dict() for p in proveedores])

# GET - Obtener proveedor por ID
@app.route('/api/proveedores/<int:id>', methods=['GET'])
@login_required
def get_proveedor(id):
    proveedor = Proveedor.query.get(id)
    if not proveedor:
        return jsonify({'error': 'Proveedor no encontrado'}), 404
    return jsonify(proveedor.to_dict())

# POST - Crear proveedor
@app.route('/api/proveedores', methods=['POST'])
@login_required
def crear_proveedor():
    data = request.get_json()
    
    if not data or not data.get('nombre'):
        return jsonify({'error': 'Nombre es obligatorio'}), 400
    
    # Verificar si ya existe
    existente = Proveedor.query.filter_by(nombre=data.get('nombre')).first()
    if existente:
        return jsonify({'error': 'El proveedor ya existe'}), 400
    
    nuevo_proveedor = Proveedor(
        nombre=data.get('nombre'),
        telefono=data.get('telefono', ''),
        rfc=data.get('rfc', ''),
        domicilio=data.get('domicilio', ''),
        correo=data.get('correo', ''),
        contacto=data.get('contacto', ''),
        notas=data.get('notas', '')
    )
    
    db.session.add(nuevo_proveedor)
    db.session.commit()
    
    return jsonify(nuevo_proveedor.to_dict()), 201

# PUT - Actualizar proveedor
@app.route('/api/proveedores/<int:id>', methods=['PUT'])
@login_required
def actualizar_proveedor(id):
    proveedor = Proveedor.query.get(id)
    
    if not proveedor:
        return jsonify({'error': 'Proveedor no encontrado'}), 404
    
    data = request.get_json()
    
    proveedor.nombre = data.get('nombre', proveedor.nombre)
    proveedor.telefono = data.get('telefono', proveedor.telefono)
    proveedor.rfc = data.get('rfc', proveedor.rfc)
    proveedor.domicilio = data.get('domicilio', proveedor.domicilio)
    proveedor.correo = data.get('correo', proveedor.correo)
    proveedor.contacto = data.get('contacto', proveedor.contacto)
    proveedor.notas = data.get('notas', proveedor.notas)
    
    db.session.commit()
    
    return jsonify(proveedor.to_dict())

# DELETE - Eliminar proveedor
@app.route('/api/proveedores/<int:id>', methods=['DELETE'])
@login_required
def eliminar_proveedor(id):
    proveedor = Proveedor.query.get(id)
    
    if not proveedor:
        return jsonify({'error': 'Proveedor no encontrado'}), 404
    
    db.session.delete(proveedor)
    db.session.commit()
    
    return jsonify({'mensaje': 'Proveedor eliminado correctamente'})

# ==================== ENDPOINTS DE PRODUCTO-PROVEEDOR ====================

# GET - Obtener proveedores de un producto
@app.route('/api/productos/<int:producto_id>/proveedores', methods=['GET'])
@login_required
def get_proveedores_producto(producto_id):
    producto = Producto.query.get(producto_id)
    if not producto:
        return jsonify({'error': 'Producto no encontrado'}), 404
    
    return jsonify([pp.to_dict() for pp in producto.proveedores])

# POST - Asignar proveedor a producto
@app.route('/api/productos/<int:producto_id>/proveedores', methods=['POST'])
@login_required
def asignar_proveedor(producto_id):
    data = request.get_json()
    
    producto = Producto.query.get(producto_id)
    if not producto:
        return jsonify({'error': 'Producto no encontrado'}), 404
    
    proveedor = Proveedor.query.get(data.get('proveedor_id'))
    if not proveedor:
        return jsonify({'error': 'Proveedor no encontrado'}), 404
    
    # Verificar si ya existe la relación
    existente = ProductoProveedor.query.filter_by(
        producto_id=producto_id,
        proveedor_id=data.get('proveedor_id')
    ).first()
    
    if existente:
        # Actualizar precio si existe
        existente.precio_proveedor = float(data.get('precio_proveedor', existente.precio_proveedor))
        if data.get('fecha_precio'):
            existente.fecha_precio = datetime.strptime(data.get('fecha_precio'), '%Y-%m-%d').date()
        db.session.commit()
        return jsonify(existente.to_dict()), 200
    
    # Crear nueva relación
    from datetime import datetime as dt
    fecha_precio = data.get('fecha_precio')
    if fecha_precio:
        fecha_precio = datetime.strptime(fecha_precio, '%Y-%m-%d').date()
    else:
        fecha_precio = dt.now().date()
    
    nuevo_asignar = ProductoProveedor(
        producto_id=producto_id,
        proveedor_id=data.get('proveedor_id'),
        precio_proveedor=float(data.get('precio_proveedor')),
        fecha_precio=fecha_precio,
        cantidad_minima=int(data.get('cantidad_minima', 1))
    )
    
    db.session.add(nuevo_asignar)
    db.session.commit()
    
    return jsonify(nuevo_asignar.to_dict()), 201

# DELETE - Desasignar proveedor de producto
@app.route('/api/productos/<int:producto_id>/proveedores/<int:proveedor_id>', methods=['DELETE'])
@login_required
def desasignar_proveedor(producto_id, proveedor_id):
    asignacion = ProductoProveedor.query.filter_by(
        producto_id=producto_id,
        proveedor_id=proveedor_id
    ).first()
    
    if not asignacion:
        return jsonify({'error': 'Asignación no encontrada'}), 404
    
    db.session.delete(asignacion)
    db.session.commit()
    
    return jsonify({'mensaje': 'Proveedor desasignado correctamente'})


# ==================== SYNC PRECIOS COMPRA (PLANTA) ====================

@app.route('/api/precios_compra_sync', methods=['POST'])
def precios_compra_sync():
    ok, err = _require_sync_key()
    if not ok:
        return err

    payload = request.get_json(silent=True)
    if payload is None:
        return jsonify({'error': 'JSON requerido'}), 400

    items = payload
    if isinstance(payload, dict):
        items = payload.get('items', [])

    if not isinstance(items, list):
        return jsonify({'error': 'Formato invalido, se espera lista'}), 400

    stats = {
        'creados_producto': 0,
        'creados_proveedor': 0,
        'actualizados_precio': 0,
        'creados_historial': 0,
        'ignorados': 0,
        'errores': 0,
        'errores_muestra': []
    }

    for idx, item in enumerate(items, start=1):
        try:
            product_key = (item.get('ProductKey') or item.get('product_key') or '').strip()
            descripcion = (item.get('Description') or item.get('descripcion') or '').strip()
            proveedor_nombre = (item.get('BusinessEntityName') or item.get('proveedor') or '').strip()
            precio = item.get('UnitPrice') if 'UnitPrice' in item else item.get('precio')
            divisa = (item.get('Currency') or item.get('Divisa') or item.get('divisa') or '').strip()
            fecha_documento = _parse_date(item.get('DateDocument') or item.get('fecha_documento'))

            if not product_key or precio is None:
                stats['ignorados'] += 1
                continue

            # Enforce column sizes
            if len(product_key) > 100:
                product_key = product_key[:100]
            if len(proveedor_nombre) > 255:
                proveedor_nombre = proveedor_nombre[:255]
            if len(divisa) > 10:
                divisa = divisa[:10]

            try:
                precio = float(precio)
            except (TypeError, ValueError):
                stats['errores'] += 1
                continue
            if not math.isfinite(precio):
                stats['errores'] += 1
                continue

            producto = Producto.query.filter_by(clave=product_key).first()
            if not producto:
                producto = Producto(
                    clave=product_key,
                    nombre=descripcion or product_key,
                    descripcion=descripcion,
                    precio=precio,
                    cantidad=0,
                    categoria='Compras'
                )
                db.session.add(producto)
                stats['creados_producto'] += 1

            proveedor = None
            if proveedor_nombre:
                proveedor = Proveedor.query.filter_by(nombre=proveedor_nombre).first()
                if not proveedor:
                    proveedor = Proveedor(nombre=proveedor_nombre)
                    db.session.add(proveedor)
                    stats['creados_proveedor'] += 1

            if not proveedor:
                stats['ignorados'] += 1
                continue

            asignacion = ProductoProveedor.query.filter_by(
                producto_id=producto.id,
                proveedor_id=proveedor.id
            ).first()

            if not asignacion:
                asignacion = ProductoProveedor(
                    producto_id=producto.id,
                    proveedor_id=proveedor.id,
                    precio_proveedor=precio,
                    fecha_precio=fecha_documento or datetime.utcnow().date(),
                    divisa=divisa or None,
                    cantidad_minima=1
                )
                db.session.add(asignacion)
                stats['actualizados_precio'] += 1
            else:
                if fecha_documento and asignacion.fecha_precio and fecha_documento < asignacion.fecha_precio:
                    # Only fill divisa if missing, do not downgrade price/date
                    if divisa and not asignacion.divisa:
                        asignacion.divisa = divisa
                        stats['actualizados_precio'] += 1
                    else:
                        stats['ignorados'] += 1
                    continue
                asignacion.precio_proveedor = precio
                if fecha_documento:
                    asignacion.fecha_precio = fecha_documento
                if divisa:
                    asignacion.divisa = divisa
                stats['actualizados_precio'] += 1

            fecha_hist = fecha_documento or asignacion.fecha_precio or datetime.utcnow().date()
            existente_hist = HistorialPreciosProveedor.query.filter_by(
                producto_proveedor_id=asignacion.id,
                precio=precio,
                fecha_precio=fecha_hist
            ).first()
            if not existente_hist:
                db.session.add(HistorialPreciosProveedor(
                    producto_proveedor_id=asignacion.id,
                    precio=precio,
                    fecha_precio=fecha_hist,
                    notas='sync_planta',
                    divisa=divisa or None
                ))
                stats['creados_historial'] += 1

        except Exception as e:
            try:
                db.session.rollback()
            except Exception:
                pass
            stats['errores'] += 1
            if len(stats['errores_muestra']) < 5:
                stats['errores_muestra'].append(str(e))
            continue

    db.session.commit()
    return jsonify({'ok': True, 'stats': stats})

# ==================== ENDPOINTS DE HISTORIAL DE PRECIOS ====================

# GET - Obtener historial de precios de un proveedor en un producto
@app.route('/api/productos/<int:producto_id>/proveedores/<int:proveedor_id>/historial', methods=['GET'])
@login_required
def get_historial_precios(producto_id, proveedor_id):
    asignacion = ProductoProveedor.query.filter_by(
        producto_id=producto_id,
        proveedor_id=proveedor_id
    ).first()
    
    if not asignacion:
        return jsonify({'error': 'Asignación no encontrada'}), 404
    
    historial = HistorialPreciosProveedor.query.filter_by(
        producto_proveedor_id=asignacion.id
    ).order_by(HistorialPreciosProveedor.fecha_precio.desc()).all()
    
    return jsonify([h.to_dict() for h in historial])

# POST - Agregar precio histórico
@app.route('/api/productos/<int:producto_id>/proveedores/<int:proveedor_id>/historial', methods=['POST'])
@login_required
def agregar_precio_historico(producto_id, proveedor_id):
    asignacion = ProductoProveedor.query.filter_by(
        producto_id=producto_id,
        proveedor_id=proveedor_id
    ).first()
    
    if not asignacion:
        return jsonify({'error': 'Asignación no encontrada'}), 404
    
    data = request.get_json()
    
    if not data.get('precio'):
        return jsonify({'error': 'El precio es requerido'}), 400
    
    if not data.get('fecha_precio'):
        return jsonify({'error': 'La fecha es requerida'}), 400
    
    try:
        fecha_precio = datetime.strptime(data.get('fecha_precio'), '%Y-%m-%d').date()
        divisa = data.get('divisa')
        
        # Crear nuevo registro de precio histórico
        nuevo_precio = HistorialPreciosProveedor(
            producto_proveedor_id=asignacion.id,
            precio=float(data.get('precio')),
            fecha_precio=fecha_precio,
            notas=data.get('notas', ''),
            divisa=divisa
        )
        
        # Actualizar el precio actual en ProductoProveedor
        asignacion.precio_proveedor = float(data.get('precio'))
        asignacion.fecha_precio = fecha_precio
        if divisa:
            asignacion.divisa = divisa
        
        db.session.add(nuevo_precio)
        db.session.commit()
        
        return jsonify({
            'mensaje': 'Precio agregado al historial',
            'precio_historico': nuevo_precio.to_dict()
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Error: {str(e)}'}), 500

# DELETE - Eliminar precio histórico
@app.route('/api/historial-precios/<int:precio_id>', methods=['DELETE'])
@login_required
def eliminar_precio_historico(precio_id):
    precio = HistorialPreciosProveedor.query.get(precio_id)
    
    if not precio:
        return jsonify({'error': 'Registro de precio no encontrado'}), 404
    
    db.session.delete(precio)
    db.session.commit()
    
    return jsonify({'mensaje': 'Precio histórico eliminado correctamente'})

# ==================== ENDPOINTS DE CARGA DE IMÁGENES ====================

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# POST - Subir imagen para producto
@app.route('/api/productos/upload-imagen', methods=['POST'])
@login_required
def upload_imagen():
    if 'imagen' not in request.files:
        return jsonify({'error': 'No se envió archivo'}), 400
    
    file = request.files['imagen']
    
    if file.filename == '':
        return jsonify({'error': 'No se seleccionó archivo'}), 400
    
    if not allowed_file(file.filename):
        return jsonify({'error': 'Formato de archivo no permitido. Usa: png, jpg, jpeg, gif, webp'}), 400
    
    try:
        # Generar nombre seguro con timestamp
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_')
        filename = timestamp + filename
        
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        # Retornar path relativo para acceso
        image_url = f'/uploads/productos/{filename}'
        
        return jsonify({
            'mensaje': 'Imagen subida exitosamente',
            'url': image_url,
            'filename': filename
        }), 201
    
    except Exception as e:
        return jsonify({'error': f'Error al subir imagen: {str(e)}'}), 500

# GET - Servir imagen (ruta para acceso de archivos)
@app.route('/uploads/productos/<filename>')
def descargar_imagen(filename):
    from flask import send_from_directory
    try:
        return send_from_directory(app.config['UPLOAD_FOLDER'], filename)
    except Exception as e:
        return jsonify({'error': 'Imagen no encontrada'}), 404


# ==================== RUTAS DE TICKETS ====================

# 1. POST /api/tickets (PÚBLICO - sin login)
@app.route('/api/tickets', methods=['POST'])
def crear_ticket():
    """Crear ticket (nombre_solicitante, email, departamento, titulo, descripcion)"""
    try:
        data = request.get_json()
        
        # Validar campos requeridos
        campos_requeridos = ['nombre_solicitante', 'email', 'departamento', 'titulo', 'descripcion']
        for campo in campos_requeridos:
            if not data.get(campo):
                return jsonify({'error': f'Campo requerido: {campo}'}), 400
        
        # Generar número único de ticket
        numero_ticket = f"TKT-{int(time() * 1000)}"
        
        # Crear ticket
        nuevo_ticket = Ticket(
            numero_ticket=numero_ticket,
            nombre_solicitante=data.get('nombre_solicitante'),
            email_solicitante=data.get('email'),
            departamento=data.get('departamento'),
            titulo=data.get('titulo'),
            descripcion=data.get('descripcion'),
            estado='nuevo',
            prioridad=data.get('prioridad', 'media'),
            categoria=data.get('categoria', 'general'),
            fecha_creacion=datetime.utcnow()
        )
        
        db.session.add(nuevo_ticket)
        db.session.commit()
        
        logger.info(f"✓ Ticket creado: {numero_ticket} por {data.get('nombre_solicitante')}")
        
        return jsonify({
            'id': nuevo_ticket.id,
            'numero_ticket': numero_ticket,
            'mensaje': 'Ticket creado exitosamente'
        }), 201
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error al crear ticket: {e}")
        return jsonify({'error': str(e)}), 500


# 2. GET /api/bandeja-entrada (LOGIN REQUERIDO)
@app.route('/api/bandeja-entrada', methods=['GET'])
@ingeniero_login_required
def bandeja_entrada():
    """Ver todos los tickets con estado 'nuevo' (sin ingeniero_id)"""
    try:
        # Obtener tickets sin asignar
        tickets = Ticket.query.filter_by(estado='nuevo', ingeniero_id=None).order_by(Ticket.fecha_creacion.desc()).all()
        
        return jsonify({
            'tickets': [
                {
                    'id': t.id,
                    'numero_ticket': t.numero_ticket,
                    'titulo': t.titulo,
                    'nombre_solicitante': t.nombre_solicitante,
                    'email_solicitante': t.email_solicitante,
                    'departamento': t.departamento,
                    'estado': t.estado,
                    'prioridad': t.prioridad,
                    'categoria': t.categoria,
                    'fecha_creacion': t.fecha_creacion.isoformat() if t.fecha_creacion else None
                }
                for t in tickets
            ]
        }), 200
        
    except Exception as e:
        logger.error(f"Error en bandeja_entrada: {e}")
        return jsonify({'error': str(e)}), 500


# 3. POST /api/tickets/<id>/tomar (LOGIN REQUERIDO)
@app.route('/api/tickets/<int:ticket_id>/tomar', methods=['POST'])
@ingeniero_login_required
def tomar_ticket(ticket_id):
    """Ingeniero 'toma' el ticket"""
    try:
        usuario = Usuario.query.filter_by(username=session.get('ingeniero_user')).first()
        if not usuario:
            return jsonify({'error': 'Usuario no encontrado'}), 401
        
        ticket = Ticket.query.get(ticket_id)
        if not ticket:
            return jsonify({'error': 'Ticket no encontrado'}), 404
        
        if ticket.estado != 'nuevo':
            return jsonify({'error': 'El ticket no está disponible'}), 400
        
        # Asignar al usuario actual
        ticket.ingeniero_id = usuario.id
        ticket.estado = 'en_progreso'
        ticket.fecha_asignacion = datetime.utcnow()
        
        db.session.commit()
        
        logger.info(f"✓ Ticket {ticket.numero_ticket} asignado a {usuario.username}")
        
        return jsonify({
            'mensaje': 'Ticket asignado exitosamente',
            'numero_ticket': ticket.numero_ticket
        }), 200
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error al tomar ticket: {e}")
        return jsonify({'error': str(e)}), 500


# 4. GET /api/mis-tickets (LOGIN REQUERIDO)
@app.route('/api/mis-tickets', methods=['GET'])
@ingeniero_login_required
def mis_tickets():
    """Ver tickets donde ingeniero_id = usuario actual (en_progreso o resuelto)"""
    try:
        usuario = Usuario.query.filter_by(username=session.get('ingeniero_user')).first()
        if not usuario:
            return jsonify({'error': 'Usuario no encontrado'}), 401
        
        tickets = Ticket.query.filter(
            Ticket.ingeniero_id == usuario.id,
            Ticket.estado.in_(['en_progreso', 'resuelto'])
        ).order_by(Ticket.fecha_creacion.desc()).all()
        
        resultado = []
        for t in tickets:
            comentarios = ComentarioTicket.query.filter_by(ticket_id=t.id).all()
            resultado.append({
                'id': t.id,
                'numero_ticket': t.numero_ticket,
                'titulo': t.titulo,
                'nombre_solicitante': t.nombre_solicitante,
                'email_solicitante': t.email_solicitante,
                'departamento': t.departamento,
                'descripcion': t.descripcion,
                'estado': t.estado,
                'prioridad': t.prioridad,
                'categoria': t.categoria,
                'fecha_creacion': t.fecha_creacion.isoformat() if t.fecha_creacion else None,
                'fecha_asignacion': t.fecha_asignacion.isoformat() if t.fecha_asignacion else None,
                'fecha_resolucion': t.fecha_resolucion.isoformat() if t.fecha_resolucion else None,
                'comentarios_count': len(comentarios),
                'comentarios': [
                    {
                        'id': c.id,
                        'contenido': c.contenido,
                        'imagen_url': c.imagen_url,
                        'fecha_creacion': c.fecha_creacion.isoformat() if c.fecha_creacion else None
                    }
                    for c in comentarios
                ]
            })
        
        return jsonify({'tickets': resultado}), 200
        
    except Exception as e:
        logger.error(f"Error en mis_tickets: {e}")
        return jsonify({'error': str(e)}), 500


# 5. GET /api/tickets/<id> (LOGIN REQUERIDO)
@app.route('/api/tickets/<int:ticket_id>', methods=['GET'])
@ingeniero_login_required
def obtener_ticket(ticket_id):
    """Ver detalles de un ticket con todos sus comentarios"""
    try:
        usuario = Usuario.query.filter_by(username=session.get('ingeniero_user')).first()
        if not usuario:
            return jsonify({'error': 'Usuario no encontrado'}), 401
        
        ticket = Ticket.query.get(ticket_id)
        if not ticket:
            return jsonify({'error': 'Ticket no encontrado'}), 404
        
        # Verificar permisos: solicitante, ingeniero o admin
        es_solicitante = ticket.email_solicitante
        es_ingeniero = ticket.ingeniero_id == usuario.id
        es_admin = is_admin_user()
        
        if not (es_solicitante or es_ingeniero or es_admin):
            return jsonify({'error': 'Acceso denegado'}), 403
        
        comentarios = ComentarioTicket.query.filter_by(ticket_id=ticket_id).all()
        
        # Obtener nombre del ingeniero asignado
        ingeniero_nombre = None
        if ticket.ingeniero_id:
            ing = Usuario.query.get(ticket.ingeniero_id)
            ingeniero_nombre = ing.username if ing else None
        
        return jsonify({
            'ticket': {
                'id': ticket.id,
                'numero_ticket': ticket.numero_ticket,
                'titulo': ticket.titulo,
                'nombre_solicitante': ticket.nombre_solicitante,
                'email_solicitante': ticket.email_solicitante,
                'departamento': ticket.departamento,
                'descripcion': ticket.descripcion,
                'estado': ticket.estado,
                'prioridad': ticket.prioridad,
                'categoria': ticket.categoria,
                'ingeniero_id': ticket.ingeniero_id,
                'ingeniero_nombre': ingeniero_nombre,
                'fecha_creacion': ticket.fecha_creacion.isoformat() if ticket.fecha_creacion else None,
                'fecha_asignacion': ticket.fecha_asignacion.isoformat() if ticket.fecha_asignacion else None,
                'fecha_resolucion': ticket.fecha_resolucion.isoformat() if ticket.fecha_resolucion else None,
                'comentarios': [
                    {
                        'id': c.id,
                        'contenido': c.contenido,
                        'imagen_url': c.imagen_url,
                        'fecha_creacion': c.fecha_creacion.isoformat() if c.fecha_creacion else None,
                        'ingeniero_nombre': Usuario.query.get(c.ingeniero_id).username if c.ingeniero_id else 'Desconocido'
                    }
                    for c in comentarios
                ]
            }
        }), 200
        
    except Exception as e:
        logger.error(f"Error al obtener ticket: {e}")
        return jsonify({'error': str(e)}), 500


# 6. POST /api/tickets/<id>/comentario (LOGIN REQUERIDO)
@app.route('/api/tickets/<int:ticket_id>/comentario', methods=['POST'])
@ingeniero_login_required
def agregar_comentario(ticket_id):
    """Agregar comentario/documentación con imagen opcional"""
    try:
        usuario = Usuario.query.filter_by(username=session.get('ingeniero_user')).first()
        if not usuario:
            return jsonify({'error': 'Usuario no encontrado'}), 401
        
        ticket = Ticket.query.get(ticket_id)
        if not ticket:
            return jsonify({'error': 'Ticket no encontrado'}), 404
        
        # Verificar que es el ingeniero asignado
        if ticket.ingeniero_id != usuario.id:
            return jsonify({'error': 'Acceso denegado'}), 403
        
        data = request.get_json()
        contenido = data.get('contenido')
        
        if not contenido:
            return jsonify({'error': 'Contenido requerido'}), 400
        
        imagen_url = data.get('imagen_url')
        
        comentario = ComentarioTicket(
            ticket_id=ticket_id,
            ingeniero_id=usuario.id,
            contenido=contenido,
            imagen_url=imagen_url,
            fecha_creacion=datetime.utcnow()
        )
        
        db.session.add(comentario)
        db.session.commit()
        
        logger.info(f"✓ Comentario agregado al ticket {ticket.numero_ticket}")
        
        return jsonify({
            'id': comentario.id,
            'mensaje': 'Comentario agregado exitosamente'
        }), 201
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error al agregar comentario: {e}")
        return jsonify({'error': str(e)}), 500


# 7. PUT /api/tickets/<id>/estado (LOGIN REQUERIDO)
@app.route('/api/tickets/<int:ticket_id>/estado', methods=['PUT'])
@ingeniero_login_required
def cambiar_estado_ticket(ticket_id):
    """Cambiar estado a 'resuelto'"""
    try:
        usuario = Usuario.query.filter_by(username=session.get('ingeniero_user')).first()
        if not usuario:
            return jsonify({'error': 'Usuario no encontrado'}), 401
        
        ticket = Ticket.query.get(ticket_id)
        if not ticket:
            return jsonify({'error': 'Ticket no encontrado'}), 404
        
        # Solo el ingeniero asignado puede cambiar el estado
        if ticket.ingeniero_id != usuario.id:
            return jsonify({'error': 'Acceso denegado'}), 403
        
        data = request.get_json()
        nuevo_estado = data.get('estado', 'resuelto')
        
        if nuevo_estado not in ['en_progreso', 'resuelto']:
            return jsonify({'error': 'Estado inválido'}), 400
        
        ticket.estado = nuevo_estado
        
        if nuevo_estado == 'resuelto':
            ticket.fecha_resolucion = datetime.utcnow()
        
        db.session.commit()
        
        logger.info(f"✓ Ticket {ticket.numero_ticket} cambiado a {nuevo_estado}")
        
        return jsonify({
            'mensaje': 'Estado actualizado',
            'numero_ticket': ticket.numero_ticket,
            'estado': nuevo_estado
        }), 200
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error al cambiar estado: {e}")
        return jsonify({'error': str(e)}), 500


# 7.5 PUT /api/tickets/<id>/devolver (LOGIN REQUERIDO - devolver a bandeja)
@app.route('/api/tickets/<int:ticket_id>/devolver', methods=['PUT'])
@ingeniero_login_required
def devolver_ticket(ticket_id):
    """Devolver ticket a la bandeja (poner como 'nuevo' sin ingeniero)"""
    try:
        usuario = Usuario.query.filter_by(username=session.get('ingeniero_user')).first()
        if not usuario:
            return jsonify({'error': 'Usuario no encontrado'}), 401
        
        ticket = Ticket.query.get(ticket_id)
        if not ticket:
            return jsonify({'error': 'Ticket no encontrado'}), 404
        
        # Solo el ingeniero asignado puede devolverlo
        if ticket.ingeniero_id != usuario.id:
            return jsonify({'error': 'Solo el ingeniero asignado puede devolver este ticket'}), 403
        
        # Devolver a estado nuevo y limpiar ingeniero_id
        ticket.estado = 'nuevo'
        ticket.ingeniero_id = None
        ticket.fecha_asignacion = None
        
        db.session.commit()
        logger.info(f"✓ Ticket {ticket.numero_ticket} devuelto a bandeja por {usuario.username}")
        
        return jsonify({
            'mensaje': 'Ticket devuelto a la bandeja',
            'numero_ticket': ticket.numero_ticket,
            'estado': ticket.estado
        }), 200
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error al devolver ticket: {e}")
        return jsonify({'error': str(e)}), 500


# 8. POST /api/tickets/<id>/imagen (LOGIN REQUERIDO)
@app.route('/api/tickets/<int:ticket_id>/imagen', methods=['POST'])
@ingeniero_login_required
def subir_imagen_ticket(ticket_id):
    """Subir imagen para comentario"""
    try:
        usuario = Usuario.query.filter_by(username=session.get('ingeniero_user')).first()
        if not usuario:
            return jsonify({'error': 'Usuario no encontrado'}), 401
        
        ticket = Ticket.query.get(ticket_id)
        if not ticket:
            return jsonify({'error': 'Ticket no encontrado'}), 404
        
        if ticket.ingeniero_id != usuario.id:
            return jsonify({'error': 'Acceso denegado'}), 403
        
        if 'imagen' not in request.files:
            return jsonify({'error': 'No se proporcionó imagen'}), 400
        
        archivo = request.files['imagen']
        
        if archivo.filename == '':
            return jsonify({'error': 'No se seleccionó archivo'}), 400
        
        # Validar extensión
        ext = archivo.filename.rsplit('.', 1)[1].lower() if '.' in archivo.filename else ''
        if ext not in ALLOWED_EXTENSIONS:
            return jsonify({'error': f'Extensión no permitida. Permitidas: {ALLOWED_EXTENSIONS}'}), 400
        
        # Guardar archivo
        nombre_seguro = secure_filename(f"ticket_{ticket_id}_{uuid.uuid4()}.{ext}")
        ruta_archivo = os.path.join(UPLOAD_FOLDER, nombre_seguro)
        archivo.save(ruta_archivo)
        
        # Retornar URL
        url_imagen = f"/uploads/productos/{nombre_seguro}"
        
        logger.info(f"✓ Imagen subida para ticket {ticket.numero_ticket}")
        
        return jsonify({
            'url': url_imagen,
            'mensaje': 'Imagen subida exitosamente'
        }), 201
        
    except Exception as e:
        logger.error(f"Error al subir imagen: {e}")
        return jsonify({'error': str(e)}), 500


# 9. GET /api/tickets/descargar/excel (LOGIN REQUERIDO)
@app.route('/api/tickets/descargar/excel', methods=['GET'])
@ingeniero_login_required
def descargar_tickets_excel():
    """Descargar en Excel los 'mis-tickets' con: numero, titulo, solicitante, estado, fecha_creacion, comentarios_count"""
    try:
        usuario = Usuario.query.filter_by(username=session.get('ingeniero_user')).first()
        if not usuario:
            return jsonify({'error': 'Usuario no encontrado'}), 401
        
        # Obtener tickets del usuario
        tickets = Ticket.query.filter(
            Ticket.ingeniero_id == usuario.id,
            Ticket.estado.in_(['en_progreso', 'resuelto'])
        ).order_by(Ticket.fecha_creacion.desc()).all()
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Mis Tickets"
        
        # Encabezados
        encabezados = ['Número', 'Título', 'Solicitante', 'Estado', 'Fecha Creación', 'Comentarios']
        ws.append(encabezados)
        
        # Datos
        for ticket in tickets:
            comentarios_count = len(ComentarioTicket.query.filter_by(ticket_id=ticket.id).all())
            ws.append([
                ticket.numero_ticket,
                ticket.titulo,
                ticket.nombre_solicitante,
                ticket.estado,
                ticket.fecha_creacion.strftime('%d/%m/%Y %H:%M') if ticket.fecha_creacion else '',
                comentarios_count
            ])
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Guardar en memoria
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"mis_tickets_{datetime.now().strftime('%d%m%Y_%H%M%S')}.xlsx"
        
        logger.info(f"✓ Tickets descargados en Excel por {usuario.username}")
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logger.error(f"Error al descargar tickets en Excel: {e}")
        return jsonify({'error': str(e)}), 500


# ==================== RUTAS DE INGENIEROS ====================

# GET - Obtener lista de ingenieros (Admin)
@app.route('/api/ingenieros', methods=['GET'])
@login_required
def obtener_ingenieros():
    """Obtiene lista de ingenieros"""
    try:
        ingenieros = Ingeniero.query.all()
        
        return jsonify({
            'ingenieros': [i.to_dict() for i in ingenieros]
        }), 200
        
    except Exception as e:
        logger.error(f"Error al obtener ingenieros: {e}")
        return jsonify({'error': str(e)}), 500


# POST - Registrar nuevo ingeniero (Admin)
@app.route('/api/ingenieros', methods=['POST'])
@login_required
def crear_ingeniero():
    """Crea un nuevo registro de ingeniero"""
    if not is_admin_user():
        return jsonify({'error': 'Acceso denegado'}), 403
    
    try:
        data = request.get_json()
        usuario_id = data.get('usuario_id')
        
        usuario = Usuario.query.get(usuario_id)
        if not usuario:
            return jsonify({'error': 'Usuario no encontrado'}), 404
        
        # Verificar si ya es ingeniero
        if Ingeniero.query.filter_by(usuario_id=usuario_id).first():
            return jsonify({'error': 'Este usuario ya es ingeniero'}), 400
        
        nuevo_ingeniero = Ingeniero(
            usuario_id=usuario_id,
            especialidad=data.get('especialidad'),
            telefono=data.get('telefono'),
            disponible=True
        )
        
        db.session.add(nuevo_ingeniero)
        db.session.commit()
        
        logger.info(f"✓ Nuevo ingeniero registrado: {usuario.username}")
        
        return jsonify({
            'mensaje': 'Ingeniero registrado exitosamente',
            'ingeniero': nuevo_ingeniero.to_dict()
        }), 201
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error al crear ingeniero: {e}")
        return jsonify({'error': str(e)}), 500


# PUT - Actualizar datos de ingeniero
@app.route('/api/ingenieros/<int:ingeniero_id>', methods=['PUT'])
@login_required
def actualizar_ingeniero(ingeniero_id):
    """Actualiza datos de un ingeniero"""
    if not is_admin_user():
        return jsonify({'error': 'Acceso denegado'}), 403
    
    try:
        ingeniero = Ingeniero.query.get(ingeniero_id)
        if not ingeniero:
            return jsonify({'error': 'Ingeniero no encontrado'}), 404
        
        data = request.get_json()
        
        if 'especialidad' in data:
            ingeniero.especialidad = data['especialidad']
        if 'telefono' in data:
            ingeniero.telefono = data['telefono']
        if 'disponible' in data:
            ingeniero.disponible = data['disponible']
        
        db.session.commit()
        
        logger.info(f"✓ Ingeniero {ingeniero.usuario.username} actualizado")
        
        return jsonify({
            'mensaje': 'Ingeniero actualizado exitosamente',
            'ingeniero': ingeniero.to_dict()
        }), 200
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error al actualizar ingeniero: {e}")
        return jsonify({'error': str(e)}), 500


# ==================== PROCESOS Y CLAVES - ADMIN PANEL ====================

def _build_claves_procesos_export_rows(solo_activas=False):
    """Filas para exportar clave + descripción + cada proceso en secuencia."""
    q = ClaveProducto.query.order_by(ClaveProducto.clave.asc())
    if solo_activas:
        q = q.filter_by(activo=True)
    rows = []
    for clave in q.all():
        procesos = (
            ClaveProceso.query.filter_by(clave_id=clave.id)
            .order_by(ClaveProceso.orden.asc(), ClaveProceso.id.asc())
            .all()
        )
        if not procesos:
            rows.append({
                'clave': clave.clave,
                'descripcion_clave': _clean_nullable_text(clave.nombre) or '',
                'notas_clave': _clean_nullable_text(clave.notas) or '',
                'clave_activa': 'SI' if clave.activo else 'NO',
                'orden': '',
                'proceso_codigo': '',
                'proceso_nombre': '',
                'proceso_descripcion': '',
                'operacion': '',
                'centro_trabajo': '',
                't_e': '',
                't_tct': '',
                't_tco': '',
                't_to': '',
                'notas_proceso': '',
            })
            continue
        for cp in procesos:
            proc = cp.proceso
            rows.append({
                'clave': clave.clave,
                'descripcion_clave': _clean_nullable_text(clave.nombre) or '',
                'notas_clave': _clean_nullable_text(clave.notas) or '',
                'clave_activa': 'SI' if clave.activo else 'NO',
                'orden': cp.orden,
                'proceso_codigo': _clean_nullable_text(proc.codigo) if proc else '',
                'proceso_nombre': _clean_nullable_text(proc.nombre) if proc else '',
                'proceso_descripcion': _clean_nullable_text(proc.descripcion) if proc else '',
                'operacion': _clean_nullable_text(cp.operacion) or (_clean_nullable_text(proc.operacion) if proc else ''),
                'centro_trabajo': _clean_nullable_text(cp.centro_trabajo) or (_clean_nullable_text(proc.centro_trabajo) if proc else ''),
                't_e': _clean_nullable_text(cp.t_e) or '',
                't_tct': _clean_nullable_text(cp.t_tct) or '',
                't_tco': _clean_nullable_text(cp.t_tco) or '',
                't_to': _clean_nullable_text(cp.t_to) or '',
                'notas_proceso': _clean_nullable_text(cp.notas) or '',
            })
    return rows


@app.route('/procesos')
@login_required
@requires_any_permission([('procesos', 'view'), ('procesos', 'edit'), ('procesos', 'update'), ('procesos', 'create'), ('procesos', 'delete')])
def procesos_panel():
    """Panel de administración para procesos (catálogo) y claves (productos)."""
    try:
        clave_id = request.args.get('clave_id', type=int)
        claves = ClaveProducto.query.order_by(ClaveProducto.clave.asc()).all()
        procesos = ProcesoCatalogo.query.filter_by(activo=True).order_by(ProcesoCatalogo.nombre.asc()).all()
        clave_sel = ClaveProducto.query.get(clave_id) if clave_id else None
        return render_template('procesos_admin.html', claves=claves, procesos=procesos, clave_sel=clave_sel)
    except Exception as e:
        logger.error(f"Error cargando panel de procesos: {e}")
        return render_template('procesos_admin.html', claves=[], procesos=[], clave_sel=None, error=str(e))


@app.route('/procesos/export.csv', methods=['GET'])
@login_required
@requires_any_permission([('procesos', 'view'), ('procesos', 'edit'), ('procesos', 'update'), ('procesos', 'create'), ('procesos', 'delete')])
def procesos_claves_export_csv():
    """Exporta informe completo: claves, descripción y todos sus procesos."""
    try:
        solo_activas = (request.args.get('solo_activas') or '').strip().lower() in ('1', 'true', 'yes', 'on')
        rows = _build_claves_procesos_export_rows(solo_activas=solo_activas)
        columns = [
            'clave', 'descripcion_clave', 'notas_clave', 'clave_activa', 'orden',
            'proceso_codigo', 'proceso_nombre', 'proceso_descripcion',
            'operacion', 'centro_trabajo', 't_e', 't_tct', 't_tco', 't_to', 'notas_proceso',
        ]
        df = pd.DataFrame(rows, columns=columns)
        csv_data = df.to_csv(index=False, encoding='utf-8-sig')
        stamp = datetime.utcnow().strftime('%Y%m%d')
        suffix = '_activas' if solo_activas else '_todas'
        response = make_response(csv_data)
        response.headers['Content-Type'] = 'text/csv; charset=utf-8'
        response.headers['Content-Disposition'] = f'attachment; filename=claves_procesos{suffix}_{stamp}.csv'
        return response
    except Exception as exc:
        logger.error(f'Error exportando claves/procesos CSV: {exc}', exc_info=True)
        return jsonify({'error': str(exc)}), 500


@app.route('/procesos/clave/save', methods=['POST'])
@login_required
@requires_any_permission([('procesos', 'create'), ('procesos', 'edit'), ('procesos', 'update')])
def procesos_clave_save():
    """Crear o actualizar una clave producto."""
    try:
        clave_id = request.form.get('id', type=int)
        clave = _clean_nullable_text(request.form.get('clave', ''))
        nombre = _clean_nullable_text(request.form.get('nombre', ''))
        notas = _clean_nullable_text(request.form.get('notas', ''))
        activo = request.form.get('activo') == 'on'

        if not clave:
            return redirect(url_for('procesos_panel'))

        if clave_id:
            obj = ClaveProducto.query.get_or_404(clave_id)
            obj.clave = clave
            obj.nombre = nombre
            obj.notas = notas
            obj.activo = activo
        else:
            obj = ClaveProducto(clave=clave, nombre=nombre, notas=notas, activo=activo)
            db.session.add(obj)
        db.session.commit()
        return redirect(url_for('procesos_panel', clave_id=obj.id))
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error guardando clave: {e}")
        return redirect(url_for('procesos_panel'))


@app.route('/procesos/clave/<int:clave_id>/secuencia/save', methods=['POST'])
@login_required
@requires_any_permission([('procesos', 'edit'), ('procesos', 'update')])
def procesos_clave_secuencia_save(clave_id):
    """Guardar la secuencia de procesos para una clave."""
    try:
        clave = ClaveProducto.query.get_or_404(clave_id)
        # Arrays enviados desde el formulario
        procesos_ids = request.form.getlist('proceso_id[]')
        ordenes = request.form.getlist('orden[]')
        ct_list = request.form.getlist('ct[]')
        oper_list = request.form.getlist('operacion[]')
        t_e_list = request.form.getlist('t_e[]')
        t_tct_list = request.form.getlist('t_tct[]')
        t_tco_list = request.form.getlist('t_tco[]')
        t_to_list = request.form.getlist('t_to[]')
        tiempos = request.form.getlist('tiempo_est[]')  # legacy opcional
        notas_list = request.form.getlist('notas[]')

        # Limpiar actuales
        ClaveProceso.query.filter_by(clave_id=clave.id).delete()
        db.session.flush()

        # Crear nuevos en orden
        for idx, pid in enumerate(procesos_ids):
            try:
                p_id = int(pid)
            except ValueError:
                continue
            orden = int(ordenes[idx]) if idx < len(ordenes) and str(ordenes[idx]).isdigit() else idx + 1
            ct = ct_list[idx] if idx < len(ct_list) else None
            oper = oper_list[idx] if idx < len(oper_list) else None
            t_e = t_e_list[idx] if idx < len(t_e_list) else None
            t_tct = t_tct_list[idx] if idx < len(t_tct_list) else None
            t_tco = t_tco_list[idx] if idx < len(t_tco_list) else None
            t_to = t_to_list[idx] if idx < len(t_to_list) else None
            tiempo = tiempos[idx] if idx < len(tiempos) else None
            nota = notas_list[idx] if idx < len(notas_list) else None
            cp = ClaveProceso(
                clave_id=clave.id,
                proceso_id=p_id,
                orden=orden,
                centro_trabajo=ct,
                operacion=oper,
                t_e=t_e,
                t_tct=t_tct,
                t_tco=t_tco,
                t_to=t_to,
                tiempo_estimado=tiempo,
                notas=nota
            )
            db.session.add(cp)

        db.session.commit()
        return redirect(url_for('procesos_panel', clave_id=clave.id))
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error guardando secuencia de clave {clave_id}: {e}")
        return redirect(url_for('procesos_panel', clave_id=clave_id))


@app.route('/procesos/clave/<int:clave_id>/delete', methods=['POST'])
@login_required
@requires_any_permission([('procesos', 'delete')])
def procesos_clave_delete(clave_id):
    try:
        obj = ClaveProducto.query.get_or_404(clave_id)
        db.session.delete(obj)
        db.session.commit()
        return redirect(url_for('procesos_panel'))
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error eliminando clave {clave_id}: {e}")
        return redirect(url_for('procesos_panel'))


@app.route('/procesos/base/save', methods=['POST'])
@login_required
@requires_any_permission([('procesos', 'create'), ('procesos', 'edit'), ('procesos', 'update')])
def procesos_base_save():
    """Crear o actualizar un proceso del catálogo."""
    try:
        proc_id = request.form.get('id', type=int)
        codigo = request.form.get('codigo', '').strip()
        nombre = request.form.get('nombre', '').strip()
        operacion = request.form.get('operacion', '').strip()
        descripcion = request.form.get('descripcion', '').strip()
        centro_trabajo = request.form.get('centro_trabajo', '').strip()
        tiempo_est = request.form.get('tiempo_estimado', '').strip()
        activo = request.form.get('activo') == 'on'

        # Si no quieren código, lo permitimos vacío. Si no se envía nombre, usar operacion como nombre
        if not nombre:
            nombre = operacion or 'Operacion'

        if proc_id:
            p = ProcesoCatalogo.query.get_or_404(proc_id)
            p.codigo = codigo
            p.nombre = nombre
            p.operacion = operacion
            p.descripcion = descripcion
            p.centro_trabajo = centro_trabajo
            p.tiempo_estimado = tiempo_est
            p.activo = activo
        else:
            p = ProcesoCatalogo(codigo=codigo or None, nombre=nombre, operacion=operacion, descripcion=descripcion, centro_trabajo=centro_trabajo, tiempo_estimado=tiempo_est, activo=activo)
            db.session.add(p)
        db.session.commit()
        return redirect(url_for('procesos_panel'))
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error guardando proceso base: {e}")
        return redirect(url_for('procesos_panel'))


@app.route('/procesos/base/<int:proc_id>/delete', methods=['POST'])
@login_required
@requires_any_permission([('procesos', 'delete')])
def procesos_base_delete(proc_id):
    try:
        uso = ClaveProceso.query.filter_by(proceso_id=proc_id).count()
        if uso > 0:
            # Evitar borrar si está en uso
            return redirect(url_for('procesos_panel'))
        p = ProcesoCatalogo.query.get_or_404(proc_id)
        db.session.delete(p)
        db.session.commit()
        return redirect(url_for('procesos_panel'))
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error eliminando proceso base {proc_id}: {e}")
        return redirect(url_for('procesos_panel'))


# ==================== CONTPAQI CONCILIACION (URL PRIVADA) ====================

@app.route('/contpaq/conciliacion')
@login_required
@requires_permission('contpaq', 'view')
def contpaq_conciliacion_page():
    """Vista privada para consultar pedidos Odoo en conciliacion (sin menu)."""
    last_run = None
    last_sync_label = ''
    try:
        _ensure_odoo_tables()
        last_run = OdooSyncRun.query.order_by(OdooSyncRun.id.desc()).first()
        if last_run:
            sync_dt = last_run.finished_at or last_run.started_at
            if sync_dt:
                last_sync_label = sync_dt.strftime('%Y-%m-%d %H:%M:%S')
    except Exception as exc:
        logger.error(f'Error cargando estado Odoo en conciliacion: {exc}', exc_info=True)
    return render_template(
        'contpaq_conciliacion.html',
        last_run=last_run,
        last_sync_label=last_sync_label,
        odoo_sync_auto=ODOO_SYNC_ENABLED,
        odoo_sync_interval=ODOO_SYNC_INTERVAL_MINUTES,
    )


@app.route('/contpaq/maximos-minimos')
@login_required
@requires_permission('contpaq', 'view')
def contpaq_maximos_minimos_page():
    """Vista privada para máximos y mínimos calculados con demanda de pedidos CONTPAQ."""
    return render_template('contpaq_maximos_minimos.html')


@app.route('/api/contpaq/maximos-minimos', methods=['GET'])
@login_required
@requires_permission('contpaq', 'view')
def api_contpaq_maximos_minimos():
    try:
        payload = _contpaq_max_min_rows(
            q=(request.args.get('q') or '').strip(),
            sucursal=(request.args.get('sucursal') or '').strip(),
            category1=(request.args.get('category1') or '').strip(),
            category2=(request.args.get('category2') or '').strip(),
            only_alert=(request.args.get('only_alert') or '').strip().lower() in ('1', 'true', 'yes', 'on'),
            period_type=(request.args.get('period_type') or 'week').strip(),
            period_value=request.args.get('period_value'),
            period_from=request.args.get('period_from'),
            period_to=request.args.get('period_to'),
            period_year=request.args.get('period_year'),
            date_from=request.args.get('date_from'),
            date_to=request.args.get('date_to'),
            limit=request.args.get('limit', default=200, type=int),
            page=request.args.get('page', default=1, type=int),
        )
        return jsonify(payload), 200
    except Exception as exc:
        logger.error(f"Error generando maximos/minimos CONTPAQ: {exc}", exc_info=True)
        return jsonify({'error': str(exc)}), 500


@app.route('/api/contpaq/maximos-minimos/export.csv', methods=['GET'])
@login_required
@requires_permission('contpaq', 'view')
def api_contpaq_maximos_minimos_export_csv():
    try:
        payload = _contpaq_max_min_rows(
            q=(request.args.get('q') or '').strip(),
            sucursal=(request.args.get('sucursal') or '').strip(),
            category1=(request.args.get('category1') or '').strip(),
            category2=(request.args.get('category2') or '').strip(),
            only_alert=(request.args.get('only_alert') or '').strip().lower() in ('1', 'true', 'yes', 'on'),
            period_type=(request.args.get('period_type') or 'week').strip(),
            period_value=request.args.get('period_value'),
            period_from=request.args.get('period_from'),
            period_to=request.args.get('period_to'),
            period_year=request.args.get('period_year'),
            date_from=request.args.get('date_from'),
            date_to=request.args.get('date_to'),
            limit=200000,
            page=1,
        )
        items = payload.get('items') or []

        columns = [
            'product_key', 'product_name', 'category1', 'category2', 'unit',
            *list(payload.get('summary', {}).get('period_columns') or []),
            'cantidad_periodo', 'pedidos_periodo', 'promedio', 'existencia', 'maximo', 'minimo',
            'sugerido_compra', 'status', 'detalle_periodos', 'qty_available', 'qty_to_deliver_customer',
            'qty_to_receive_supplier', 'qty_on_transit', 'qty_to_receive', 'period_type', 'period_value',
            'period_from', 'period_to', 'period_year', 'date_from', 'date_to', 'updated_at'
        ]
        export_rows = []
        for item in items:
            row = dict(item)
            for label in payload.get('summary', {}).get('period_columns') or []:
                row[label] = float((item.get('period_values') or {}).get(label, 0.0) or 0.0)
            export_rows.append(row)
        df = pd.DataFrame(export_rows, columns=columns)
        csv_data = df.to_csv(index=False, encoding='utf-8-sig')

        response = make_response(csv_data)
        response.headers['Content-Type'] = 'text/csv; charset=utf-8'
        response.headers['Content-Disposition'] = 'attachment; filename=contpaq_maximos_minimos.csv'
        return response
    except Exception as exc:
        logger.error(f"Error exportando maximos/minimos CSV: {exc}", exc_info=True)
        return jsonify({'error': str(exc)}), 500


def _conciliacion_odoo_periodo_semana(pedido):
    """Semana de negocio para un pedido Odoo (titulo o fecha del pedido)."""
    titulo_key = _contpaq_title_week_key(pedido.titulo, None, pedido.date_order)
    if titulo_key:
        return titulo_key
    return _contpaq_week_label_from_date(pedido.date_order) or 'SIN SEMANA'


def _build_conciliacion_odoo_response(
    *,
    q='',
    folio='',
    cliente='',
    sucursal='',
    titulo='',
    fecha_desde_raw='',
    fecha_hasta_raw='',
    limit=120,
    page=1,
    collect_all=False,
):
    """Arma la respuesta de conciliacion leyendo pedidos sincronizados desde Odoo."""
    _ensure_odoo_tables()

    limit = max(1, min(int(limit or 120), 500))
    page = max(1, int(page or 1))

    pedidos_q = OdooPedidoVenta.query.filter(
        OdooPedidoVenta.sucursal.isnot(None),
        func.length(func.trim(OdooPedidoVenta.sucursal)) > 0,
    )

    if cliente:
        pedidos_q = pedidos_q.filter(OdooPedidoVenta.partner_name.ilike(f"%{cliente}%"))

    if folio:
        like_folio = f"%{folio}%"
        pedidos_q = pedidos_q.filter(
            db.or_(
                OdooPedidoVenta.name.ilike(like_folio),
                OdooPedidoVenta.client_order_ref.ilike(like_folio),
            )
        )

    if sucursal:
        pedidos_q = pedidos_q.filter(OdooPedidoVenta.sucursal.ilike(f"%{sucursal}%"))

    if fecha_desde_raw:
        fecha_desde = _parse_date(fecha_desde_raw)
        if not fecha_desde:
            return {'error': 'fecha_desde invalida. Usa YYYY-MM-DD'}, 400
        pedidos_q = pedidos_q.filter(
            OdooPedidoVenta.date_order >= datetime.combine(fecha_desde, datetime.min.time())
        )

    if fecha_hasta_raw:
        fecha_hasta = _parse_date(fecha_hasta_raw)
        if not fecha_hasta:
            return {'error': 'fecha_hasta invalida. Usa YYYY-MM-DD'}, 400
        pedidos_q = pedidos_q.filter(
            OdooPedidoVenta.date_order < datetime.combine(fecha_hasta + timedelta(days=1), datetime.min.time())
        )

    if q:
        terms = [t.strip() for t in q.split() if t.strip()]
        for term in terms:
            like = f"%{term}%"
            pedidos_q = pedidos_q.filter(
                db.or_(
                    OdooPedidoVenta.titulo.ilike(like),
                    OdooPedidoVenta.name.ilike(like),
                    OdooPedidoVenta.sucursal.ilike(like),
                    OdooPedidoVenta.partner_name.ilike(like),
                    OdooPedidoVenta.client_order_ref.ilike(like),
                    OdooPedidoVenta.origin.ilike(like),
                )
            )

    ordered_q = pedidos_q.order_by(OdooPedidoVenta.date_order.desc(), OdooPedidoVenta.id.desc())
    compare_pedidos = ordered_q.all()

    if titulo:
        compare_pedidos = [
            p for p in compare_pedidos
            if _contpaq_period_matches(
                titulo,
                p.titulo,
                _conciliacion_odoo_periodo_semana(p),
                _contpaq_title_week_key(p.titulo, None, p.date_order),
            )
        ]

    display_pedidos = list(compare_pedidos)

    semana_options = _contpaq_unique_semana_options([
        _contpaq_title_week_key(p.titulo, _conciliacion_odoo_periodo_semana(p), p.date_order)
        for p in display_pedidos
    ])
    sucursal_options = sorted({
        str(p.sucursal or '').strip()
        for p in display_pedidos
        if str(p.sucursal or '').strip()
    })

    total_records = len(display_pedidos)
    compare_map = {p.id: p for p in compare_pedidos}
    compare_pedido_ids = [p.id for p in compare_pedidos]
    if collect_all:
        page = 1
        offset = 0
        pedidos = display_pedidos
    else:
        offset = (page - 1) * limit
        pedidos = display_pedidos[offset:offset + limit]

    if not pedidos:
        return {
            'items': [],
            'total': 0,
            'page': page,
            'limit': limit,
            'total_records': total_records,
            'total_pages': (total_records + limit - 1) // limit if total_records else 0,
            'has_prev': page > 1,
            'has_next': offset + limit < total_records,
            'filter_options': {
                'semanas': semana_options,
                'sucursales': sucursal_options,
            },
        }, 200

    def _query_odoo_lines(pedido_ids):
        if not pedido_ids:
            return []
        rows = []
        chunk_size = 500
        for i in range(0, len(pedido_ids), chunk_size):
            chunk = pedido_ids[i:i + chunk_size]
            rows.extend(
                OdooPedidoVentaLinea.query.filter(
                    OdooPedidoVentaLinea.pedido_id.in_(chunk)
                ).order_by(
                    OdooPedidoVentaLinea.pedido_id.asc(),
                    OdooPedidoVentaLinea.sequence.asc(),
                    OdooPedidoVentaLinea.id.asc(),
                ).all()
            )
        return rows

    pedido_ids = [p.id for p in pedidos]
    detalles = _query_odoo_lines(pedido_ids)
    compare_detalles = _query_odoo_lines(compare_pedido_ids) if compare_pedido_ids else []

    detalles_map = {}
    for d in detalles:
        detalles_map.setdefault(d.pedido_id, []).append(d)

    visible_claves = {
        _contpaq_norm_text(d.product_key)
        for d in (detalles + compare_detalles)
        if _contpaq_norm_text(d.product_key)
    }
    public_price_map = {}
    if visible_claves:
        public_price_map = {
            _contpaq_norm_text(row.clave_producto): float(row.precio_publico)
            for row in ContpaqPrecioPublico.query.filter(
                ContpaqPrecioPublico.clave_producto.in_(list(visible_claves))
            ).all()
            if row.clave_producto is not None and row.precio_publico is not None
        }

    def _norm_txt(v):
        return str(v or '').strip().upper()

    def _norm_num(v):
        try:
            f = float(str(v).replace(',', '').strip())
            if f.is_integer():
                return str(int(f))
            return f"{f:.4f}".rstrip('0').rstrip('.')
        except Exception:
            return _norm_txt(v)

    import re as _re_sem

    def _sem_year_strip(s):
        v = _contpaq_norm_text(str(s or ''))
        return _re_sem.sub(r'\s+DE\s+\d{4}\s*$', '', v).strip()

    p_set = set()
    for d in compare_detalles:
        p_cmp = compare_map.get(d.pedido_id)
        if not p_cmp:
            continue
        semana_cmp_norm = _sem_year_strip(
            _contpaq_title_week_key(p_cmp.titulo, _conciliacion_odoo_periodo_semana(p_cmp), p_cmp.date_order)
        )
        sucursal_cmp_norm = _norm_txt(p_cmp.sucursal)
        row_key_cmp = (
            _norm_txt(d.product_key),
            _norm_num(d.product_uom_qty),
            semana_cmp_norm,
            sucursal_cmp_norm,
        )
        p_set.add(row_key_cmp)

    items = []
    total_partidas = 0
    total_importe = 0.0
    total_remisiones = 0
    semana_totales = {}

    for p in pedidos:
        pedido_rows = []
        pedido_total = 0.0
        semana_norm = _contpaq_title_week_key(p.titulo, _conciliacion_odoo_periodo_semana(p), p.date_order)
        periodo_semana = _conciliacion_odoo_periodo_semana(p)

        for d in detalles_map.get(p.id, []):
            cantidad_num = _to_float(d.product_uom_qty)
            precio_unit = _to_float(d.price_unit)
            partida_total = _to_float(d.price_subtotal)
            if partida_total is None and cantidad_num is not None and precio_unit is not None:
                partida_total = round(cantidad_num * precio_unit, 2)
            partida_total = float(partida_total or 0)
            precio_publico = public_price_map.get(_norm_txt(d.product_key))
            total_publico = round(cantidad_num * precio_publico, 2) if precio_publico is not None and cantidad_num is not None else None
            diferencia_publico = round(total_publico - partida_total, 2) if total_publico is not None else None
            pedido_total += partida_total
            total_partidas += 1

            pedido_rows.append({
                'line_number': d.sequence,
                'clave_producto': d.product_key,
                'descripcion': d.description or d.product_name,
                'cantidad': d.product_uom_qty,
                'precio_unitario': precio_unit,
                'total_partida': partida_total,
                'precio_publico_unitario': precio_publico,
                'total_partida_precio_publico': total_publico,
                'diferencia_precio_publico': diferencia_publico,
                'serie': '',
                'folio': p.name,
                'fecha_documento': p.date_order.isoformat() if p.date_order else None,
                'es_inyectado': False,
            })

        total_importe += pedido_total
        sem = periodo_semana or 'SIN SEMANA'
        if sem not in semana_totales:
            semana_totales[sem] = {'pedidos': 0, 'total': 0.0}
        semana_totales[sem]['pedidos'] += 1
        semana_totales[sem]['total'] += pedido_total

        items.append({
            'document_id': p.odoo_id,
            'doc_folio': p.name,
            'serie': '',
            'cliente': p.partner_name,
            'sucursal': p.sucursal,
            'titulo': p.titulo,
            'periodo_semana': periodo_semana,
            'semana_match_key': semana_norm,
            'fecha_documento': p.date_order.isoformat() if p.date_order else None,
            'pedido_total': round(pedido_total, 2),
            'detalles': pedido_rows,
            'remisiones': [],
            'fuente': 'odoo',
        })

    indice_q = ContpaqSucursalIndice.query
    if sucursal:
        indice_q = indice_q.filter(ContpaqSucursalIndice.sucursal.ilike(f"%{sucursal}%"))
    if titulo:
        indice_q = indice_q.filter(ContpaqSucursalIndice.semana.ilike(f"%{titulo}%"))

    indice_rows = indice_q.all()
    faltantes_indice = []
    for idx in indice_rows:
        row_key = (
            _contpaq_norm_text(idx.clave_producto),
            _contpaq_norm_qty(idx.cantidad),
            _sem_year_strip(idx.semana),
            _contpaq_norm_text(idx.sucursal),
        )
        if row_key in p_set:
            continue
        faltantes_indice.append({
            'source_document_id': None,
            'folio': idx.folio or '',
            'semana': idx.semana,
            'sucursal': idx.sucursal,
            'clave_producto': idx.clave_producto,
            'descripcion': idx.descripcion,
            'cantidad': idx.cantidad,
            'precio_unitario': None,
            'origen': 'INDICE_SUCURSALES',
            'line_number_origen': None,
        })

    faltantes_inyectados = 0
    if page == 1:
        ws_targets = {}
        faltantes_line_counter = {}
        for i, item in enumerate(items):
            sem_k = _sem_year_strip(item.get('semana_match_key') or item.get('periodo_semana') or '')
            key = (sem_k, _norm_txt(item.get('sucursal')))
            if key not in ws_targets:
                ws_targets[key] = i

        for f in faltantes_indice:
            sem_f = _sem_year_strip(str(f.get('semana') or '').strip())
            key = (sem_f, _norm_txt(str(f.get('sucursal') or '').strip()))
            target_idx = ws_targets.get(key)
            if target_idx is None:
                continue

            cantidad = _to_float(f.get('cantidad'))
            precio = _to_float(f.get('precio_unitario'))
            total_partida = round(cantidad * precio, 2) if cantidad is not None and precio is not None else 0.0
            precio_publico = public_price_map.get(_norm_txt(f.get('clave_producto')))
            total_publico = round(cantidad * precio_publico, 2) if precio_publico is not None and cantidad is not None else None
            diferencia_publico = round(total_publico - total_partida, 2) if total_publico is not None else None

            target_item = items[target_idx]
            target_item.setdefault('detalles', [])
            faltantes_line_counter[target_idx] = faltantes_line_counter.get(target_idx, 0) + 1
            target_item['detalles'].append({
                'line_number': f"F-{faltantes_line_counter[target_idx]}",
                'clave_producto': f.get('clave_producto'),
                'descripcion': f.get('descripcion') or 'FALTANTE DESDE INDICE',
                'cantidad': f.get('cantidad'),
                'precio_unitario': precio,
                'total_partida': total_partida,
                'precio_publico_unitario': precio_publico,
                'total_partida_precio_publico': total_publico,
                'diferencia_precio_publico': diferencia_publico,
                'origen_faltante': f.get('origen'),
                'folio_origen': f.get('folio') or '',
                'serie': '',
                'folio': f.get('folio') or '',
                'fecha_documento': None,
                'es_inyectado': True,
            })

            target_item['pedido_total'] = round(_to_float(target_item.get('pedido_total')) + total_partida, 2)
            target_item['es_faltante'] = True
            total_partidas += 1
            total_importe += total_partida
            sem_target = target_item.get('periodo_semana') or sem_f
            if sem_target not in semana_totales:
                semana_totales[sem_target] = {'pedidos': 0, 'total': 0.0}
            semana_totales[sem_target]['total'] += total_partida
            faltantes_inyectados += 1

    items.sort(
        key=lambda x: (
            x.get('periodo_semana') or '',
            x.get('fecha_documento') or '',
            str(x.get('doc_folio') or ''),
        ),
        reverse=True,
    )

    resumen = {
        'total_pedidos': len(items),
        'total_partidas': total_partidas,
        'total_importe': round(total_importe, 2),
        'total_remisiones': total_remisiones,
        'total_faltantes_desde_indice': len(faltantes_indice),
        'faltantes_inyectados': faltantes_inyectados,
        'totales_por_semana': [
            {
                'semana': semana,
                'pedidos': vals['pedidos'],
                'total': round(vals['total'], 2),
            }
            for semana, vals in sorted(semana_totales.items(), key=lambda x: str(x[0]), reverse=True)
        ],
    }

    return {
        'items': items,
        'total': len(items),
        'resumen': resumen,
        'filter_options': {
            'semanas': semana_options,
            'sucursales': sucursal_options,
        },
        'page': page,
        'limit': limit,
        'total_records': total_records,
        'total_pages': (total_records + limit - 1) // limit if total_records else 0,
        'has_prev': page > 1,
        'has_next': offset + limit < total_records,
        'fuente': 'odoo',
    }, 200


def _merge_conciliacion_resumen(res_a, res_b):
    """Combina dos resumenes de conciliacion en uno solo."""
    res_a = res_a or {}
    res_b = res_b or {}
    semana_totales = {}
    for res in (res_a, res_b):
        for row in (res.get('totales_por_semana') or []):
            sem = row.get('semana') or 'SIN SEMANA'
            if sem not in semana_totales:
                semana_totales[sem] = {'pedidos': 0, 'total': 0.0}
            semana_totales[sem]['pedidos'] += int(row.get('pedidos') or 0)
            semana_totales[sem]['total'] += float(row.get('total') or 0)
    return {
        'total_pedidos': int(res_a.get('total_pedidos') or 0) + int(res_b.get('total_pedidos') or 0),
        'total_partidas': int(res_a.get('total_partidas') or 0) + int(res_b.get('total_partidas') or 0),
        'total_importe': round(float(res_a.get('total_importe') or 0) + float(res_b.get('total_importe') or 0), 2),
        'total_remisiones': int(res_a.get('total_remisiones') or 0) + int(res_b.get('total_remisiones') or 0),
        'total_faltantes_desde_indice': int(res_a.get('total_faltantes_desde_indice') or 0) + int(res_b.get('total_faltantes_desde_indice') or 0),
        'faltantes_inyectados': int(res_a.get('faltantes_inyectados') or 0) + int(res_b.get('faltantes_inyectados') or 0),
        'totales_por_semana': [
            {'semana': sem, 'pedidos': vals['pedidos'], 'total': round(vals['total'], 2)}
            for sem, vals in sorted(semana_totales.items(), key=lambda x: str(x[0]), reverse=True)
        ],
    }


def _sucursal_canonical_key(raw):
    """Clave canonica de sucursal: sin acentos, mayusculas y sin prefijo SUC/SUCURSAL.

    Asi 'TOLUCA', 'Toluca' y 'SUC TOLUCA' quedan bajo la misma clave, igual que
    'QUERETARO' y 'Queretaro'.
    """
    import unicodedata
    s = str(raw or '').strip()
    if not s:
        return ''
    s = ''.join(c for c in unicodedata.normalize('NFKD', s) if not unicodedata.combining(c))
    s = s.upper()
    s = re.sub(r'^\s*SUC(?:URSAL)?\.?\s+', '', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s


# Nombres "bonitos" con acento para claves conocidas. Cualquier otra clave se
# muestra en Title Case. Esto es DETERMINISTA: una misma clave siempre da el
# mismo nombre, sin importar que variantes lleguen del sync (evita duplicados).
_SUCURSAL_DISPLAY_OVERRIDES = {
    'QUERETARO': 'Querétaro',
    'TULTITLAN': 'Tultitlán',
}


def _sucursal_pretty(key):
    """Nombre visible determinista a partir de la clave canonica de sucursal."""
    if not key:
        return key
    if key in _SUCURSAL_DISPLAY_OVERRIDES:
        return _SUCURSAL_DISPLAY_OVERRIDES[key]
    return key.title()


# Cache en memoria del dataset combinado de conciliacion. Como armar todo
# (CONTPAQ + Odoo con detalles) es costoso, guardamos el resultado ya combinado
# y unificado, y solo paginamos/filtramos por sucursal en cada request. La firma
# por conteos + TTL invalida la cache cuando cambian los datos.
_CONCILIACION_CACHE = {}
_CONCILIACION_CACHE_TTL = 120
_CONCILIACION_CACHE_MAX = 24


def _conciliacion_cache_get(key):
    entry = _CONCILIACION_CACHE.get(key)
    if not entry:
        return None
    if time() - entry['ts'] > _CONCILIACION_CACHE_TTL:
        _CONCILIACION_CACHE.pop(key, None)
        return None
    return entry['data']


def _conciliacion_cache_set(key, data):
    _CONCILIACION_CACHE[key] = {'ts': time(), 'data': data}
    if len(_CONCILIACION_CACHE) > _CONCILIACION_CACHE_MAX:
        oldest = min(_CONCILIACION_CACHE.items(), key=lambda kv: kv[1]['ts'])[0]
        _CONCILIACION_CACHE.pop(oldest, None)


def _conciliacion_dataset_signature():
    """Firma barata: si cambian los conteos, la cache se invalida sola."""
    try:
        return (
            db.session.query(func.count(ContpaqPedido.id)).scalar() or 0,
            db.session.query(func.count(OdooPedidoVenta.id)).scalar() or 0,
            db.session.query(func.count(ContpaqSucursalIndice.id)).scalar() or 0,
            db.session.query(func.count(ContpaqPrecioPublico.id)).scalar() or 0,
        )
    except Exception:
        return None


_CONCILIACION_YEARS_CACHE = {'ts': 0.0, 'val': []}


def _conciliacion_available_years():
    """Años con pedidos (CONTPAQ + Odoo). Cacheado 5 min porque cambia poco."""
    if _CONCILIACION_YEARS_CACHE['val'] and (time() - _CONCILIACION_YEARS_CACHE['ts'] < 300):
        return _CONCILIACION_YEARS_CACHE['val']
    years = set()
    try:
        for (y,) in db.session.query(
            func.distinct(func.extract('year', ContpaqPedido.fecha_documento))
        ).all():
            if y:
                years.add(int(y))
    except Exception:
        pass
    try:
        for (y,) in db.session.query(
            func.distinct(func.extract('year', OdooPedidoVenta.date_order))
        ).all():
            if y:
                years.add(int(y))
    except Exception:
        pass
    result = sorted(years, reverse=True)
    _CONCILIACION_YEARS_CACHE['ts'] = time()
    _CONCILIACION_YEARS_CACHE['val'] = result
    return result


def _compute_conciliacion_resumen_from_items(items):
    """Calcula el resumen (totales) a partir de los items ya filtrados.

    Asi los totales SIEMPRE respetan los filtros del usuario (sucursal, semana,
    año) y coinciden con lo que se muestra en las tarjetas y en el paginador.
    """
    total_pedidos = len(items)
    total_partidas = 0
    total_importe = 0.0
    total_remisiones = 0
    total_faltantes = 0
    semana_totales = {}

    for it in items:
        pedido_total = _to_float(it.get('pedido_total')) or 0.0
        detalles = it.get('detalles') or []
        total_partidas += len(detalles)
        for d in detalles:
            if d.get('es_inyectado'):
                total_faltantes += 1
        total_remisiones += len(it.get('remisiones') or [])
        total_importe += pedido_total

        sem = it.get('periodo_semana') or 'SIN SEMANA'
        if sem not in semana_totales:
            semana_totales[sem] = {'pedidos': 0, 'total': 0.0}
        semana_totales[sem]['pedidos'] += 1
        semana_totales[sem]['total'] += pedido_total

    return {
        'total_pedidos': total_pedidos,
        'total_partidas': total_partidas,
        'total_importe': round(total_importe, 2),
        'total_remisiones': total_remisiones,
        'total_faltantes_desde_indice': total_faltantes,
        'faltantes_inyectados': total_faltantes,
        'totales_por_semana': [
            {'semana': sem, 'pedidos': vals['pedidos'], 'total': round(vals['total'], 2)}
            for sem, vals in sorted(semana_totales.items(), key=lambda x: str(x[0]), reverse=True)
        ],
    }


def _build_conciliacion_combined_dataset(base_params):
    """Arma el dataset combinado (CONTPAQ + Odoo) ya unificado por sucursal.

    Devuelve (dict_dataset, None) en exito o (payload_error, status) en error.
    El dataset NO esta filtrado por sucursal ni paginado.
    """
    builder_params = dict(base_params, sucursal='')

    contpaq_payload, contpaq_status = _build_conciliacion_contpaq_response(collect_all=True, **builder_params)
    if isinstance(contpaq_payload, dict) and 'error' in contpaq_payload:
        return contpaq_payload, contpaq_status

    odoo_payload, odoo_status = _build_conciliacion_odoo_response(collect_all=True, **builder_params)
    if isinstance(odoo_payload, dict) and 'error' in odoo_payload:
        return odoo_payload, odoo_status

    items = list(contpaq_payload.get('items') or []) + list(odoo_payload.get('items') or [])

    # Unificar sucursales de forma DETERMINISTA: cada variante se convierte a su
    # clave canonica y se muestra siempre con el mismo nombre. Asi el sync nunca
    # vuelve a duplicar sucursales.
    sucursal_keys = set()
    for it in items:
        key = _sucursal_canonical_key(it.get('sucursal'))
        if key:
            it['sucursal'] = _sucursal_pretty(key)
            sucursal_keys.add(key)

    items.sort(
        key=lambda x: (
            x.get('periodo_semana') or '',
            x.get('fecha_documento') or '',
            str(x.get('doc_folio') or ''),
        ),
        reverse=True,
    )

    resumen = _merge_conciliacion_resumen(
        contpaq_payload.get('resumen'),
        odoo_payload.get('resumen'),
    )

    contpaq_opts = (contpaq_payload.get('filter_options') or {})
    odoo_opts = (odoo_payload.get('filter_options') or {})
    semanas = sorted(set((contpaq_opts.get('semanas') or [])) | set((odoo_opts.get('semanas') or [])))
    sucursales = sorted({_sucursal_pretty(k) for k in sucursal_keys})

    return {
        'items': items,
        'resumen': resumen,
        'semanas': semanas,
        'sucursales': sucursales,
    }, None


@app.route('/api/contpaq/conciliacion', methods=['GET'])
@login_required
@requires_permission('contpaq', 'view')
def api_contpaq_conciliacion():
    """Conciliacion combinada: muestra pedidos de CONTPAQ y Odoo juntos en una sola lista."""
    try:
        sucursal_param = (request.args.get('sucursal') or '').strip()
        # Filtro por año: acota los pedidos al año elegido para que los totales no
        # sumen años anteriores. Se traduce a un rango de fechas que ya manejan los
        # builders. Si no se manda 'anio', respeta fecha_desde/fecha_hasta (compat).
        anio = (request.args.get('anio') or '').strip()
        fecha_desde_raw = (request.args.get('fecha_desde') or '').strip()
        fecha_hasta_raw = (request.args.get('fecha_hasta') or '').strip()
        if anio:
            try:
                y = int(anio)
                fecha_desde_raw = f"{y:04d}-01-01"
                fecha_hasta_raw = f"{y:04d}-12-31"
            except (TypeError, ValueError):
                pass
        # La sucursal se filtra por clave canonica en Python (no por SQL) para
        # que atrape variantes con acentos/mayusculas/prefijo SUC. No forma parte
        # de la cache: asi cambiar de sucursal es instantaneo sobre el mismo dataset.
        base_params = dict(
            q=(request.args.get('q') or '').strip(),
            folio=(request.args.get('folio') or '').strip(),
            cliente=(request.args.get('cliente') or '').strip(),
            titulo=(request.args.get('titulo') or '').strip(),
            fecha_desde_raw=fecha_desde_raw,
            fecha_hasta_raw=fecha_hasta_raw,
        )
        limit = request.args.get('limit', default=120, type=int)
        limit = max(1, min(int(limit or 120), 500))
        page = request.args.get('page', default=1, type=int)
        page = max(1, int(page or 1))

        signature = _conciliacion_dataset_signature()
        cache_key = None
        dataset = None
        if signature is not None:
            cache_key = (signature, tuple(sorted(base_params.items())))
            dataset = _conciliacion_cache_get(cache_key)

        if dataset is None:
            dataset, err_status = _build_conciliacion_combined_dataset(base_params)
            if err_status is not None:
                return jsonify(dataset), err_status
            if cache_key is not None:
                _conciliacion_cache_set(cache_key, dataset)

        items = dataset['items']

        # Filtrar por sucursal seleccionada (comparando claves canonicas).
        if sucursal_param:
            want_key = _sucursal_canonical_key(sucursal_param)
            if want_key:
                items = [
                    it for it in items
                    if _sucursal_canonical_key(it.get('sucursal')) == want_key
                ]

        total_records = len(items)
        offset = (page - 1) * limit
        page_items = items[offset:offset + limit]

        # El resumen se recalcula sobre los items ya filtrados (incluye sucursal),
        # para que los totales coincidan siempre con lo que se muestra.
        resumen = _compute_conciliacion_resumen_from_items(items)

        return jsonify({
            'items': page_items,
            'total': len(page_items),
            'resumen': resumen,
            'filter_options': {
                'semanas': dataset['semanas'],
                'sucursales': dataset['sucursales'],
                'anios': _conciliacion_available_years(),
            },
            'page': page,
            'limit': limit,
            'total_records': total_records,
            'total_pages': (total_records + limit - 1) // limit if total_records else 0,
            'has_prev': page > 1,
            'has_next': offset + limit < total_records,
            'fuente': 'combinada',
        }), 200
    except Exception as e:
        logger.error(f"Error consultando conciliacion combinada: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500


def _build_conciliacion_contpaq_response(
    *,
    q='',
    folio='',
    cliente='',
    sucursal='',
    titulo='',
    fecha_desde_raw='',
    fecha_hasta_raw='',
    limit=120,
    page=1,
    collect_all=False,
):
    """Arma la respuesta de conciliacion leyendo pedidos CONTPAQ. Devuelve (dict, status)."""
    try:
        d_fallback_days = max(1, min(int(os.getenv('CONTPAQ_D_FALLBACK_DAYS', '14') or '14'), 31))

        limit = max(1, min(int(limit or 120), 500))
        page = max(1, int(page or 1))

        pedidos_q = ContpaqPedido.query

        # Mostrar todos los clientes por defecto.
        # Si el usuario especifica cliente, filtrar por ese valor.
        if cliente:
            pedidos_q = pedidos_q.filter(ContpaqPedido.cliente.ilike(f"%{cliente}%"))

        if folio:
            like_folio = f"%{folio}%"
            pedidos_q = pedidos_q.filter(
                db.or_(
                    ContpaqPedido.doc_folio.ilike(like_folio),
                    ContpaqPedido.serie.ilike(like_folio),
                )
            )

        if sucursal:
            pedidos_q = pedidos_q.filter(ContpaqPedido.sucursal.ilike(f"%{sucursal}%"))

        if fecha_desde_raw:
            fecha_desde = _parse_date(fecha_desde_raw)
            if not fecha_desde:
                return {'error': 'fecha_desde invalida. Usa YYYY-MM-DD'}, 400
            pedidos_q = pedidos_q.filter(ContpaqPedido.fecha_documento >= datetime.combine(fecha_desde, datetime.min.time()))

        if fecha_hasta_raw:
            fecha_hasta = _parse_date(fecha_hasta_raw)
            if not fecha_hasta:
                return {'error': 'fecha_hasta invalida. Usa YYYY-MM-DD'}, 400
            pedidos_q = pedidos_q.filter(ContpaqPedido.fecha_documento < datetime.combine(fecha_hasta + timedelta(days=1), datetime.min.time()))

        # Busqueda global por terminos combinables (AND por termino, OR por columna).
        if q:
            terms = [t.strip() for t in q.split() if t.strip()]
            for term in terms:
                like = f"%{term}%"
                pedidos_q = pedidos_q.filter(
                    db.or_(
                        ContpaqPedido.titulo.ilike(like),
                        ContpaqPedido.periodo_semana.ilike(like),
                        ContpaqPedido.doc_folio.ilike(like),
                        ContpaqPedido.serie.ilike(like),
                        ContpaqPedido.sucursal.ilike(like),
                        ContpaqPedido.cliente.ilike(like),
                    )
                )

        ordered_q = pedidos_q.order_by(ContpaqPedido.fecha_documento.desc(), ContpaqPedido.id.desc())
        compare_pedidos = ordered_q.all()

        if titulo:
            compare_pedidos = [
                p for p in compare_pedidos
                if _contpaq_period_matches(
                    titulo,
                    p.titulo,
                    p.periodo_semana,
                    _contpaq_title_week_key(p.titulo, p.periodo_semana, p.fecha_documento),
                )
            ]

        # Serie D solo para comparar y como fuente de faltantes; NO se muestra como fila propia.
        display_pedidos = [
            p for p in compare_pedidos
            if not str(p.doc_folio or '').strip().upper().startswith('D')
        ]

        semana_options = _contpaq_unique_semana_options([
            _contpaq_title_week_key(p.titulo, p.periodo_semana, p.fecha_documento)
            for p in display_pedidos
        ])
        sucursal_options = sorted({
            str(p.sucursal or '').strip()
            for p in display_pedidos
            if str(p.sucursal or '').strip()
        })

        total_records = len(display_pedidos)
        compare_map = {p.document_id: p for p in compare_pedidos}
        compare_doc_ids = [p.document_id for p in compare_pedidos]
        if collect_all:
            page = 1
            offset = 0
            pedidos = display_pedidos
        else:
            offset = (page - 1) * limit
            pedidos = display_pedidos[offset:offset + limit]
        if not pedidos:
            return {
                'items': [],
                'total': 0,
                'page': page,
                'limit': limit,
                'total_records': total_records,
                'total_pages': (total_records + limit - 1) // limit if total_records else 0,
                'has_prev': page > 1,
                'has_next': offset + limit < total_records,
                'filter_options': {
                    'semanas': semana_options,
                    'sucursales': sucursal_options,
                },
                'fuente': 'contpaq',
            }, 200

        def _query_in_chunks(model_class, id_column, id_list, order_by_cols, chunk_size=500):
            """Evita IN(...) gigantes que saturan el temp space de PostgreSQL."""
            results = []
            for i in range(0, len(id_list), chunk_size):
                chunk = id_list[i:i + chunk_size]
                rows = model_class.query.filter(id_column.in_(chunk)).order_by(*order_by_cols).all()
                results.extend(rows)
            return results

        pedido_doc_ids = [p.document_id for p in pedidos]
        detalles = _query_in_chunks(
            ContpaqPedidoDetalle, ContpaqPedidoDetalle.document_id, pedido_doc_ids,
            [ContpaqPedidoDetalle.document_id.asc(), ContpaqPedidoDetalle.line_number.asc()]
        )
        compare_detalles = []
        if compare_doc_ids:
            compare_detalles = _query_in_chunks(
                ContpaqPedidoDetalle, ContpaqPedidoDetalle.document_id, compare_doc_ids,
                [ContpaqPedidoDetalle.document_id.asc(), ContpaqPedidoDetalle.line_number.asc()]
            )

        detalles_map = {}
        for d in detalles:
            detalles_map.setdefault(d.document_id, []).append(d)

        remisiones = []
        if compare_doc_ids:
            remisiones = _query_in_chunks(
                ContpaqRemision, ContpaqRemision.source_document_id, compare_doc_ids,
                [ContpaqRemision.fecha_documento.asc(), ContpaqRemision.id.asc()]
            )
        remision_doc_ids = [r.document_id for r in remisiones]

        remision_detalles_map = {}
        if remision_doc_ids:
            rem_details = _query_in_chunks(
                ContpaqRemisionDetalle, ContpaqRemisionDetalle.document_id, remision_doc_ids,
                [ContpaqRemisionDetalle.document_id.asc(), ContpaqRemisionDetalle.line_number.asc()]
            )
            for d in rem_details:
                remision_detalles_map.setdefault(d.document_id, []).append(d)

        remisiones_map = {}
        for r in remisiones:
            remisiones_map.setdefault(r.source_document_id, []).append(r)

        visible_claves = {
            _contpaq_norm_text(d.clave_producto)
            for d in (detalles + compare_detalles)
            if _contpaq_norm_text(d.clave_producto)
        }
        public_price_map = {}
        if visible_claves:
            public_price_map = {
                _contpaq_norm_text(row.clave_producto): float(row.precio_publico)
                for row in ContpaqPrecioPublico.query.filter(ContpaqPrecioPublico.clave_producto.in_(list(visible_claves))).all()
                if row.clave_producto is not None and row.precio_publico is not None
            }

        items = []
        total_partidas = 0
        total_importe = 0.0
        total_remisiones = 0
        semana_totales = {}

        def _norm_txt(v):
            return str(v or '').strip().upper()

        def _norm_num(v):
            try:
                f = float(str(v).replace(',', '').strip())
                if f.is_integer():
                    return str(int(f))
                return f"{f:.4f}".rstrip('0').rstrip('.')
            except Exception:
                return _norm_txt(v)

        import re as _re_sem

        def _sem_year_strip(s):
            """Quita el anno de 4 digitos al final: '13 AL 19 DE ENERO DE 2026' -> '13 AL 19 DE ENERO'."""
            v = _contpaq_norm_text(str(s or ''))
            return _re_sem.sub(r'\s+DE\s+\d{4}\s*$', '', v).strip()

        p_set = set()
        d_rows = []

        for d in compare_detalles:
            p_cmp = compare_map.get(d.document_id)
            if not p_cmp:
                continue

            semana_cmp_norm = _sem_year_strip(
                _contpaq_title_week_key(p_cmp.titulo, p_cmp.periodo_semana, p_cmp.fecha_documento)
            )
            sucursal_cmp_norm = _norm_txt(p_cmp.sucursal)
            folio_cmp_norm = _norm_txt(p_cmp.doc_folio)

            row_key_cmp = (
                _norm_txt(d.clave_producto),
                _norm_num(d.cantidad),
                semana_cmp_norm,
                sucursal_cmp_norm,
            )

            # Solo los pedidos P confirman existencia; D solo sirven como fuente de inyeccion.
            if not folio_cmp_norm.startswith('D'):
                p_set.add(row_key_cmp)

            if folio_cmp_norm.startswith('D'):
                d_rows.append({
                    'source_document_id': p_cmp.document_id,
                    'folio': p_cmp.doc_folio,
                    'semana': p_cmp.periodo_semana,
                    'semana_norm': semana_cmp_norm,
                    'sucursal': p_cmp.sucursal,
                    'clave_producto': d.clave_producto,
                    'descripcion': d.descripcion,
                    'cantidad': d.cantidad,
                    'precio_unitario': d.precio_unitario,
                    'line_number': d.line_number,
                    'fecha_documento': p_cmp.fecha_documento.date() if p_cmp.fecha_documento else None,
                    'clave_norm': _norm_txt(d.clave_producto),
                    'cantidad_norm': _norm_num(d.cantidad),
                    'sucursal_norm': sucursal_cmp_norm,
                    'key': row_key_cmp,
                })

        for p in pedidos:
            pedido_rows = []
            pedido_total = 0.0
            semana_norm = _contpaq_title_week_key(p.titulo, p.periodo_semana, p.fecha_documento)
            sucursal_norm = _norm_txt(p.sucursal)
            folio_norm = _norm_txt(p.doc_folio)

            for d in detalles_map.get(p.document_id, []):
                partida_total = float(d.total_partida or 0)
                cantidad_num = _to_float(d.cantidad)
                precio_publico = public_price_map.get(_norm_txt(d.clave_producto))
                total_publico = round(cantidad_num * precio_publico, 2) if precio_publico is not None else None
                diferencia_publico = round(total_publico - partida_total, 2) if total_publico is not None else None
                pedido_total += partida_total
                total_partidas += 1

                pedido_rows.append({
                    'line_number': d.line_number,
                    'clave_producto': d.clave_producto,
                    'descripcion': d.descripcion,
                    'cantidad': d.cantidad,
                    'precio_unitario': d.precio_unitario,
                    'total_partida': d.total_partida,
                    'precio_publico_unitario': precio_publico,
                    'total_partida_precio_publico': total_publico,
                    'diferencia_precio_publico': diferencia_publico,
                    'serie': p.serie,
                    'folio': p.doc_folio,
                    'fecha_documento': p.fecha_documento.isoformat() if p.fecha_documento else None,
                    'es_inyectado': False,
                })

            remisiones_rows = []
            for r in remisiones_map.get(p.document_id, []):
                total_remisiones += 1
                rr = {
                    'document_id': r.document_id,
                    'doc_folio': r.doc_folio,
                    'sucursal': r.sucursal,
                    'fecha_documento': r.fecha_documento.isoformat() if r.fecha_documento else None,
                    'detalles': [],
                }
                for d in remision_detalles_map.get(r.document_id, []):
                    rr['detalles'].append({
                        'line_number': d.line_number,
                        'clave_producto': d.clave_producto,
                        'descripcion': d.descripcion,
                        'cantidad': d.cantidad,
                        'precio_unitario': d.precio_unitario,
                        'total_partida': d.total_partida,
                    })
                remisiones_rows.append(rr)

            total_importe += pedido_total
            sem = p.periodo_semana or 'SIN SEMANA'
            if sem not in semana_totales:
                semana_totales[sem] = {'pedidos': 0, 'total': 0.0}
            semana_totales[sem]['pedidos'] += 1
            semana_totales[sem]['total'] += pedido_total

            items.append({
                'document_id': p.document_id,
                'doc_folio': p.doc_folio,
                'serie': p.serie,
                'cliente': p.cliente,
                'sucursal': p.sucursal,
                'titulo': p.titulo,
                'periodo_semana': p.periodo_semana,
                'semana_match_key': semana_norm,
                'fecha_documento': p.fecha_documento.isoformat() if p.fecha_documento else None,
                'pedido_total': round(pedido_total, 2),
                'detalles': pedido_rows,
                'remisiones': remisiones_rows,
                'fuente': 'contpaq',
            })

        indice_q = ContpaqSucursalIndice.query
        if sucursal:
            indice_q = indice_q.filter(ContpaqSucursalIndice.sucursal.ilike(f"%{sucursal}%"))
        if titulo:
            indice_q = indice_q.filter(ContpaqSucursalIndice.semana.ilike(f"%{titulo}%"))
        if fecha_desde_raw or fecha_hasta_raw:
            # El indice principal esta basado en semana; no forzamos fecha aqui.
            pass

        indice_rows = indice_q.all()
        faltantes_indice = []
        for idx in indice_rows:
            row_key = (
                _contpaq_norm_text(idx.clave_producto),
                _contpaq_norm_qty(idx.cantidad),
                _sem_year_strip(idx.semana),
                _contpaq_norm_text(idx.sucursal),
            )
            if row_key in p_set:
                continue

            d_match = next((row for row in d_rows if row['key'] == row_key), None)

            # Si no hay match exacto por semana, buscar en serie D por fecha dentro de rango
            # con misma clave, cantidad y sucursal.
            if d_match is None:
                idx_fecha = None
                if idx.fecha_documento:
                    try:
                        idx_fecha = idx.fecha_documento.date() if hasattr(idx.fecha_documento, 'date') else idx.fecha_documento
                    except Exception:
                        idx_fecha = None

                if idx_fecha is not None:
                    clave_norm = _contpaq_norm_text(idx.clave_producto)
                    cantidad_norm = _contpaq_norm_qty(idx.cantidad)
                    sucursal_norm = _contpaq_norm_text(idx.sucursal)

                    d_match = next(
                        (
                            row for row in d_rows
                            if row.get('fecha_documento')
                            and row.get('clave_norm') == clave_norm
                            and row.get('cantidad_norm') == cantidad_norm
                            and row.get('sucursal_norm') == sucursal_norm
                            and abs((row['fecha_documento'] - idx_fecha).days) <= d_fallback_days
                        ),
                        None,
                    )

            if d_match is None:
                clave_norm = _contpaq_norm_text(idx.clave_producto)
                cantidad_norm = _contpaq_norm_qty(idx.cantidad)
                sucursal_norm = _contpaq_norm_text(idx.sucursal)
                d_match = next(
                    (
                        row for row in d_rows
                        if row.get('clave_norm') == clave_norm
                        and row.get('cantidad_norm') == cantidad_norm
                        and row.get('sucursal_norm') == sucursal_norm
                        and not (row.get('semana_norm') or '').strip()
                    ),
                    None,
                )

            faltantes_indice.append({
                'source_document_id': d_match['source_document_id'] if d_match else None,
                'folio': d_match['folio'] if d_match else (idx.folio or ''),
                'semana': idx.semana,
                'sucursal': idx.sucursal,
                'clave_producto': idx.clave_producto,
                'descripcion': d_match['descripcion'] if d_match else idx.descripcion,
                'cantidad': d_match['cantidad'] if d_match else idx.cantidad,
                'precio_unitario': d_match['precio_unitario'] if d_match else None,
                'origen': 'PEDIDO_D' if d_match else 'INDICE_SUCURSALES',
                'line_number_origen': d_match['line_number'] if d_match else None,
            })

        faltantes_inyectados = 0
        injected_remision_detail_keys = set()
        # Inyectamos faltantes solo en la primera pagina para mantener navegacion estable.
        if page == 1:
            # Mapa de destino por (semana_sin_año, sucursal). Preferimos folios P- cuando existan.
            ws_targets = {}
            faltantes_line_counter = {}
            for i, item in enumerate(items):
                sem_k = _sem_year_strip(item.get('semana_match_key') or item.get('periodo_semana') or '')
                key = (sem_k, _norm_txt(item.get('sucursal')))
                folio_norm = _norm_txt(item.get('doc_folio'))
                if key not in ws_targets:
                    ws_targets[key] = i
                if folio_norm.startswith('P-'):
                    ws_targets[key] = i

            for f in faltantes_indice:
                sem_f = _sem_year_strip(str(f.get('semana') or '').strip())
                suc_f = str(f.get('sucursal') or '').strip()
                key = (sem_f, _norm_txt(suc_f))

                target_idx = ws_targets.get(key)
                if target_idx is None:
                    # No crear tarjetas artificiales; solo inyectar en pedidos visibles.
                    continue

                cantidad = _to_float(f.get('cantidad'))
                precio = _to_float(f.get('precio_unitario'))
                total_partida = round(cantidad * precio, 2)
                precio_publico = public_price_map.get(_norm_txt(f.get('clave_producto')))
                total_publico = round(cantidad * precio_publico, 2) if precio_publico is not None else None
                diferencia_publico = round(total_publico - total_partida, 2) if total_publico is not None else None

                target_item = items[target_idx]
                target_item.setdefault('detalles', [])
                faltantes_line_counter[target_idx] = faltantes_line_counter.get(target_idx, 0) + 1
                target_item['detalles'].append({
                    'line_number': f"F-{faltantes_line_counter[target_idx]}",
                    'clave_producto': f.get('clave_producto'),
                    'descripcion': f.get('descripcion') or 'FALTANTE DESDE INDICE',
                    'cantidad': f.get('cantidad'),
                    'precio_unitario': precio,
                    'total_partida': total_partida,
                    'precio_publico_unitario': precio_publico,
                    'total_partida_precio_publico': total_publico,
                    'diferencia_precio_publico': diferencia_publico,
                    'origen_faltante': f.get('origen'),
                    'folio_origen': f.get('folio') or '',
                    'serie': 'D' if f.get('origen') == 'PEDIDO_D' else '',
                    'folio': f.get('folio') or '',
                    'fecha_documento': None,
                    'es_inyectado': True,
                })

                target_item['pedido_total'] = round(_to_float(target_item.get('pedido_total')) + total_partida, 2)
                target_item['es_faltante'] = True

                # Si el faltante vino de un pedido D, inyectar solo las partidas de remision
                # que correspondan a esa partida, no la remision completa.
                source_doc_id = f.get('source_document_id')
                if source_doc_id:
                    faltante_clave_norm = _norm_txt(f.get('clave_producto'))
                    faltante_cantidad_norm = _norm_num(f.get('cantidad'))
                    target_item.setdefault('remisiones', [])
                    for r in remisiones_map.get(source_doc_id, []):
                        exact_qty_details = []
                        clave_only_details = []
                        for rd in remision_detalles_map.get(r.document_id, []):
                            if _norm_txt(rd.clave_producto) != faltante_clave_norm:
                                continue

                            detail_payload = {
                                'line_number': rd.line_number,
                                'clave_producto': rd.clave_producto,
                                'descripcion': rd.descripcion,
                                'cantidad': rd.cantidad,
                                'precio_unitario': rd.precio_unitario,
                                'total_partida': rd.total_partida,
                                'es_inyectado': True,
                            }

                            if _norm_num(rd.cantidad) == faltante_cantidad_norm:
                                exact_qty_details.append(detail_payload)
                            else:
                                clave_only_details.append(detail_payload)

                        matched_details = exact_qty_details or clave_only_details
                        filtered_details = []
                        for md in matched_details:
                            detail_key = (target_idx, r.document_id, md.get('line_number'))
                            if detail_key in injected_remision_detail_keys:
                                continue
                            injected_remision_detail_keys.add(detail_key)
                            filtered_details.append(md)

                        if filtered_details:
                            target_item['remisiones'].append({
                                'document_id': r.document_id,
                                'doc_folio': r.doc_folio,
                                'sucursal': r.sucursal,
                                'fecha_documento': r.fecha_documento.isoformat() if r.fecha_documento else None,
                                'detalles': filtered_details,
                                'es_inyectado': True,
                            })

                total_partidas += 1
                total_importe += total_partida
                sem_target = target_item.get('periodo_semana') or sem_f
                if sem_target not in semana_totales:
                    semana_totales[sem_target] = {'pedidos': 0, 'total': 0.0}
                semana_totales[sem_target]['total'] += total_partida
                faltantes_inyectados += 1

        items.sort(
            key=lambda x: (
                x.get('periodo_semana') or '',
                x.get('fecha_documento') or '',
                str(x.get('doc_folio') or ''),
            ),
            reverse=True,
        )

        resumen = {
            'total_pedidos': len(items),
            'total_partidas': total_partidas,
            'total_importe': round(total_importe, 2),
            'total_remisiones': total_remisiones,
            'total_faltantes_desde_indice': len(faltantes_indice),
            'faltantes_inyectados': faltantes_inyectados,
            'totales_por_semana': [
                {
                    'semana': semana,
                    'pedidos': vals['pedidos'],
                    'total': round(vals['total'], 2),
                }
                for semana, vals in sorted(semana_totales.items(), key=lambda x: str(x[0]), reverse=True)
            ],
        }

        return {
            'items': items,
            'total': len(items),
            'resumen': resumen,
            'filter_options': {
                'semanas': semana_options,
                'sucursales': sucursal_options,
            },
            'page': page,
            'limit': limit,
            'total_records': total_records,
            'total_pages': (total_records + limit - 1) // limit if total_records else 0,
            'has_prev': page > 1,
            'has_next': offset + limit < total_records,
            'fuente': 'contpaq',
        }, 200
    except Exception as e:
        logger.error(f"Error consultando conciliacion CONTPAQ: {e}", exc_info=True)
        return {'error': str(e)}, 500


@app.route('/api/contpaq/conciliacion-contpaq-legacy', methods=['GET'])
@login_required
@requires_permission('contpaq', 'view')
def api_contpaq_conciliacion_legacy():
    """Respaldo: conciliacion leyendo solo pedidos CONTPAQ."""
    payload, status = _build_conciliacion_contpaq_response(
        q=(request.args.get('q') or '').strip(),
        folio=(request.args.get('folio') or '').strip(),
        cliente=(request.args.get('cliente') or '').strip(),
        sucursal=(request.args.get('sucursal') or '').strip(),
        titulo=(request.args.get('titulo') or '').strip(),
        fecha_desde_raw=(request.args.get('fecha_desde') or '').strip(),
        fecha_hasta_raw=(request.args.get('fecha_hasta') or '').strip(),
        limit=request.args.get('limit', default=120, type=int),
        page=request.args.get('page', default=1, type=int),
    )
    return jsonify(payload), status


@app.route('/api/contpaq/sync', methods=['POST'])
@login_required
@requires_permission('contpaq', 'edit')
def api_contpaq_sync():
    """Ejecuta sincronizacion manual de pedidos Odoo para conciliacion."""
    result = _run_odoo_sync_guarded(trigger=f"manual_conciliacion:{session.get('user')}")
    status = 200 if result.get('ok') else 500
    return jsonify(result), status


@app.route('/api/contpaq/odoo/sync/last', methods=['GET'])
@login_required
@requires_permission('contpaq', 'view')
def api_contpaq_odoo_sync_last():
    """Estado de la ultima sincronizacion Odoo (para la vista de conciliacion)."""
    try:
        _ensure_odoo_tables()
        run = OdooSyncRun.query.order_by(OdooSyncRun.id.desc()).first()
        return jsonify({
            'ok': True,
            'last_sync': run.to_dict() if run else None,
            'auto_enabled': ODOO_SYNC_ENABLED,
            'interval_minutes': ODOO_SYNC_INTERVAL_MINUTES,
        })
    except Exception as exc:
        return jsonify({'ok': False, 'error': str(exc)}), 500


@app.route('/api/contpaq/sync/push', methods=['POST'])
def api_contpaq_sync_push():
    """Recibe payload desde un agente local (inhouse) y actualiza la BD nube."""
    ok, err = _require_sync_key()
    if not ok:
        return err

    payload = request.get_json(silent=True) or {}
    if not isinstance(payload, dict):
        return jsonify({'error': 'Payload invalido'}), 400

    sync_run = None
    try:
        _ensure_contpaq_notas_venta_schema()
        sync_run = ContpaqSyncRun(
            status='running',
            started_at=datetime.utcnow(),
            message='started_by=push_api'
        )
        db.session.add(sync_run)
        db.session.commit()

        stats = _upsert_contpaq_data(payload)
        sync_run.status = 'success'
        sync_run.finished_at = datetime.utcnow()
        sync_run.message = 'push_api completed'
        sync_run.pedidos_upserted = stats['pedidos_upserted']
        sync_run.pedido_detalles_upserted = stats['pedido_detalles_upserted']
        sync_run.remisiones_upserted = stats['remisiones_upserted']
        sync_run.remision_detalles_upserted = stats['remision_detalles_upserted']
        sync_run.notas_venta_upserted = stats['notas_venta_upserted']
        db.session.commit()
        return jsonify({'ok': True, 'run_id': sync_run.id, 'stats': stats}), 200
    except Exception as exc:
        logger.error('[CONTPAQ] Error push agente: %s', exc, exc_info=True)
        db.session.rollback()
        run_id = getattr(sync_run, 'id', None)
        if run_id:
            try:
                run = ContpaqSyncRun.query.get(run_id)
                if run:
                    run.status = 'error'
                    run.finished_at = datetime.utcnow()
                    run.message = str(exc)[:2000]
                    db.session.commit()
            except Exception:
                db.session.rollback()
        return jsonify({'ok': False, 'error': str(exc), 'run_id': run_id}), 500


@app.route('/api/empaque/sync/push', methods=['POST'])
def api_empaque_sync_push():
    """Recibe snapshot de Empaque360 (planta MySQL) y actualiza espejo en nube."""
    ok, err = _require_sync_key()
    if not ok:
        return err

    payload = request.get_json(silent=True) or {}
    if not isinstance(payload, dict):
        return jsonify({'error': 'Payload invalido'}), 400

    try:
        if not _ensure_empaque_tables():
            return jsonify({'ok': False, 'error': 'No se pudieron crear tablas Empaque'}), 500
        stats = _upsert_empaque_data(payload)
        # Cierra huecos: pedidos viejos sin ficha también reciben clave sola
        recon = _reconcile_empaque_clientes_from_pedidos()
        stats['clientes_reconcile'] = recon
        db.session.commit()
        return jsonify({'ok': True, 'stats': stats}), 200
    except Exception as exc:
        db.session.rollback()
        logger.error('[EMPAQUE] Error push agente: %s', exc, exc_info=True)
        return jsonify({'ok': False, 'error': str(exc)}), 500


@app.route('/empaque/clientes', methods=['GET'])
@login_required
@requires_any_permission([('empaque', 'view'), ('empaque', 'edit')])
def empaque_clientes_page():
    """Módulo interno: claves únicas de acceso al portal /seguimiento."""
    _ensure_empaque_tables()
    # Auto: si hay pedidos sin ficha, genera clave sin botón
    try:
        recon = _reconcile_empaque_clientes_from_pedidos()
        if recon.get('created'):
            db.session.commit()
        else:
            db.session.rollback()
    except Exception:
        db.session.rollback()
    q = (request.args.get('q') or '').strip()
    clientes_q = EmpaqueCliente.query
    if q:
        like = f'%{q}%'
        clientes_q = clientes_q.filter(db.or_(
            EmpaqueCliente.customer_name.ilike(like),
            EmpaqueCliente.customer_code.ilike(like),
            EmpaqueCliente.access_code.ilike(like),
        ))
    clientes = clientes_q.order_by(
        EmpaqueCliente.customer_name.asc().nullslast(),
        EmpaqueCliente.id.asc(),
    ).all()

    # Conteos de pedidos por customer_code
    counts = dict(
        db.session.query(
            EmpaquePedido.customer_code,
            func.count(EmpaquePedido.id),
        ).group_by(EmpaquePedido.customer_code).all()
    )
    rows = []
    for c in clientes:
        rows.append({
            **c.to_dict(),
            'pedidos_count': int(counts.get(c.customer_code) or 0),
        })

    portal_url = request.url_root.rstrip('/') + '/seguimiento'
    return render_template(
        'empaque_clientes.html',
        clientes=rows,
        q=q,
        portal_url=portal_url,
    )


@app.route('/api/empaque/clientes/sync-from-pedidos', methods=['POST'])
@login_required
@requires_permission('empaque', 'edit')
def api_empaque_clientes_sync_from_pedidos():
    """Compat: las claves ya se generan solas al sync / al abrir el módulo."""
    _ensure_empaque_tables()
    recon = _reconcile_empaque_clientes_from_pedidos()
    db.session.commit()
    return jsonify({'ok': True, **recon})


@app.route('/api/empaque/clientes/<int:cliente_id>/regenerar', methods=['POST'])
@login_required
@requires_permission('empaque', 'edit')
def api_empaque_cliente_regenerar(cliente_id):
    _ensure_empaque_tables()
    cliente = EmpaqueCliente.query.get_or_404(cliente_id)
    cliente.access_code = _generate_empaque_access_code()
    cliente.updated_at = datetime.utcnow()
    db.session.commit()
    return jsonify({'ok': True, 'cliente': cliente.to_dict()})


@app.route('/api/empaque/clientes/<int:cliente_id>/toggle', methods=['POST'])
@login_required
@requires_permission('empaque', 'edit')
def api_empaque_cliente_toggle(cliente_id):
    _ensure_empaque_tables()
    cliente = EmpaqueCliente.query.get_or_404(cliente_id)
    cliente.activo = not bool(cliente.activo)
    cliente.updated_at = datetime.utcnow()
    db.session.commit()
    return jsonify({'ok': True, 'cliente': cliente.to_dict()})


@app.route('/api/empaque/clientes', methods=['POST'])
@login_required
@requires_permission('empaque', 'edit')
def api_empaque_cliente_crear():
    """Alta manual de cliente + generación de clave."""
    _ensure_empaque_tables()
    data = request.get_json(silent=True) or {}
    customer_code = _normalize_empaque_clave(data.get('customer_code'))
    customer_name = str(data.get('customer_name') or '').strip()[:255] or None
    if not customer_code:
        return jsonify({'ok': False, 'error': 'customer_code requerido'}), 400
    existing = EmpaqueCliente.query.filter(
        func.lower(EmpaqueCliente.customer_code) == customer_code.lower()
    ).first()
    if existing:
        return jsonify({'ok': False, 'error': 'Ya existe un cliente con ese código fuente', 'cliente': existing.to_dict()}), 409
    cliente, _ = _ensure_empaque_cliente(customer_code, customer_name)
    db.session.commit()
    return jsonify({'ok': True, 'cliente': cliente.to_dict()}), 201


@app.route('/api/contpaq/supplier_ot/sync/push', methods=['POST'])
def api_contpaq_supplier_ot_sync_push():
    """Recibe OTs de compra pendientes desde un agente local independiente."""
    ok, err = _require_sync_key()
    if not ok:
        return err

    payload = request.get_json(silent=True) or {}
    if not isinstance(payload, dict):
        return jsonify({'error': 'Payload invalido'}), 400

    try:
        stats = _upsert_contpaq_supplier_ot_data(payload)
        db.session.commit()
        return jsonify({'ok': True, 'stats': stats}), 200
    except Exception as exc:
        db.session.rollback()
        logger.error(f'[CONTPAQ-OT] Error de sincronizacion: {exc}', exc_info=True)
        return jsonify({'ok': False, 'error': str(exc)}), 500


@app.route('/api/contpaq/supplier_ot/disponibles', methods=['GET'])
@login_required
@requires_any_permission([('hojas_entregas', 'view'), ('hojas', 'view'), ('catalog', 'view')])
def api_contpaq_supplier_ot_disponibles():
    clave = (request.args.get('clave') or '').strip().upper()
    cantidad = request.args.get('cantidad')
    if not clave:
        return jsonify({'items': []})

    items = _contpaq_supplier_ot_options_for_product_key(clave, requested_qty=cantidad)
    response = {
        'items': items,
        'total': len(items),
        'key_status': _contpaq_supplier_ot_key_status(clave),
    }
    if not items:
        response['suggested_keys'] = _contpaq_supplier_ot_top_keys(limit=10)
    return jsonify(response)


@app.route('/api/contpaq/maquinaria/sync/push', methods=['POST'])
def api_contpaq_maquinaria_sync_push():
    """Recibe solo pedidos de maquinaria desde el agente local, separado de conciliacion."""
    ok, err = _require_sync_key()
    if not ok:
        return err

    payload = request.get_json(silent=True) or {}
    if not isinstance(payload, dict):
        return jsonify({'error': 'Payload invalido'}), 400

    try:
        stats = _upsert_maquinaria_contpaq_data(payload)
        db.session.commit()
        return jsonify({'ok': True, 'stats': stats}), 200
    except Exception as exc:
        db.session.rollback()
        logger.error(f'[CONTPAQ-MAQUINARIA] Error de sincronizacion: {exc}', exc_info=True)
        return jsonify({'ok': False, 'error': str(exc)}), 500


@app.route('/api/contpaq/existencia/sync/push', methods=['POST'])
def api_contpaq_existencia_sync_push():
    """Recibe existencias por sucursal desde agente local CONTPAQ."""
    ok, err = _require_sync_key()
    if not ok:
        return err

    payload = request.get_json(silent=True) or {}
    if not isinstance(payload, dict):
        return jsonify({'error': 'Payload invalido'}), 400

    try:
        stats = _upsert_contpaq_existencia_data(payload)
        db.session.commit()
        return jsonify({'ok': True, 'stats': stats}), 200
    except Exception as exc:
        db.session.rollback()
        logger.error(f'[CONTPAQ-EXISTENCIA] Error de sincronizacion: {exc}', exc_info=True)
        return jsonify({'ok': False, 'error': str(exc)}), 500


@app.route('/api/contpaq/sync/last', methods=['GET'])
@login_required
@requires_permission('contpaq', 'view')
def api_contpaq_sync_last():
    """Devuelve estado de la ultima sincronizacion para mostrar en la vista."""
    run = ContpaqSyncRun.query.order_by(ContpaqSyncRun.id.desc()).first()
    return jsonify({'last_sync': run.to_dict() if run else None})


MAQUINARIA_TABLES = [
    'maquinaria_pedidos',
    'maquinaria_contpaq_pedidos',
    'maquinaria_contpaq_pedidos_detalle',
    'maquinaria_boms',
    'maquinaria_bom_componentes',
    'maquinaria_bom_procesos',
    'maquinaria_ordenes_trabajo',
    'maquinaria_orden_bom_items',
    'maquinaria_orden_procesos',
    'maquinaria_calidad_registros',
    'maquinaria_series',
    'maquinaria_almacen_resguardos',
]


def _maquinaria_tables_status_raw():
    try:
        inspector = inspect(db.engine)
        missing = [table for table in MAQUINARIA_TABLES if not inspector.has_table(table)]
        return len(missing) == 0, missing
    except Exception:
        return False, list(MAQUINARIA_TABLES)


def _run_maquinaria_base_ddl():
    """Ejecuta DDL idempotente del modulo Maquinaria (mismo script del repo)."""
    try:
        from create_maquinaria_ensamble_tables import SQL as maquinaria_ddl
        with db.engine.begin() as conn:
            conn.execute(text(maquinaria_ddl))
        return True
    except Exception as exc:
        logger.error(f'No se pudo ejecutar DDL base de Maquinaria: {exc}', exc_info=True)
        try:
            db.session.rollback()
        except Exception:
            pass
        return False


def _ensure_maquinaria_tables():
    """Crea tablas/columnas faltantes del modulo y devuelve estado actualizado."""
    ready, missing = _maquinaria_tables_status_raw()
    if ready:
        return True, []

    try:
        base_core = {
            'maquinaria_pedidos', 'maquinaria_boms', 'maquinaria_ordenes_trabajo',
            'maquinaria_contpaq_pedidos', 'maquinaria_contpaq_pedidos_detalle',
        }
        if missing and any(table in base_core for table in missing):
            _run_maquinaria_base_ddl()
        else:
            # Tablas base ya existen: solo extensiones nuevas (OT snapshot, procesos BOM).
            _ensure_maquinaria_ordenes_extension_tables()
            _ensure_maquinaria_bom_procesos_table()

        ready_after, missing_after = _maquinaria_tables_status_raw()
        if not ready_after:
            # Segundo intento: DDL completo por si faltan varias tablas a la vez.
            _run_maquinaria_base_ddl()
            _ensure_maquinaria_ordenes_extension_tables()
            _ensure_maquinaria_bom_procesos_table()
            ready_after, missing_after = _maquinaria_tables_status_raw()
        return ready_after, missing_after
    except Exception as exc:
        logger.error(f'No se pudieron asegurar tablas de Maquinaria: {exc}', exc_info=True)
        try:
            db.session.rollback()
        except Exception:
            pass
        return _maquinaria_tables_status_raw()


def _maquinaria_tables_status():
    return _ensure_maquinaria_tables()


def _ensure_maquinaria_ordenes_extension_tables():
    """Asegura tablas extendidas de OT (snapshot BOM + procesos OT) sin depender de migración manual."""
    try:
        MaquinariaOrdenProceso.__table__.create(bind=db.engine, checkfirst=True)
        MaquinariaOrdenBOMItem.__table__.create(bind=db.engine, checkfirst=True)
        # Columnas nuevas del flujo de OT por solicitud (idempotente en Postgres)
        alter_stmts = [
            "ALTER TABLE maquinaria_ordenes_trabajo ADD COLUMN IF NOT EXISTS solicitud_id INTEGER",
            "ALTER TABLE maquinaria_ordenes_trabajo ADD COLUMN IF NOT EXISTS descripcion TEXT",
            "ALTER TABLE maquinaria_ordenes_trabajo ADD COLUMN IF NOT EXISTS bom_id INTEGER",
            "ALTER TABLE maquinaria_ordenes_trabajo ADD COLUMN IF NOT EXISTS orden_compra_name VARCHAR(120)",
            "ALTER TABLE maquinaria_ordenes_trabajo ADD COLUMN IF NOT EXISTS orden_compra_odoo_id BIGINT",
            "ALTER TABLE maquinaria_ordenes_trabajo ADD COLUMN IF NOT EXISTS cliente VARCHAR(255)",
            "ALTER TABLE maquinaria_orden_bom_items ADD COLUMN IF NOT EXISTS proceso_id INTEGER",
        ]
        with db.engine.begin() as conn:
            for stmt in alter_stmts:
                try:
                    conn.execute(text(stmt))
                except Exception as col_exc:
                    logger.warning(f"OT extensión columna omitida: {col_exc}")
        return True
    except Exception as exc:
        logger.error(f"No se pudieron asegurar tablas OT extensión: {exc}", exc_info=True)
        try:
            db.session.rollback()
        except Exception:
            pass
        return False


def _ensure_maquinaria_bom_procesos_table():
    try:
        MaquinariaBOMProceso.__table__.create(bind=db.engine, checkfirst=True)
        return True
    except Exception as exc:
        logger.error(f'No se pudo asegurar tabla maquinaria_bom_procesos: {exc}', exc_info=True)
        try:
            db.session.rollback()
        except Exception:
            pass
        return False


def _maquinaria_bom_for_clave(clave_maquina):
    clave_norm = (clave_maquina or '').strip().upper()
    if not clave_norm:
        return None
    return MaquinariaBOM.query.filter(func.upper(func.trim(MaquinariaBOM.clave_maquina)) == clave_norm).first()


@app.route('/maquinaria/pedidos')
@login_required
def maquinaria_pedidos_page():
    # Modulo de pedidos CONTPAQ retirado: ahora los pedidos vienen de Odoo.
    return redirect(url_for('maquinaria_odoo_page'))


@app.route('/maquinaria/pedidos-contpaq-legacy')
@login_required
@requires_any_permission([('maquinaria_pedidos', 'view'), ('maquinaria_pedidos', 'edit'), ('maquinaria_pedidos', 'create'), ('maquinaria_pedidos', 'update'), ('maquinaria_pedidos', 'delete')])
def maquinaria_pedidos_legacy_page():
    ready, missing = _maquinaria_tables_status()
    pedidos_locales = []
    if ready:
        pedidos_locales = MaquinariaPedido.query.order_by(MaquinariaPedido.id.desc()).limit(50).all()
    pedidos_locales_data = [p.to_dict() for p in pedidos_locales]
    pedidos_contpaq = []
    if ready:
        pedidos_contpaq = MaquinariaContpaqPedido.query.filter(
            MaquinariaContpaqPedido.cancelled.is_(False),
            MaquinariaContpaqPedido.deleted.is_(False),
        ).order_by(MaquinariaContpaqPedido.date_document.desc(), MaquinariaContpaqPedido.id.desc()).limit(80).all()
    pedidos_contpaq_selector_data = []
    for p in pedidos_contpaq:
        detalles = []
        for d in (p.detalles or []):
            detalles.append({
                'line_number': d.line_number,
                'product_key': d.product_key or '',
                'description': d.description or '',
                'quantity': float(d.quantity or 0),
                'unit_price': float(d.unit_price or 0),
                'total_item': float(d.total_item or 0),
            })
        pedidos_contpaq_selector_data.append({
            'document_id': int(p.document_id or 0),
            'folio': p.folio or '',
            'serie': p.serie or 'E',
            'cliente': p.business_entity_name or '',
            'sucursal': p.sucursal or p.depot_name or '',
            'fecha_documento': p.date_document.strftime('%Y-%m-%d') if p.date_document else '',
            'titulo': p.title or '',
            'comentarios': p.comments or '',
            'detalles': detalles,
        })
    boms = MaquinariaBOM.query.order_by(MaquinariaBOM.clave_maquina.asc()).all() if ready else []
    return render_template(
        'maquinaria_pedidos.html',
        setup_required=not ready,
        missing_tables=missing,
        pedidos_locales=pedidos_locales,
        pedidos_locales_data=pedidos_locales_data,
        pedidos_contpaq=pedidos_contpaq,
        pedidos_contpaq_selector_data=pedidos_contpaq_selector_data,
        boms=boms,
    )


@app.route('/maquinaria/pedidos/create', methods=['POST'])
@login_required
@requires_any_permission([('maquinaria_pedidos', 'create'), ('maquinaria_pedidos', 'edit'), ('maquinaria_pedidos', 'update')])
def maquinaria_pedidos_create():
    ready, _missing = _maquinaria_tables_status()
    if not ready:
        return redirect(url_for('maquinaria_pedidos_page'))

    pedido_id = request.form.get('pedido_id', type=int)
    pedido = MaquinariaPedido.query.get(pedido_id) if pedido_id else None

    folio = _clean_nullable_text(request.form.get('folio_interno', ''))
    if not folio:
        folio = pedido.folio_interno if pedido else f"MEP-{datetime.utcnow().strftime('%Y%m%d-%H%M%S')}"

    clave_maquina = _clean_nullable_text(request.form.get('clave_maquina', ''))
    if not clave_maquina:
        return redirect(url_for('maquinaria_pedidos_page'))

    existing = MaquinariaPedido.query.filter_by(folio_interno=folio).first()
    if existing and (not pedido or existing.id != pedido.id):
        return redirect(url_for('maquinaria_pedidos_page'))

    if pedido is None:
        pedido = MaquinariaPedido(
            folio_interno=folio,
            created_by=session.get('user'),
        )
        db.session.add(pedido)

    pedido.folio_interno = folio
    pedido.contpaq_document_id = request.form.get('contpaq_document_id', type=int)
    pedido.cliente = _clean_nullable_text(request.form.get('cliente', ''))
    pedido.clave_maquina = clave_maquina
    pedido.descripcion_maquina = _clean_nullable_text(request.form.get('descripcion_maquina', ''))
    pedido.cantidad = max(1, request.form.get('cantidad', type=int) or 1)
    pedido.estado = _clean_nullable_text(request.form.get('estado', 'abierto')) or 'abierto'
    pedido.notas = _clean_nullable_text(request.form.get('notas', ''))

    db.session.commit()
    return redirect(url_for('maquinaria_pedidos_page'))


@app.route('/maquinaria/pedidos/<int:pedido_id>/delete', methods=['POST'])
@login_required
@requires_any_permission([('maquinaria_pedidos', 'delete'), ('maquinaria_pedidos', 'edit'), ('maquinaria_pedidos', 'update')])
def maquinaria_pedidos_delete(pedido_id):
    ready, _missing = _maquinaria_tables_status()
    if not ready:
        return redirect(url_for('maquinaria_pedidos_page'))

    pedido = MaquinariaPedido.query.get_or_404(pedido_id)
    if pedido.ordenes_trabajo or pedido.series:
        return redirect(url_for('maquinaria_pedidos_page'))

    db.session.delete(pedido)
    db.session.commit()
    return redirect(url_for('maquinaria_pedidos_page'))


@app.route('/maquinaria/ordenes-trabajo')
@login_required
@requires_any_permission([('maquinaria_ordenes', 'view'), ('maquinaria_ordenes', 'edit'), ('maquinaria_ordenes', 'create'), ('maquinaria_ordenes', 'update'), ('maquinaria_ordenes', 'delete')])
def maquinaria_ordenes_page():
    ready, missing = _maquinaria_tables_status()
    ordenes = MaquinariaOrdenTrabajo.query.order_by(MaquinariaOrdenTrabajo.id.desc()).limit(200).all() if ready else []
    return render_template(
        'maquinaria_ordenes_trabajo.html',
        setup_required=not ready,
        missing_tables=missing,
        ordenes=ordenes,
    )


@app.route('/api/maquinaria/boms/by-clave/<string:clave_maquina>')
@login_required
@requires_any_permission([('maquinaria_ordenes', 'view'), ('maquinaria_boms', 'view'), ('maquinaria_ordenes', 'edit'), ('maquinaria_ordenes', 'create')])
def api_maquinaria_bom_by_clave(clave_maquina):
    ready, _missing = _maquinaria_tables_status()
    if not ready:
        return jsonify({'found': False, 'error': 'Tablas de Maquinaria no disponibles'}), 503

    bom = _maquinaria_bom_for_clave(clave_maquina)
    if not bom:
        return jsonify({'found': False, 'clave_maquina': (clave_maquina or '').strip().upper(), 'bom': None, 'componentes': []})

    componentes = [c.to_dict() for c in sorted((bom.componentes or []), key=lambda x: x.id)]
    return jsonify({'found': True, 'clave_maquina': bom.clave_maquina, 'bom': bom.to_dict(), 'componentes': componentes})


@app.route('/api/maquinaria/ordenes-trabajo/<int:orden_id>')
@login_required
@requires_any_permission([('maquinaria_ordenes', 'view'), ('maquinaria_ordenes', 'edit'), ('maquinaria_ordenes', 'create')])
def api_maquinaria_orden_detalle(orden_id):
    ready, _missing = _maquinaria_tables_status()
    if not ready:
        return jsonify({'error': 'Tablas de Maquinaria no disponibles'}), 503

    _ensure_maquinaria_ordenes_extension_tables()

    orden = MaquinariaOrdenTrabajo.query.get_or_404(orden_id)
    bom_items = [i.to_dict() for i in sorted((orden.bom_items_ot or []), key=lambda x: x.id)]
    procesos = [p.to_dict() for p in sorted((orden.procesos_ot or []), key=lambda x: (x.orden or 0, x.id or 0))]
    return jsonify({'ok': True, 'orden': orden.to_dict(), 'bom_items': bom_items, 'procesos': procesos})


@app.route('/maquinaria/ordenes-trabajo/<int:orden_id>/editar')
@login_required
@requires_any_permission([('maquinaria_ordenes', 'view'), ('maquinaria_ordenes', 'edit'), ('maquinaria_ordenes', 'create'), ('maquinaria_ordenes', 'update')])
def maquinaria_orden_builder_page(orden_id):
    ready, missing = _maquinaria_tables_status()
    if not ready:
        return redirect(url_for('maquinaria_ordenes_page'))
    _ensure_maquinaria_ordenes_extension_tables()
    orden = MaquinariaOrdenTrabajo.query.get_or_404(orden_id)
    boms = MaquinariaBOM.query.filter_by(estado='activo').order_by(MaquinariaBOM.clave_maquina.asc()).all()
    return render_template('maquinaria_orden_builder.html', orden=orden, boms=boms)


@app.route('/api/maquinaria/ordenes-trabajo/<int:orden_id>/asignar-bom', methods=['POST'])
@login_required
@requires_any_permission([('maquinaria_ordenes', 'edit'), ('maquinaria_ordenes', 'create'), ('maquinaria_ordenes', 'update')])
def api_maquinaria_orden_asignar_bom(orden_id):
    ready, _missing = _maquinaria_tables_status()
    if not ready:
        return jsonify({'ok': False, 'error': 'Tablas de Maquinaria no disponibles'}), 503
    _ensure_maquinaria_ordenes_extension_tables()
    orden = MaquinariaOrdenTrabajo.query.get_or_404(orden_id)
    payload = request.get_json(silent=True) or {}
    bom_id = payload.get('bom_id')
    try:
        bom_id = int(bom_id)
    except Exception:
        return jsonify({'ok': False, 'error': 'bom_id invalido'}), 400
    bom = MaquinariaBOM.query.get(bom_id)
    if not bom:
        return jsonify({'ok': False, 'error': 'BOM no encontrado'}), 404
    try:
        # Reemplaza el snapshot de piezas con el BOM seleccionado (se pierden asignaciones previas)
        MaquinariaOrdenBOMItem.query.filter_by(orden_trabajo_id=orden.id).delete()
        for comp in sorted((bom.componentes or []), key=lambda x: x.id):
            db.session.add(MaquinariaOrdenBOMItem(
                orden_trabajo_id=orden.id,
                bom_id=bom.id,
                proceso_id=None,
                codigo_componente=comp.codigo_componente,
                nombre_componente=comp.nombre_componente,
                cantidad=comp.cantidad,
                unidad=comp.unidad,
                proceso_base=comp.proceso_base,
            ))
        orden.bom_id = bom.id
        orden.updated_at = datetime.utcnow()
        # Copiar procesos plantilla del BOM si la OT aun no tiene procesos
        tiene_procesos = MaquinariaOrdenProceso.query.filter_by(orden_trabajo_id=orden.id).count() > 0
        if not tiene_procesos:
            _ensure_maquinaria_bom_procesos_table()
            plantilla = MaquinariaBOMProceso.query.filter_by(bom_id=bom.id).order_by(
                MaquinariaBOMProceso.orden.asc(), MaquinariaBOMProceso.id.asc()
            ).all()
            for pp in plantilla:
                db.session.add(MaquinariaOrdenProceso(
                    orden_trabajo_id=orden.id,
                    orden=pp.orden,
                    nombre=pp.nombre,
                    centro_trabajo=pp.centro_trabajo,
                    operacion=pp.operacion,
                    t_e=pp.t_e,
                    t_tct=pp.t_tct,
                    t_tco=pp.t_tco,
                    t_to=pp.t_to,
                    notas=pp.notas,
                ))
        db.session.commit()
        return jsonify({'ok': True, 'bom': bom.to_dict(), 'procesos_copiados': not tiene_procesos})
    except Exception as exc:
        db.session.rollback()
        logger.error(f'[OT] Error al asignar BOM: {exc}', exc_info=True)
        return jsonify({'ok': False, 'error': str(exc)}), 500


@app.route('/api/maquinaria/ordenes-trabajo/<int:orden_id>/cargar-procesos-bom', methods=['POST'])
@login_required
@requires_any_permission([('maquinaria_ordenes', 'edit'), ('maquinaria_ordenes', 'create'), ('maquinaria_ordenes', 'update')])
def api_maquinaria_orden_cargar_procesos_bom(orden_id):
    """Copia los procesos plantilla del BOM asignado a la OT."""
    ready, _missing = _maquinaria_tables_status()
    if not ready:
        return jsonify({'ok': False, 'error': 'Tablas de Maquinaria no disponibles'}), 503
    _ensure_maquinaria_ordenes_extension_tables()
    _ensure_maquinaria_bom_procesos_table()
    orden = MaquinariaOrdenTrabajo.query.get_or_404(orden_id)
    if not orden.bom_id:
        return jsonify({'ok': False, 'error': 'Primero asigna un BOM a la orden'}), 400
    payload = request.get_json(silent=True) or {}
    reemplazar = bool(payload.get('reemplazar'))
    try:
        existentes = MaquinariaOrdenProceso.query.filter_by(orden_trabajo_id=orden.id).count()
        if existentes and not reemplazar:
            return jsonify({'ok': False, 'error': 'La OT ya tiene procesos. Envia reemplazar:true para sustituirlos.'}), 400
        if reemplazar and existentes:
            MaquinariaOrdenBOMItem.query.filter_by(orden_trabajo_id=orden.id).update({'proceso_id': None})
            MaquinariaOrdenProceso.query.filter_by(orden_trabajo_id=orden.id).delete()
        plantilla = MaquinariaBOMProceso.query.filter_by(bom_id=orden.bom_id).order_by(
            MaquinariaBOMProceso.orden.asc(), MaquinariaBOMProceso.id.asc()
        ).all()
        if not plantilla:
            return jsonify({'ok': False, 'error': 'El BOM no tiene procesos plantilla cargados'}), 400
        for pp in plantilla:
            db.session.add(MaquinariaOrdenProceso(
                orden_trabajo_id=orden.id,
                orden=pp.orden,
                nombre=pp.nombre,
                centro_trabajo=pp.centro_trabajo,
                operacion=pp.operacion,
                t_e=pp.t_e,
                t_tct=pp.t_tct,
                t_tco=pp.t_tco,
                t_to=pp.t_to,
                notas=pp.notas,
            ))
        orden.updated_at = datetime.utcnow()
        db.session.commit()
        return jsonify({'ok': True, 'procesos_copiados': len(plantilla)})
    except Exception as exc:
        db.session.rollback()
        logger.error(f'[OT] Error al cargar procesos BOM: {exc}', exc_info=True)
        return jsonify({'ok': False, 'error': str(exc)}), 500


@app.route('/api/maquinaria/ordenes-trabajo/<int:orden_id>/procesos', methods=['POST'])
@login_required
@requires_any_permission([('maquinaria_ordenes', 'edit'), ('maquinaria_ordenes', 'create'), ('maquinaria_ordenes', 'update')])
def api_maquinaria_orden_proceso_add(orden_id):
    ready, _missing = _maquinaria_tables_status()
    if not ready:
        return jsonify({'ok': False, 'error': 'Tablas de Maquinaria no disponibles'}), 503
    _ensure_maquinaria_ordenes_extension_tables()
    orden = MaquinariaOrdenTrabajo.query.get_or_404(orden_id)
    payload = request.get_json(silent=True) or {}
    nombre = (payload.get('nombre') or '').strip()
    if not nombre:
        return jsonify({'ok': False, 'error': 'El nombre del proceso es obligatorio'}), 400
    try:
        max_orden = db.session.query(func.max(MaquinariaOrdenProceso.orden)).filter_by(orden_trabajo_id=orden.id).scalar() or 0
        proc = MaquinariaOrdenProceso(
            orden_trabajo_id=orden.id,
            orden=int(max_orden) + 1,
            nombre=nombre,
            centro_trabajo=(payload.get('centro_trabajo') or '').strip() or None,
            operacion=(payload.get('operacion') or '').strip() or None,
        )
        db.session.add(proc)
        db.session.commit()
        return jsonify({'ok': True, 'proceso': proc.to_dict()})
    except Exception as exc:
        db.session.rollback()
        logger.error(f'[OT] Error al agregar proceso: {exc}', exc_info=True)
        return jsonify({'ok': False, 'error': str(exc)}), 500


@app.route('/api/maquinaria/orden-procesos/<int:proceso_id>/update', methods=['POST'])
@login_required
@requires_any_permission([('maquinaria_ordenes', 'edit'), ('maquinaria_ordenes', 'update')])
def api_maquinaria_orden_proceso_update(proceso_id):
    ready, _missing = _maquinaria_tables_status()
    if not ready:
        return jsonify({'ok': False, 'error': 'Tablas de Maquinaria no disponibles'}), 503
    proc = MaquinariaOrdenProceso.query.get_or_404(proceso_id)
    payload = request.get_json(silent=True) or {}
    try:
        if 'nombre' in payload:
            nombre = (payload.get('nombre') or '').strip()
            if nombre:
                proc.nombre = nombre
        if 'centro_trabajo' in payload:
            proc.centro_trabajo = (payload.get('centro_trabajo') or '').strip() or None
        if 'operacion' in payload:
            proc.operacion = (payload.get('operacion') or '').strip() or None
        proc.updated_at = datetime.utcnow()
        db.session.commit()
        return jsonify({'ok': True, 'proceso': proc.to_dict()})
    except Exception as exc:
        db.session.rollback()
        return jsonify({'ok': False, 'error': str(exc)}), 500


@app.route('/api/maquinaria/orden-procesos/<int:proceso_id>/delete', methods=['POST'])
@login_required
@requires_any_permission([('maquinaria_ordenes', 'delete'), ('maquinaria_ordenes', 'edit'), ('maquinaria_ordenes', 'update')])
def api_maquinaria_orden_proceso_delete(proceso_id):
    ready, _missing = _maquinaria_tables_status()
    if not ready:
        return jsonify({'ok': False, 'error': 'Tablas de Maquinaria no disponibles'}), 503
    proc = MaquinariaOrdenProceso.query.get_or_404(proceso_id)
    try:
        # Las piezas asignadas a este proceso vuelven a "sin asignar"
        MaquinariaOrdenBOMItem.query.filter_by(proceso_id=proc.id).update({'proceso_id': None})
        db.session.delete(proc)
        db.session.commit()
        return jsonify({'ok': True})
    except Exception as exc:
        db.session.rollback()
        return jsonify({'ok': False, 'error': str(exc)}), 500


@app.route('/api/maquinaria/orden-bom-items/<int:item_id>/asignar-proceso', methods=['POST'])
@login_required
@requires_any_permission([('maquinaria_ordenes', 'edit'), ('maquinaria_ordenes', 'update')])
def api_maquinaria_orden_item_asignar_proceso(item_id):
    ready, _missing = _maquinaria_tables_status()
    if not ready:
        return jsonify({'ok': False, 'error': 'Tablas de Maquinaria no disponibles'}), 503
    item = MaquinariaOrdenBOMItem.query.get_or_404(item_id)
    payload = request.get_json(silent=True) or {}
    proceso_id = payload.get('proceso_id')
    try:
        if proceso_id in (None, '', 0, '0'):
            item.proceso_id = None
        else:
            proceso_id = int(proceso_id)
            proc = MaquinariaOrdenProceso.query.get(proceso_id)
            if not proc or proc.orden_trabajo_id != item.orden_trabajo_id:
                return jsonify({'ok': False, 'error': 'Proceso invalido'}), 400
            item.proceso_id = proc.id
        db.session.commit()
        return jsonify({'ok': True, 'item': item.to_dict()})
    except Exception as exc:
        db.session.rollback()
        return jsonify({'ok': False, 'error': str(exc)}), 500


@app.route('/maquinaria/ordenes-trabajo/<int:orden_id>/imprimir')
@login_required
@requires_any_permission([('maquinaria_ordenes', 'view'), ('maquinaria_ordenes', 'edit'), ('maquinaria_ordenes', 'create')])
def maquinaria_ordenes_print(orden_id):
    ready, _missing = _maquinaria_tables_status()
    if not ready:
        return redirect(url_for('maquinaria_ordenes_page'))

    _ensure_maquinaria_ordenes_extension_tables()

    orden = MaquinariaOrdenTrabajo.query.get_or_404(orden_id)
    bom_items = sorted((orden.bom_items_ot or []), key=lambda x: x.id)
    procesos = sorted((orden.procesos_ot or []), key=lambda x: (x.orden or 0, x.id or 0))
    return render_template(
        'maquinaria_ordenes_trabajo_print.html',
        orden=orden,
        bom_items=bom_items,
        procesos=procesos,
    )


@app.route('/api/maquinaria/ordenes-trabajo/<int:orden_id>/delete', methods=['POST'])
@login_required
@requires_any_permission([('maquinaria_ordenes', 'delete'), ('maquinaria_ordenes', 'edit'), ('maquinaria_ordenes', 'update')])
def api_maquinaria_orden_delete(orden_id):
    ready, _missing = _maquinaria_tables_status()
    if not ready:
        return jsonify({'ok': False, 'message': 'Tablas de Maquinaria no disponibles'}), 503

    _ensure_maquinaria_ordenes_extension_tables()

    orden = MaquinariaOrdenTrabajo.query.get_or_404(orden_id)
    folio = orden.folio_ot or f'OT #{orden.id}'

    db.session.delete(orden)
    db.session.commit()
    return jsonify({'ok': True, 'message': f'Orden {folio} eliminada'})


@app.route('/maquinaria/ordenes-trabajo/create', methods=['POST'])
@login_required
@requires_any_permission([('maquinaria_ordenes', 'create'), ('maquinaria_ordenes', 'edit'), ('maquinaria_ordenes', 'update')])
def maquinaria_ordenes_create():
    ready, _missing = _maquinaria_tables_status()
    if not ready:
        return redirect(url_for('maquinaria_ordenes_page'))

    _ensure_maquinaria_ordenes_extension_tables()

    def _new_ot_folio(suffix_idx=None):
        base = datetime.utcnow().strftime('%Y%m%d-%H%M%S')
        raw = f"MOT-{base}" if suffix_idx is None else f"MOT-{base}-{suffix_idx:02d}"
        folio = raw
        bump = 1
        while MaquinariaOrdenTrabajo.query.filter_by(folio_ot=folio).first() is not None:
            folio = f"{raw}-{bump}"
            bump += 1
        return folio

    fecha_objetivo = _parse_datetime(request.form.get('fecha_objetivo'))
    selected_lines_raw = _clean_nullable_text(request.form.get('selected_line_numbers', ''))
    selected_line_numbers = set()
    if selected_lines_raw:
        for token in selected_lines_raw.split(','):
            token = (token or '').strip()
            if not token:
                continue
            try:
                selected_line_numbers.add(int(token))
            except Exception:
                continue
    if not fecha_objetivo:
        fecha_objetivo = datetime.utcnow()

    procesos_json_raw = _clean_nullable_text(request.form.get('procesos_json', ''))
    procesos_payload = []
    if procesos_json_raw:
        try:
            parsed = json.loads(procesos_json_raw)
            if isinstance(parsed, list):
                procesos_payload = parsed
        except Exception:
            procesos_payload = []

    pedido_id = request.form.get('pedido_id', type=int)
    lineas_a_crear = []
    pedido_ref_id = pedido_id
    if pedido_id:
        pedido_base = MaquinariaPedido.query.get(pedido_id)
        if pedido_base:
            if pedido_base.contpaq_document_id:
                contpaq_rows = MaquinariaContpaqPedidoDetalle.query.filter_by(
                    document_id=pedido_base.contpaq_document_id
                ).order_by(MaquinariaContpaqPedidoDetalle.line_number.asc()).all()
                if contpaq_rows:
                    for row in contpaq_rows:
                        if selected_line_numbers and int(row.line_number or 0) not in selected_line_numbers:
                            continue
                        clave = _clean_nullable_text(row.product_key or '')
                        if not clave:
                            continue
                        lineas_a_crear.append({
                            'pedido_id': pedido_base.id,
                            'clave': clave,
                            'cantidad': max(1, int(row.quantity or 1)),
                            'descripcion': _clean_nullable_text(row.description or ''),
                        })
                else:
                    lineas = MaquinariaPedido.query.filter_by(contpaq_document_id=pedido_base.contpaq_document_id).order_by(MaquinariaPedido.id.asc()).all()
                    for p in lineas:
                        if (p.clave_maquina or '').strip():
                            lineas_a_crear.append({
                                'pedido_id': p.id,
                                'clave': p.clave_maquina,
                                'cantidad': max(1, int(p.cantidad or 1)),
                                'descripcion': p.descripcion_maquina or '',
                            })
            else:
                lineas = [pedido_base]
                for p in lineas:
                    if (p.clave_maquina or '').strip():
                        lineas_a_crear.append({
                            'pedido_id': p.id,
                            'clave': p.clave_maquina,
                            'cantidad': max(1, int(p.cantidad or 1)),
                            'descripcion': p.descripcion_maquina or '',
                        })

    if not lineas_a_crear:
        clave = _clean_nullable_text(request.form.get('clave_maquina', ''))
        if not clave:
            return redirect(url_for('maquinaria_ordenes_page'))
        lineas_a_crear.append({
            'pedido_id': pedido_id,
            'clave': clave,
            'cantidad': max(1, request.form.get('cantidad', type=int) or 1),
            'descripcion': _clean_nullable_text(request.form.get('descripcion_maquina', '')),
        })

    notas_base = _clean_nullable_text(request.form.get('notas', ''))
    for idx, line in enumerate(lineas_a_crear, start=1):
        notas_linea = notas_base or ''
        if line.get('descripcion'):
            notas_linea = f"{notas_linea}\nDesc: {line['descripcion']}".strip()

        orden = MaquinariaOrdenTrabajo(
            folio_ot=_new_ot_folio(None if len(lineas_a_crear) == 1 else idx),
            pedido_id=line.get('pedido_id') or pedido_ref_id,
            clave_maquina=_clean_nullable_text(line.get('clave', '')),
            cantidad=max(1, int(line.get('cantidad') or 1)),
            estado=_clean_nullable_text(request.form.get('estado', 'planeacion')) or 'planeacion',
            fecha_objetivo=fecha_objetivo,
            notas=notas_linea,
            created_by=session.get('user'),
        )
        db.session.add(orden)
        db.session.flush()

        bom_match = _maquinaria_bom_for_clave(line.get('clave', ''))
        if bom_match:
            for comp in (bom_match.componentes or []):
                db.session.add(MaquinariaOrdenBOMItem(
                    orden_trabajo_id=orden.id,
                    bom_id=bom_match.id,
                    codigo_componente=_clean_nullable_text(comp.codigo_componente or ''),
                    nombre_componente=_clean_nullable_text(comp.nombre_componente or ''),
                    cantidad=float(comp.cantidad or 0),
                    unidad=_clean_nullable_text(comp.unidad or ''),
                    proceso_base=_clean_nullable_text(comp.proceso_base or ''),
                    notas=_clean_nullable_text(comp.notas or ''),
                ))

        for p_idx, proc in enumerate(procesos_payload, start=1):
            nombre = _clean_nullable_text((proc or {}).get('nombre', ''))
            if not nombre:
                continue
            db.session.add(MaquinariaOrdenProceso(
                orden_trabajo_id=orden.id,
                orden=max(1, int((proc or {}).get('orden') or p_idx)),
                nombre=nombre,
                centro_trabajo=_clean_nullable_text((proc or {}).get('centro_trabajo', '')),
                operacion=_clean_nullable_text((proc or {}).get('operacion', '')),
                t_e=_clean_nullable_text((proc or {}).get('t_e', '')),
                t_tct=_clean_nullable_text((proc or {}).get('t_tct', '')),
                t_tco=_clean_nullable_text((proc or {}).get('t_tco', '')),
                t_to=_clean_nullable_text((proc or {}).get('t_to', '')),
                notas=_clean_nullable_text((proc or {}).get('notas', '')),
                estado='pendiente',
            ))

    db.session.commit()
    return redirect(url_for('maquinaria_ordenes_page'))


@app.route('/maquinaria/boms')
@login_required
@requires_any_permission([('maquinaria_boms', 'view'), ('maquinaria_boms', 'edit'), ('maquinaria_boms', 'create'), ('maquinaria_boms', 'update'), ('maquinaria_boms', 'delete')])
def maquinaria_boms_page():
    ready, missing = _maquinaria_tables_status()
    boms_index = []
    bom_sel = None
    componentes = []
    if ready:
        all_boms = MaquinariaBOM.query.order_by(MaquinariaBOM.clave_maquina.asc()).all()
        for bom in all_boms:
            boms_index.append({
                'id': bom.id,
                'clave_maquina': bom.clave_maquina,
                'nombre_maquina': bom.nombre_maquina,
                'version': bom.version or '',
                'estado': bom.estado or 'activo',
                'componentes_count': len(bom.componentes or []),
                'procesos_count': len(bom.procesos_plantilla or []),
            })
        bom_id = request.args.get('bom_id', type=int)
        if bom_id:
            bom_sel = MaquinariaBOM.query.get(bom_id)
            if bom_sel:
                componentes = sorted(bom_sel.componentes or [], key=lambda c: c.id)
    return render_template(
        'maquinaria_boms.html',
        setup_required=not ready,
        missing_tables=missing,
        boms_index=boms_index,
        bom_sel=bom_sel,
        componentes=componentes,
    )


def _maquinaria_boms_redirect(bom_id=None):
    if bom_id:
        return redirect(url_for('maquinaria_boms_page', bom_id=int(bom_id)))
    return redirect(url_for('maquinaria_boms_page'))


@app.route('/maquinaria/boms/create', methods=['POST'])
@login_required
@requires_any_permission([('maquinaria_boms', 'create'), ('maquinaria_boms', 'edit'), ('maquinaria_boms', 'update')])
def maquinaria_boms_create():
    ready, _missing = _maquinaria_tables_status()
    if not ready:
        return redirect(url_for('maquinaria_boms_page'))

    clave_maquina = (_clean_nullable_text(request.form.get('clave_maquina', '')) or '').upper()
    nombre_maquina = _clean_nullable_text(request.form.get('nombre_maquina', ''))
    if not clave_maquina or not nombre_maquina:
        flash('Clave y nombre de maquina son obligatorios.', 'error')
        return redirect(url_for('maquinaria_boms_page'))

    existente = MaquinariaBOM.query.filter(
        func.upper(func.trim(MaquinariaBOM.clave_maquina)) == clave_maquina
    ).first()
    if existente:
        flash(f'Ya existe un BOM con la clave {clave_maquina}. Abrelo desde el buscador para editarlo.', 'error')
        return _maquinaria_boms_redirect(existente.id)

    bom = MaquinariaBOM(
        clave_maquina=clave_maquina,
        nombre_maquina=nombre_maquina,
        version=_clean_nullable_text(request.form.get('version', '')),
        notas=_clean_nullable_text(request.form.get('notas', '')),
        estado=_clean_nullable_text(request.form.get('estado', '')) or 'activo',
    )
    try:
        db.session.add(bom)
        db.session.commit()
        flash(f'BOM {clave_maquina} creado.', 'success')
    except Exception as exc:
        db.session.rollback()
        flash(f'No se pudo crear el BOM: {exc}', 'error')
        return redirect(url_for('maquinaria_boms_page'))
    return _maquinaria_boms_redirect(bom.id)


@app.route('/maquinaria/boms/<int:bom_id>/update', methods=['POST'])
@login_required
@requires_any_permission([('maquinaria_boms', 'edit'), ('maquinaria_boms', 'update')])
def maquinaria_boms_update(bom_id):
    ready, _missing = _maquinaria_tables_status()
    if not ready:
        return redirect(url_for('maquinaria_boms_page'))

    bom = MaquinariaBOM.query.get_or_404(bom_id)
    nombre_maquina = _clean_nullable_text(request.form.get('nombre_maquina', ''))
    if not nombre_maquina:
        flash('El nombre de maquina es obligatorio.', 'error')
        return _maquinaria_boms_redirect(bom.id)

    bom.nombre_maquina = nombre_maquina
    bom.version = _clean_nullable_text(request.form.get('version', ''))
    bom.notas = _clean_nullable_text(request.form.get('notas', ''))
    bom.estado = _clean_nullable_text(request.form.get('estado', '')) or bom.estado or 'activo'
    bom.updated_at = datetime.utcnow()
    try:
        db.session.commit()
        flash(f'BOM {bom.clave_maquina} actualizado.', 'success')
    except Exception as exc:
        db.session.rollback()
        flash(f'No se pudo actualizar: {exc}', 'error')
    return _maquinaria_boms_redirect(bom.id)


@app.route('/maquinaria/boms/<int:bom_id>/componentes/create', methods=['POST'])
@login_required
@requires_any_permission([('maquinaria_boms', 'create'), ('maquinaria_boms', 'edit'), ('maquinaria_boms', 'update')])
def maquinaria_boms_componentes_create(bom_id):
    ready, _missing = _maquinaria_tables_status()
    if not ready:
        return redirect(url_for('maquinaria_boms_page'))

    bom = MaquinariaBOM.query.get_or_404(bom_id)
    comp = MaquinariaBOMComponente(
        bom_id=bom.id,
        codigo_componente=_clean_nullable_text(request.form.get('codigo_componente', '')),
        nombre_componente=_clean_nullable_text(request.form.get('nombre_componente', '')),
        cantidad=max(0.01, request.form.get('cantidad', type=float) or 1.0),
        unidad=_clean_nullable_text(request.form.get('unidad', '')),
        proceso_base=_clean_nullable_text(request.form.get('proceso_base', '')),
        notas=_clean_nullable_text(request.form.get('notas', '')),
    )
    if not comp.codigo_componente or not comp.nombre_componente:
        flash('Codigo y nombre del componente son obligatorios.', 'error')
        return _maquinaria_boms_redirect(bom.id)

    db.session.add(comp)
    bom.updated_at = datetime.utcnow()
    db.session.commit()
    flash('Componente agregado.', 'success')
    return _maquinaria_boms_redirect(bom.id)


@app.route('/maquinaria/boms/<int:bom_id>/componentes/<int:comp_id>/update', methods=['POST'])
@login_required
@requires_any_permission([('maquinaria_boms', 'edit'), ('maquinaria_boms', 'update')])
def maquinaria_boms_componentes_update(bom_id, comp_id):
    ready, _missing = _maquinaria_tables_status()
    if not ready:
        return redirect(url_for('maquinaria_boms_page'))

    comp = MaquinariaBOMComponente.query.filter_by(id=comp_id, bom_id=bom_id).first_or_404()
    comp.codigo_componente = _clean_nullable_text(request.form.get('codigo_componente', '')) or comp.codigo_componente
    comp.nombre_componente = _clean_nullable_text(request.form.get('nombre_componente', '')) or comp.nombre_componente
    comp.cantidad = max(0.01, request.form.get('cantidad', type=float) or comp.cantidad or 1.0)
    comp.unidad = _clean_nullable_text(request.form.get('unidad', ''))
    comp.proceso_base = _clean_nullable_text(request.form.get('proceso_base', ''))
    comp.notas = _clean_nullable_text(request.form.get('notas', ''))
    if comp.bom:
        comp.bom.updated_at = datetime.utcnow()
    db.session.commit()
    flash('Componente actualizado.', 'success')
    return _maquinaria_boms_redirect(bom_id)


@app.route('/maquinaria/boms/<int:bom_id>/componentes/<int:comp_id>/delete', methods=['POST'])
@login_required
@requires_any_permission([('maquinaria_boms', 'delete'), ('maquinaria_boms', 'edit'), ('maquinaria_boms', 'update')])
def maquinaria_boms_componentes_delete(bom_id, comp_id):
    ready, _missing = _maquinaria_tables_status()
    if not ready:
        return redirect(url_for('maquinaria_boms_page'))

    comp = MaquinariaBOMComponente.query.filter_by(id=comp_id, bom_id=bom_id).first_or_404()
    bom = comp.bom
    db.session.delete(comp)
    if bom:
        bom.updated_at = datetime.utcnow()
    db.session.commit()
    flash('Componente eliminado.', 'success')
    return _maquinaria_boms_redirect(bom_id)


@app.route('/maquinaria/procesos-maquina')
@login_required
@requires_any_permission([('maquinaria_procesos', 'view'), ('maquinaria_procesos', 'edit'), ('maquinaria_procesos', 'create'), ('maquinaria_procesos', 'update'), ('maquinaria_procesos', 'delete'), ('maquinaria_boms', 'view'), ('maquinaria_boms', 'edit')])
def maquinaria_procesos_maquina_page():
    ready, missing = _maquinaria_tables_status()
    if ready:
        _ensure_maquinaria_bom_procesos_table()
    boms = MaquinariaBOM.query.order_by(MaquinariaBOM.clave_maquina.asc()).all() if ready else []
    bom_id = request.args.get('bom_id', type=int)
    bom_sel = MaquinariaBOM.query.get(bom_id) if ready and bom_id else None
    procesos = []
    if bom_sel:
        procesos = MaquinariaBOMProceso.query.filter_by(bom_id=bom_sel.id).order_by(
            MaquinariaBOMProceso.orden.asc(), MaquinariaBOMProceso.id.asc()
        ).all()
    return render_template(
        'maquinaria_procesos_maquina.html',
        setup_required=not ready,
        missing_tables=missing,
        boms=boms,
        bom_sel=bom_sel,
        procesos=procesos,
    )


@app.route('/maquinaria/procesos-maquina/<int:bom_id>/create', methods=['POST'])
@login_required
@requires_any_permission([('maquinaria_procesos', 'create'), ('maquinaria_procesos', 'edit'), ('maquinaria_boms', 'create'), ('maquinaria_boms', 'edit'), ('maquinaria_boms', 'update')])
def maquinaria_procesos_maquina_create(bom_id):
    ready, _missing = _maquinaria_tables_status()
    if not ready:
        return redirect(url_for('maquinaria_procesos_maquina_page'))
    _ensure_maquinaria_bom_procesos_table()
    bom = MaquinariaBOM.query.get_or_404(bom_id)
    nombre = _clean_nullable_text(request.form.get('nombre', ''))
    if not nombre:
        flash('El nombre del proceso es obligatorio.', 'error')
        return redirect(url_for('maquinaria_procesos_maquina_page', bom_id=bom.id))

    max_orden = db.session.query(func.max(MaquinariaBOMProceso.orden)).filter_by(bom_id=bom.id).scalar() or 0
    proc = MaquinariaBOMProceso(
        bom_id=bom.id,
        orden=int(max_orden) + 1,
        nombre=nombre,
        centro_trabajo=_clean_nullable_text(request.form.get('centro_trabajo', '')),
        operacion=_clean_nullable_text(request.form.get('operacion', '')),
        t_e=_clean_nullable_text(request.form.get('t_e', '')),
        t_tct=_clean_nullable_text(request.form.get('t_tct', '')),
        t_tco=_clean_nullable_text(request.form.get('t_tco', '')),
        t_to=_clean_nullable_text(request.form.get('t_to', '')),
        notas=_clean_nullable_text(request.form.get('notas', '')),
    )
    db.session.add(proc)
    bom.updated_at = datetime.utcnow()
    db.session.commit()
    flash('Proceso agregado.', 'success')
    return redirect(url_for('maquinaria_procesos_maquina_page', bom_id=bom.id))


@app.route('/maquinaria/procesos-maquina/proceso/<int:proceso_id>/update', methods=['POST'])
@login_required
@requires_any_permission([('maquinaria_procesos', 'edit'), ('maquinaria_procesos', 'update'), ('maquinaria_boms', 'edit'), ('maquinaria_boms', 'update')])
def maquinaria_procesos_maquina_update(proceso_id):
    ready, _missing = _maquinaria_tables_status()
    if not ready:
        return redirect(url_for('maquinaria_procesos_maquina_page'))
    proc = MaquinariaBOMProceso.query.get_or_404(proceso_id)
    nombre = _clean_nullable_text(request.form.get('nombre', ''))
    if nombre:
        proc.nombre = nombre
    proc.centro_trabajo = _clean_nullable_text(request.form.get('centro_trabajo', ''))
    proc.operacion = _clean_nullable_text(request.form.get('operacion', ''))
    proc.t_e = _clean_nullable_text(request.form.get('t_e', ''))
    proc.t_tct = _clean_nullable_text(request.form.get('t_tct', ''))
    proc.t_tco = _clean_nullable_text(request.form.get('t_tco', ''))
    proc.t_to = _clean_nullable_text(request.form.get('t_to', ''))
    proc.notas = _clean_nullable_text(request.form.get('notas', ''))
    proc.orden = request.form.get('orden', type=int) or proc.orden
    proc.updated_at = datetime.utcnow()
    if proc.bom:
        proc.bom.updated_at = datetime.utcnow()
    db.session.commit()
    flash('Proceso actualizado.', 'success')
    return redirect(url_for('maquinaria_procesos_maquina_page', bom_id=proc.bom_id))


@app.route('/maquinaria/procesos-maquina/proceso/<int:proceso_id>/delete', methods=['POST'])
@login_required
@requires_any_permission([('maquinaria_procesos', 'delete'), ('maquinaria_procesos', 'edit'), ('maquinaria_boms', 'delete'), ('maquinaria_boms', 'edit')])
def maquinaria_procesos_maquina_delete(proceso_id):
    ready, _missing = _maquinaria_tables_status()
    if not ready:
        return redirect(url_for('maquinaria_procesos_maquina_page'))
    proc = MaquinariaBOMProceso.query.get_or_404(proceso_id)
    bom_id = proc.bom_id
    db.session.delete(proc)
    db.session.commit()
    flash('Proceso eliminado.', 'success')
    return redirect(url_for('maquinaria_procesos_maquina_page', bom_id=bom_id))


@app.route('/api/maquinaria/boms/<int:bom_id>/procesos')
@login_required
@requires_any_permission([('maquinaria_ordenes', 'view'), ('maquinaria_procesos', 'view'), ('maquinaria_boms', 'view')])
def api_maquinaria_bom_procesos(bom_id):
    ready, _missing = _maquinaria_tables_status()
    if not ready:
        return jsonify({'ok': False, 'error': 'Tablas no disponibles'}), 503
    _ensure_maquinaria_bom_procesos_table()
    bom = MaquinariaBOM.query.get_or_404(bom_id)
    procesos = MaquinariaBOMProceso.query.filter_by(bom_id=bom.id).order_by(
        MaquinariaBOMProceso.orden.asc(), MaquinariaBOMProceso.id.asc()
    ).all()
    return jsonify({'ok': True, 'bom': bom.to_dict(), 'procesos': [p.to_dict() for p in procesos]})


@app.route('/maquinaria/claves-procesos')
@login_required
def maquinaria_claves_procesos_page():
    # Retirado del flujo de Maquinaria y Ensamble: el catalogo de produccion ya no
    # forma parte de este modulo. Se conserva la ruta solo para redirigir enlaces viejos.
    return redirect(url_for('maquinaria_pedidos_page'))


@app.route('/api/maquinaria/odoo/test')
@login_required
def api_maquinaria_odoo_test():
    """Diagnostico de conexion a Odoo (solo admin).

    Verifica credenciales, version del servidor y existencia de los modelos clave
    (sale.order / mrp.production / mrp.bom o los configurados via env).
    """
    user = get_current_user()
    if not (user and user.es_admin):
        return jsonify({'ok': False, 'error': 'Solo administradores'}), 403
    if not OdooClient.is_configured():
        return jsonify({
            'ok': False,
            'configured': False,
            'error': 'Faltan variables de entorno de Odoo (ODOO_URL, ODOO_DB, '
                     'ODOO_USERNAME y ODOO_API_KEY o ODOO_PASSWORD).',
        }), 400
    try:
        client = OdooClient.from_env()
        # Permite probar nombres de base sin editar el .env: ...?db=NOMBRE
        db_override = (request.args.get('db') or '').strip()
        if db_override:
            client.db = db_override
            client._uid = None
        info = client.test_connection()
        info['configured'] = True
        return jsonify(info)
    except OdooError as exc:
        return jsonify({'ok': False, 'configured': True, 'error': str(exc)}), 502


@app.route('/api/maquinaria/odoo/databases')
@login_required
def api_maquinaria_odoo_databases():
    """Lista las bases de datos del servidor Odoo (solo admin).

    Util cuando Odoo es auto-alojado y no se conoce el nombre exacto de la base.
    Algunos servidores lo tienen deshabilitado (list_db = False).
    """
    user = get_current_user()
    if not (user and user.es_admin):
        return jsonify({'ok': False, 'error': 'Solo administradores'}), 403
    if not OdooClient.is_configured():
        return jsonify({'ok': False, 'error': 'Odoo no configurado'}), 400
    try:
        client = OdooClient.from_env()
        dbs = client.list_databases()
        return jsonify({'ok': True, 'databases': dbs, 'detected_db': client.db})
    except OdooError as exc:
        return jsonify({
            'ok': False,
            'error': str(exc),
            'hint': 'Si el servidor tiene list_db deshabilitado, pide el nombre '
                    'exacto de la base al administrador de Odoo y ponlo en ODOO_DB.',
        }), 502


@app.route('/api/maquinaria/odoo/discover')
@login_required
def api_maquinaria_odoo_discover():
    """Devuelve campos + un registro de ejemplo de un modelo Odoo (solo admin).

    Sirve para mapear los campos reales de tu Odoo antes de programar el sync.
    Uso: /api/maquinaria/odoo/discover?model=sale.order&sample=1
    """
    user = get_current_user()
    if not (user and user.es_admin):
        return jsonify({'ok': False, 'error': 'Solo administradores'}), 403
    if not OdooClient.is_configured():
        return jsonify({'ok': False, 'error': 'Odoo no configurado'}), 400
    model = (request.args.get('model') or '').strip()
    if not model:
        return jsonify({'ok': False, 'error': 'Falta parametro model'}), 400
    try:
        sample = max(1, min(5, request.args.get('sample', type=int) or 1))
    except Exception:
        sample = 1
    try:
        client = OdooClient.from_env()
        data = client.discover(model, sample=sample)
        data['ok'] = True
        return jsonify(data)
    except OdooError as exc:
        return jsonify({'ok': False, 'error': str(exc)}), 502


ODOO_TABLES = [
    'odoo_sync_runs',
    'odoo_pedidos_venta',
    'odoo_pedidos_venta_lineas',
    'odoo_ordenes_compra',
    'odoo_ordenes_compra_lineas',
    'maquinaria_solicitudes',
    'maquinaria_solicitud_items',
]


def _ensure_odoo_tables():
    """Crea las tablas de Odoo y solicitudes si no existen (sin migracion manual)."""
    try:
        for model_cls in (OdooSyncRun, OdooPedidoVenta, OdooPedidoVentaLinea,
                          OdooOrdenCompra, OdooOrdenCompraLinea,
                          MaquinariaSolicitud, MaquinariaSolicitudItem):
            model_cls.__table__.create(bind=db.engine, checkfirst=True)
        return True
    except Exception as exc:
        logger.error(f"No se pudieron asegurar tablas de Odoo: {exc}", exc_info=True)
        try:
            db.session.rollback()
        except Exception:
            pass
        return False


def _run_odoo_sync_guarded(trigger='manual'):
    """Ejecuta el sync de Odoo evitando corridas solapadas."""
    if not _ODOO_SYNC_LOCK.acquire(blocking=False):
        return {'ok': False, 'error': 'Ya hay una sincronizacion de Odoo en curso.'}
    try:
        _ensure_odoo_tables()
        return run_odoo_sync(trigger=trigger)
    finally:
        _ODOO_SYNC_LOCK.release()


@app.route('/api/maquinaria/odoo/sync', methods=['POST'])
@login_required
def api_maquinaria_odoo_sync():
    """Fuerza una sincronizacion con Odoo (solo admin)."""
    user = get_current_user()
    if not (user and user.es_admin):
        return jsonify({'ok': False, 'error': 'Solo administradores'}), 403
    if not OdooClient.is_configured():
        return jsonify({'ok': False, 'error': 'Odoo no configurado'}), 400
    result = _run_odoo_sync_guarded(trigger='manual')
    return jsonify(result), (200 if result.get('ok') else 502)


@app.route('/api/maquinaria/odoo/sync/last')
@login_required
def api_maquinaria_odoo_sync_last():
    """Devuelve el estado de la ultima sincronizacion con Odoo."""
    user = get_current_user()
    if not (user and user.es_admin):
        return jsonify({'ok': False, 'error': 'Solo administradores'}), 403
    try:
        _ensure_odoo_tables()
        run = OdooSyncRun.query.order_by(OdooSyncRun.id.desc()).first()
        return jsonify({'ok': True, 'last_sync': run.to_dict() if run else None})
    except Exception as exc:
        return jsonify({'ok': False, 'error': str(exc)}), 500


@app.route('/maquinaria/odoo')
@login_required
def maquinaria_odoo_page():
    """Vista de pedidos de venta y ordenes de compra (OT) traidos de Odoo."""
    user = get_current_user()
    if not (user and (user.es_admin or user.has_permission('maquinaria_pedidos', 'view')
                      or user.has_permission('maquinaria_ordenes', 'view'))):
        return render_template('403.html'), 403
    _ensure_odoo_tables()
    pedidos = OdooPedidoVenta.query.order_by(
        OdooPedidoVenta.date_order.desc(), OdooPedidoVenta.id.desc()
    ).limit(100).all()
    ordenes = OdooOrdenCompra.query.order_by(
        OdooOrdenCompra.date_order.desc(), OdooOrdenCompra.id.desc()
    ).limit(100).all()
    last_run = OdooSyncRun.query.order_by(OdooSyncRun.id.desc()).first()
    solicitudes = MaquinariaSolicitud.query.order_by(MaquinariaSolicitud.id.desc()).limit(100).all()
    return render_template(
        'maquinaria_odoo.html',
        pedidos=[p.to_dict() for p in pedidos],
        ordenes=[o.to_dict() for o in ordenes],
        solicitudes=[s.to_dict() for s in solicitudes],
        last_run=last_run.to_dict() if last_run else None,
        configured=OdooClient.is_configured(),
        is_admin=bool(user and user.es_admin),
    )


def _can_use_maquinaria_solicitudes(user):
    return bool(user and (user.es_admin
                          or user.has_permission('maquinaria_pedidos', 'view')
                          or user.has_permission('maquinaria_ordenes', 'view')))


def _new_solicitud_folio():
    base = datetime.utcnow().strftime('%Y%m%d-%H%M%S')
    folio = f"SOL-{base}"
    bump = 1
    while MaquinariaSolicitud.query.filter_by(folio=folio).first() is not None:
        folio = f"SOL-{base}-{bump}"
        bump += 1
    return folio


@app.route('/api/maquinaria/odoo/pedidos/<int:odoo_id>/detalle')
@login_required
def api_maquinaria_odoo_pedido_detalle(odoo_id):
    """Detalle de un pedido de venta Odoo: solo clave, descripcion y cantidad."""
    user = get_current_user()
    if not _can_use_maquinaria_solicitudes(user):
        return jsonify({'ok': False, 'error': 'Sin permiso'}), 403
    _ensure_odoo_tables()
    pedido = OdooPedidoVenta.query.filter_by(odoo_id=odoo_id).first()
    if not pedido:
        return jsonify({'ok': False, 'error': 'Pedido no encontrado'}), 404
    items = []
    for ln in sorted(pedido.lineas, key=lambda x: (x.sequence or 0, x.id)):
        items.append({
            'clave': ln.product_key or '',
            'descripcion': ln.description or ln.product_name or '',
            'cantidad': ln.product_uom_qty or 0,
            'odoo_linea_id': ln.odoo_id,
        })
    return jsonify({'ok': True, 'pedido': pedido.to_dict(), 'items': items})


@app.route('/api/maquinaria/odoo/pedidos/<int:odoo_id>/tomar', methods=['POST'])
@login_required
def api_maquinaria_odoo_pedido_tomar(odoo_id):
    """Toma un pedido de venta Odoo y crea una solicitud en la plataforma."""
    user = get_current_user()
    if not _can_use_maquinaria_solicitudes(user):
        return jsonify({'ok': False, 'error': 'Sin permiso'}), 403
    _ensure_odoo_tables()
    pedido = OdooPedidoVenta.query.filter_by(odoo_id=odoo_id).first()
    if not pedido:
        return jsonify({'ok': False, 'error': 'Pedido no encontrado'}), 404
    try:
        solicitud = MaquinariaSolicitud(
            folio=_new_solicitud_folio(),
            tipo='odoo',
            origen_pedido_odoo_id=pedido.odoo_id,
            origen_pedido_name=pedido.name,
            cliente=pedido.partner_name,
            sucursal=pedido.sucursal,
            estado='tomado',
            created_by=session.get('user'),
        )
        db.session.add(solicitud)
        db.session.flush()
        for ln in sorted(pedido.lineas, key=lambda x: (x.sequence or 0, x.id)):
            if not (ln.product_key or ln.description or ln.product_name):
                continue
            db.session.add(MaquinariaSolicitudItem(
                solicitud_id=solicitud.id,
                clave=(ln.product_key or '')[:120],
                descripcion=ln.description or ln.product_name or '',
                cantidad=float(ln.product_uom_qty or 0) or 1,
                origen_linea_odoo_id=ln.odoo_id,
            ))
        db.session.commit()
        return jsonify({'ok': True, 'solicitud': solicitud.to_dict(include_items=True)})
    except Exception as exc:
        db.session.rollback()
        logger.error(f'[SOLICITUD] Error al tomar pedido: {exc}', exc_info=True)
        return jsonify({'ok': False, 'error': str(exc)}), 500


@app.route('/api/maquinaria/odoo/ordenes-compra')
@login_required
def api_maquinaria_odoo_ordenes_compra():
    """Lista ordenes de compra de Odoo para la ventana de relacion (con busqueda)."""
    user = get_current_user()
    if not _can_use_maquinaria_solicitudes(user):
        return jsonify({'ok': False, 'error': 'Sin permiso'}), 403
    _ensure_odoo_tables()
    q = (request.args.get('q') or '').strip()
    query = OdooOrdenCompra.query
    if q:
        like = f"%{q}%"
        query = query.filter(db.or_(
            OdooOrdenCompra.name.ilike(like),
            OdooOrdenCompra.partner_name.ilike(like),
            OdooOrdenCompra.origin.ilike(like),
        ))
    ordenes = query.order_by(OdooOrdenCompra.date_order.desc(), OdooOrdenCompra.id.desc()).limit(80).all()
    return jsonify({'ok': True, 'ordenes': [o.to_dict() for o in ordenes]})


@app.route('/api/maquinaria/odoo/ordenes-compra/<int:odoo_id>/detalle')
@login_required
def api_maquinaria_odoo_orden_compra_detalle(odoo_id):
    """Detalle especifico de una orden de compra Odoo."""
    user = get_current_user()
    if not _can_use_maquinaria_solicitudes(user):
        return jsonify({'ok': False, 'error': 'Sin permiso'}), 403
    _ensure_odoo_tables()
    orden = OdooOrdenCompra.query.filter_by(odoo_id=odoo_id).first()
    if not orden:
        return jsonify({'ok': False, 'error': 'Orden de compra no encontrada'}), 404
    return jsonify({'ok': True, 'orden': orden.to_dict(include_lines=True)})


@app.route('/api/maquinaria/solicitudes', methods=['GET'])
@login_required
def api_maquinaria_solicitudes_list():
    user = get_current_user()
    if not _can_use_maquinaria_solicitudes(user):
        return jsonify({'ok': False, 'error': 'Sin permiso'}), 403
    _ensure_odoo_tables()
    sols = MaquinariaSolicitud.query.order_by(MaquinariaSolicitud.id.desc()).limit(100).all()
    return jsonify({'ok': True, 'solicitudes': [s.to_dict() for s in sols]})


@app.route('/api/maquinaria/solicitudes/<int:solicitud_id>', methods=['GET'])
@login_required
def api_maquinaria_solicitud_detalle(solicitud_id):
    user = get_current_user()
    if not _can_use_maquinaria_solicitudes(user):
        return jsonify({'ok': False, 'error': 'Sin permiso'}), 403
    _ensure_odoo_tables()
    sol = MaquinariaSolicitud.query.get(solicitud_id)
    if not sol:
        return jsonify({'ok': False, 'error': 'Solicitud no encontrada'}), 404
    return jsonify({'ok': True, 'solicitud': sol.to_dict(include_items=True)})


@app.route('/api/maquinaria/solicitudes/stock', methods=['POST'])
@login_required
def api_maquinaria_solicitud_stock_create():
    """Crea una solicitud de pedido de stock capturada en la plataforma."""
    user = get_current_user()
    if not _can_use_maquinaria_solicitudes(user):
        return jsonify({'ok': False, 'error': 'Sin permiso'}), 403
    _ensure_odoo_tables()
    payload = request.get_json(silent=True) or {}
    items = payload.get('items') or []
    items_validos = []
    for it in items:
        clave = (str(it.get('clave') or '')).strip()
        descripcion = (str(it.get('descripcion') or '')).strip()
        try:
            cantidad = float(it.get('cantidad') or 0)
        except Exception:
            cantidad = 0
        if not clave and not descripcion:
            continue
        items_validos.append((clave[:120], descripcion, cantidad if cantidad > 0 else 1))
    if not items_validos:
        return jsonify({'ok': False, 'error': 'Agrega al menos un producto con clave o descripcion'}), 400
    try:
        solicitud = MaquinariaSolicitud(
            folio=_new_solicitud_folio(),
            tipo='stock',
            cliente=(str(payload.get('cliente') or 'STOCK')).strip()[:255] or 'STOCK',
            notas=(str(payload.get('notas') or '')).strip() or None,
            estado='tomado',
            created_by=session.get('user'),
        )
        db.session.add(solicitud)
        db.session.flush()
        for clave, descripcion, cantidad in items_validos:
            db.session.add(MaquinariaSolicitudItem(
                solicitud_id=solicitud.id,
                clave=clave,
                descripcion=descripcion,
                cantidad=cantidad,
            ))
        db.session.commit()
        return jsonify({'ok': True, 'solicitud': solicitud.to_dict(include_items=True)})
    except Exception as exc:
        db.session.rollback()
        logger.error(f'[SOLICITUD] Error al crear stock: {exc}', exc_info=True)
        return jsonify({'ok': False, 'error': str(exc)}), 500


@app.route('/api/maquinaria/solicitudes/<int:solicitud_id>/relacionar-oc', methods=['POST'])
@login_required
def api_maquinaria_solicitud_relacionar_oc(solicitud_id):
    """Relaciona una orden de compra de Odoo a la solicitud (solo referencia)."""
    user = get_current_user()
    if not _can_use_maquinaria_solicitudes(user):
        return jsonify({'ok': False, 'error': 'Sin permiso'}), 403
    _ensure_odoo_tables()
    sol = MaquinariaSolicitud.query.get(solicitud_id)
    if not sol:
        return jsonify({'ok': False, 'error': 'Solicitud no encontrada'}), 404
    payload = request.get_json(silent=True) or {}
    oc_odoo_id = payload.get('orden_compra_odoo_id')
    try:
        oc_odoo_id = int(oc_odoo_id)
    except Exception:
        return jsonify({'ok': False, 'error': 'orden_compra_odoo_id invalido'}), 400
    orden = OdooOrdenCompra.query.filter_by(odoo_id=oc_odoo_id).first()
    if not orden:
        return jsonify({'ok': False, 'error': 'Orden de compra no encontrada'}), 404
    try:
        sol.orden_compra_odoo_id = orden.odoo_id
        sol.orden_compra_name = orden.name
        sol.updated_at = datetime.utcnow()
        db.session.commit()
        return jsonify({'ok': True, 'solicitud': sol.to_dict()})
    except Exception as exc:
        db.session.rollback()
        return jsonify({'ok': False, 'error': str(exc)}), 500


@app.route('/api/maquinaria/solicitudes/<int:solicitud_id>/generar', methods=['POST'])
@login_required
def api_maquinaria_solicitud_generar(solicitud_id):
    """Genera la solicitud (la deja lista para continuar el proceso en la plataforma)."""
    user = get_current_user()
    if not _can_use_maquinaria_solicitudes(user):
        return jsonify({'ok': False, 'error': 'Sin permiso'}), 403
    _ensure_odoo_tables()
    sol = MaquinariaSolicitud.query.get(solicitud_id)
    if not sol:
        return jsonify({'ok': False, 'error': 'Solicitud no encontrada'}), 404
    if not sol.orden_compra_odoo_id:
        return jsonify({'ok': False, 'error': 'Primero relaciona una orden de compra'}), 400
    if not sol.items:
        return jsonify({'ok': False, 'error': 'La solicitud no tiene productos'}), 400
    try:
        _ensure_maquinaria_ordenes_extension_tables()

        def _new_ot_folio(idx):
            base = f"OT-{sol.folio}-{idx:02d}"
            folio = base
            bump = 1
            while MaquinariaOrdenTrabajo.query.filter_by(folio_ot=folio).first() is not None:
                folio = f"{base}-{bump}"
                bump += 1
            return folio

        creadas = 0
        # Evita duplicar OT si la solicitud ya se habia generado antes
        ya_generadas = MaquinariaOrdenTrabajo.query.filter_by(solicitud_id=sol.id).count()
        if ya_generadas == 0:
            for idx, item in enumerate(sorted(sol.items, key=lambda x: x.id), start=1):
                ot = MaquinariaOrdenTrabajo(
                    folio_ot=_new_ot_folio(idx),
                    solicitud_id=sol.id,
                    clave_maquina=(item.clave or '').strip().upper() or 'SIN-CLAVE',
                    descripcion=item.descripcion or '',
                    cantidad=max(1, int(item.cantidad or 1)),
                    estado='planeacion',
                    orden_compra_name=sol.orden_compra_name,
                    orden_compra_odoo_id=sol.orden_compra_odoo_id,
                    cliente=sol.cliente,
                    created_by=getattr(user, 'usuario', None) or getattr(user, 'nombre', None),
                )
                db.session.add(ot)
                creadas += 1

        sol.estado = 'solicitado'
        sol.updated_at = datetime.utcnow()
        db.session.commit()
        return jsonify({
            'ok': True,
            'solicitud': sol.to_dict(include_items=True),
            'ordenes_creadas': creadas,
        })
    except Exception as exc:
        db.session.rollback()
        logger.error(f'[SOLICITUD] Error al generar OT: {exc}', exc_info=True)
        return jsonify({'ok': False, 'error': str(exc)}), 500


@app.route('/api/maquinaria/solicitudes/<int:solicitud_id>/delete', methods=['POST'])
@login_required
def api_maquinaria_solicitud_delete(solicitud_id):
    user = get_current_user()
    if not _can_use_maquinaria_solicitudes(user):
        return jsonify({'ok': False, 'error': 'Sin permiso'}), 403
    _ensure_odoo_tables()
    sol = MaquinariaSolicitud.query.get(solicitud_id)
    if not sol:
        return jsonify({'ok': False, 'error': 'Solicitud no encontrada'}), 404
    try:
        db.session.delete(sol)
        db.session.commit()
        return jsonify({'ok': True})
    except Exception as exc:
        db.session.rollback()
        return jsonify({'ok': False, 'error': str(exc)}), 500


def _odoo_scheduler_loop():
    sleep(ODOO_SYNC_STARTUP_DELAY_SECONDS)
    while not _ODOO_SYNC_STOP.is_set():
        try:
            with app.app_context():
                _run_odoo_sync_guarded(trigger='scheduler')
        except Exception as exc:
            logger.error(f'[ODOO] Error en scheduler: {exc}', exc_info=True)
        total_wait = ODOO_SYNC_INTERVAL_MINUTES * 60
        waited = 0
        while waited < total_wait and not _ODOO_SYNC_STOP.is_set():
            sleep(1)
            waited += 1


def _start_odoo_scheduler_once():
    global _ODOO_SYNC_THREAD
    global _ODOO_SCHEDULER_INIT
    if _ODOO_SCHEDULER_INIT:
        return
    _ODOO_SCHEDULER_INIT = True
    if not ODOO_SYNC_ENABLED:
        logger.info('[ODOO] Scheduler deshabilitado por ODOO_SYNC_ENABLED=0')
        return
    _ODOO_SYNC_THREAD = threading.Thread(
        target=_odoo_scheduler_loop,
        name='odoo-sync-scheduler',
        daemon=True,
    )
    _ODOO_SYNC_THREAD.start()
    logger.info('[ODOO] Scheduler iniciado (cada %s min)', ODOO_SYNC_INTERVAL_MINUTES)


@app.route('/maquinaria/estaciones')
@login_required
@requires_any_permission([('maquinaria_estaciones', 'view'), ('maquinaria_estaciones', 'edit'), ('maquinaria_estaciones', 'create'), ('maquinaria_estaciones', 'update'), ('maquinaria_estaciones', 'delete')])
def maquinaria_estaciones_page():
    ready, missing = _maquinaria_tables_status()
    estaciones_ref = EstacionPlantilla.query.order_by(EstacionPlantilla.orden.asc()).all()
    ordenes = MaquinariaOrdenTrabajo.query.order_by(MaquinariaOrdenTrabajo.id.desc()).limit(180).all() if ready else []
    catalog = _mye_read_catalog()

    estaciones_catalogo = []
    for e in estaciones_ref:
        label = (e.operacion or e.centro_trabajo or '').strip()
        if not label:
            continue
        estaciones_catalogo.append(label)
    estaciones_catalogo = sorted(set(estaciones_catalogo))

    operadores_payload = sorted(
        [
            {
                'id': int(o.get('id') or 0),
                'username': _clean_nullable_text(o.get('username')) or '',
                'role': _clean_nullable_text(o.get('role')) or 'Operador',
            }
            for o in (catalog.get('operators') or [])
            if int(o.get('id') or 0) > 0 and (_clean_nullable_text(o.get('username')) or '')
        ],
        key=lambda x: (x.get('username') or '').lower()
    )

    maquinas_payload = sorted(
        [
            {
                'id': int(m.get('id') or 0),
                'clave_maquina': _clean_nullable_text(m.get('clave_maquina')) or '',
                'nombre_maquina': _clean_nullable_text(m.get('nombre_maquina')) or '',
                'icon': _clean_nullable_text(m.get('icon')) or _mye_machine_icon(m.get('clave_maquina')),
            }
            for m in (catalog.get('machines') or [])
            if int(m.get('id') or 0) > 0 and (_clean_nullable_text(m.get('clave_maquina')) or '')
        ],
        key=lambda x: (x.get('clave_maquina') or '').lower()
    )

    procesos_payload = sorted(
        [
            {
                'id': int(p.get('id') or 0),
                'nombre': _clean_nullable_text(p.get('nombre')) or '',
                'duracion_default': float(p.get('duracion_default') or 1),
                'created_at': p.get('created_at') or '',
            }
            for p in (catalog.get('processes') or [])
            if int(p.get('id') or 0) > 0 and (_clean_nullable_text(p.get('nombre')) or '')
        ],
        key=lambda x: (x.get('nombre') or '').lower()
    )

    status_catalogo = [
        {'id': 'por_iniciar', 'label': 'Por iniciar', 'color': '#94a3b8'},
        {'id': 'configuracion', 'label': 'Configurando', 'color': '#f59e0b'},
        {'id': 'en_proceso', 'label': 'En proceso', 'color': '#3b82f6'},
        {'id': 'revision', 'label': 'Revision', 'color': '#8b5cf6'},
        {'id': 'terminado', 'label': 'Terminado', 'color': '#10b981'},
        {'id': 'pausado', 'label': 'Pausado', 'color': '#ef4444'},
    ]

    ordenes_payload = []
    for o in ordenes:
        plan = _mye_parse_plan_state(o.notas)
        order_notes = _mye_strip_plan_state_block(o.notas)
        procesos_ot = sorted((o.procesos_ot or []), key=lambda x: (x.orden or 0, x.id or 0))
        default_steps = []
        for proc in procesos_ot:
            proc_name = _clean_nullable_text(proc.nombre) or ''
            if not proc_name:
                continue
            dur_txt = _clean_nullable_text(proc.t_to) or _clean_nullable_text(proc.t_tct) or _clean_nullable_text(proc.t_e) or ''
            dur_seconds = _parse_time_to_seconds(dur_txt)
            dur_hours = round(max(1, dur_seconds) / 3600.0, 4)
            default_steps.append({
                'name': proc_name,
                'duration_hours': dur_hours,
                'started_at': None,
                'elapsed_seconds': 0,
            })

        default_process_name = ', '.join([s.get('name') for s in default_steps if s.get('name')])
        default_duration_hours = round(sum(float(s.get('duration_hours') or 0) for s in default_steps), 4)

        order_status = (o.estado or '').strip().lower() or 'planeacion'
        plan_status = (plan.get('process_status') or '').strip().lower()
        if not plan_status:
            if order_status in ('produccion', 'en_proceso'):
                plan_status = 'en_proceso'
            elif order_status in ('cerrada', 'terminada', 'finalizada', 'completa'):
                plan_status = 'terminado'
            else:
                plan_status = 'por_iniciar'

        ordenes_payload.append({
            'id': o.id,
            'folio_ot': o.folio_ot,
            'clave_maquina': o.clave_maquina,
            'cantidad': int(o.cantidad or 1),
            'estado': o.estado or 'planeacion',
            'fecha_objetivo': o.fecha_objetivo.isoformat() if o.fecha_objetivo else None,
            'notas': order_notes or '',
            'default_process_name': default_process_name,
            'default_duration_hours': default_duration_hours,
            'default_steps': default_steps,
            'plan': {
                'operator_id': plan.get('operator_id'),
                'operator_username': plan.get('operator_username') or '',
                'station': plan.get('station') or '',
                'process_name': plan.get('process_name') or default_process_name,
                'process_status': plan_status,
                'start_at': plan.get('start_at') or None,
                'timer_started_at': plan.get('timer_started_at') or None,
                'elapsed_seconds': float(plan.get('elapsed_seconds') or 0),
                'duration_hours': float(plan.get('duration_hours') or default_duration_hours or 0),
                'machine_icon': plan.get('machine_icon') or _mye_machine_icon(o.clave_maquina),
                'steps': plan.get('steps') or default_steps,
                'plans': plan.get('plans') or [],
            },
        })

    resp = make_response(render_template(
        'maquinaria_estaciones.html',
        setup_required=not ready,
        missing_tables=missing,
        estaciones_ref=estaciones_ref,
        estaciones_catalogo=estaciones_catalogo,
        operadores=operadores_payload,
        maquinas=maquinas_payload,
        procesos=procesos_payload,
        status_catalogo=status_catalogo,
        ordenes=ordenes_payload,
    ))
    # Force fresh HTML for this highly interactive board to avoid stale JS/template cache.
    resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    resp.headers['Pragma'] = 'no-cache'
    resp.headers['Expires'] = '0'
    return resp


@app.route('/api/maquinaria/estaciones/plan/update', methods=['POST'])
@login_required
@requires_any_permission([('maquinaria_estaciones', 'edit'), ('maquinaria_estaciones', 'create'), ('maquinaria_estaciones', 'update')])
def api_maquinaria_estaciones_plan_update():
    ready, missing = _maquinaria_tables_status()
    if not ready:
        return jsonify({'ok': False, 'message': f'Faltan tablas: {", ".join(missing)}'}), 400

    data = request.get_json(silent=True) or {}
    ot_id = data.get('ot_id')
    try:
        ot_id = int(ot_id)
    except Exception:
        return jsonify({'ok': False, 'message': 'OT invalida'}), 422

    orden = MaquinariaOrdenTrabajo.query.get_or_404(ot_id)
    base_notes = _clean_nullable_text(data.get('notas'))
    action = (str(data.get('action') or 'save')).strip().lower()
    if action == 'clear':
        orden.notas = base_notes or _mye_strip_plan_state_block(orden.notas) or None
        orden.updated_at = datetime.utcnow()
        db.session.commit()
        return jsonify({'ok': True, 'message': 'Asignacion eliminada'})

    operator_id = data.get('operator_id')
    if operator_id in (None, ''):
        operator_id = None
    else:
        try:
            operator_id = int(operator_id)
        except Exception:
            return jsonify({'ok': False, 'message': 'Operador invalido'}), 422

    operator_username = ''
    if operator_id:
        catalog = _mye_read_catalog()
        operator_ref = next(
            (o for o in (catalog.get('operators') or []) if int(o.get('id') or 0) == int(operator_id)),
            None,
        )
        if not operator_ref:
            return jsonify({'ok': False, 'message': 'Operador no existe en catalogo'}), 422
        operator_username = _clean_nullable_text(operator_ref.get('username')) or ''

    status_value = (str(data.get('process_status') or 'por_iniciar')).strip().lower()
    allowed_status = {'por_iniciar', 'configuracion', 'en_proceso', 'revision', 'terminado', 'pausado'}
    if status_value not in allowed_status:
        status_value = 'por_iniciar'

    duration_hours = 0
    try:
        duration_hours = max(0.0, float(data.get('duration_hours') or 0))
    except Exception:
        duration_hours = 0

    start_at = (str(data.get('start_at') or '')).strip() or None
    timer_started_at = (str(data.get('timer_started_at') or '')).strip() or None
    station = _clean_nullable_text(data.get('station'))
    process_name = _clean_nullable_text(data.get('process_name'))
    machine_icon = _clean_nullable_text(data.get('machine_icon')) or _mye_machine_icon(orden.clave_maquina)
    try:
        elapsed_seconds = max(0.0, float(data.get('elapsed_seconds') or 0))
    except Exception:
        elapsed_seconds = 0.0

    steps_raw = data.get('steps')
    steps = steps_raw if isinstance(steps_raw, list) else None
    plans_raw = data.get('plans')
    plans = plans_raw if isinstance(plans_raw, list) else None

    current_state = _mye_parse_plan_state(orden.notas)
    current_state.update({
        'operator_id': operator_id,
        'operator_username': operator_username,
        'station': station,
        'process_name': process_name,
        'process_status': status_value,
        'start_at': start_at,
        'timer_started_at': timer_started_at,
        'elapsed_seconds': elapsed_seconds,
        'duration_hours': duration_hours,
        'machine_icon': machine_icon,
    })
    if steps is not None:
        current_state['steps'] = steps
    if plans is not None:
        current_state['plans'] = plans

    notes_src = base_notes if base_notes is not None else _mye_strip_plan_state_block(orden.notas)
    orden.notas = _mye_upsert_plan_state_block(notes_src, current_state)
    if status_value == 'terminado':
        orden.estado = 'cerrada'
    elif status_value == 'en_proceso':
        orden.estado = 'produccion'
    elif status_value == 'pausado':
        orden.estado = 'pausada'
    elif status_value == 'configuracion':
        orden.estado = 'configuracion'
    else:
        orden.estado = 'planeacion'

    orden.updated_at = datetime.utcnow()
    db.session.commit()
    return jsonify({'ok': True, 'message': 'Planeacion guardada'})


@app.route('/api/maquinaria/estaciones/operators/create', methods=['POST'])
@login_required
@requires_any_permission([('maquinaria_estaciones', 'create'), ('maquinaria_estaciones', 'edit'), ('maquinaria_estaciones', 'update')])
def api_maquinaria_estaciones_operators_create():
    data = request.get_json(silent=True) or {}
    username = _clean_nullable_text(data.get('username'))
    role = _clean_nullable_text(data.get('role')) or 'Operador'

    if not username:
        return jsonify({'ok': False, 'message': 'Nombre del operador es obligatorio'}), 422

    with _MYE_CATALOG_LOCK:
        catalog = _mye_read_catalog()
        operators = catalog.get('operators') or []

        duplicate = next(
            (o for o in operators if (str(o.get('username') or '').strip().lower() == username.lower())),
            None,
        )
        if duplicate:
            return jsonify({'ok': False, 'message': 'Ese operador ya existe'}), 409

        new_operator = {
            'id': _mye_next_id(operators),
            'username': username,
            'role': role,
            'created_at': datetime.utcnow().isoformat(),
        }
        operators.append(new_operator)
        catalog['operators'] = operators
        _mye_write_catalog(catalog)

    return jsonify({'ok': True, 'message': 'Operador agregado', 'operator': new_operator})


@app.route('/api/maquinaria/estaciones/machines/create', methods=['POST'])
@login_required
@requires_any_permission([('maquinaria_estaciones', 'create'), ('maquinaria_estaciones', 'edit'), ('maquinaria_estaciones', 'update')])
def api_maquinaria_estaciones_machines_create():
    data = request.get_json(silent=True) or {}
    clave_maquina = _clean_nullable_text(data.get('clave_maquina'))
    nombre_maquina = _clean_nullable_text(data.get('nombre_maquina'))
    icon = _clean_nullable_text(data.get('icon'))

    if not clave_maquina:
        return jsonify({'ok': False, 'message': 'Clave de maquina es obligatoria'}), 422

    if not icon:
        icon = _mye_machine_icon(clave_maquina)

    with _MYE_CATALOG_LOCK:
        catalog = _mye_read_catalog()
        machines = catalog.get('machines') or []

        duplicate = next(
            (m for m in machines if (str(m.get('clave_maquina') or '').strip().lower() == clave_maquina.lower())),
            None,
        )
        if duplicate:
            return jsonify({'ok': False, 'message': 'Esa maquina ya existe'}), 409

        new_machine = {
            'id': _mye_next_id(machines),
            'clave_maquina': clave_maquina,
            'nombre_maquina': nombre_maquina or clave_maquina,
            'icon': icon,
            'created_at': datetime.utcnow().isoformat(),
        }
        machines.append(new_machine)
        catalog['machines'] = machines
        _mye_write_catalog(catalog)

    return jsonify({'ok': True, 'message': 'Maquina agregada', 'machine': new_machine})


@app.route('/api/maquinaria/estaciones/processes/create', methods=['POST'])
@login_required
@requires_any_permission([('maquinaria_estaciones', 'create'), ('maquinaria_estaciones', 'edit'), ('maquinaria_estaciones', 'update')])
def api_maquinaria_estaciones_processes_create():
    data = request.get_json(silent=True) or {}
    nombre = _clean_nullable_text(data.get('nombre'))
    duracion_default = float(data.get('duracion_default') or 1)

    if not nombre:
        return jsonify({'ok': False, 'message': 'Nombre del proceso es obligatorio'}), 422

    with _MYE_CATALOG_LOCK:
        catalog = _mye_read_catalog()
        processes = catalog.get('processes') or []

        duplicate = next(
            (p for p in processes if (str(p.get('nombre') or '').strip().lower() == nombre.lower())),
            None,
        )
        if duplicate:
            return jsonify({'ok': False, 'message': 'Ese proceso ya existe'}), 409

        new_process = {
            'id': _mye_next_id(processes),
            'nombre': nombre,
            'duracion_default': max(0.5, duracion_default),
            'created_at': datetime.utcnow().isoformat(),
        }
        processes.append(new_process)
        catalog['processes'] = processes
        _mye_write_catalog(catalog)

    return jsonify({'ok': True, 'message': 'Proceso creado', 'process': new_process})


@app.route('/api/maquinaria/estaciones/operators/delete', methods=['POST'])
@login_required
@requires_any_permission([('maquinaria_estaciones', 'delete'), ('maquinaria_estaciones', 'edit'), ('maquinaria_estaciones', 'update')])
def api_maquinaria_estaciones_operators_delete():
    data = request.get_json(silent=True) or {}
    operator_id = int(data.get('id') or 0)
    if operator_id <= 0:
        return jsonify({'ok': False, 'message': 'ID de operador invalido'}), 422

    in_use = _mye_catalog_in_use('operator', operator_id)
    if in_use:
        folio = in_use[0].get('folio_ot') or f"OT #{in_use[0].get('id')}"
        return jsonify({'ok': False, 'message': f'No se puede borrar: operador asignado en {folio}', 'in_use': in_use[:10]}), 409

    with _MYE_CATALOG_LOCK:
        catalog = _mye_read_catalog()
        operators = catalog.get('operators') or []
        keep = [o for o in operators if int(o.get('id') or 0) != operator_id]
        if len(keep) == len(operators):
            return jsonify({'ok': False, 'message': 'Operador no encontrado'}), 404
        catalog['operators'] = keep
        _mye_write_catalog(catalog)

    return jsonify({'ok': True, 'message': 'Operador eliminado'})


@app.route('/api/maquinaria/estaciones/machines/delete', methods=['POST'])
@login_required
@requires_any_permission([('maquinaria_estaciones', 'delete'), ('maquinaria_estaciones', 'edit'), ('maquinaria_estaciones', 'update')])
def api_maquinaria_estaciones_machines_delete():
    data = request.get_json(silent=True) or {}
    clave_maquina = _clean_nullable_text(data.get('clave_maquina'))
    if not clave_maquina:
        return jsonify({'ok': False, 'message': 'Clave de maquina invalida'}), 422

    in_use = _mye_catalog_in_use('machine', clave_maquina)
    if in_use:
        folio = in_use[0].get('folio_ot') or f"OT #{in_use[0].get('id')}"
        return jsonify({'ok': False, 'message': f'No se puede borrar: estacion asignada en {folio}', 'in_use': in_use[:10]}), 409

    with _MYE_CATALOG_LOCK:
        catalog = _mye_read_catalog()
        machines = catalog.get('machines') or []
        keep = [m for m in machines if str(m.get('clave_maquina') or '').strip().lower() != clave_maquina.lower()]
        if len(keep) == len(machines):
            return jsonify({'ok': False, 'message': 'Estacion no encontrada'}), 404
        catalog['machines'] = keep
        _mye_write_catalog(catalog)

    return jsonify({'ok': True, 'message': 'Estacion eliminada'})


@app.route('/api/maquinaria/estaciones/processes/delete', methods=['POST'])
@login_required
@requires_any_permission([('maquinaria_estaciones', 'delete'), ('maquinaria_estaciones', 'edit'), ('maquinaria_estaciones', 'update')])
def api_maquinaria_estaciones_processes_delete():
    data = request.get_json(silent=True) or {}
    process_id = int(data.get('id') or 0)
    if process_id <= 0:
        return jsonify({'ok': False, 'message': 'ID de proceso invalido'}), 422

    with _MYE_CATALOG_LOCK:
        catalog = _mye_read_catalog()
        processes = catalog.get('processes') or []
        target = next((p for p in processes if int(p.get('id') or 0) == process_id), None)
        if not target:
            return jsonify({'ok': False, 'message': 'Proceso no encontrado'}), 404

    process_name = _clean_nullable_text(target.get('nombre'))
    in_use = _mye_catalog_in_use('process', process_name)
    if in_use:
        folio = in_use[0].get('folio_ot') or f"OT #{in_use[0].get('id')}"
        return jsonify({'ok': False, 'message': f'No se puede borrar: proceso asignado en {folio}', 'in_use': in_use[:10]}), 409

    with _MYE_CATALOG_LOCK:
        catalog = _mye_read_catalog()
        processes = catalog.get('processes') or []
        keep = [p for p in processes if int(p.get('id') or 0) != process_id]
        if len(keep) == len(processes):
            return jsonify({'ok': False, 'message': 'Proceso no encontrado'}), 404
        catalog['processes'] = keep
        _mye_write_catalog(catalog)

    return jsonify({'ok': True, 'message': 'Proceso eliminado'})


@app.route('/maquinaria/calidad')
@login_required
@requires_any_permission([('maquinaria_calidad', 'view'), ('maquinaria_calidad', 'edit'), ('maquinaria_calidad', 'create'), ('maquinaria_calidad', 'update'), ('maquinaria_calidad', 'delete')])
def maquinaria_calidad_page():
    ready, missing = _maquinaria_tables_status()
    registros = MaquinariaCalidadRegistro.query.order_by(MaquinariaCalidadRegistro.id.desc()).limit(100).all() if ready else []
    return render_template('maquinaria_calidad.html', setup_required=not ready, missing_tables=missing, registros=registros)


@app.route('/maquinaria/calidad/create', methods=['POST'])
@login_required
@requires_any_permission([('maquinaria_calidad', 'create'), ('maquinaria_calidad', 'edit'), ('maquinaria_calidad', 'update')])
def maquinaria_calidad_create():
    ready, _missing = _maquinaria_tables_status()
    if not ready:
        return redirect(url_for('maquinaria_calidad_page'))

    folio_ot = _clean_nullable_text(request.form.get('folio_ot', ''))
    if not folio_ot:
        return redirect(url_for('maquinaria_calidad_page'))

    registro = MaquinariaCalidadRegistro(
        folio_ot=folio_ot,
        funcionalidad_ok=request.form.get('funcionalidad_ok') == 'on',
        seguridad_ok=request.form.get('seguridad_ok') == 'on',
        acabado_ok=request.form.get('acabado_ok') == 'on',
        observaciones=_clean_nullable_text(request  .form.get('observaciones', '')),
        evaluado_por=session.get('user'),
    )
    db.session.add(registro)
    db.session.commit()
    return redirect(url_for('maquinaria_calidad_page'))


@app.route('/maquinaria/seriado')
@login_required
@requires_any_permission([('maquinaria_seriado', 'view'), ('maquinaria_seriado', 'edit'), ('maquinaria_seriado', 'create'), ('maquinaria_seriado', 'update'), ('maquinaria_seriado', 'delete')])
def maquinaria_seriado_page():
    ready, missing = _maquinaria_tables_status()
    series = MaquinariaSerie.query.order_by(MaquinariaSerie.id.desc()).limit(120).all() if ready else []
    pedidos = MaquinariaPedido.query.order_by(MaquinariaPedido.id.desc()).limit(120).all() if ready else []
    ordenes = MaquinariaOrdenTrabajo.query.order_by(MaquinariaOrdenTrabajo.id.desc()).limit(120).all() if ready else []
    return render_template(
        'maquinaria_seriado.html',
        setup_required=not ready,
        missing_tables=missing,
        series=series,
        pedidos=pedidos,
        ordenes=ordenes,
    )


@app.route('/maquinaria/seriado/create', methods=['POST'])
@login_required
@requires_any_permission([('maquinaria_seriado', 'create'), ('maquinaria_seriado', 'edit'), ('maquinaria_seriado', 'update')])
def maquinaria_seriado_create():
    ready, _missing = _maquinaria_tables_status()
    if not ready:
        return redirect(url_for('maquinaria_seriado_page'))

    serie_txt = _clean_nullable_text(request.form.get('serie', ''))
    clave_maquina = _clean_nullable_text(request.form.get('clave_maquina', ''))
    if not serie_txt or not clave_maquina:
        return redirect(url_for('maquinaria_seriado_page'))

    serie_obj = MaquinariaSerie(
        serie=serie_txt,
        clave_maquina=clave_maquina,
        anio=request.form.get('anio', type=int),
        pedido_id=request.form.get('pedido_id', type=int),
        orden_trabajo_id=request.form.get('orden_trabajo_id', type=int),
        estado=_clean_nullable_text(request.form.get('estado', 'ensamble')) or 'ensamble',
        notas=_clean_nullable_text(request.form.get('notas', '')),
    )
    db.session.add(serie_obj)
    db.session.commit()
    return redirect(url_for('maquinaria_seriado_page'))


@app.route('/maquinaria/almacen')
@login_required
@requires_any_permission([('maquinaria_almacen', 'view'), ('maquinaria_almacen', 'edit'), ('maquinaria_almacen', 'create'), ('maquinaria_almacen', 'update'), ('maquinaria_almacen', 'delete')])
def maquinaria_almacen_page():
    ready, missing = _maquinaria_tables_status()
    series = MaquinariaSerie.query.order_by(MaquinariaSerie.id.desc()).limit(120).all() if ready else []
    resguardos = MaquinariaAlmacenResguardo.query.order_by(MaquinariaAlmacenResguardo.id.desc()).limit(120).all() if ready else []
    return render_template('maquinaria_almacen.html', setup_required=not ready, missing_tables=missing, series=series, resguardos=resguardos)


@app.route('/maquinaria/almacen/create', methods=['POST'])
@login_required
@requires_any_permission([('maquinaria_almacen', 'create'), ('maquinaria_almacen', 'edit'), ('maquinaria_almacen', 'update')])
def maquinaria_almacen_create():
    ready, _missing = _maquinaria_tables_status()
    if not ready:
        return redirect(url_for('maquinaria_almacen_page'))

    serie_id = request.form.get('serie_id', type=int)
    ubicacion = _clean_nullable_text(request.form.get('ubicacion', ''))
    if not serie_id or not ubicacion:
        return redirect(url_for('maquinaria_almacen_page'))

    reg = MaquinariaAlmacenResguardo(
        serie_id=serie_id,
        ubicacion=ubicacion,
        estatus=_clean_nullable_text(request.form.get('estatus', 'resguardo')) or 'resguardo',
        observaciones=_clean_nullable_text(request.form.get('observaciones', '')),
    )
    db.session.add(reg)
    db.session.commit()
    return redirect(url_for('maquinaria_almacen_page'))


@app.errorhandler(404)
def not_found(error):
    return jsonify({'error': 'Recurso no encontrado'}), 404

@app.errorhandler(500)
def internal_error(error):
    return jsonify({'error': 'Error interno del servidor'}), 500

if __name__ == '__main__':
    _start_contpaq_scheduler_once()
    _start_machine_schedule_scheduler_once()
    _start_odoo_scheduler_once()
    app.run(host='0.0.0.0', port=5000, debug=False)
