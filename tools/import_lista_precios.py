
import os
import sys
import unicodedata
from decimal import Decimal, InvalidOperation
from openpyxl import load_workbook

BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), os.pardir))
if BASE_DIR not in sys.path:
    sys.path.insert(0, BASE_DIR)

from app import app
from models import db, Producto


def _norm_header(value):
    if value is None:
        return ''
    text = str(value).strip().upper()
    text = ''.join(
        c for c in unicodedata.normalize('NFKD', text)
        if not unicodedata.combining(c)
    )
    return text


def _to_decimal(value):
    if value is None:
        return None
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    text = str(value).strip().replace(',', '')
    if not text:
        return None
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def _find_header_row(ws, max_rows=5):
    for row_num, row in enumerate(ws.iter_rows(min_row=1, max_row=max_rows, values_only=True), start=1):
        normalized = {_norm_header(v) for v in row if v is not None}
        if 'CLAVE' in normalized and 'PRECIO DE LISTA' in normalized:
            return row_num
    return None


def _merge_clasificacion(*values):
    items = []
    for value in values:
        if not value:
            continue
        for part in str(value).split(','):
            clean = part.strip().upper()
            if clean and clean not in items:
                items.append(clean)
    return ','.join(items) if items else None


def import_lista_precios(path, sheet_name='Hoja1', clasificacion_tag='VENTAS'):
    wb = load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Hoja no encontrada: {sheet_name}")

    ws = wb[sheet_name]
    header_row_num = _find_header_row(ws)
    if not header_row_num:
        raise ValueError('No se encontro encabezado con CLAVE y PRECIO DE LISTA')
    header_row = list(ws.iter_rows(min_row=header_row_num, max_row=header_row_num, values_only=True))[0]

    header_map = {}
    for idx, value in enumerate(header_row):
        key = _norm_header(value)
        if key and key not in header_map:
            header_map[key] = idx

    clave_idx = header_map.get('CLAVE')
    precio_idx = header_map.get('PRECIO DE LISTA')
    unidad_idx = header_map.get('UNIDAD')
    linea_idx = header_map.get('LINEA')
    clasif_idx = header_map.get('CLASIFICACION')
    clasif_dep_idx = header_map.get('CLASIFICACION POR DEPARTAMENTO')

    if clave_idx is None or precio_idx is None:
        raise ValueError('No se encontraron columnas CLAVE y PRECIO DE LISTA')

    stats = {
        'actualizados': 0,
        'no_encontrados': 0,
        'sin_precio': 0,
        'sin_clave': 0,
    }

    with app.app_context():
        data_start = header_row_num + 1
        for row in ws.iter_rows(min_row=data_start, values_only=True):
            clave = row[clave_idx] if clave_idx < len(row) else None
            if not clave:
                stats['sin_clave'] += 1
                continue
            clave = str(clave).strip()
            if not clave:
                stats['sin_clave'] += 1
                continue

            precio = row[precio_idx] if precio_idx < len(row) else None
            precio_dec = _to_decimal(precio)
            if precio_dec is None:
                stats['sin_precio'] += 1
                continue

            producto = Producto.query.filter_by(clave=clave).first()
            if not producto:
                stats['no_encontrados'] += 1
                continue

            producto.precio = float(precio_dec)
            producto.divisa_venta = 'MXN'

            if unidad_idx is not None and unidad_idx < len(row):
                producto.unidad = str(row[unidad_idx]).strip() if row[unidad_idx] else producto.unidad
            if linea_idx is not None and linea_idx < len(row):
                producto.linea = str(row[linea_idx]).strip() if row[linea_idx] else producto.linea
            if clasif_idx is not None and clasif_idx < len(row):
                producto.clasificacion = _merge_clasificacion(producto.clasificacion, row[clasif_idx], clasificacion_tag)
            else:
                producto.clasificacion = _merge_clasificacion(producto.clasificacion, clasificacion_tag)
            if clasif_dep_idx is not None and clasif_dep_idx < len(row):
                producto.clasificacion_departamento = str(row[clasif_dep_idx]).strip() if row[clasif_dep_idx] else producto.clasificacion_departamento
            stats['actualizados'] += 1

        db.session.commit()

    return stats


def main():
    if len(sys.argv) > 1:
        path = sys.argv[1]
    else:
        path = 'LISTA DE PRECIOS.xlsx'

    sheet_name = 'Hoja1'
    if len(sys.argv) > 2:
        sheet_name = sys.argv[2]

    clasificacion_tag = 'VENTAS'
    if len(sys.argv) > 3:
        clasificacion_tag = sys.argv[3]

    stats = import_lista_precios(path, sheet_name=sheet_name, clasificacion_tag=clasificacion_tag)
    print('Import terminado:', stats)


if __name__ == '__main__':
    main()
